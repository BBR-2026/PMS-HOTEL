"""BBR Email Campaigns service.

Lets staff import a CSV/XLSX list of recipients and schedule a bulk email
send. Reuses the master BBr luxury template from ``email_service`` so all
broadcasts keep the brand identity.

Storage: MongoDB collection ``email_campaigns`` with fields:
- id (str, uuid)
- name (str, internal)
- subject (str)
- title (str, hero title shown inside template)
- body (str, content paragraphs separated by blank lines)
- offer_type (str, drives the hero image)
- cta_label (str|None)
- cta_url (str|None)
- recipients (list of {email, name?})
- scheduled_at (ISO datetime, UTC) — when to send
- status: scheduled | sending | done | cancelled | failed
- stats: {total, sent, failed}
- created_at, created_by
"""
from __future__ import annotations

import csv
import io
import logging
import re
import uuid
from datetime import datetime, timezone
from typing import Iterable, Optional

from . import email_service

log = logging.getLogger("bbr.campaigns")

_EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")


def _is_valid_email(s: Optional[str]) -> bool:
    return bool(s and _EMAIL_RE.fullmatch(s.strip()))


# ---------- File parsing ----------

def parse_recipients_csv(content: bytes) -> list[dict]:
    """Accept text with semicolon or comma delimiter, with or without header.

    Returns ``[{"email": "...", "name": "..."}, ...]`` (name optional).
    Email column is auto-detected (any cell matching the regex). Name column
    is whichever non-email column appears in the same row.
    """
    text = content.decode("utf-8-sig", errors="ignore")
    # Sniff delimiter
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    return _extract_from_rows(rows)


def parse_recipients_xlsx(content: bytes) -> list[dict]:
    """Parse an .xlsx file using openpyxl. Reads the first sheet."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [("" if v is None else str(v)).strip() for v in row]
        if any(cells):
            rows.append(cells)
    return _extract_from_rows(rows)


def _extract_from_rows(rows: list[list[str]]) -> list[dict]:
    """Robust extraction: drops header row if no email match in row 0,
    detects email column dynamically per row, picks a name column when
    multiple non-email columns exist."""
    if not rows:
        return []
    # Drop header row if it has no emails (typical "Email","Nom" header)
    if not any(_is_valid_email(c) for c in rows[0]):
        rows = rows[1:]
    out: list[dict] = []
    seen: set[str] = set()
    for r in rows:
        email = ""
        name_parts: list[str] = []
        for cell in r:
            c = (cell or "").strip()
            if _is_valid_email(c):
                email = c.lower()
            elif c:
                name_parts.append(c)
        if not email or email in seen:
            continue
        seen.add(email)
        out.append({"email": email, "name": " ".join(name_parts).strip() or None})
    return out


# ---------- Render & send ----------

def _render_campaign_html(*, title: str, body: str, recipient_name: Optional[str],
                         cta_label: Optional[str], cta_url: Optional[str],
                         offer_type: str) -> tuple[str, str]:
    """Render the HTML + plain version of a campaign body for a recipient.

    Variables substituted in body & title: ``{prenom}`` → first name (or
    "cher invité" fallback).

    Uses the SAME ``email_service._render_template`` as transactional emails
    (booking confirmation, J-1 reminder, J+1 review) so every BBR email keeps
    the exact same visual identity. If the campaign has no CTA configured,
    we fall back to a generic "Réserver" button so the dark CTA bar is still
    rendered — guarantees identical structure for all email types.
    """
    first = email_service._first_name(recipient_name) if recipient_name else "cher invité"
    title_r = title.replace("{prenom}", first.capitalize())
    body_r = body.replace("{prenom}", first.capitalize())

    # Always provide a CTA so the dark button bar renders (identity consistency)
    final_cta_label = (cta_label or "").strip() or "Réserver"
    final_cta_url = (cta_url or "").strip() or email_service.BBR_WEBSITE_URL

    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", body_r) if p.strip()]
    hero = email_service.OFFER_HERO_IMAGES.get(offer_type, email_service.DEFAULT_HERO)
    html = email_service._render_template(
        hero_image=hero, title=title_r, paragraphs=paragraphs,
        cta_label=final_cta_label, cta_url=final_cta_url,
        preheader=title_r[:140],
    )
    plain = title_r + "\n\n" + "\n\n".join(paragraphs)
    plain += f"\n\n{final_cta_label}: {final_cta_url}"
    return html, plain


async def send_campaign_now(db, campaign_id: str) -> dict:
    """Dispatch all recipients of a campaign. Idempotent — skips recipients
    already marked as sent. Updates status & stats."""
    camp = await db.email_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not camp:
        return {"ok": False, "error": "campaign not found"}
    if camp.get("status") in ("done", "sending"):
        return {"ok": False, "error": f"campaign already {camp['status']}"}

    await db.email_campaigns.update_one(
        {"id": campaign_id},
        {"$set": {"status": "sending", "started_at": datetime.now(timezone.utc).isoformat()}},
    )

    total = len(camp.get("recipients", []))
    sent_count = 0
    failed_count = 0
    sent_emails: set[str] = set(camp.get("sent_emails", []) or [])

    for rcp in camp.get("recipients", []):
        e = (rcp.get("email") or "").lower().strip()
        if not e or e in sent_emails or not _is_valid_email(e):
            continue
        html, plain = _render_campaign_html(
            title=camp["title"], body=camp["body"],
            recipient_name=rcp.get("name"),
            cta_label=camp.get("cta_label"),
            cta_url=camp.get("cta_url"),
            offer_type=camp.get("offer_type", "pass_day"),
        )
        try:
            res = await email_service.send_email(
                db, to_email=e, to_name=rcp.get("name"),
                subject=camp["subject"], html=html, plain=plain,
                purpose=f"campaign:{campaign_id}",
            )
            if res.get("ok"):
                sent_count += 1
                sent_emails.add(e)
            else:
                failed_count += 1
        except Exception as ex:  # noqa: BLE001
            log.exception("Campaign send failed for %s: %s", e, ex)
            failed_count += 1

    final_status = "done" if failed_count == 0 else "done"  # always "done" once attempted; UI shows stats
    await db.email_campaigns.update_one(
        {"id": campaign_id},
        {"$set": {
            "status": final_status,
            "stats": {"total": total, "sent": sent_count, "failed": failed_count},
            "sent_emails": list(sent_emails),
            "finished_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    return {"ok": True, "total": total, "sent": sent_count, "failed": failed_count}


async def run_due_campaigns(db) -> int:
    """Called by APScheduler each minute. Sends any 'scheduled' campaign whose
    ``scheduled_at`` is in the past."""
    now_iso = datetime.now(timezone.utc).isoformat()
    cursor = db.email_campaigns.find(
        {"status": "scheduled", "scheduled_at": {"$lte": now_iso}},
        {"_id": 0, "id": 1},
    )
    fired = 0
    async for c in cursor:
        try:
            await send_campaign_now(db, c["id"])
            fired += 1
        except Exception as ex:  # noqa: BLE001
            log.exception("Failed to fire campaign %s: %s", c["id"], ex)
    return fired


# ---------- Helpers exposed to routes ----------

def new_campaign_doc(*, payload: dict, recipients: list[dict], created_by: str) -> dict:
    return {
        "id": str(uuid.uuid4()),
        "name": (payload.get("name") or "").strip() or "Campagne sans nom",
        "subject": (payload.get("subject") or "").strip() or "Boulay Beach Resort",
        "title": (payload.get("title") or "").strip() or "Boulay Beach Resort",
        "body": payload.get("body") or "",
        "offer_type": (payload.get("offer_type") or "pass_day").strip(),
        "cta_label": (payload.get("cta_label") or "").strip() or None,
        "cta_url": (payload.get("cta_url") or "").strip() or None,
        "recipients": recipients,
        "scheduled_at": payload.get("scheduled_at"),
        "status": "scheduled" if payload.get("scheduled_at") else "draft",
        "stats": {"total": len(recipients), "sent": 0, "failed": 0},
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": created_by,
    }
