"""Boulay Beach Resort - Reservation API (guest checkout flow)"""
import os
import io
import json
import re
import uuid
import base64
import logging
import urllib.request
from pathlib import Path
from datetime import datetime, timezone, timedelta
import secrets
from typing import List, Optional, Literal

import jwt
import qrcode
import bcrypt
import httpx
from fastapi import FastAPI, APIRouter, Depends, HTTPException, Body, Request, Query
from fastapi.responses import Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from PIL import Image, ImageDraw, ImageFont

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# Imported after load_dotenv so env vars are populated when the module reads them.
from services import twilio_service  # noqa: E402
from services import email_service  # noqa: E402
from services import campaign_service  # noqa: E402

# ----- Config -----
MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ.get("JWT_SECRET", "boulay-beach-resort-secret-key-change-me")
JWT_ALG = "HS256"
JWT_EXPIRE_HOURS = 72

# ============== FineoPay configuration ==============
FINEO_BUSINESS_CODE = os.environ.get("FINEO_BUSINESS_CODE", "")
FINEO_API_KEY = os.environ.get("FINEO_API_KEY", "")
FINEO_BASE_URL = os.environ.get("FINEO_BASE_URL", "https://dev.fineopay.com/api/v1/business/dev/").rstrip("/") + "/"
FINEO_CALLBACK_SECRET = os.environ.get("FINEO_CALLBACK_SECRET", "")
FINEO_PUBLIC_BASE_URL = os.environ.get("FINEO_PUBLIC_BASE_URL", "").rstrip("/")
FINEO_ENABLED = bool(FINEO_BUSINESS_CODE and FINEO_API_KEY and FINEO_PUBLIC_BASE_URL)

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

OFFERS = {
    "pass_day": {
        "id": "pass_day",
        "name_fr": "Day Pass",
        "name_en": "Day Pass",
        "schedule_fr": "Du Lundi au Vendredi",
        "schedule_en": "Monday to Friday",
        "tagline_fr": "Offrez-vous une parenthèse d'évasion du lundi au vendredi avec le Day Pass BBr, une expérience alliant détente, confort et plaisir dans un cadre exclusif. Idéal pour se relaxer au soleil, profiter de la piscine et savourer une ambiance paisible en bord de lagune.\n\nL'expérience inclut :\n• Traversée lagunaire aller-retour\n• Sélection de tapas\n• Cocktail signature\n\nBBr Day Pass, votre escapade en toute simplicité.",
        "tagline_en": "Treat yourself to an escape from Monday to Friday with the BBr Day Pass, an experience blending relaxation, comfort and pleasure in an exclusive setting. Perfect to unwind in the sun, enjoy the pool and savor a peaceful lagoon atmosphere.\n\nThe experience includes:\n• Round-trip lagoon crossing\n• Tapas selection\n• Signature cocktail\n\nBBr Day Pass, your escape made simple.",
        "price_adult": 50000,
        "price_child": 25000,
        "max_capacity": 250,
    },
    "sunset": {
        "id": "sunset",
        "name_fr": "The Sunset experience",
        "name_en": "The Sunset Experience",
        "schedule_fr": "Tous les samedis · 12h — 17h",
        "schedule_en": "Every Saturday · 12pm — 5pm",
        "tagline_fr": "Chaque samedi, vivez l'expérience iconique du Sunset à BBr, un moment où le temps ralentit et où l'énergie monte doucement avec le coucher du soleil. Entre ciel doré, ambiance musicale envoûtante et cadre élégant en bord de lagune, laissez-vous porter par une atmosphère chic et conviviale.\n\nL'expérience inclut :\n• Traversée lagunaire aller-retour\n• Welcome drink à l'arrivée\n• Repas complet\n• Accès piscine\n\nThe Sunset experience, bien plus qu'un moment, une signature.",
        "tagline_en": "Every Saturday, live the iconic Sunset experience at BBr, a moment when time slows down and energy gently rises with the setting sun. Between golden skies, captivating music and an elegant lagoon-side setting, let yourself be carried away by a chic and convivial atmosphere.\n\nThe experience includes:\n• Round-trip lagoon crossing\n• Welcome drink on arrival\n• Full meal\n• Pool access\n\nThe Sunset experience, much more than a moment, a signature.",
        "price_adult": 60000,
        "price_child": 30000,
        "max_capacity": 250,
    },
    "brunch": {
        "id": "brunch",
        "name_fr": "B Brunch",
        "name_en": "B Brunch",
        "schedule_fr": "Tous les dimanches · 12h — 16h",
        "schedule_en": "Every Sunday · 12pm — 4pm",
        "tagline_fr": "Vivez l'escapade dominicale ultime au Boulay Beach Resort.\n\nChaque dimanche, BBr vous invite à une expérience brunch premium sur les rives de l'Île Boulay. Savourez un menu soigneusement élaboré, dégustez du champagne et laissez-vous porter par une sélection musicale curatée.\n\nLife Is Here Every Sunday.",
        "tagline_en": "Experience the ultimate Sunday escape at Boulay Beach Resort.\n\nEvery Sunday, BBr invites you to a premium brunch experience on the shores of Île Boulay. Savor a carefully crafted menu, enjoy champagne and let yourself be carried away by a curated musical selection.\n\nLife Is Here Every Sunday.",
        "price_adult": 60000,
        "price_child": 30000,
        "max_capacity": 250,
    },
    "le_kaai": {
        "id": "le_kaai",
        "name_fr": "Le Kaai",
        "name_en": "Le Kaai",
        "schedule_fr": "Tous les jours · 11h — 22h30",
        "schedule_en": "Every day · 11am — 10:30pm",
        "tagline_fr": "Le KAAÏ est le nouveau restaurant du BBr. Une table à l'ambition gastronomique affirmée, portée par des saveurs d'inspiration africaine contemporaine, dans une atmosphère élégante et chaleureuse.",
        "tagline_en": "KAAÏ is BBr's new restaurant. A dining destination with a bold gastronomic ambition, driven by contemporary African-inspired flavors, in an elegant and warm atmosphere.",
        "price_adult": 0,
        "price_child": 0,
        "max_capacity": 250,
    },
    "hebergement": {
        "id": "hebergement",
        "name_fr": "Hébergement",
        "name_en": "Accommodation",
        "schedule_fr": "Du lundi au dimanche · Séjour à l'hôtel",
        "schedule_en": "Monday to Sunday · Hotel stay",
        "tagline_fr": "Une nuit en suspens entre lagune et océan, dans nos suites signature.",
        "tagline_en": "A night suspended between lagoon and ocean, in our signature suites.",
        "price_adult": 0,
        "price_child": 0,
        "max_capacity": 60,
        "is_overnight": True,
        "room_tiers": [
            {
                "id": "superieure",
                "name_fr": "Chambre Supérieure",
                "name_en": "Superior Room",
                "price": 200000,
                "inventory": 20,
            },
            {
                "id": "suite_jardin",
                "name_fr": "Suite côté jardin",
                "name_en": "Garden-view Suite",
                "price": 420000,
                "inventory": 3,
            },
            {
                "id": "suite_lagune",
                "name_fr": "Suite côté lagune",
                "name_en": "Lagoon-view Suite",
                "price": 470000,
                "inventory": 3,
            },
        ],
    },
    "lounge": {
        "id": "lounge",
        "name_fr": "Lounge",
        "name_en": "Lounge",
        "schedule_fr": "Tous les jours · 16h – 23h · Sur demande",
        "schedule_en": "Every day · 4 PM – 11 PM · On request",
        "tagline_fr": "Espace lounge intimiste face à la lagune, pensé pour les longues escales et les retrouvailles. Ouvert chaque jour de 16h à 23h. Salon privé, ambiance feutrée et service signature à la demande.",
        "tagline_en": "Intimate lounge facing the lagoon, designed for long stopovers and reunions. Open daily from 4 PM to 11 PM. Private lounge, refined atmosphere and signature service on demand.",
        "price_adult": 0,
        "price_child": 0,
        "max_capacity": 30,
        "price_on_request": True,
        "opens_at": "16:00",
        "closes_at": "23:00",
    },
    "spa_wellness": {
        "id": "spa_wellness",
        "name_fr": "Spa & Wellness",
        "name_en": "Spa & Wellness",
        "schedule_fr": "Tous les jours · Soins 10h — 18h",
        "schedule_en": "Every day · Treatments 10am — 6pm",
        "tagline_fr": "Soins signature et rituels bien-être au bord de la lagune.",
        "tagline_en": "Signature treatments and wellness rituals by the lagoon.",
        "price_adult": 80000,
        "price_child": 0,
        "max_capacity": 12,
    },
    "seminaire": {
        "id": "seminaire",
        "name_fr": "Séminaire résidentiel",
        "name_en": "Residential Seminar",
        "schedule_fr": "Tous les jours · Sur demande",
        "schedule_en": "Every day · On request",
        "tagline_fr": "Salles équipées, vue océan, hébergement et pauses gastronomiques pour vos séminaires résidentiels.",
        "tagline_en": "Equipped rooms, ocean view, accommodation and gourmet breaks for your residential seminars.",
        "price_adult": 0,
        "price_child": 0,
        "max_capacity": 300,
        "price_on_request": True,
    },
    "journee_etude": {
        "id": "journee_etude",
        "name_fr": "Journée d'étude",
        "name_en": "Study Day",
        "schedule_fr": "Tous les jours · Sur demande",
        "schedule_en": "Every day · On request",
        "tagline_fr": "Une journée pro complète : salle équipée, pauses café, déjeuner d'affaires et accès Beach Club.",
        "tagline_en": "A complete business day: equipped room, coffee breaks, business lunch and Beach Club access.",
        "price_adult": 0,
        "price_child": 0,
        "max_capacity": 300,
        "price_on_request": True,
    },
    "team_building": {
        "id": "team_building",
        "name_fr": "Team Building",
        "name_en": "Team Building",
        "schedule_fr": "Tous les jours · Sur demande",
        "schedule_en": "Every day · On request",
        "tagline_fr": "Activités lagunaires, défis et expériences sur-mesure pour fédérer vos équipes.",
        "tagline_en": "Lagoon activities, challenges and tailor-made experiences to unite your teams.",
        "price_adult": 0,
        "price_child": 0,
        "max_capacity": 300,
        "price_on_request": True,
    },
    "dejeuner_diner_entreprise": {
        "id": "dejeuner_diner_entreprise",
        "name_fr": "Déjeuner et dîner entreprise",
        "name_en": "Corporate Lunch & Dinner",
        "schedule_fr": "Tous les jours · Sur demande",
        "schedule_en": "Every day · On request",
        "tagline_fr": "Repas d'affaires signature au Kaai ou en privatisation, dans un cadre lagune d'exception.",
        "tagline_en": "Signature business meals at Le Kaai or in private settings, in an exceptional lagoon atmosphere.",
        "price_adult": 0,
        "price_child": 0,
        "max_capacity": 300,
        "price_on_request": True,
    },
    "formule_personnalisee": {
        "id": "formule_personnalisee",
        "name_fr": "Formule personnalisée",
        "name_en": "Tailor-Made Package",
        "schedule_fr": "Tous les jours · Sur devis",
        "schedule_en": "Every day · On quote",
        "tagline_fr": "Construisons ensemble votre événement corporate sur-mesure, du brief à la mise en scène signature BBr.",
        "tagline_en": "Let's build together your bespoke corporate event, from brief to BBr signature production.",
        "price_adult": 0,
        "price_child": 0,
        "max_capacity": 300,
        "price_on_request": True,
    },
    "offres_loisirs": {
        "id": "offres_loisirs",
        "name_fr": "Offres Loisirs",
        "name_en": "Leisure Packs",
        "schedule_fr": "Tous les jours · Forfait découverte",
        "schedule_en": "Every day · Discovery pack",
        "tagline_fr": "Jet ski, paddle, kayak et plus — une journée d'activités lagunaires.",
        "tagline_en": "Jet ski, paddle, kayak and more — a day of lagoon activities.",
        "price_adult": 30000,
        "price_child": 15000,
        "max_capacity": 80,
    },
}

# ============== PHYSICAL ROOM INVENTORY (Hébergement) ==============
# Each physical room belongs to a tier and has a stable identifier that the
# staff sees (room number for "superieure", suite name for "suite_jardin" /
# "suite_lagune" — split by view: 3 garden-side + 3 lagoon-side).
HEBERGEMENT_ROOMS = [
    # Supérieures — aile A (étage 1)
    *[{"id": f"R{n}", "label": str(n), "tier": "superieure"} for n in range(1001, 1011)],
    # Supérieures — aile B (étage 2)
    *[{"id": f"R{n}", "label": str(n), "tier": "superieure"} for n in range(1011, 1021)],
    # Suites côté jardin — 3 chambres signature (vue jardin tropical)
    {"id": "SUITE_MAKENA", "label": "Makena", "tier": "suite_jardin"},
    {"id": "SUITE_MOHELI", "label": "Moheli", "tier": "suite_jardin"},
    {"id": "SUITE_KALEMA", "label": "Kalema", "tier": "suite_jardin"},
    # Suites côté lagune — 3 chambres signature (vue lagune premium)
    {"id": "SUITE_MAUPITI", "label": "Maupiti", "tier": "suite_lagune"},
    {"id": "SUITE_NZURI", "label": "N'Zuri", "tier": "suite_lagune"},
    {"id": "SUITE_MANDA", "label": "Manda", "tier": "suite_lagune"},
]
HEBERGEMENT_ROOMS_BY_ID = {r["id"]: r for r in HEBERGEMENT_ROOMS}
HEBERGEMENT_DEFAULT_CHECKIN = "14:00"   # 2pm hotel-wide check-in
HEBERGEMENT_DEFAULT_CHECKOUT = "12:00"  # 12pm hotel-wide check-out

# Pôles d'entrée — taxonomie publique
POLES = {
    "beach_club": {
        "id": "beach_club",
        "name_fr": "Beach Club",
        "name_en": "Beach Club",
        "tagline_fr": "Le club de plage signature, du lundi au dimanche.",
        "tagline_en": "The signature beach club, Monday to Sunday.",
        "offers": ["pass_day", "sunset", "brunch"],
        "sort_order": 1,
    },
    "hebergement": {
        "id": "hebergement",
        "name_fr": "Hébergement",
        "name_en": "Accommodation",
        "tagline_fr": "Suites signature et soins bien-être au cœur de la lagune.",
        "tagline_en": "Signature suites and wellness treatments at the heart of the lagoon.",
        "offers": ["hebergement", "spa_wellness", "lounge"],
        "sort_order": 2,
    },
    "corporate": {
        "id": "corporate",
        "name_fr": "Corporate",
        "name_en": "Corporate",
        "tagline_fr": "Séminaires et team buildings haut de gamme, en bord d'océan.",
        "tagline_en": "Premium seminars and team buildings, by the ocean.",
        "offers": ["seminaire", "journee_etude", "team_building", "dejeuner_diner_entreprise", "formule_personnalisee"],
        "sort_order": 3,
    },
    "activites_events": {
        "id": "activites_events",
        "name_fr": "Activités & Événements",
        "name_en": "Activities & Events",
        "tagline_fr": "Loisirs lagunaires et événements maison signés Boulay.",
        "tagline_en": "Lagoon leisure and signature in-house events by Boulay.",
        "offers": ["offres_loisirs", "events_maison"],  # 'events_maison' = special_events
        "sort_order": 4,
    },
    "le_kaai": {
        "id": "le_kaai",
        "name_fr": "Le Kaai",
        "name_en": "Le Kaai",
        "tagline_fr": "Le restaurant signature — gastronomie entre lagune et océan.",
        "tagline_en": "The signature restaurant — gastronomy between lagoon and ocean.",
        "offers": ["le_kaai"],
        "sort_order": 5,
    },
}

OFFER_TO_POLE = {offer_id: pid for pid, p in POLES.items() for offer_id in p["offers"]}
# Special events live under 'activites_events' as the 'events_maison' sub-offer
OFFER_TO_POLE["special_event"] = "activites_events"


def _pole_for_offer(offer_id: str) -> str:
    return OFFER_TO_POLE.get(offer_id, "")


OfferType = Literal[
    "pass_day", "sunset", "brunch", "le_kaai", "hebergement", "special_event",
    "spa_wellness", "seminaire", "team_building", "offres_loisirs", "lounge",
    "journee_etude", "dejeuner_diner_entreprise", "formule_personnalisee",
]
BookingStatus = Literal["pending", "confirmed", "arrived", "completed", "cancelled"]

# Weekday boat times (every 2 hours) and weekend boat times (hourly)
BOAT_TIMES_WEEKDAY = ["10H", "12H", "14H", "16H", "18H", "20H"]
BOAT_TIMES_WEEKEND = [f"{h}H" for h in range(10, 21)]

# Boat departure times available per offer
BOAT_TIMES_BY_OFFER = {
    "pass_day": BOAT_TIMES_WEEKDAY,
    "sunset": BOAT_TIMES_WEEKEND,
    "brunch": BOAT_TIMES_WEEKEND,
    "spa_wellness": ["10H", "12H", "14H", "16H", "18H"],
    "lounge": ["16H", "17H", "18H", "19H", "20H"],
    "seminaire": ["8H", "9H", "10H"],
    "team_building": ["8H", "9H", "10H"],
    "journee_etude": ["8H", "9H", "10H"],
    "dejeuner_diner_entreprise": ["10H", "12H", "14H", "16H", "18H", "20H"],
    "formule_personnalisee": ["8H", "9H", "10H", "12H", "14H", "16H", "18H", "20H"],
    "offres_loisirs": ["10H", "12H", "14H", "16H"],
    # le_kaai + hebergement are day-dependent — resolved via _boat_times_for_date()
}

# Python weekday(): Monday=0, Sunday=6
ALLOWED_WEEKDAYS_BY_OFFER = {
    "pass_day": [0, 1, 2, 3, 4],     # Monday to Friday
    "sunset": [5],                     # Saturday only
    "brunch": [6],                     # Sunday only
    "le_kaai": [0, 1, 2, 3, 4, 5, 6],  # Every day
    "hebergement": [0, 1, 2, 3, 4, 5, 6],  # Every day
    "spa_wellness": [0, 1, 2, 3, 4, 5, 6],  # Every day
    "lounge": [0, 1, 2, 3, 4, 5, 6],  # Every day
    "seminaire": [0, 1, 2, 3, 4, 5, 6],     # Every day
    "team_building": [0, 1, 2, 3, 4, 5, 6], # Every day
    "journee_etude": [0, 1, 2, 3, 4, 5, 6], # Every day
    "dejeuner_diner_entreprise": [0, 1, 2, 3, 4, 5, 6], # Every day
    "formule_personnalisee": [0, 1, 2, 3, 4, 5, 6], # Every day
    "offres_loisirs": [0, 1, 2, 3, 4, 5, 6],  # Every day
}


def _boat_times_for_date(offer_id: str, weekday: int) -> list:
    """Return valid boat times for the given offer + weekday (Python Mon=0..Sun=6)."""
    if offer_id in ("le_kaai", "hebergement"):
        return BOAT_TIMES_WEEKEND if weekday in (5, 6) else BOAT_TIMES_WEEKDAY
    return BOAT_TIMES_BY_OFFER.get(offer_id, [])


# ----- Models -----
class StaffLogin(BaseModel):
    email: EmailStr
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict


class Participant(BaseModel):
    name: str
    surname: str
    # email/phone only required on the booker (first adult) — backend enforces.
    # Children are no longer collected as participants (counted via booking.children).
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    whatsapp: Optional[str] = None  # optional WhatsApp contact (defaults to phone)
    nationality: str
    kind: Literal["adult", "child"] = "adult"


class PackageSelection(BaseModel):
    """One package picked by the customer for a given event date."""
    date: str  # YYYY-MM-DD
    package_id: str
    adults: int = Field(ge=0, le=50)
    children: int = Field(ge=0, le=50)


class BookingCreate(BaseModel):
    offer_type: OfferType
    date: str  # YYYY-MM-DD (arrival date for overnight stays)
    checkout_date: Optional[str] = None  # YYYY-MM-DD, required if offer is_overnight
    room_tier: Optional[str] = None  # required if offer has room_tiers
    rooms: int = Field(default=1, ge=1, le=20)
    adults: int = Field(ge=0, le=20)
    children: int = Field(ge=0, le=20)
    # NEW (iter-30) — distinguishes children 6–12 (billed at child rate) vs <6 (free).
    # Backward compat: if a client still sends only `children`, the create
    # endpoint maps it to `children_paid` (legacy behaviour preserved).
    children_paid: Optional[int] = Field(default=None, ge=0, le=20)
    children_free: Optional[int] = Field(default=None, ge=0, le=20)
    boat_time: str
    return_boat_time: Optional[str] = None  # required for overnight stays (departure from resort on checkout day)
    participants: List[Participant]
    special_requests: Optional[str] = ""
    # Required when offer_type == "special_event". Identifies which event the booking targets.
    special_event_id: Optional[str] = None
    # Multi-day special events: pass the full list of dates the customer
    # selected on the event detail page. When set (≥2 dates), backend bills
    # a cumulative total = Σ (adults*priceA[d] + children*priceC[d]) and
    # generates one ticket per (adult × date).
    multi_day_dates: Optional[List[str]] = None
    # Premium package add-ons (per event day). When set, each entry's
    # (adults*pkg.price_adult + children*pkg.price_child) is ADDED on top of
    # the base event price. Multiple packages per day are allowed (mix).
    package_selections: Optional[List[PackageSelection]] = None
    # Optional private boat charter — adds boat_charter_amount to the total.
    charter_boat_id: Optional[str] = None
    # Beach Club only — numbered transats / balinés the customer wants to
    # reserve for `body.date`. Validated server-side against active spaces +
    # existing bookings (uniqueness per date) inside create_booking().
    vip_space_ids: Optional[List[str]] = None
    # Optional room add-on — let the customer add a hotel night on top of any
    # offer (except `hebergement` which is a room booking already). When set
    # the requested tier price * nights * rooms is added to the total and the
    # booking persists `room_addon_*` metadata for the staff dashboards.
    room_addon_tier: Optional[str] = None  # tier id in OFFERS["hebergement"].room_tiers
    room_addon_checkin: Optional[str] = None  # YYYY-MM-DD (defaults to body.date)
    room_addon_checkout: Optional[str] = None  # YYYY-MM-DD (defaults to checkin + 1 day)
    room_addon_rooms: Optional[int] = Field(default=None, ge=1, le=10)


class PayBooking(BaseModel):
    reference_token: str
    payment_method: Optional[Literal["fineo", "card", "mobile_money", "cash", "deposit"]] = "fineo"
    # When payment_method = "deposit" (Hébergement only): % of total paid as deposit.
    deposit_pct: Optional[Literal[10, 30, 70]] = None


class EventPrivatization(BaseModel):
    name: str
    surname: str
    phone: str
    email: EmailStr
    event_type: str
    event_date: str
    guest_count: int
    message: Optional[str] = ""


class EventMatch(BaseModel):
    """A sports match displayed in the day's "Calendrier des matchs" modal.
    Optional and unique to event days that carry one (e.g., World Cup, CAN).
    """
    time: str = Field(min_length=1, max_length=10)  # "17H00", "21:00"
    team_home: str = Field(min_length=1, max_length=80)
    team_away: str = Field(min_length=1, max_length=80)
    stage: Optional[str] = Field(default="", max_length=80)  # "Match d'ouverture", "1/4 finale"
    flag_home: Optional[str] = ""  # URL or emoji
    flag_away: Optional[str] = ""


class EventPackage(BaseModel):
    """Premium pass / package offered for one specific day of an event.

    Customers buy a package as a flat forfait (regardless of head-count, up to
    `max_persons`). `stock` lets the staff cap how many copies of THIS package
    can be sold for the day (0 = unlimited).
    """
    id: str = Field(default_factory=lambda: uuid.uuid4().hex[:12])
    label: str = Field(min_length=1, max_length=120)
    description: Optional[str] = ""  # rich text shown in the "Voir le contenu" modal
    price_adult: int = Field(default=0, ge=0)  # flat forfait price (legacy field name)
    price_child: int = Field(default=0, ge=0)  # kept for schema backward compat
    max_persons: int = Field(default=2, ge=1, le=50)
    stock: int = Field(default=0, ge=0, le=10_000)  # 0 = unlimited copies/day


class ProgrammeItem(BaseModel):
    """A mini-event within a multi-day special event (one entry per day).
    `date` must fall between the event's `start_date` and `end_date` inclusive.
    """
    date: str  # YYYY-MM-DD
    title: str = Field(min_length=1, max_length=120)
    description: Optional[str] = ""
    price_adult: int = Field(default=0, ge=0)
    price_child: int = Field(default=0, ge=0)
    packages: List[EventPackage] = Field(default_factory=list)
    matches: List[EventMatch] = Field(default_factory=list)


class SpecialEventCreate(BaseModel):
    """Bookable themed event (e.g. NYE, Valentine's, Easter Brunch).
    Only one event can be `is_featured=True` at a time — see /staff/special-events/{id}/feature.

    Two flavours:
      • `event_kind="single_day"` — uses `event_dates` (legacy behaviour). Customer
        picks any of the listed dates and pays `price_adult`/`price_child`.
      • `event_kind="multi_day"` — uses `start_date`, `end_date` and `programme`.
        Each programme item is a mini-event on a specific date with its own
        title/description/price. Customer picks any programme date to book.
    """
    title: str = Field(min_length=1, max_length=120)
    subtitle: Optional[str] = ""
    description: Optional[str] = ""
    image_url: Optional[str] = ""  # http URL or "data:image/...;base64,..."
    event_dates: List[str] = Field(default_factory=list)  # YYYY-MM-DD (single_day mode)
    boat_times: List[str] = Field(default_factory=list)
    return_boat_times: List[str] = Field(default_factory=list)
    price_adult: int = Field(default=0, ge=0)
    price_child: int = Field(default=0, ge=0)
    capacity: int = Field(default=100, ge=1, le=2000)
    active_from: Optional[str] = None  # YYYY-MM-DD (visibility window start)
    active_to: Optional[str] = None  # YYYY-MM-DD
    cta_label: Optional[str] = "Réserver ma place"
    status: Literal["draft", "published", "archived"] = "draft"
    # Multi-day support
    event_kind: Literal["single_day", "multi_day"] = "single_day"
    start_date: Optional[str] = None  # YYYY-MM-DD (multi_day mode)
    end_date: Optional[str] = None    # YYYY-MM-DD (multi_day mode)
    programme: List[ProgrammeItem] = Field(default_factory=list)


class SpecialEventUpdate(BaseModel):
    title: Optional[str] = None
    subtitle: Optional[str] = None
    description: Optional[str] = None
    image_url: Optional[str] = None
    event_dates: Optional[List[str]] = None
    boat_times: Optional[List[str]] = None
    return_boat_times: Optional[List[str]] = None
    price_adult: Optional[int] = Field(default=None, ge=0)
    price_child: Optional[int] = Field(default=None, ge=0)
    capacity: Optional[int] = Field(default=None, ge=1, le=2000)
    active_from: Optional[str] = None
    active_to: Optional[str] = None
    cta_label: Optional[str] = None
    status: Optional[Literal["draft", "published", "archived"]] = None
    # Multi-day support
    event_kind: Optional[Literal["single_day", "multi_day"]] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    programme: Optional[List[ProgrammeItem]] = None


# ===== Exclusivity feature (homepage spotlight) =====
class ExclusivityFeature(BaseModel):
    """Single document that controls the "En exclusivité" hero card displayed
    above the Beach Club pôle on the public landing.
    """
    enabled: bool = False
    title: str = ""
    subtitle: Optional[str] = ""
    description: Optional[str] = ""
    image_url: Optional[str] = ""
    cta_label: Optional[str] = "Découvrir"
    # Type of resource the card points to:
    #   • special_event : link_target_id = id of a published special_event
    #   • offer         : link_target_id = offer key (pass_day, sunset, ...)
    #   • activity      : link_target_id = id of an activity catalog entry
    #   • custom        : link_url is used verbatim
    link_type: Literal["special_event", "offer", "activity", "custom"] = "special_event"
    link_target_id: Optional[str] = None
    link_url: Optional[str] = None


# ----- Helpers -----
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def verify_password(password: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(password.encode(), hashed.encode())
    except Exception:
        return False


def create_token(payload: dict) -> str:
    data = payload.copy()
    data["exp"] = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS)
    return jwt.encode(data, JWT_SECRET, algorithm=JWT_ALG)


def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")


def make_qr(payload: str, styled: bool = False) -> str:
    """Generate a QR code as a base64 PNG data URL.

    When ``styled`` is True, the QR is rendered with rounded gold modules on a
    white background and a small white square in the centre containing the
    "BBr" mark — to match the luxury ticket template used for card / mobile
    money receipts. Otherwise a plain black-and-white QR is returned.
    """
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H if styled else qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(payload)
    qr.make(fit=True)
    if styled:
        from qrcode.image.styledpil import StyledPilImage
        from qrcode.image.styles.moduledrawers.pil import RoundedModuleDrawer
        from qrcode.image.styles.colormasks import SolidFillColorMask
        gold = (140, 95, 38)  # warm brown-gold matching the ticket palette
        img = qr.make_image(
            image_factory=StyledPilImage,
            module_drawer=RoundedModuleDrawer(),
            color_mask=SolidFillColorMask(back_color=(255, 255, 255), front_color=gold),
        ).convert("RGB")
    else:
        img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


# --- Ticket image generation (printable PNG matching the brand template) ---
OFFER_HERO_URLS = {
    "pass_day": "https://customer-assets.emergentagent.com/job_reserve-bbr/artifacts/4kr4z5g1_DAY%20PASS.jpeg",
    "sunset": "https://customer-assets.emergentagent.com/job_reserve-bbr/artifacts/3g3onmkg_THE%20SUNSET.jpeg",
    "brunch": "https://customer-assets.emergentagent.com/job_reserve-bbr/artifacts/1txrnqdp_B%20BRUNCH.jpeg",
    "le_kaai": "https://customer-assets.emergentagent.com/job_reserve-bbr/artifacts/kgqk46mw_LE%20KAAI.jpeg",
    "hebergement": "https://images.unsplash.com/photo-1582719508461-905c673771fd?auto=format&fit=crop&w=1600&q=80",
}

BBR_LOGO_URL = "https://customer-assets.emergentagent.com/job_reserve-bbr/artifacts/2p8ulkeu_LOGO_BBr_VF_Plan_de_travail_1-removebg-preview.png"
_LOGO_CACHE: dict = {}
_LOGO_BYTES_CACHE: dict = {}


def _fetch_logo_bytes() -> Optional[bytes]:
    """Fetch + cache the raw PNG bytes of the BBR logo for ReportLab usage."""
    if BBR_LOGO_URL in _LOGO_BYTES_CACHE:
        return _LOGO_BYTES_CACHE[BBR_LOGO_URL]
    try:
        req = urllib.request.Request(BBR_LOGO_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = resp.read()
        _LOGO_BYTES_CACHE[BBR_LOGO_URL] = data
        return data
    except Exception as e:
        logging.warning("Failed to fetch BBr logo bytes: %s", e)
        return None


def _fetch_logo():
    """Fetch + cache the BBr logo (RGBA, transparent background)."""
    if BBR_LOGO_URL in _LOGO_CACHE:
        return _LOGO_CACHE[BBR_LOGO_URL].copy()
    try:
        req = urllib.request.Request(BBR_LOGO_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = resp.read()
        logo = Image.open(io.BytesIO(data)).convert("RGBA")
        _LOGO_CACHE[BBR_LOGO_URL] = logo
        return logo.copy()
    except Exception as e:
        logging.warning("Failed to fetch BBr logo: %s", e)
        return None


def _paste_logo(canvas, top: int, max_h: int = 110, max_w_ratio: float = 1.0):
    """Paste the BBr logo centred horizontally on ``canvas`` at vertical ``top``,
    sized so its height does not exceed ``max_h`` and width does not exceed
    ``max_w_ratio * canvas.width``. Returns the actual rendered height."""
    logo = _fetch_logo()
    if logo is None:
        return 0
    w0, h0 = logo.size
    new_h = max_h
    new_w = int(w0 * (new_h / h0))
    max_w = int(canvas.width * max_w_ratio)
    if new_w > max_w:
        new_w = max_w
        new_h = int(h0 * (new_w / w0))
    logo = logo.resize((new_w, new_h))
    canvas.paste(logo, ((canvas.width - new_w) // 2, top), logo)
    return new_h

_HERO_CACHE: dict = {}


def _fetch_hero(offer_id: str, hero_url: Optional[str] = None):
    """Fetch + cache the hero image. If ``hero_url`` is provided, it overrides
    the static OFFER_HERO_URLS lookup (used for special-event tickets with a
    staff-uploaded image). Accepts http(s) URLs and base64 ``data:`` URIs.

    Hardened against slow CDNs and very large source files (event hero images
    are frequently 5-10 MB JPEGs straight from a photographer): timeout bumped
    to 25s and the result is down-sampled to a max of 1600px on the long edge
    BEFORE the bottom-band trim so PIL doesn't OOM on a 5000×3500 source.
    """
    url = hero_url or OFFER_HERO_URLS.get(offer_id)
    if not url:
        return None
    if url in _HERO_CACHE:
        return _HERO_CACHE[url].copy()
    try:
        if url.startswith("data:"):
            try:
                head, b64 = url.split(",", 1)
            except ValueError:
                return None
            data = base64.b64decode(b64)
        else:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=25) as resp:
                data = resp.read()
        img = Image.open(io.BytesIO(data)).convert("RGB")
        # Downscale extreme sources so the per-ticket compositing stays fast
        # and uses bounded memory. 1600px on the longest edge is well above
        # the 780×440 hero target size, so detail is preserved.
        max_side = 1600
        if max(img.width, img.height) > max_side:
            scale = max_side / max(img.width, img.height)
            img = img.resize((int(img.width * scale), int(img.height * scale)))
        # Trim the bottom white/cream "footer band" that brand marketing assets
        # commonly include below the chevron decoration.
        trim = max(2, int(img.height * 0.06))
        if img.height - trim > 50:
            img = img.crop((0, 0, img.width, img.height - trim))
        _HERO_CACHE[url] = img
        return img.copy()
    except Exception as e:
        logging.warning("Failed to fetch hero (%s) %s: %s", offer_id, url, e)
        return None


def _load_font(size: int, bold: bool = False):
    candidates = (
        [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        ]
        if bold
        else [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        ]
    )
    for path in candidates:
        try:
            return ImageFont.truetype(path, size=size)
        except Exception:
            continue
    return ImageFont.load_default()


def _format_date_long(date_iso: str, lang: str = "fr") -> str:
    try:
        d = datetime.strptime(date_iso, "%Y-%m-%d").date()
        if lang == "fr":
            days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
            months = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                      "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            return f"{days[d.weekday()]} {d.day:02d} {months[d.month]} {d.year}"
        return d.strftime("%A %B %d %Y")
    except Exception:
        return date_iso


def _format_dates_list(dates_iso: List[str], lang: str = "fr") -> str:
    """Compact human-readable list of dates for the passport ticket.
    - 2 dates  → "3 août · 5 août 2026"
    - 3-5 dates same month → "3, 4, 5 août 2026"
    - 3-5 dates spanning months → "3 août · 5 sept · 7 oct 2026"
    - 6+ dates → "3 août → 7 octobre 2026 (5 dates)"
    """
    if not dates_iso:
        return ""
    try:
        ds = sorted({datetime.strptime(s, "%Y-%m-%d").date() for s in dates_iso})
    except Exception:
        return ", ".join(dates_iso)
    if len(ds) == 1:
        return _format_date_long(ds[0].isoformat(), lang)
    months_fr = ["", "janv.", "févr.", "mars", "avr.", "mai", "juin",
                 "juill.", "août", "sept.", "oct.", "nov.", "déc."]
    months_en = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    months = months_fr if lang == "fr" else months_en
    year = ds[-1].year
    if len(ds) >= 6:
        sep = " → "
        return f"{ds[0].day} {months[ds[0].month]}{sep}{ds[-1].day} {months[ds[-1].month]} {year} ({len(ds)} dates)"
    # If all in the same month/year → compact "3, 4, 5 août 2026"
    if all(d.month == ds[0].month and d.year == ds[0].year for d in ds):
        days = ", ".join(str(d.day) for d in ds)
        return f"{days} {months[ds[0].month]} {year}"
    # Mixed months → "3 août · 5 sept · 7 oct 2026"
    parts = [f"{d.day} {months[d.month]}" for d in ds]
    return " · ".join(parts) + f" {year}"




def make_ticket_image(
    offer_id: str, offer_name: str, date_iso: str, boat_time: str,
    owner_name: str, qr_payload: str, ref_code: str,
    lang: str = "fr", hero_url: Optional[str] = None,
    dates_list: Optional[List[str]] = None,
    *, party_size: Optional[int] = None, composition: Optional[dict] = None,
) -> str:
    """Iter-40 — Premium image-forward boarding pass (1080×1920, 9:16).

    Layout:
      • Top hero image (full-width, ~46% of card height) showcasing the offer.
      • White card body with offer title, passenger, date/time, party, venue.
      • Side notches + dashed perforation (true ticket feel).
      • Classic BLACK QR code on white at the bottom for max scan reliability.
      • Subtle gold accent line + booking reference footer.
    """
    from PIL import Image, ImageDraw, ImageFont, ImageFilter
    import io
    import base64
    import qrcode

    W, H = 1080, 1920
    BG = (12, 14, 18)              # near-black device backdrop
    CARD = (255, 255, 255)
    INK = (15, 18, 22)
    SUB = (95, 102, 112)
    GOLD = (184, 146, 42)
    GOLD_SOFT = (212, 178, 86)
    DIVIDER = (220, 222, 226)

    img = Image.new("RGB", (W, H), BG)
    draw = ImageDraw.Draw(img)

    # ---- Background glow (subtle radial via blurred ellipse) ----
    try:
        glow = Image.new("RGB", (W, H), BG)
        gd = ImageDraw.Draw(glow)
        gd.ellipse((-300, -400, W + 300, 700), fill=(60, 48, 22))
        glow = glow.filter(ImageFilter.GaussianBlur(180))
        img = Image.blend(img, glow, 0.55)
        draw = ImageDraw.Draw(img)
    except Exception:
        pass

    # Card geometry
    CX, CY = 60, 150
    CW, CH = W - 120, H - 240
    R = 36

    def font(sz, bold=False):
        path = ("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold
                else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
        try:
            return ImageFont.truetype(path, sz)
        except Exception:
            return ImageFont.load_default()

    f_brand  = font(30, bold=True)
    f_kicker = font(26, bold=True)
    f_label  = font(30, bold=True)
    f_title  = font(58, bold=True)
    f_title_sm = font(46, bold=True)
    f_value  = font(50, bold=True)
    f_small  = font(40, bold=True)
    f_mono   = font(46, bold=True)

    def round_rect(xy, radius, fill, outline=None, width=1):
        draw.rounded_rectangle(xy, radius=radius, fill=fill,
                               outline=outline, width=width)

    # ---- Top brand strip (above card) ----
    brand_y = 70
    draw.text((CX + 10, brand_y), "BOULAY BEACH RESORT",
              fill=(245, 235, 210), font=f_brand)
    # Right side: "BOARDING PASS" kicker
    bp_text = "BOARDING PASS" if lang != "fr" else "PASS D'EMBARQUEMENT"
    bb = draw.textbbox((0, 0), bp_text, font=f_kicker)
    draw.text((CX + CW - (bb[2] - bb[0]) - 10, brand_y + 3),
              bp_text, fill=GOLD_SOFT, font=f_kicker)

    # ---- Card base (white rounded) ----
    round_rect((CX, CY, CX + CW, CY + CH), R, CARD)

    # ---- Hero image (top of card) ----
    HERO_H = 600

    # Build hero with rounded top corners
    hero_layer = None
    # Use the dedicated cached fetcher which also handles OFFER_HERO_URLS
    # fallback per offer_id (so a static offer like sunset/brunch always shows
    # its branded image even when no per-event image_url is provided).
    src_img = _fetch_hero(offer_id, hero_url=hero_url)
    if src_img is not None:
        try:
            hero = src_img.convert("RGB")
            target_w, target_h = CW, HERO_H
            src_ratio = hero.width / hero.height
            tgt_ratio = target_w / target_h
            if src_ratio > tgt_ratio:
                # too wide → crop sides
                new_w = int(hero.height * tgt_ratio)
                lx = (hero.width - new_w) // 2
                hero = hero.crop((lx, 0, lx + new_w, hero.height))
            else:
                # too tall → crop top/bottom
                new_h = int(hero.width / tgt_ratio)
                t = (hero.height - new_h) // 2
                hero = hero.crop((0, t, hero.width, t + new_h))
            hero = hero.resize((target_w, target_h), Image.LANCZOS)
            hero_layer = hero
        except Exception:
            hero_layer = None

    if hero_layer is None:
        # gradient fallback (gold to dark)
        hero_layer = Image.new("RGB", (CW, HERO_H), (40, 36, 28))
        gd = ImageDraw.Draw(hero_layer)
        for i in range(HERO_H):
            t = i / HERO_H
            c = (int(64 + (184 - 64) * (1 - t)),
                 int(48 + (146 - 48) * (1 - t)),
                 int(20 + (42 - 20) * (1 - t)))
            gd.line([(0, i), (CW, i)], fill=c)
        gd.text((CW // 2 - 80, HERO_H // 2 - 30), "BBR", fill=(245, 235, 210),
                font=font(120, bold=True))

    # Paste hero with rounded top corners (bottom stays straight for the
    # divider to feel like a real ticket cut)
    mask = Image.new("L", (CW, HERO_H), 0)
    md = ImageDraw.Draw(mask)
    md.rounded_rectangle((0, 0, CW, HERO_H), R, fill=255)
    # square off bottom corners so the perforation looks crisp
    md.rectangle((0, HERO_H - R, CW, HERO_H), fill=255)
    img.paste(hero_layer, (CX, CY), mask)

    # Dark overlay at bottom of hero for text legibility (gradient)
    try:
        overlay = Image.new("RGBA", (CW, HERO_H), (0, 0, 0, 0))
        od = ImageDraw.Draw(overlay)
        for i in range(HERO_H):
            # transparent at top, ~80% black at bottom
            a = max(0, int(220 * (i / HERO_H) ** 3))
            od.line([(0, i), (CW, i)], fill=(0, 0, 0, a))
        img.paste(overlay, (CX, CY), overlay)
    except Exception:
        pass

    # Offer title over the hero (bottom-left)
    title_lines = []
    words = (offer_name or "Réservation BBR").split()
    line = ""
    available_w = CW - 80
    use_font = f_title
    for w in words:
        test = (line + " " + w).strip()
        bb = draw.textbbox((0, 0), test, font=use_font)
        if bb[2] - bb[0] > available_w:
            title_lines.append(line)
            line = w
        else:
            line = test
    if line:
        title_lines.append(line)
    if len(title_lines) > 2:
        # reflow with smaller font
        title_lines = []
        line = ""
        use_font = f_title_sm
        for w in words:
            test = (line + " " + w).strip()
            bb = draw.textbbox((0, 0), test, font=use_font)
            if bb[2] - bb[0] > available_w:
                title_lines.append(line)
                line = w
            else:
                line = test
        if line:
            title_lines.append(line)
        title_lines = title_lines[:3]

    # Gold kicker above the title
    kicker = "OFFRE RÉSERVÉE" if lang == "fr" else "BOOKED EXPERIENCE"
    line_h = 70 if use_font is f_title else 56
    tt_h = line_h * len(title_lines)
    tt_start_y = CY + HERO_H - 60 - tt_h
    draw.text((CX + 40, tt_start_y - 36), kicker, fill=GOLD_SOFT, font=f_kicker)
    ty = tt_start_y
    for ln in title_lines:
        draw.text((CX + 40, ty), ln, fill=(255, 255, 255), font=use_font)
        ty += line_h

    # ---- Perforation row (between hero and info) ----
    PERF_Y = CY + HERO_H
    notch_r = 36
    # Erase ellipses into BG to simulate cutouts
    draw.ellipse((CX - notch_r, PERF_Y - notch_r,
                  CX + notch_r, PERF_Y + notch_r), fill=BG)
    draw.ellipse((CX + CW - notch_r, PERF_Y - notch_r,
                  CX + CW + notch_r, PERF_Y + notch_r), fill=BG)
    # Dashed perforation line
    x = CX + notch_r + 18
    while x < CX + CW - notch_r - 18:
        draw.line([(x, PERF_Y), (x + 18, PERF_Y)], fill=DIVIDER, width=3)
        x += 32

    # ---- Info block (between perforation and QR) ----
    INFO_PAD = 56
    info_top = PERF_Y + 50

    # Passenger row (full width)
    draw.text((CX + INFO_PAD, info_top), "PASSAGER" if lang == "fr" else "PASSENGER",
              fill=SUB, font=f_label)
    draw.text((CX + INFO_PAD, info_top + 46), (owner_name or "—"),
              fill=INK, font=f_value)

    # Thin gold accent under passenger
    accent_y = info_top + 130
    draw.line([(CX + INFO_PAD, accent_y),
               (CX + INFO_PAD + 100, accent_y)], fill=GOLD, width=4)

    # 2-column grid: Date / Embarquement, Convives / Lieu
    grid_top = accent_y + 36
    col_w = (CW - 2 * INFO_PAD) // 2
    col1_x = CX + INFO_PAD
    col2_x = CX + INFO_PAD + col_w

    # Date
    try:
        d_obj = datetime.fromisoformat((dates_list or [date_iso])[0])
        months = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin",
                  "Juil", "Août", "Sep", "Oct", "Nov", "Déc"] if lang == "fr" else \
                 ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        date_label = f"{d_obj.day:02d} {months[d_obj.month - 1]} {d_obj.year}"
    except Exception:
        date_label = date_iso or "—"

    draw.text((col1_x, grid_top), "DATE", fill=SUB, font=f_label)
    draw.text((col1_x, grid_top + 42), date_label, fill=INK, font=f_value)

    draw.text((col2_x, grid_top), "EMBARQUEMENT" if lang == "fr" else "BOARDING",
              fill=SUB, font=f_label)
    draw.text((col2_x, grid_top + 42), (boat_time or "—"),
              fill=INK, font=f_value)

    # Convives / Lieu row
    grid_top2 = grid_top + 134
    if composition:
        a = int(composition.get("adults") or 0)
        cp = int(composition.get("children_paid") or 0)
        cf = int(composition.get("children_free") or 0)
        parts = []
        if a:
            parts.append(f"{a} adulte{'s' if a > 1 else ''}")
        if cp:
            parts.append(f"{cp} enf. 6-12")
        if cf:
            parts.append(f"{cf} enf. <6")
        ps_text = "  ·  ".join(parts) or "1 personne"
    elif party_size:
        ps_text = f"{party_size} personne{'s' if party_size > 1 else ''}"
    else:
        ps_text = "1 personne"

    draw.text((col1_x, grid_top2), "CONVIVES" if lang == "fr" else "GUESTS",
              fill=SUB, font=f_label)
    draw.text((col1_x, grid_top2 + 42), ps_text, fill=INK, font=f_small)

    draw.text((col2_x, grid_top2), "LIEU" if lang == "fr" else "VENUE",
              fill=SUB, font=f_label)
    draw.text((col2_x, grid_top2 + 42), "Boulay Beach Resort",
              fill=INK, font=f_small)

    # Divider before QR
    div_y = grid_top2 + 150
    draw.line([(CX + INFO_PAD, div_y), (CX + CW - INFO_PAD, div_y)],
              fill=DIVIDER, width=2)

    # ---- QR section (bottom) ----
    # ECC=M (15% redundancy) keeps module density LOW for fast phone-camera
    # decode, and QR_SIZE=440 ensures the QR is large enough even when the
    # ticket is viewed on a phone screen (the scanner camera-viewfinder
    # qrbox is 240×240 — we want the QR to fill more than that).
    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=20, border=2,
    )
    qr.add_data(qr_payload)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    QR_SIZE = 440
    qr_img = qr_img.resize((QR_SIZE, QR_SIZE), Image.NEAREST)
    qr_y = div_y + 32
    qr_x = (W - QR_SIZE) // 2
    img.paste(qr_img, (qr_x, qr_y))

    # Booking ID under QR (monospace-style)
    bid_y = qr_y + QR_SIZE + 28
    label = "RÉFÉRENCE" if lang == "fr" else "REFERENCE"
    bb = draw.textbbox((0, 0), label, font=f_label)
    draw.text(((W - (bb[2] - bb[0])) / 2, bid_y), label, fill=SUB, font=f_label)
    bb2 = draw.textbbox((0, 0), ref_code or "", font=f_mono)
    draw.text(((W - (bb2[2] - bb2[0])) / 2, bid_y + 44),
              ref_code or "", fill=INK, font=f_mono)

    # ---- Signature below card ----
    sig_y = CY + CH + 16
    sig_text = "Life is Here  ·  boulaybeachresort.com"
    bb = draw.textbbox((0, 0), sig_text, font=f_small)
    draw.text(((W - (bb[2] - bb[0])) / 2, sig_y),
              sig_text, fill=GOLD_SOFT, font=f_small)

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


def make_cash_receipt_image(
    offer_id: str,
    offer_name: str,
    date_iso: str,
    boat_time: str,
    owner_name: str,
    ref_code: str,
    lang: str = "fr",
    hero_url: Optional[str] = None,
    dates_list: Optional[List[str]] = None,
) -> str:
    """Render the *temporary cash receipt* template as a base64 PNG data URL.

    Layout: gold-bordered logo header on white, hero image of the offer, then a
    cream/beige body holding a greeting (left) and the four reservation fields
    (right) separated by thin grey dividers. The reference code is printed in
    the bottom-right corner. No QR code (cash payments don't get a QR-as-pass).
    """
    W = 900
    GOLD = (140, 95, 38)
    CREAM = (245, 238, 219)
    DARK = (50, 38, 28)
    LIGHT_DARK = (70, 58, 48)
    LINE = (180, 170, 150)

    H_PAD = 36
    H_HEADER = 230
    hero_w = W - 120
    hero_h = int(hero_w * 9 / 16)
    H_BODY = 520
    H = H_PAD + H_HEADER + 20 + hero_h + H_BODY + H_PAD

    img = Image.new("RGB", (W, H), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    # --- Header: gold-bordered box with BBr logo image ---
    y = H_PAD
    box_x0, box_x1 = 60, W - 60
    box_y0 = y
    box_y1 = y + H_HEADER
    draw.rectangle([box_x0, box_y0, box_x1, box_y1], outline=GOLD, width=2)

    # Paste the logo centred inside the bordered box
    logo = _fetch_logo()
    if logo is not None:
        inner_h = H_HEADER - 24
        w0, h0 = logo.size
        new_h = inner_h
        new_w = int(w0 * (new_h / h0))
        # Cap width so the logo doesn't bleed over the gold border
        max_w = (box_x1 - box_x0) - 24
        if new_w > max_w:
            new_w = max_w
            new_h = int(h0 * (new_w / w0))
        logo_r = logo.resize((new_w, new_h))
        cx = box_x0 + (box_x1 - box_x0 - new_w) // 2
        cy = box_y0 + (H_HEADER - new_h) // 2
        img.paste(logo_r, (cx, cy), logo_r)
    else:
        f_logo = _load_font(60, bold=True)
        bbox = draw.textbbox((0, 0), "BBr", font=f_logo)
        draw.text((box_x0 + (box_x1 - box_x0 - (bbox[2] - bbox[0])) / 2 - bbox[0], box_y0 + 30), "BBr", fill=GOLD, font=f_logo)

    y = box_y1 + 20

    # --- Hero image ---
    hero_x = 60
    hero = _fetch_hero(offer_id, hero_url=hero_url)
    if hero is not None:
        ratio_src = hero.width / hero.height
        ratio_dst = hero_w / hero_h
        if ratio_src > ratio_dst:
            new_w = int(hero.height * ratio_dst)
            left = (hero.width - new_w) // 2
            hero = hero.crop((left, 0, left + new_w, hero.height))
        else:
            new_h = int(hero.width / ratio_dst)
            top = (hero.height - new_h) // 2
            hero = hero.crop((0, top, hero.width, top + new_h))
        hero = hero.resize((hero_w, hero_h))
        img.paste(hero, (hero_x, y))
    else:
        draw.rectangle([hero_x, y, hero_x + hero_w, y + hero_h], fill=(220, 215, 205))
    y += hero_h

    # --- Cream body: greeting on left, 4 fields on right ---
    body_x0, body_x1 = 60, W - 60
    body_y0 = y
    body_y1 = y + H_BODY
    draw.rectangle([body_x0, body_y0, body_x1, body_y1], fill=CREAM)

    pad = 40
    col_left_x = body_x0 + pad
    col_right_x = body_x0 + (body_x1 - body_x0) // 2 + 10
    col_right_end = body_x1 - pad

    f_h = _load_font(20, bold=True)
    f_body = _load_font(16)
    f_label = _load_font(16)
    f_value = _load_font(16, bold=True)
    f_ref = _load_font(26, bold=True)

    if lang == "fr":
        bold_block = (
            "Voici votre réçu temporaire, émis "
            "suite à votre réservation avec paiement en espèces."
        )
        body_block = "Nous vous souhaitons une expérience inoubliable."
        signoff = "Life is Here."
        labels = ("Propriétaire", "Offre", "Date", "Heure d'embarquement")
    else:
        bold_block = (
            "Here is your temporary receipt, "
            "issued upon your reservation with cash payment."
        )
        body_block = "We wish you an unforgettable experience."
        signoff = "Life is Here."
        labels = ("Owner", "Offer", "Date", "Boarding time")

    def _wrap(text: str, font, max_w: int) -> list:
        """Greedy word-wrap into a list of lines that fit within max_w pixels."""
        words = text.split()
        lines: list = []
        cur = ""
        for w in words:
            test = (cur + " " + w).strip()
            tw = draw.textbbox((0, 0), test, font=font)[2]
            if tw <= max_w or not cur:
                cur = test
            else:
                lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        return lines

    left_max_w = (body_x0 + (body_x1 - body_x0) // 2) - col_left_x - 30
    cy = body_y0 + 40
    for line in _wrap(bold_block, f_h, left_max_w):
        draw.text((col_left_x, cy), line, fill=DARK, font=f_h)
        cy += 28
    cy += 22
    for line in _wrap(body_block, f_body, left_max_w):
        draw.text((col_left_x, cy), line, fill=LIGHT_DARK, font=f_body)
        cy += 24
    cy += 18
    draw.text((col_left_x, cy), signoff, fill=LIGHT_DARK, font=f_body)

    # Right column: 4 fields with thin grey dividers
    field_y = body_y0 + 40
    if dates_list and len(dates_list) > 1:
        date_value = _format_dates_list(dates_list, lang)
    else:
        date_value = _format_date_long(date_iso, lang)
    fields = (
        (labels[0], owner_name),
        (labels[1], offer_name),
        (labels[2], date_value),
        (labels[3], boat_time),
    )
    for label, value in fields:
        label_text = label + " : "
        draw.text((col_right_x, field_y), label_text, fill=LIGHT_DARK, font=f_label)
        bbox = draw.textbbox((0, 0), label_text, font=f_label)
        lw = bbox[2] - bbox[0]
        # Shrink value font if it would overflow the column.
        col_w_avail = col_right_end - (col_right_x + lw)
        val_font = f_value
        if value:
            bbox_v = draw.textbbox((0, 0), value, font=val_font)
            if (bbox_v[2] - bbox_v[0]) > col_w_avail:
                val_font = _load_font(13, bold=True)
                bbox_v = draw.textbbox((0, 0), value, font=val_font)
                if (bbox_v[2] - bbox_v[0]) > col_w_avail:
                    val_font = _load_font(11, bold=True)
        draw.text((col_right_x + lw, field_y), value, fill=DARK, font=val_font)
        draw.line(
            [(col_right_x, field_y + 30), (col_right_end, field_y + 30)],
            fill=LINE,
            width=1,
        )
        field_y += 64

    # Reference code in the bottom-right corner of the cream body
    bbox = draw.textbbox((0, 0), ref_code, font=f_ref)
    rw = bbox[2] - bbox[0]
    draw.text(
        (col_right_end - rw - bbox[0], body_y1 - 65),
        ref_code,
        fill=DARK,
        font=f_ref,
    )

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


# ---------- Activities catalog (jet ski, quad, ...) ----------
DEFAULT_ACTIVITIES = [
    # Activités & Loisirs — Sport & Terrains
    {"id": "multisport", "name_fr": "Terrain multisport (1h)", "name_en": "Multisport Field (1h)", "category": "Activités & Loisirs", "subcategory": "Sport & Terrains", "price": 50000, "active": True},
    {"id": "padel", "name_fr": "Terrain de padel (1h)", "name_en": "Padel Court (1h)", "category": "Activités & Loisirs", "subcategory": "Sport & Terrains", "price": 20000, "active": True},
    {"id": "beach_volley", "name_fr": "Beach Volley (30 min)", "name_en": "Beach Volley (30 min)", "category": "Activités & Loisirs", "subcategory": "Sport & Terrains", "price": 10000, "active": True},
    {"id": "tir_arc", "name_fr": "Tir à l'arc (30 min)", "name_en": "Archery (30 min)", "category": "Activités & Loisirs", "subcategory": "Sport & Terrains", "price": 10000, "active": True},
    # Activités & Loisirs — Activités Nautiques
    {"id": "jetski", "name_fr": "Jet Ski (10 min)", "name_en": "Jet Ski (10 min)", "category": "Activités & Loisirs", "subcategory": "Activités Nautiques", "price": 15000, "active": True},
    {"id": "pedalo", "name_fr": "Pédalo (20 min)", "name_en": "Pedal Boat (20 min)", "category": "Activités & Loisirs", "subcategory": "Activités Nautiques", "price": 5000, "active": True},
    {"id": "kayak", "name_fr": "Canoë-Kayak (20 min)", "name_en": "Canoe-Kayak (20 min)", "category": "Activités & Loisirs", "subcategory": "Activités Nautiques", "price": 5000, "active": True},
    {"id": "paddle", "name_fr": "Paddle (20 min)", "name_en": "Paddle (20 min)", "category": "Activités & Loisirs", "subcategory": "Activités Nautiques", "price": 5000, "active": True},
    # Activités & Loisirs — Randonnées & Mobilité
    {"id": "quad", "name_fr": "Quad (30 min)", "name_en": "Quad (30 min)", "category": "Activités & Loisirs", "subcategory": "Randonnées & Mobilité", "price": 30000, "active": True},
    {"id": "buggy", "name_fr": "Buggy (30 min)", "name_en": "Buggy (30 min)", "category": "Activités & Loisirs", "subcategory": "Randonnées & Mobilité", "price": 50000, "active": True},
    {"id": "golfette", "name_fr": "Randonnée en golfette (30 min)", "name_en": "Golf Cart Tour (30 min)", "category": "Activités & Loisirs", "subcategory": "Randonnées & Mobilité", "price": 50000, "active": True},
    {"id": "rando_pied", "name_fr": "Randonnée à pied (1h)", "name_en": "Hiking (1h)", "category": "Activités & Loisirs", "subcategory": "Randonnées & Mobilité", "price": 10000, "active": True},
    {"id": "vtt", "name_fr": "VTT (1h)", "name_en": "Mountain Bike (1h)", "category": "Activités & Loisirs", "subcategory": "Randonnées & Mobilité", "price": 5000, "active": True},
    # Activités & Loisirs — Bien-être
    {"id": "massage", "name_fr": "Massage Signature (60 min)", "name_en": "Signature Massage (60 min)", "category": "Activités & Loisirs", "subcategory": "Bien-être", "price": 45000, "active": True},
    {"id": "spa_day", "name_fr": "Forfait Spa Journée", "name_en": "Spa Day Pass", "category": "Activités & Loisirs", "subcategory": "Bien-être", "price": 60000, "active": True},
    # Menus
    {"id": "menu_kaai", "name_fr": "Menu Le Kaai", "name_en": "Le Kaai Menu", "category": "Menus", "subcategory": "Kaai", "price": 35000, "active": True},
    {"id": "menu_beach_club", "name_fr": "Menu Beach Club", "name_en": "Beach Club Menu", "category": "Menus", "subcategory": "Beach Club", "price": 28000, "active": True},
    {"id": "menu_lounge", "name_fr": "Menu Lounge", "name_en": "Lounge Menu", "category": "Menus", "subcategory": "Lounge", "price": 22000, "active": True},
    # Espace privatif — Plage
    {"id": "espace_plage_balinais_6", "name_fr": "Salon balinais (6 places)", "name_en": "Balinese Lounge (6 seats)", "category": "Espace privatif", "subcategory": "Plage", "price": 50000, "active": True},
    {"id": "espace_plage_transat", "name_fr": "Transat (1 place)", "name_en": "Sun Lounger (1 seat)", "category": "Espace privatif", "subcategory": "Plage", "price": 10000, "active": True},
    # Espace privatif — Terrasse N°3
    {"id": "espace_t3_balinais_5", "name_fr": "Salon balinais (5 places)", "name_en": "Balinese Lounge (5 seats)", "category": "Espace privatif", "subcategory": "Terrasse 3", "price": 50000, "active": True},
    {"id": "espace_t3_transat", "name_fr": "Transat (1 place)", "name_en": "Sun Lounger (1 seat)", "category": "Espace privatif", "subcategory": "Terrasse 3", "price": 10000, "active": True},
    # Espace privatif — Terrasse N°2
    {"id": "espace_t2_balinais_5", "name_fr": "Salon balinais (5 places)", "name_en": "Balinese Lounge (5 seats)", "category": "Espace privatif", "subcategory": "Terrasse 2", "price": 50000, "active": True},
    {"id": "espace_t2_transat", "name_fr": "Transat (1 place)", "name_en": "Sun Lounger (1 seat)", "category": "Espace privatif", "subcategory": "Terrasse 2", "price": 10000, "active": True},
    # Espace privatif — Terrasse N°1
    {"id": "espace_t1_balinais_6", "name_fr": "Salon balinais (6 places)", "name_en": "Balinese Lounge (6 seats)", "category": "Espace privatif", "subcategory": "Terrasse 1", "price": 50000, "active": True},
    {"id": "espace_t1_transat", "name_fr": "Transat (1 place)", "name_en": "Sun Lounger (1 seat)", "category": "Espace privatif", "subcategory": "Terrasse 1", "price": 10000, "active": True},
    {"id": "espace_t1_cosy_2", "name_fr": "Salon cosy (2 places)", "name_en": "Cosy Lounge (2 seats)", "category": "Espace privatif", "subcategory": "Terrasse 1", "price": 25000, "active": True},
    {"id": "espace_t1_jacuzzi", "name_fr": "Jacuzzi piscine", "name_en": "Pool Jacuzzi", "category": "Espace privatif", "subcategory": "Terrasse 1", "price": 100000, "active": True},
    {"id": "espace_t1_salon_sec_15", "name_fr": "Salon sec (15 places)", "name_en": "Dry Lounge (15 seats)", "category": "Espace privatif", "subcategory": "Terrasse 1", "price": 150000, "active": True},
    {"id": "espace_t1_salon_sec_10", "name_fr": "Salon sec (10 places)", "name_en": "Dry Lounge (10 seats)", "category": "Espace privatif", "subcategory": "Terrasse 1", "price": 100000, "active": True},
]


async def _seed_default_activities():
    """Seed the activities collection on first run. On subsequent runs, idempotently
    add any new built-in activity that didn't exist yet (so adding new defaults
    propagates without wiping admin-customised prices)."""
    if await db.activities.count_documents({}) == 0:
        await db.activities.insert_many([dict(a) for a in DEFAULT_ACTIVITIES])
        return
    # Idempotent top-up: insert defaults whose id is missing
    existing_ids = {d["id"] async for d in db.activities.find({}, {"_id": 0, "id": 1})}
    to_add = [dict(a) for a in DEFAULT_ACTIVITIES if a["id"] not in existing_ids]
    if to_add:
        await db.activities.insert_many(to_add)
    # Backfill category/subcategory for the original defaults — keeps user-defined prices.
    for a in DEFAULT_ACTIVITIES:
        if a["id"] in existing_ids:
            await db.activities.update_one(
                {
                    "id": a["id"],
                    "$or": [
                        {"category": {"$in": ["Nautique", "Terrestre", "Bien-être", "Activité"]}},
                        {"subcategory": {"$exists": False}},
                        {"subcategory": ""},
                    ],
                },
                {"$set": {"category": a["category"], "subcategory": a["subcategory"]}},
            )
    # Retire the v1 generic Espace privatif entries now superseded by per-item zones.
    # They stay in DB (historical wallet charges keep referencing them) but disappear from the picker.
    LEGACY_PRIVATIF_IDS = ["espace_plage", "espace_terrasse_1", "espace_terrasse_2", "espace_terrasse_3"]
    await db.activities.update_many(
        {"id": {"$in": LEGACY_PRIVATIF_IDS}, "active": True},
        {"$set": {"active": False}},
    )

    # ====== Versioned catalog upgrade: v2 — refreshed Activités & Loisirs taxonomy ======
    # Force-aligns ids with their new names/prices/subcategories ONCE. Subsequent boots
    # are no-ops because the flag flips. Admin price edits made AFTER the upgrade are preserved.
    flag = await db.app_state.find_one({"key": "activities_catalog_v2"})
    if not flag:
        REFRESH_IDS = {
            "multisport", "padel", "beach_volley", "tir_arc",
            "jetski", "pedalo", "kayak", "paddle",
            "quad", "buggy", "golfette", "rando_pied", "vtt",
            "massage", "spa_day",
        }
        for a in DEFAULT_ACTIVITIES:
            if a["id"] in REFRESH_IDS:
                await db.activities.update_one(
                    {"id": a["id"]},
                    {"$set": {
                        "name_fr": a["name_fr"],
                        "name_en": a["name_en"],
                        "category": a["category"],
                        "subcategory": a["subcategory"],
                        "price": a["price"],
                        "active": True,
                    }},
                    upsert=True,
                )
        # Retire activities no longer in the v2 catalog (preserves historical charges)
        DEPRECATED_IDS = ["jetski_60", "ski_nautique", "boat_tour"]
        await db.activities.update_many(
            {"id": {"$in": DEPRECATED_IDS}, "active": True},
            {"$set": {"active": False}},
        )
        await db.app_state.insert_one({"key": "activities_catalog_v2", "applied_at": now_iso()})
        logging.info("Activities catalog v2 migration applied")


# ---------- Wallet QR card (sandstone cream styling — distinct from gold ticket) ----------
def make_wallet_image(
    owner_name: str,
    wallet_token: str,
    booking_ref: str,
    lang: str = "fr",
) -> str:
    """Build a printable wallet card image (PNG data URL) shown next to the
    travel tickets. The wallet QR is scanned at point of sale (jet ski / quad /
    massage etc.) to charge activities to the guest's stay."""
    from PIL import Image, ImageDraw, ImageFont

    W, H = 980, 1320
    CREAM = (250, 246, 232)
    GOLD = (184, 146, 42)
    DARK = (10, 10, 10)
    MUTED = (90, 80, 60)

    img = Image.new("RGB", (W, H), CREAM)
    draw = ImageDraw.Draw(img)

    # Gold ornamental border
    draw.rectangle([(24, 24), (W - 24, H - 24)], outline=GOLD, width=2)
    draw.rectangle([(40, 40), (W - 40, H - 40)], outline=GOLD, width=1)

    # Logo
    logo_h = _paste_logo(img, top=64, max_h=160, max_w_ratio=0.6)

    # Fonts
    try:
        f_eyebrow = ImageFont.truetype("DejaVuSans.ttf", 22)
        f_title = ImageFont.truetype("DejaVuSerif-Bold.ttf", 50)
        f_body = ImageFont.truetype("DejaVuSans.ttf", 22)
        f_small = ImageFont.truetype("DejaVuSans.ttf", 18)
        f_ref = ImageFont.truetype("DejaVuSansMono-Bold.ttf", 22)
    except Exception:
        f_eyebrow = f_title = f_body = f_small = f_ref = ImageFont.load_default()

    y = 64 + logo_h + 30

    eyebrow_text = "CARTE ACTIVITÉS" if lang == "fr" else "ACTIVITIES CARD"
    w_eb = draw.textlength(eyebrow_text, font=f_eyebrow)
    draw.text(((W - w_eb) / 2, y), eyebrow_text, fill=GOLD, font=f_eyebrow)
    y += 36

    title_text = "Boulay Beach Resort" if lang == "fr" else "Boulay Beach Resort"
    w_t = draw.textlength(title_text, font=f_title)
    draw.text(((W - w_t) / 2, y), title_text, fill=DARK, font=f_title)
    y += 70

    # Owner line
    draw.line([(W * 0.18, y), (W * 0.82, y)], fill=GOLD, width=1)
    y += 18
    name_text = f"Au nom de · {owner_name}" if lang == "fr" else f"In the name of · {owner_name}"
    w_n = draw.textlength(name_text, font=f_body)
    draw.text(((W - w_n) / 2, y), name_text, fill=DARK, font=f_body)
    y += 36
    ref_text = f"Réservation #{booking_ref}" if lang == "fr" else f"Booking #{booking_ref}"
    w_r = draw.textlength(ref_text, font=f_small)
    draw.text(((W - w_r) / 2, y), ref_text, fill=MUTED, font=f_small)
    y += 24
    draw.line([(W * 0.18, y), (W * 0.82, y)], fill=GOLD, width=1)
    y += 30

    # QR
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=2,
    )
    qr_payload = json.dumps({"type": "wallet", "token": wallet_token, "booking_ref": booking_ref})
    qr.add_data(qr_payload)
    qr.make(fit=True)
    from qrcode.image.styledpil import StyledPilImage
    from qrcode.image.styles.moduledrawers.pil import RoundedModuleDrawer
    from qrcode.image.styles.colormasks import SolidFillColorMask
    qr_img = qr.make_image(
        image_factory=StyledPilImage,
        module_drawer=RoundedModuleDrawer(),
        color_mask=SolidFillColorMask(back_color=CREAM, front_color=(120, 90, 36)),
    ).convert("RGB")
    qr_size = 460
    qr_img = qr_img.resize((qr_size, qr_size))
    qx = (W - qr_size) // 2
    img.paste(qr_img, (qx, y))
    y += qr_size + 24

    # Caption
    lines = (
        [
            "Présentez ce QR aux points d'activités du resort",
            "(Jet Ski, Quad, Paddle, Spa, Excursions…)",
            "pour ajouter une prestation à votre séjour.",
            "Le solde est réglé au moment du check-out.",
        ]
        if lang == "fr"
        else [
            "Show this QR at any resort activity point",
            "(Jet Ski, Quad, Paddle, Spa, Excursions…)",
            "to add a service to your stay.",
            "Balance is settled at check-out.",
        ]
    )
    for line in lines:
        w_l = draw.textlength(line, font=f_small)
        draw.text(((W - w_l) / 2, y), line, fill=MUTED, font=f_small)
        y += 24

    # Footer ref
    short = wallet_token[:10].upper()
    foot_text = f"WALLET · {short}"
    w_f = draw.textlength(foot_text, font=f_ref)
    draw.text(((W - w_f) / 2, H - 80), foot_text, fill=GOLD, font=f_ref)

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()



def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


bearer = HTTPBearer(auto_error=False)


async def get_current_staff(creds: HTTPAuthorizationCredentials = Depends(bearer)) -> dict:
    if not creds:
        raise HTTPException(status_code=401, detail="Authentication required")
    payload = decode_token(creds.credentials)
    if payload.get("type") != "staff":
        raise HTTPException(status_code=403, detail="Staff account required")
    user = await db.staff.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Staff not found")
    return user


# ----- App -----
app = FastAPI(title="Boulay Beach Resort API")
api = APIRouter(prefix="/api")


@api.get("/")
async def root():
    return {"name": "Boulay Beach Resort API", "status": "ok"}


# ----- Auth: Staff (kept for back-office) -----
@api.post("/auth/staff/login", response_model=TokenResponse)
async def login_staff(body: StaffLogin, request: Request):
    email = body.email.lower()
    # Honor X-Forwarded-For (k8s/cloudflare) before falling back to the direct peer
    fwd = request.headers.get("x-forwarded-for", "")
    client_ip = (fwd.split(",")[0].strip() if fwd else None) or (request.client.host if request.client else "unknown")
    identifier = f"{client_ip}:{email}"

    # Brute-force lockout: 5 failed attempts within 15 minutes -> 423 Locked
    now = datetime.now(timezone.utc)
    window_start = now - timedelta(minutes=15)
    attempt_doc = await db.login_attempts.find_one({"identifier": identifier})
    if attempt_doc:
        recent = [a for a in attempt_doc.get("attempts", []) if a >= window_start.isoformat()]
        if len(recent) >= 5:
            # Compute retry-after from oldest recent failure
            oldest = min(recent)
            retry_after = max(1, int((datetime.fromisoformat(oldest) + timedelta(minutes=15) - now).total_seconds()))
            raise HTTPException(
                status_code=423,
                detail=f"Trop de tentatives. Réessayez dans {retry_after // 60 + 1} minute(s).",
                headers={"Retry-After": str(retry_after)},
            )

    user = await db.staff.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        # Log failure (keep only attempts within the window)
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$push": {"attempts": now.isoformat()}, "$set": {"last_failure_at": now.isoformat()}},
            upsert=True,
        )
        # Trim to last 10 attempts to avoid unbounded growth
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$push": {"attempts": {"$each": [], "$slice": -10}}},
        )
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Success: clear all attempts for this identifier
    await db.login_attempts.delete_one({"identifier": identifier})

    token = create_token({"sub": user["id"], "type": "staff", "role": user["role"]})
    public = {
        "id": user["id"], "name": user["name"], "email": user["email"],
        "role": user["role"], "pole_id": user.get("pole_id"),
        "nav_sections": user.get("nav_sections") or None,
    }
    return TokenResponse(access_token=token, user=public)


# ----- Offers -----
def _with_boat_times(offer: dict) -> dict:
    oid = offer["id"]
    extra = {"allowed_weekdays": ALLOWED_WEEKDAYS_BY_OFFER.get(oid, [])}
    if oid in ("le_kaai", "hebergement"):
        extra["boat_times_weekday"] = BOAT_TIMES_WEEKDAY
        extra["boat_times_weekend"] = BOAT_TIMES_WEEKEND
        extra["boat_times"] = BOAT_TIMES_WEEKDAY  # default fallback
    else:
        extra["boat_times"] = BOAT_TIMES_BY_OFFER.get(oid, [])
    return {**offer, **extra}


@api.get("/offers")
async def list_offers():
    out = []
    for o in OFFERS.values():
        merged = await _apply_overrides(o)
        out.append(_with_boat_times(merged))
    return out


@api.get("/offers/{offer_id}")
async def get_offer(offer_id: str):
    if offer_id not in OFFERS:
        raise HTTPException(status_code=404, detail="Offer not found")
    merged = await _apply_overrides(OFFERS[offer_id])
    out = _with_boat_times(merged)
    out["pole"] = _pole_for_offer(offer_id)
    return out


# ----- Poles (taxonomy of public entry points) -----
@api.get("/poles")
async def list_poles():
    """Public — returns the 5 entry-point pôles with their sub-offers hydrated.
    The frontend uses this to build the landing page (5 pôle cards) and the
    per-pôle landing pages (sub-offer mini-cards)."""
    out = []
    for pid, p in sorted(POLES.items(), key=lambda kv: kv[1].get("sort_order", 99)):
        sub_offers = []
        for oid in p["offers"]:
            if oid == "events_maison":
                # Synthetic sub-offer pointing at the special_events module
                sub_offers.append({
                    "id": "events_maison",
                    "name_fr": "Events Maison",
                    "name_en": "In-house Events",
                    "schedule_fr": "Événements spéciaux signature",
                    "schedule_en": "Signature special events",
                    "tagline_fr": "Découvrez les événements spéciaux à venir.",
                    "tagline_en": "Discover the upcoming signature events.",
                    "price_adult": 0,
                    "price_child": 0,
                    "max_capacity": 0,
                    "is_synthetic": True,
                    "kind": "events_list",
                })
            elif oid in OFFERS:
                o = await _apply_overrides(OFFERS[oid])
                o = dict(o)
                o.update(_with_boat_times(o))
                sub_offers.append(o)
        out.append({**p, "sub_offers": sub_offers})
    return out


@api.get("/poles/{pole_id}")
async def get_pole(pole_id: str):
    if pole_id not in POLES:
        raise HTTPException(status_code=404, detail="Pôle non trouvé")
    p = POLES[pole_id]
    sub_offers = []
    for oid in p["offers"]:
        if oid == "events_maison":
            # Hydrate with all published+active special events (not just the featured one)
            today = datetime.now(timezone.utc).date().isoformat()
            cursor = db.special_events.find({"status": "published"}, {"_id": 0}).sort("created_at", -1)
            evs = []
            async for ev in cursor:
                if not _event_is_currently_active(ev, today):
                    continue
                evs.append({
                    **_public_event(ev),
                    "event_dates": [d for d in (ev.get("event_dates") or []) if d >= today],
                })
            sub_offers.append({
                "id": "events_maison",
                "name_fr": "Events Maison",
                "name_en": "In-house Events",
                "schedule_fr": "Événements spéciaux signature",
                "schedule_en": "Signature special events",
                "tagline_fr": "Découvrez les événements spéciaux à venir.",
                "tagline_en": "Discover the upcoming signature events.",
                "is_synthetic": True,
                "kind": "events_list",
                "events": evs,
            })
        elif oid in OFFERS:
            o = await _apply_overrides(OFFERS[oid])
            o = _with_boat_times(o)
            sub_offers.append(o)
    return {**p, "sub_offers": sub_offers}


@api.get("/staff/consumption/analytics")
async def staff_consumption_analytics(period: str = "30d", staff=Depends(get_current_staff)):
    """Wallet-level consumption analytics (charges added on-site via /staff/activites).
    Aggregates ACTIVE (non-voided) wallet transactions over the period.
    Returns:
      - kpis: total_charges, total_revenue, active_count, voided_count, voided_amount
      - by_category / by_subcategory : count + revenue per group
      - top_items : top 8 most billed activities
      - daily_trend : revenue per day
    """
    await _require_role(staff, ["manager", "admin"])

    days_map = {"today": 0, "7d": 7, "30d": 30, "90d": 90}
    days = days_map.get(period, 30)
    cutoff_dt = datetime.now(timezone.utc) - timedelta(days=days)
    cutoff_iso = cutoff_dt.isoformat()

    # Pull the activities catalog once for category/subcategory lookup
    cat_map = {}
    async for a in db.activities.find({}, {"_id": 0, "id": 1, "category": 1, "subcategory": 1}):
        cat_map[a["id"]] = {"category": a.get("category") or "Autre", "subcategory": a.get("subcategory") or "—"}

    # Flatten all transactions across wallets where any tx falls in the window
    pipeline = [
        {"$match": {"transactions": {"$exists": True, "$ne": []}}},
        {"$unwind": "$transactions"},
        {"$match": {"transactions.created_at": {"$gte": cutoff_iso}}},
        {"$replaceRoot": {"newRoot": "$transactions"}},
    ]
    txs = [t async for t in db.wallets.aggregate(pipeline)]

    active = [t for t in txs if t.get("status") != "voided"]
    voided = [t for t in txs if t.get("status") == "voided"]

    total_revenue = sum(int(t.get("amount", 0) or 0) for t in active)
    voided_amount = sum(int(t.get("amount", 0) or 0) for t in voided)

    by_category: dict = {}
    by_subcategory: dict = {}
    by_item: dict = {}
    daily: dict = {}

    for t in active:
        aid = t.get("activity_id") or "custom"
        meta = cat_map.get(aid, {"category": "Offres spéciales", "subcategory": "—"})
        cat = meta["category"]
        sub = meta["subcategory"]
        amount = int(t.get("amount", 0) or 0)
        qty = int(t.get("quantity", 0) or 0)
        label = t.get("label") or aid

        by_category.setdefault(cat, {"category": cat, "count": 0, "revenue": 0, "quantity": 0})
        by_category[cat]["count"] += 1
        by_category[cat]["revenue"] += amount
        by_category[cat]["quantity"] += qty

        key = f"{cat}||{sub}"
        by_subcategory.setdefault(key, {"category": cat, "subcategory": sub, "count": 0, "revenue": 0, "quantity": 0})
        by_subcategory[key]["count"] += 1
        by_subcategory[key]["revenue"] += amount
        by_subcategory[key]["quantity"] += qty

        by_item.setdefault(aid, {"activity_id": aid, "label": label, "category": cat, "subcategory": sub, "count": 0, "revenue": 0, "quantity": 0})
        by_item[aid]["count"] += 1
        by_item[aid]["revenue"] += amount
        by_item[aid]["quantity"] += qty

        date_str = (t.get("created_at") or "")[:10]
        if date_str:
            daily.setdefault(date_str, {"date": date_str, "revenue": 0, "count": 0})
            daily[date_str]["revenue"] += amount
            daily[date_str]["count"] += 1

    return {
        "period": period,
        "kpis": {
            "active_count": len(active),
            "total_revenue": total_revenue,
            "voided_count": len(voided),
            "voided_amount": voided_amount,
            "avg_charge": int(total_revenue / len(active)) if active else 0,
        },
        "by_category": sorted(by_category.values(), key=lambda x: x["revenue"], reverse=True),
        "by_subcategory": sorted(by_subcategory.values(), key=lambda x: x["revenue"], reverse=True),
        "top_items": sorted(by_item.values(), key=lambda x: x["revenue"], reverse=True)[:8],
        "daily_trend": sorted(daily.values(), key=lambda x: x["date"]),
    }


@api.get("/staff/poles/{pole_id}/overview")
async def staff_pole_overview(pole_id: str, staff=Depends(get_current_staff)):
    """Return everything needed to render the pôle-focused staff page:
    pôle metadata, sub-offer breakdown, KPIs (today / 30d), recent bookings."""
    await _require_role(staff, ["manager", "admin"])
    if pole_id not in POLES:
        raise HTTPException(status_code=404, detail="Pôle non trouvé")
    pole = POLES[pole_id]
    offer_ids = list(pole.get("offers", []))
    # 'events_maison' is the synthetic sub-offer mapped to special_event bookings
    mongo_offer_filter = []
    has_events_maison = "events_maison" in offer_ids
    static_offers = [o for o in offer_ids if o != "events_maison"]
    if static_offers:
        mongo_offer_filter.append({"offer_type": {"$in": static_offers}})
    if has_events_maison:
        mongo_offer_filter.append({"offer_type": "special_event"})
    base_or = mongo_offer_filter + [{"pole": pole_id}]

    today = datetime.now(timezone.utc).date().isoformat()
    from datetime import timedelta as _td
    cutoff_30d = (datetime.now(timezone.utc).date() - _td(days=30)).isoformat()

    # Sub-offer breakdown (last 30d)
    sub_offer_stats = {oid: {"id": oid, "count": 0, "revenue": 0, "guests": 0} for oid in offer_ids}
    cursor_30d = db.bookings.find(
        {"$or": base_or, "date": {"$gte": cutoff_30d}, "status": {"$ne": "cancelled"}},
        {"_id": 0, "offer_type": 1, "total_amount": 1, "adults": 1, "children": 1},
    )
    today_count = 0
    today_revenue = 0
    today_guests = 0
    total_30d_count = 0
    total_30d_revenue = 0
    async for b in cursor_30d:
        oid = b.get("offer_type") or ""
        bucket = "events_maison" if oid == "special_event" else oid
        if bucket in sub_offer_stats:
            sub_offer_stats[bucket]["count"] += 1
            sub_offer_stats[bucket]["revenue"] += int(b.get("total_amount", 0) or 0)
            sub_offer_stats[bucket]["guests"] += int(b.get("adults", 0)) + int(b.get("children", 0))
        total_30d_count += 1
        total_30d_revenue += int(b.get("total_amount", 0) or 0)

    cursor_today = db.bookings.find(
        {"$or": base_or, "date": today, "status": {"$ne": "cancelled"}},
        {"_id": 0, "adults": 1, "children": 1, "total_amount": 1, "status": 1},
    )
    async for b in cursor_today:
        today_count += 1
        today_guests += int(b.get("adults", 0)) + int(b.get("children", 0))
        if b.get("status") in ("confirmed", "arrived", "completed"):
            today_revenue += int(b.get("total_amount", 0) or 0)

    # Recent bookings (last 20, most recent first by created_at)
    recent_cursor = db.bookings.find(
        {"$or": base_or, "status": {"$ne": "cancelled"}},
        {
            "_id": 0, "id": 1, "offer_type": 1, "offer_name": 1, "date": 1,
            "adults": 1, "children": 1, "boat_time": 1, "total_amount": 1,
            "status": 1, "phone": 1, "participants": 1, "created_at": 1, "paid_at": 1,
        },
    ).sort("created_at", -1).limit(20)
    recent = await recent_cursor.to_list(length=20)

    # ============== ANALYTICS (30 days) ==============
    # Pull all bookings of the pôle over the last 30d, including cancelled (to compute the rate)
    all_30d = await db.bookings.find(
        {"$or": base_or, "date": {"$gte": cutoff_30d}},
        {
            "_id": 0, "id": 1, "offer_type": 1, "date": 1, "status": 1,
            "adults": 1, "children": 1, "boat_time": 1, "total_amount": 1,
            "payment_method": 1, "paid_at": 1, "created_at": 1,
            "phone": 1, "email": 1, "participants": 1,
        },
    ).to_list(length=5000)

    weekday_labels = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    daily_revenue: dict = {}
    by_status: dict = {"pending": 0, "confirmed": 0, "arrived": 0, "completed": 0, "cancelled": 0}
    by_payment_method: dict = {}
    by_weekday: dict = {lbl: {"count": 0, "revenue": 0} for lbl in weekday_labels}
    by_boat_time: dict = {}
    by_client: dict = {}
    total_adults = 0
    total_children = 0
    lead_times: list = []
    revenue_paid = 0
    bookings_paid = 0

    for b in all_30d:
        st = b.get("status") or "pending"
        if st in by_status:
            by_status[st] += 1

        if st == "cancelled":
            continue  # All other analytics exclude cancelled bookings

        date_str = b.get("date") or ""
        amount = int(b.get("total_amount", 0) or 0)
        daily_revenue.setdefault(date_str, {"date": date_str, "revenue": 0, "count": 0})
        daily_revenue[date_str]["revenue"] += amount
        daily_revenue[date_str]["count"] += 1

        if b.get("paid_at"):
            revenue_paid += amount
            bookings_paid += 1

        method = b.get("payment_method") or "unknown"
        by_payment_method.setdefault(method, {"count": 0, "total": 0})
        by_payment_method[method]["count"] += 1
        by_payment_method[method]["total"] += amount

        try:
            wd = datetime.strptime(date_str, "%Y-%m-%d").weekday()
            wl = weekday_labels[wd]
            by_weekday[wl]["count"] += 1
            by_weekday[wl]["revenue"] += amount
        except (ValueError, IndexError):
            pass

        bt = b.get("boat_time") or "—"
        by_boat_time.setdefault(bt, 0)
        by_boat_time[bt] += 1

        adults = int(b.get("adults", 0) or 0)
        children = int(b.get("children", 0) or 0)
        total_adults += adults
        total_children += children

        created = b.get("created_at")
        if created and date_str:
            try:
                # created_at is iso, date is yyyy-mm-dd
                cdt = datetime.fromisoformat(created.replace("Z", "+00:00")).date()
                bdt = datetime.strptime(date_str, "%Y-%m-%d").date()
                lead = (bdt - cdt).days
                if lead >= 0:
                    lead_times.append(lead)
            except (ValueError, AttributeError):
                pass

        # Top clients: aggregate by phone (more stable than email for this pôle)
        phone = b.get("phone") or ""
        if phone:
            primary = (b.get("participants") or [{}])[0]
            by_client.setdefault(phone, {
                "phone": phone,
                "name": f"{primary.get('surname', '')} {primary.get('name', '')}".strip() or phone,
                "email": b.get("email") or "",
                "count": 0,
                "total": 0,
            })
            by_client[phone]["count"] += 1
            by_client[phone]["total"] += amount

    daily_trend = sorted(daily_revenue.values(), key=lambda x: x["date"])
    payment_method_list = [{"method": k, **v} for k, v in by_payment_method.items()]
    weekday_list = [{"day": lbl, **by_weekday[lbl]} for lbl in weekday_labels]
    boat_time_list = sorted(
        [{"boat_time": k, "count": v} for k, v in by_boat_time.items()],
        key=lambda x: x["count"], reverse=True,
    )[:8]
    top_clients = sorted(by_client.values(), key=lambda c: c["total"], reverse=True)[:5]

    active_30d = total_30d_count  # not cancelled
    cancelled_30d = by_status.get("cancelled", 0)
    total_with_cancelled = active_30d + cancelled_30d
    cancellation_rate = round((cancelled_30d / total_with_cancelled) * 100, 1) if total_with_cancelled else 0
    avg_basket = round(total_30d_revenue / active_30d) if active_30d else 0
    avg_lead_time = round(sum(lead_times) / len(lead_times), 1) if lead_times else 0
    paid_rate = round((bookings_paid / active_30d) * 100, 1) if active_30d else 0

    analytics = {
        "daily_trend": daily_trend,
        "by_status": [{"status": k, "count": v} for k, v in by_status.items()],
        "by_payment_method": payment_method_list,
        "by_weekday": weekday_list,
        "by_boat_time": boat_time_list,
        "top_clients": top_clients,
        "avg_basket": avg_basket,
        "avg_lead_time_days": avg_lead_time,
        "cancellation_rate": cancellation_rate,
        "paid_rate": paid_rate,
        "guests_breakdown": {"adults": total_adults, "children": total_children},
        "revenue_paid_30d": revenue_paid,
    }

    # Hydrate sub_offers with metadata + stats
    sub_offers_out = []
    for oid in offer_ids:
        if oid == "events_maison":
            sub_offers_out.append({
                "id": "events_maison",
                "name_fr": "Events Maison",
                "name_en": "In-house Events",
                "is_synthetic": True,
                "stats": sub_offer_stats[oid],
            })
        elif oid in OFFERS:
            o = dict(OFFERS[oid])
            sub_offers_out.append({
                "id": oid,
                "name_fr": o.get("name_fr"),
                "name_en": o.get("name_en"),
                "schedule_fr": o.get("schedule_fr"),
                "price_adult": o.get("price_adult"),
                "price_child": o.get("price_child"),
                "max_capacity": o.get("max_capacity"),
                "stats": sub_offer_stats[oid],
            })
    # Occupancy per sub-offer: count / (capacity * 30 days), capped at 100%
    for s in sub_offers_out:
        capacity = s.get("max_capacity") or 0
        if capacity > 0:
            s["occupancy_pct"] = round(min(100, ((s["stats"]["count"] / (capacity * 30)) * 100)), 1)
        else:
            s["occupancy_pct"] = None

    # ============== WALLET / CONSOMMATION SUR PLACE (activites_events only) ==============
    wallet_stats = None
    if pole_id == "activites_events":
        # Aggregate all wallet transactions in the last 30 days (active only,
        # voided ones counted apart for transparency).
        cutoff_dt = datetime.now(timezone.utc) - timedelta(days=30)
        cutoff_iso = cutoff_dt.isoformat()

        cat_map = {}
        async for a in db.activities.find({}, {"_id": 0, "id": 1, "category": 1, "subcategory": 1}):
            cat_map[a["id"]] = {
                "category": a.get("category") or "Autre",
                "subcategory": a.get("subcategory") or "—",
            }

        pipeline = [
            {"$match": {"transactions": {"$exists": True, "$ne": []}}},
            {"$unwind": "$transactions"},
            {"$match": {"transactions.created_at": {"$gte": cutoff_iso}}},
            {"$replaceRoot": {"newRoot": "$transactions"}},
        ]
        txs = [t async for t in db.wallets.aggregate(pipeline)]
        active_txs = [t for t in txs if t.get("status") != "voided"]
        voided_txs = [t for t in txs if t.get("status") == "voided"]

        w_total_revenue = sum(int(t.get("amount", 0) or 0) for t in active_txs)
        w_voided_amount = sum(int(t.get("amount", 0) or 0) for t in voided_txs)

        w_by_category: dict = {}
        w_by_item: dict = {}
        w_daily: dict = {}
        for t in active_txs:
            aid = t.get("activity_id") or "custom"
            meta = cat_map.get(aid, {"category": "Offres spéciales", "subcategory": "—"})
            cat = meta["category"]
            amount = int(t.get("amount", 0) or 0)
            qty = int(t.get("quantity", 0) or 0)
            label = t.get("label") or aid

            w_by_category.setdefault(cat, {"category": cat, "count": 0, "revenue": 0, "quantity": 0})
            w_by_category[cat]["count"] += 1
            w_by_category[cat]["revenue"] += amount
            w_by_category[cat]["quantity"] += qty

            w_by_item.setdefault(aid, {
                "activity_id": aid, "label": label,
                "category": cat, "subcategory": meta["subcategory"],
                "count": 0, "revenue": 0, "quantity": 0,
            })
            w_by_item[aid]["count"] += 1
            w_by_item[aid]["revenue"] += amount
            w_by_item[aid]["quantity"] += qty

            date_str = (t.get("created_at") or "")[:10]
            if date_str:
                w_daily.setdefault(date_str, {"date": date_str, "revenue": 0, "count": 0})
                w_daily[date_str]["revenue"] += amount
                w_daily[date_str]["count"] += 1

        wallet_stats = {
            "kpis": {
                "active_count": len(active_txs),
                "total_revenue": w_total_revenue,
                "voided_count": len(voided_txs),
                "voided_amount": w_voided_amount,
                "avg_charge": int(w_total_revenue / len(active_txs)) if active_txs else 0,
            },
            "by_category": sorted(w_by_category.values(), key=lambda x: x["revenue"], reverse=True),
            "top_items": sorted(w_by_item.values(), key=lambda x: x["revenue"], reverse=True)[:8],
            "daily_trend": sorted(w_daily.values(), key=lambda x: x["date"]),
        }

    return {
        "pole": {
            "id": pole_id,
            "name_fr": pole["name_fr"],
            "name_en": pole["name_en"],
            "tagline_fr": pole.get("tagline_fr"),
            "sort_order": pole.get("sort_order"),
        },
        "kpis": {
            "today": {"count": today_count, "guests": today_guests, "revenue": today_revenue},
            "last_30d": {"count": total_30d_count, "revenue": total_30d_revenue},
        },
        "sub_offers": sub_offers_out,
        "recent_bookings": recent,
        "analytics": analytics,
        "wallet_stats": wallet_stats,
    }


# ----- Availability -----
@api.get("/availability/{offer_id}/{when}")
async def availability(offer_id: str, when: str):
    if offer_id not in OFFERS:
        raise HTTPException(status_code=404, detail="Offer not found")
    max_cap = OFFERS[offer_id]["max_capacity"]
    cursor = db.bookings.find(
        {"offer_type": offer_id, "date": when, "status": {"$ne": "cancelled"}},
        {"_id": 0, "adults": 1, "children": 1},
    )
    booked = 0
    async for b in cursor:
        booked += int(b.get("adults", 0)) + int(b.get("children", 0))
    return {
        "offer_id": offer_id,
        "date": when,
        "max_capacity": max_cap,
        "booked": booked,
        "remaining": max(max_cap - booked, 0),
    }


# ----- Bookings (guest flow, no auth) -----
async def _resolve_special_event_offer(event_id: str, booking_date: Optional[str] = None) -> dict:
    """Load the event from db.special_events and return an OFFERS-shaped dict.
    Raises 400/404 if the event is missing, archived, or out of activation window.

    For multi-day events, when `booking_date` is provided and matches a
    programme item, the per-day prices and label override the event-level
    fallback. This keeps the booking total consistent with what's shown on
    the public booking tunnel for that specific day.
    """
    if not event_id:
        raise HTTPException(status_code=400, detail="special_event_id is required for special_event bookings")
    ev = await db.special_events.find_one({"id": event_id}, {"_id": 0})
    if not ev:
        raise HTTPException(status_code=404, detail="Special event not found")
    if ev.get("status") != "published":
        raise HTTPException(status_code=400, detail="Special event is not currently published")
    today = datetime.now(timezone.utc).date().isoformat()
    if ev.get("active_from") and today < ev["active_from"]:
        raise HTTPException(status_code=400, detail="Special event booking is not yet open")
    if ev.get("active_to") and today > ev["active_to"]:
        raise HTTPException(status_code=400, detail="Special event booking window is closed")
    # Per-day overrides for multi-day programmes
    title = ev.get("title") or "Événement Spécial"
    price_adult = int(ev.get("price_adult", 0))
    price_child = int(ev.get("price_child", 0))
    description = ev.get("description") or ""
    if ev.get("event_kind") == "multi_day" and booking_date:
        item = next((p for p in (ev.get("programme") or []) if p.get("date") == booking_date), None)
        if item:
            title = f"{title} — {item.get('title') or ''}".strip(" —")
            description = (item.get("description") or description)
            price_adult = int(item.get("price_adult", price_adult))
            price_child = int(item.get("price_child", price_child))
    # Compute the bookable dates list — single_day uses event_dates as-is,
    # multi_day uses the programme dates (fallback to event_dates).
    bookable_dates = list(ev.get("event_dates") or [])
    if ev.get("event_kind") == "multi_day":
        prog_dates = sorted({p.get("date") for p in (ev.get("programme") or []) if p.get("date")})
        bookable_dates = prog_dates or bookable_dates
    return {
        "id": "special_event",
        "event_id": ev["id"],
        "name_fr": title,
        "name_en": title,
        "schedule_fr": ev.get("subtitle") or "",
        "schedule_en": ev.get("subtitle") or "",
        "tagline_fr": description,
        "tagline_en": description,
        "price_adult": price_adult,
        "price_child": price_child,
        "max_capacity": int(ev.get("capacity", 0)),
        "event_dates": bookable_dates,
        "boat_times": list(ev.get("boat_times") or []),
        "return_boat_times": list(ev.get("return_boat_times") or []),
        "image_url": ev.get("image_url") or "",
    }


@api.post("/bookings")
async def create_booking(body: BookingCreate):
    is_special = body.offer_type == "special_event"
    if is_special:
        offer = await _resolve_special_event_offer(body.special_event_id or "", body.date)
    else:
        if body.offer_type not in OFFERS:
            raise HTTPException(status_code=400, detail="Invalid offer")
        offer = OFFERS[body.offer_type]

    # Validate boat_time against offer-specific allowed times (day-dependent for le_kaai)
    try:
        booking_date = datetime.strptime(body.date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format (YYYY-MM-DD)")
    if booking_date < datetime.now(timezone.utc).date():
        raise HTTPException(status_code=400, detail="Date must be in the future")

    if is_special:
        allowed_times = offer["boat_times"]
        if not allowed_times:
            raise HTTPException(status_code=400, detail="This special event has no boat schedule configured")
        if body.boat_time not in allowed_times:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid boat time for this event. Allowed: {', '.join(allowed_times)}",
            )
        event_dates = offer["event_dates"]
        if event_dates and body.date not in event_dates:
            raise HTTPException(
                status_code=400,
                detail=f"Selected date is not part of this event. Allowed: {', '.join(event_dates)}",
            )
        # Multi-day cumulative booking: validate every selected date belongs to
        # the programme and that there's still room on each one.
        if body.multi_day_dates:
            extra_dates = [d for d in body.multi_day_dates if d != body.date]
            for d in extra_dates:
                if event_dates and d not in event_dates:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Date {d} not part of this event",
                    )
    else:
        allowed_times = _boat_times_for_date(body.offer_type, booking_date.weekday())
        if body.boat_time not in allowed_times:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid boat time for {body.offer_type}. Allowed: {', '.join(allowed_times)}",
            )
        # Validate day-of-week matches the offer (Day Pass Mon-Fri, Sunset Sat, Brunch Sun)
        allowed_weekdays = ALLOWED_WEEKDAYS_BY_OFFER.get(body.offer_type, [])
        if booking_date.weekday() not in allowed_weekdays:
            names = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"}
            allowed_names = [names[d] for d in allowed_weekdays]
            raise HTTPException(
                status_code=400,
                detail=f"Selected date is not available for {body.offer_type}. Allowed days: {', '.join(allowed_names)}",
            )

    # ---- iter-30: normalize children_paid / children_free up-front.
    # Backward compat: when the caller only sends `children`, treat it as paid.
    if body.children_paid is not None or body.children_free is not None:
        children_paid_n = int(body.children_paid or 0)
        children_free_n = int(body.children_free or 0)
        body.children = children_paid_n + children_free_n
    else:
        children_paid_n = int(body.children)
        children_free_n = 0

    total_guests = body.adults + body.children
    if total_guests <= 0:
        raise HTTPException(status_code=400, detail="At least one guest required")

    # NEW booker-only flow: a single participant (the booker) is sent for any
    # head-count. We still accept the legacy shapes (adults only, or one entry
    # per guest) for backward compatibility with older API consumers. The
    # backend then expands the booker into N adult tickets (+ M child tickets
    # when packages drive the booking) when generating QRs.
    p_count = len(body.participants)
    if p_count == 1:
        if body.participants[0].kind == "child":
            raise HTTPException(status_code=400, detail="Le réservant doit être un adulte")
    elif p_count == body.adults:
        if any(p.kind == "child" for p in body.participants):
            raise HTTPException(status_code=400, detail="Children should not be sent as participants")
    elif p_count == total_guests:
        adult_count = sum(1 for p in body.participants if p.kind == "adult")
        child_count = sum(1 for p in body.participants if p.kind == "child")
        if adult_count != body.adults or child_count != body.children:
            raise HTTPException(
                status_code=400,
                detail="Participants adult/child distribution does not match",
            )
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Expected 1 booker participant (or {body.adults}/{total_guests}), received {p_count}",
        )

    for i, p in enumerate(body.participants):
        if not p.name.strip() or not p.surname.strip() or not p.nationality.strip():
            raise HTTPException(status_code=400, detail="Nom, prénom et nationalité sont obligatoires pour le réservant")
        # Booker (first adult) must provide email + phone — used for confirmation
        if i == 0 and p.kind == "adult":
            if not (p.email or "").strip() or not (p.phone or "").strip():
                raise HTTPException(status_code=400, detail="Email et téléphone obligatoires pour le réservant")

    # capacity check
    cap_filter = {"offer_type": body.offer_type, "date": body.date, "status": {"$ne": "cancelled"}}
    if is_special:
        cap_filter["special_event_id"] = offer["event_id"]
    cursor = db.bookings.find(
        cap_filter,
        {"_id": 0, "adults": 1, "children": 1},
    )
    booked = 0
    async for b in cursor:
        booked += int(b.get("adults", 0)) + int(b.get("children", 0))
    if booked + total_guests > offer["max_capacity"]:
        raise HTTPException(status_code=400, detail="Not enough availability for this date")

    bid = str(uuid.uuid4())
    reference_token = uuid.uuid4().hex

    # ---- iter-30: 5-digit numeric booking_code, unique across active bookings.
    # Used by accompanying adults to register themselves via /companion/{code}.
    booking_code = None
    for _ in range(20):  # bounded retries; collision probability ~0
        candidate = "".join(secrets.choice("0123456789") for _ in range(5))
        if not await db.bookings.find_one({"booking_code": candidate}, {"_id": 1}):
            booking_code = candidate
            break
    if not booking_code:
        booking_code = "".join(secrets.choice("0123456789") for _ in range(6))

    is_overnight = bool(offer.get("is_overnight"))
    room_tiers = offer.get("room_tiers") or []
    selected_tier = None
    nights = 0
    checkout_iso = None
    return_boat_time = None
    if is_overnight:
        if not body.checkout_date:
            raise HTTPException(status_code=400, detail="checkout_date is required for overnight stays")
        try:
            checkout = datetime.strptime(body.checkout_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid checkout_date format (YYYY-MM-DD)")
        nights = (checkout - booking_date).days
        if nights < 1:
            raise HTTPException(status_code=400, detail="Checkout date must be at least one day after arrival")
        checkout_iso = body.checkout_date
        # Validate return boat time against checkout weekday
        if not body.return_boat_time:
            raise HTTPException(status_code=400, detail="return_boat_time is required for overnight stays")
        return_allowed = _boat_times_for_date(body.offer_type, checkout.weekday())
        if body.return_boat_time not in return_allowed:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid return boat time. Allowed: {', '.join(return_allowed)}",
            )
        return_boat_time = body.return_boat_time
        if room_tiers:
            if not body.room_tier:
                raise HTTPException(status_code=400, detail="room_tier is required for this offer")
            selected_tier = next((t for t in room_tiers if t["id"] == body.room_tier), None)
            if not selected_tier:
                raise HTTPException(status_code=400, detail="Invalid room_tier")
            # Overbooking guard: for every night in the stay, sum existing rooms for this tier must
            # leave at least body.rooms slots available against the tier's `inventory`.
            tier_inventory = int(selected_tier.get("inventory", 0))
            if tier_inventory > 0:
                # Find any existing hebergement booking overlapping this date range with the same tier
                overlapping = db.bookings.find(
                    {
                        "offer_type": "hebergement",
                        "room_tier": body.room_tier,
                        "status": {"$ne": "cancelled"},
                        "date": {"$lt": body.checkout_date},
                        "checkout_date": {"$gt": body.date},
                    },
                    {"_id": 0, "date": 1, "checkout_date": 1, "rooms": 1},
                )
                # Build per-night occupancy
                night_occ: dict = {}
                async for ob in overlapping:
                    a = datetime.strptime(ob["date"], "%Y-%m-%d").date()
                    c = datetime.strptime(ob["checkout_date"], "%Y-%m-%d").date()
                    n = a
                    while n < c:
                        night_occ[n.isoformat()] = night_occ.get(n.isoformat(), 0) + int(ob.get("rooms", 1))
                        n += timedelta(days=1)
                # Check each night of our new booking
                night = booking_date
                while night < checkout:
                    used = night_occ.get(night.isoformat(), 0)
                    if used + body.rooms > tier_inventory:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Plus de chambres '{selected_tier['name_fr']}' disponibles pour la nuit du {night.isoformat()} ({tier_inventory - used} restantes).",
                        )
                    night += timedelta(days=1)
            total = nights * body.rooms * selected_tier["price"]
        else:
            total = nights * (body.adults * offer["price_adult"] + children_paid_n * offer["price_child"])
    else:
        # When a special event uses packages, the per-day base price is
        # IGNORED — each package is billed as a flat forfait, and the base
        # adult/child tariff has no meaning in that flow.
        uses_packages = bool(is_special and body.package_selections)
        if uses_packages:
            total = 0
        else:
            total = body.adults * offer["price_adult"] + children_paid_n * offer["price_child"]
        # Multi-day special event with cumulative pricing: sum each selected
        # date's per-day prices (resolved from the event programme). The
        # primary `body.date` price is already in `total` above (offer was
        # resolved with that date); add the additional dates here. Skipped
        # entirely when packages are used (forfaits-only).
        if is_special and body.multi_day_dates and body.special_event_id and not uses_packages:
            extra_dates = [d for d in body.multi_day_dates if d != body.date]
            if extra_dates:
                ev_doc = await db.special_events.find_one(
                    {"id": body.special_event_id},
                    {"_id": 0, "programme": 1, "price_adult": 1, "price_child": 1, "capacity": 1},
                )
                prog = {p.get("date"): p for p in (ev_doc.get("programme") or []) if p.get("date")} if ev_doc else {}
                ev_cap = int((ev_doc or {}).get("capacity", offer["max_capacity"]))
                for d in extra_dates:
                    item = prog.get(d) or {}
                    pa = int(item.get("price_adult", (ev_doc or {}).get("price_adult", offer["price_adult"])))
                    pc = int(item.get("price_child", (ev_doc or {}).get("price_child", offer["price_child"])))
                    total += body.adults * pa + children_paid_n * pc
                    # Capacity guard for the extra date
                    extra_booked = 0
                    async for b in db.bookings.find(
                        {"offer_type": "special_event", "special_event_id": body.special_event_id,
                         "date": d, "status": {"$ne": "cancelled"}},
                        {"_id": 0, "adults": 1, "children": 1},
                    ):
                        extra_booked += int(b.get("adults", 0)) + int(b.get("children", 0))
                    # Also count current booking's secondary dates from THIS submission
                    if extra_booked + total_guests > ev_cap:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Plus assez de places le {d} ({ev_cap - extra_booked} restantes).",
                        )

    # Premium package add-ons (events). Each selection is a flat forfait — the
    # `pkg.price_adult` field holds the package's flat price (legacy name kept
    # for schema compat). Headcount only fills the package's capacity; price
    # does NOT multiply. We still enforce `max_persons` per package per day AND
    # `stock` (max copies of this package that can be sold per day; 0 = ∞).
    package_lines: List[dict] = []
    if is_special and body.package_selections and body.special_event_id:
        ev_doc = await db.special_events.find_one(
            {"id": body.special_event_id},
            {"_id": 0, "programme": 1, "event_dates": 1, "capacity": 1},
        )
        prog = {p.get("date"): p for p in (ev_doc.get("programme") or []) if p.get("date")} if ev_doc else {}
        # Count requested copies per (date, package_id) within this booking
        req_count: dict = {}
        for sel in body.package_selections:
            if sel.adults <= 0 and sel.children <= 0:
                continue
            req_count[(sel.date, sel.package_id)] = req_count.get((sel.date, sel.package_id), 0) + 1
        for sel in body.package_selections:
            if sel.adults <= 0 and sel.children <= 0:
                continue
            day_item = prog.get(sel.date) or {}
            pkgs = {p.get("id"): p for p in (day_item.get("packages") or []) if p.get("id")}
            pkg = pkgs.get(sel.package_id)
            if not pkg:
                raise HTTPException(
                    status_code=400,
                    detail=f"Package {sel.package_id} introuvable pour le {sel.date}.",
                )
            persons = sel.adults + sel.children
            if persons > int(pkg.get("max_persons", 0)):
                raise HTTPException(
                    status_code=400,
                    detail=f"Le package « {pkg.get('label')} » accepte {pkg.get('max_persons')} personne(s) max.",
                )
            # Stock guard — count already sold copies for this (event, date, pkg).
            stock = int(pkg.get("stock", 0) or 0)
            if stock > 0:
                sold = 0
                async for b in db.bookings.find(
                    {"offer_type": "special_event", "special_event_id": body.special_event_id,
                     "status": {"$ne": "cancelled"}},
                    {"_id": 0, "package_lines": 1},
                ):
                    for line in (b.get("package_lines") or []):
                        if line.get("date") == sel.date and line.get("package_id") == sel.package_id:
                            sold += 1
                requested = req_count.get((sel.date, sel.package_id), 1)
                if sold + requested > stock:
                    remaining = max(0, stock - sold)
                    raise HTTPException(
                        status_code=409,
                        detail=(
                            f"Le forfait « {pkg.get('label')} » est limité à {stock} "
                            f"exemplaire(s) le {sel.date} (reste {remaining})."
                        ),
                    )
            # Flat forfait price — independent of headcount.
            line_amount = int(pkg.get("price_adult", pkg.get("price", 0)) or 0)
            total += line_amount
            package_lines.append({
                "date": sel.date,
                "package_id": sel.package_id,
                "label": pkg.get("label"),
                "adults": sel.adults,
                "children": sel.children,
                "amount": line_amount,
            })

    # Optional private boat charter — adds a flat amount to the total.
    charter_boat: Optional[dict] = None
    charter_amount = 0
    if body.charter_boat_id:
        charter_boat = await db.bateaux.find_one(
            {"id": body.charter_boat_id, "status": "actif"},
            {"_id": 0, "id": 1, "name": 1, "capacity": 1, "charter_price": 1},
        )
        if not charter_boat or int(charter_boat.get("charter_price", 0)) <= 0:
            raise HTTPException(status_code=400, detail="Bateau de privatisation indisponible.")
        charter_amount = int(charter_boat["charter_price"])
        total += charter_amount

    # Beach Club VIP spaces (numbered transats / balinés). Unique per date.
    vip_spaces_resolved: List[dict] = []
    vip_spaces_amount = 0
    if body.vip_space_ids:
        from routers.vip_spaces import validate_and_resolve_vip_spaces  # local import
        vip_spaces_resolved, vip_spaces_amount = await validate_and_resolve_vip_spaces(
            db,
            offer_type=body.offer_type,
            date_iso=body.date,
            vip_space_ids=body.vip_space_ids,
        )
        total += vip_spaces_amount

    # Optional room add-on (Hébergement upsell on any non-hebergement booking).
    # Validates the tier, computes nights, adds tier*nights*rooms to the total
    # and persists structured metadata for the staff dashboards.
    room_addon_doc: Optional[dict] = None
    room_addon_amount = 0
    if body.room_addon_tier:
        if body.offer_type == "hebergement":
            raise HTTPException(
                status_code=400,
                detail="L'option chambre n'est pas applicable au pôle Hébergement.",
            )
        heb_offer = OFFERS["hebergement"]
        tier = next((t for t in heb_offer["room_tiers"] if t["id"] == body.room_addon_tier), None)
        if not tier:
            raise HTTPException(status_code=400, detail="Catégorie de chambre inconnue.")
        addon_rooms = max(1, int(body.room_addon_rooms or 1))
        # Default check-in = booking date ; check-out = next day.
        ci = (body.room_addon_checkin or body.date).strip()
        try:
            ci_date = datetime.strptime(ci, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Date d'arrivée invalide.")
        if body.room_addon_checkout:
            try:
                co_date = datetime.strptime(body.room_addon_checkout, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Date de départ invalide.")
        else:
            co_date = ci_date + timedelta(days=1)
        addon_nights = (co_date - ci_date).days
        if addon_nights < 1:
            raise HTTPException(status_code=400, detail="La chambre nécessite au moins une nuit.")
        room_addon_amount = int(tier["price"]) * addon_nights * addon_rooms
        total += room_addon_amount
        room_addon_doc = {
            "tier_id": tier["id"],
            "tier_name": tier["name_fr"],
            "tier_price": int(tier["price"]),
            "checkin_date": ci_date.isoformat(),
            "checkout_date": co_date.isoformat(),
            "nights": addon_nights,
            "rooms": addon_rooms,
            "amount": room_addon_amount,
        }
    raw_participants = [
        {
            "name": p.name.strip(),
            "surname": p.surname.strip(),
            "email": (p.email or "").lower(),
            "phone": (p.phone or "").strip(),
            "nationality": p.nationality.strip(),
            "kind": p.kind,
        }
        for p in body.participants
    ]
    # Booker-only flow: when a single participant is sent, materialise the
    # full guest list by duplicating the booker into N adult tickets and M
    # child tickets. This keeps the ticket-generation code (which iterates
    # `participants`) unchanged while letting the customer fill in just one
    # form on the public site.
    if len(raw_participants) == 1 and (body.adults + body.children) > 1:
        booker = raw_participants[0]
        participants_docs = []
        for _ in range(body.adults):
            participants_docs.append({**booker, "kind": "adult"})
        for _ in range(body.children):
            # Children inherit the booker's identity for ticket personalisation
            # but their email/phone are nulled to discourage future direct
            # contact (the booker remains the single point of contact).
            participants_docs.append({**booker, "kind": "child", "email": "", "phone": ""})
    else:
        participants_docs = raw_participants
    # Primary contact = first adult (or first participant if none)
    primary = next((p for p in participants_docs if p["kind"] == "adult"), participants_docs[0])
    doc = {
        "id": bid,
        "reference_token": reference_token,
        "offer_type": body.offer_type,
        "offer_name": offer["name_fr"],
        "pole": _pole_for_offer(body.offer_type),
        "special_event_id": offer["event_id"] if is_special else None,
        # Multi-day cumulative bookings keep `date` = first day (sort key for
        # dashboards) and list every selected day in `multi_day_dates`.
        "multi_day_dates": (
            sorted(set(body.multi_day_dates)) if (is_special and body.multi_day_dates and len(body.multi_day_dates) > 1) else None
        ),
        "date": body.date,
        "checkout_date": checkout_iso,
        "nights": nights,
        "room_tier": selected_tier["id"] if selected_tier else None,
        "room_tier_name": selected_tier["name_fr"] if selected_tier else None,
        "room_tier_price": selected_tier["price"] if selected_tier else None,
        "rooms": body.rooms,
        "adults": body.adults,
        "children": body.children,
        # iter-30: explicit split — children_paid (6–12, billed) vs children_free (<6, gratuit).
        "children_paid": children_paid_n,
        "children_free": children_free_n,
        # iter-30: 5-digit code shared with accompanying adults.
        "booking_code": booking_code,
        "companion_slots_total": max(0, int(body.adults) - 1),
        "companion_slots_used": 0,
        "total_amount": total,
        "status": "pending",
        "qr_codes": [],
        "participants": participants_docs,
        "boat_time": body.boat_time,
        "return_boat_time": return_boat_time,
        "phone": primary["phone"],
        "email": primary["email"],
        "special_requests": body.special_requests or "",
        # Boat charter
        "charter_boat_id": charter_boat["id"] if charter_boat else None,
        "charter_boat_name": charter_boat["name"] if charter_boat else None,
        "charter_amount": charter_amount,
        # Beach Club numbered VIP spaces (transats / balinés)
        "vip_space_ids": [v["id"] for v in vip_spaces_resolved] or None,
        "vip_spaces": vip_spaces_resolved or None,
        "vip_spaces_amount": vip_spaces_amount,
        # Optional room add-on (Hébergement upsell)
        "room_addon": room_addon_doc,
        "room_addon_amount": room_addon_amount,
        # Premium event packages picked by the customer (flat list).
        "package_lines": package_lines or None,
        "created_at": now_iso(),
        "paid_at": None,
    }
    await db.bookings.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.post("/bookings/{booking_id}/pay")
async def pay_booking(booking_id: str, body: PayBooking):
    """FINEO placeholder - validates reference token, generates one QR per ADULT.
    Children are counted on the booker's ticket (no longer get a dedicated QR).

    Cash payments produce a TEMPORARY "EN ATTENTE" receipt — the booking stays
    in `pending_cash_payment` state until a staff member explicitly confirms
    the cash collection via /staff/bookings/{id}/confirm-cash-payment. The
    customer then receives a second email with the final styled QR ticket.
    """
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    if booking.get("reference_token") != body.reference_token:
        raise HTTPException(status_code=403, detail="Invalid reference token")
    if booking["status"] != "pending":
        raise HTTPException(status_code=400, detail="Booking already processed")

    if booking["offer_type"] == "special_event":
        offer = await _resolve_special_event_offer(booking.get("special_event_id") or "", booking.get("date"))
    else:
        offer = OFFERS[booking["offer_type"]]
    participants = booking.get("participants", [])
    # Only adults receive a ticket. Children count is appended to the booker.
    adult_participants = [p for p in participants if p.get("kind", "adult") == "adult"]
    children_count = int(booking.get("children", 0))

    # Compute paid amount (full vs deposit). Deposit is only valid for overnight offers.
    total_amount = int(booking.get("total_amount", 0))
    deposit_pct = None
    if body.payment_method == "deposit":
        if not offer.get("is_overnight"):
            raise HTTPException(status_code=400, detail="Deposit only available for overnight stays")
        if body.deposit_pct not in (10, 30, 70):
            raise HTTPException(status_code=400, detail="deposit_pct must be 10, 30 or 70")
        deposit_pct = int(body.deposit_pct)
        paid_amount = int(round(total_amount * deposit_pct / 100))
    else:
        paid_amount = total_amount
    balance_due = total_amount - paid_amount

    # Card / mobile-money / deposit / fineo → styled gold QR + immediate confirmation.
    # Cash payments → cream "EN ATTENTE" receipt without QR, booking stays
    # pending_cash_payment until staff confirms.
    is_cash_pending = body.payment_method == "cash"
    styled_qr = not is_cash_pending
    base_payload = {
        "v": 1,
        "issuer": "Boulay Beach Resort",
        "booking_id": booking_id,
        "booking_ref": booking_id[:8].upper(),
        "offer_id": booking["offer_type"],
        "offer_name": offer["name_fr"],
        "schedule": offer["schedule_fr"],
        "date": booking["date"],
        "boat_time": booking.get("boat_time", ""),
        "return_boat_time": booking.get("return_boat_time") or "",
        "adults": int(booking.get("adults", 0)),
        "children": int(booking.get("children", 0)),
        "total_amount_fcfa": int(booking["total_amount"]),
        "paid_amount_fcfa": int(paid_amount),
        "balance_due_fcfa": int(balance_due),
        "deposit_pct": deposit_pct,
        "phone": booking["phone"],
        "email": booking["email"],
        "special_requests": booking.get("special_requests", "") or "",
    }

    qr_codes = []
    # Multi-day bookings → ONE *passport* ticket per adult that covers ALL dates
    #   with a single QR code (scannable on each valid date, max 2 scans/date).
    # Single-day bookings → unchanged: one ticket per adult.
    ticket_dates: list = list(booking.get("multi_day_dates") or []) or [booking["date"]]
    is_passport = len(ticket_dates) > 1
    primary_date = ticket_dates[0]

    # iter-30: only the BOOKER gets a QR ticket at payment time. Other adults
    # register themselves via /companion/{booking_code} and receive their own
    # QR + ticket email at that moment. Children remain attached to the
    # booker's pass and never get a dedicated QR.
    children_paid_n = int(booking.get("children_paid") or 0)
    children_free_n = int(booking.get("children_free") or 0)
    booker_adults_only = adult_participants[:1]
    for adult_i, p in enumerate(booker_adults_only, start=1):
        token = uuid.uuid4().hex
        is_booker = adult_i == 1
        # Booker gets a rich label that mentions the children he/she is carrying.
        passport_suffix = " · Passeport multi-dates" if is_passport else ""
        if is_booker and children_count > 0:
            label_fr = f"Réservant · +{children_count} enfant{'s' if children_count > 1 else ''}{passport_suffix}"
            label_en = f"Booker · +{children_count} child{'ren' if children_count > 1 else ''}{passport_suffix}"
        elif is_booker:
            label_fr = f"Réservant{passport_suffix}"
            label_en = f"Booker{passport_suffix}"
        else:
            label_fr = f"Adulte #{adult_i}{passport_suffix}"
            label_en = f"Adult #{adult_i}{passport_suffix}"
        guest_payload = {
            **base_payload,
            "date": primary_date,
            "valid_dates": ticket_dates,
            "is_passport": is_passport,
            "guest_kind": "adult",
            "guest_index": adult_i,
            "guest_label": label_fr,
            "guest_name": p["name"],
            "guest_surname": p["surname"],
            "guest_email": p.get("email", "") or booking.get("email", ""),
            "guest_phone": p.get("phone", "") or booking.get("phone", ""),
            "guest_nationality": p["nationality"],
            "guest_token": token,
            "children_attached": children_count if is_booker else 0,
        }
        payload_str = json.dumps(guest_payload, ensure_ascii=False, separators=(",", ":"))
        compact_qr = json.dumps(
            {"type": "ticket", "token": token, "ref": booking_id[:8].upper()},
            ensure_ascii=False, separators=(",", ":"),
        )
        token_short = token[:10].upper()
        entry = {
            "label_fr": label_fr,
            "label_en": label_en,
            "kind": "adult",
            "event_date": primary_date,
            "valid_dates": ticket_dates,
            "is_passport": is_passport,
            "guest_name": p["name"],
            "guest_surname": p["surname"],
            "guest_email": p.get("email", "") or booking.get("email", ""),
            "guest_phone": p.get("phone", "") or booking.get("phone", ""),
            "guest_nationality": p["nationality"],
            "qr_token": token,
            "qr_payload": payload_str,
            "qr_code": make_qr(compact_qr, styled=styled_qr),
            "children_attached": children_count if is_booker else 0,
            # iter-30: explicit composition shown on scan & manifest.
            "composition": {
                "adults": 1,
                "children_paid": children_paid_n if is_booker else 0,
                "children_free": children_free_n if is_booker else 0,
            } if is_booker else None,
        }
        if styled_qr:
            entry["ticket_image"] = make_ticket_image(
                offer_id=booking["offer_type"],
                offer_name=offer["name_fr"],
                date_iso=primary_date,
                boat_time=booking.get("boat_time", ""),
                owner_name=f"{p['name']} {p['surname']}",
                qr_payload=compact_qr,
                ref_code=token_short,
                lang="fr",
                hero_url=offer.get("image_url") or None,
                dates_list=ticket_dates if is_passport else None,
                composition=({"adults": 1, "children_paid": children_paid_n, "children_free": children_free_n} if is_booker else None),
                party_size=(1 + children_count) if is_booker else 1,
            )
        else:
            entry["ticket_image"] = make_cash_receipt_image(
                offer_id=booking["offer_type"],
                offer_name=offer["name_fr"],
                date_iso=primary_date,
                boat_time=booking.get("boat_time", ""),
                owner_name=f"{p['name']} {p['surname']}",
                ref_code=token_short,
                lang="fr",
                hero_url=offer.get("image_url") or None,
                dates_list=ticket_dates if is_passport else None,
            )
        qr_codes.append(entry)

    paid_at = now_iso()

    # ---------- Wallet creation (activity payment QR) ----------
    primary = adult_participants[0] if adult_participants else (participants[0] if participants else {})
    owner_name = f"{primary.get('name','')} {primary.get('surname','')}".strip() or "Invité"
    wallet_token = str(uuid.uuid4())
    booking_ref_short = booking_id[:8].upper()
    wallet_doc = {
        "id": str(uuid.uuid4()),
        "token": wallet_token,
        "booking_id": booking_id,
        "booking_ref": booking_ref_short,
        "owner_name": owner_name,
        "phone": booking.get("phone", ""),
        "email": booking.get("email", ""),
        "transactions": [],
        "total_charged": 0,
        "status": "open",
        "created_at": paid_at,
    }
    await db.wallets.insert_one(dict(wallet_doc))
    wallet_doc.pop("_id", None)
    wallet_qr = {
        "wallet_token": wallet_token,
        "qr_code": make_qr(json.dumps({"type": "wallet", "token": wallet_token, "booking_ref": booking_ref_short}), styled=True),
        "ticket_image": make_wallet_image(
            owner_name=owner_name,
            wallet_token=wallet_token,
            booking_ref=booking_ref_short,
            lang="fr",
        ),
    }

    # Booking status branches: cash → pending validation, others → immediately confirmed.
    new_status = "pending_cash_payment" if is_cash_pending else "confirmed"
    update_doc = {
        "status": new_status,
        "qr_codes": qr_codes,
        "wallet_qr": wallet_qr,
        "wallet_token": wallet_token,
        "payment_method": body.payment_method,
        "deposit_pct": deposit_pct,
    }
    if is_cash_pending:
        # Cash collection not yet confirmed by staff. paid_at + paid_amount stay null.
        update_doc["paid_at"] = None
        update_doc["paid_amount"] = 0
        update_doc["balance_due"] = total_amount
        update_doc["cash_temp_issued_at"] = paid_at
    else:
        update_doc["paid_at"] = paid_at
        update_doc["paid_amount"] = int(paid_amount)
        update_doc["balance_due"] = int(balance_due)
    await db.bookings.update_one({"id": booking_id}, {"$set": update_doc})
    booking.update(update_doc)

    # Fiscal receipt is only emitted when money is actually collected (not for
    # cash-pending state — the staff confirmation endpoint emits it later).
    if not is_cash_pending and int(paid_amount) > 0:
        try:
            label = (
                f"Acompte {deposit_pct}% — {offer['name_fr']} ({booking['date']})"
                if body.payment_method == "deposit" and deposit_pct
                else f"{offer['name_fr']} — {booking['date']}"
            )
            line = {
                "description": label,
                "quantity": 1,
                "unit_price": int(paid_amount),
                "total": int(paid_amount),
            }
            await _create_receipt(
                source="booking",
                source_id=booking_id,
                customer_name=f"{primary.get('surname','').strip()} {primary.get('name','').strip()}".strip() or "—",
                customer_email=primary.get("email") or booking.get("email", ""),
                customer_phone=primary.get("phone") or booking.get("phone", ""),
                lines=[line],
                payment_method=body.payment_method,
                issued_by="public",
                issued_by_role="public",
                metadata={"offer_type": booking["offer_type"], "deposit_pct": deposit_pct},
            )
        except Exception as ex:
            logging.exception("Failed to create booking receipt: %s", ex)

    # Outbound Twilio notification — never blocks. The QR ticket PNG is
    # embedded as a data URI so WhatsApp can render it as a media message.
    # Skipped for cash-pending: client gets the temporary receipt via email only.
    if not is_cash_pending:
        try:
            first_qr = (qr_codes or [{}])[0]
            qr_url = first_qr.get("ticket_image", "")
            if qr_url and not qr_url.startswith(("http://", "https://")):
                qr_url = f"{FINEO_PUBLIC_BASE_URL}/api/bookings/{booking_id}/ticket.png?ref={booking['reference_token']}"
            await twilio_service.notify_booking_paid(db, booking, qr_image_url=qr_url)
        except Exception as ex:
            logging.warning("Twilio booking_paid notification failed: %s", ex)

    # SendGrid email — sent in parallel with WhatsApp/SMS.
    try:
        await _send_booking_confirmation_email(booking, temporary=is_cash_pending)
    except Exception as ex:
        logging.warning("SendGrid booking_paid email failed: %s", ex)
    return booking


@api.get("/bookings/{booking_id}")
async def get_booking(booking_id: str, ref: str):
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    if booking.get("reference_token") != ref:
        raise HTTPException(status_code=403, detail="Invalid reference token")
    return booking


@api.get("/bookings/{booking_id}/ticket.png")
async def get_booking_ticket_image(booking_id: str, ref: str):
    """Public PNG of the first ticket — used by Twilio WhatsApp media. If the
    booking was settled via FineoPay before the QR-generation fix shipped, we
    retro-generate the ticket on-the-fly so the customer can still print it."""
    from base64 import b64decode
    from fastapi.responses import Response
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking or booking.get("reference_token") != ref:
        raise HTTPException(status_code=404, detail="Not found")
    qrs = booking.get("qr_codes") or []
    img_data = (qrs[0] if qrs else {}).get("ticket_image", "")
    # Retro-fix: paid booking without QR (FineoPay legacy settle path).
    if not img_data and booking.get("paid_at") and booking.get("status") != "cancelled":
        try:
            pay_body = PayBooking(
                reference_token=booking["reference_token"],
                payment_method="fineo",
                deposit_pct=None,
            )
            # Mark status as pending so pay_booking accepts to regenerate
            await db.bookings.update_one({"id": booking_id}, {"$set": {"status": "pending"}})
            await pay_booking(booking_id, pay_body)
            booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
            qrs = (booking or {}).get("qr_codes") or []
            img_data = (qrs[0] if qrs else {}).get("ticket_image", "")
            logging.info("Ticket retro-generated for booking %s", booking_id)
        except Exception as ex:
            logging.warning("Retro QR generation failed for %s: %s", booking_id, ex)
    if not img_data:
        raise HTTPException(status_code=404, detail="No ticket")
    if img_data.startswith("data:"):
        img_data = img_data.split(",", 1)[1]
    return Response(content=b64decode(img_data), media_type="image/png")


@api.get("/bookings/{booking_id}/reservation.pdf")
async def get_booking_reservation_pdf(booking_id: str, ref: str):
    """Public PDF of the full booking confirmation (QR + all details).

    Token-protected via the booking's ``reference_token`` so unauthenticated
    customers can download their own ticket from the confirmation page after
    payment, without exposing data to the rest of the internet.
    """
    from fastapi.responses import Response
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking or booking.get("reference_token") != ref:
        raise HTTPException(status_code=404, detail="Not found")
    pdf_bytes = await _build_booking_confirmation_pdf(booking)
    if not pdf_bytes:
        raise HTTPException(status_code=500, detail="Could not generate PDF")
    ref_short = booking_id[:8].upper()
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="BBR-reservation-{ref_short}.pdf"'},
    )


# =================================================================
# SPECIAL EVENTS — bookable themed event with staff CRUD
# =================================================================
def _public_event(ev: dict) -> dict:
    """Strip internal fields before returning an event to public consumers."""
    today = datetime.now(timezone.utc).date().isoformat()
    kind = ev.get("event_kind") or "single_day"
    out = {
        "id": ev["id"],
        "title": ev.get("title", ""),
        "subtitle": ev.get("subtitle", ""),
        "description": ev.get("description", ""),
        "image_url": ev.get("image_url", ""),
        "event_dates": ev.get("event_dates") or [],
        "boat_times": ev.get("boat_times") or [],
        "return_boat_times": ev.get("return_boat_times") or [],
        "price_adult": int(ev.get("price_adult", 0)),
        "price_child": int(ev.get("price_child", 0)),
        "capacity": int(ev.get("capacity", 0)),
        "active_from": ev.get("active_from"),
        "active_to": ev.get("active_to"),
        "cta_label": ev.get("cta_label") or "Réserver ma place",
        "is_featured": bool(ev.get("is_featured")),
        "status": ev.get("status", "draft"),
        "event_kind": kind,
        "start_date": ev.get("start_date"),
        "end_date": ev.get("end_date"),
        "programme": ev.get("programme") or [],
        "today": today,
    }
    # For multi-day events, synthesize event_dates from the programme if not
    # already set, so the booking tunnel keeps working transparently.
    if kind == "multi_day" and not out["event_dates"] and out["programme"]:
        out["event_dates"] = sorted({p.get("date") for p in out["programme"] if p.get("date")})
    return out


def _event_is_currently_active(ev: dict, today: str) -> bool:
    if ev.get("status") != "published":
        return False
    if ev.get("active_from") and today < ev["active_from"]:
        return False
    if ev.get("active_to") and today > ev["active_to"]:
        return False
    return True


@api.get("/special-events/featured")
async def get_featured_special_event():
    """Public — returns the single currently-featured, published, in-window event.
    Used by the booking tunnel to inject the event card alongside the static offers.
    Returns ``{"event": null}`` when no event is featured/eligible.
    """
    today = datetime.now(timezone.utc).date().isoformat()
    ev = await db.special_events.find_one(
        {"is_featured": True, "status": "published"},
        {"_id": 0},
    )
    if not ev or not _event_is_currently_active(ev, today):
        return {"event": None}
    # Filter out past event_dates so the public UI only shows upcoming slots
    upcoming_dates = [d for d in (ev.get("event_dates") or []) if d >= today]
    out = _public_event(ev)
    out["event_dates"] = upcoming_dates
    # Compute remaining seats across all upcoming event dates (best-effort total)
    booked_cursor = db.bookings.find(
        {
            "offer_type": "special_event",
            "special_event_id": ev["id"],
            "status": {"$ne": "cancelled"},
        },
        {"_id": 0, "adults": 1, "children": 1, "date": 1},
    )
    booked_per_date: dict = {}
    async for b in booked_cursor:
        d = b.get("date") or ""
        booked_per_date[d] = booked_per_date.get(d, 0) + int(b.get("adults", 0)) + int(b.get("children", 0))
    out["seats_per_date"] = {d: max(0, int(ev.get("capacity", 0)) - booked_per_date.get(d, 0)) for d in upcoming_dates}
    return {"event": out}


# ============== Exclusivity feature (homepage spotlight) ==============
async def _resolve_exclusivity_link(doc: dict) -> dict:
    """Compute the public-facing href of the exclusivity card based on link_type.
    Always returns a `href` (possibly None) without leaking internal data.
    """
    if not doc:
        return {"href": None, "resolved": None}
    lt = doc.get("link_type", "custom")
    target = doc.get("link_target_id") or ""
    href = None
    resolved = None
    if lt == "special_event" and target:
        ev = await db.special_events.find_one(
            {"id": target, "status": "published"}, {"_id": 0, "id": 1, "title": 1, "image_url": 1},
        )
        if ev:
            href = f"/event/{ev['id']}"
            resolved = {"id": ev["id"], "title": ev.get("title"), "image_url": ev.get("image_url")}
    elif lt == "offer" and target:
        if target in OFFERS:
            href = f"/booking/{target}"
            resolved = {"id": target, "title": OFFERS[target]["name_fr"]}
    elif lt == "activity" and target:
        href = f"/accueil/paiement?activity={target}"
        resolved = {"id": target}
    elif lt == "custom":
        href = (doc.get("link_url") or "").strip() or None
    return {"href": href, "resolved": resolved}


@api.get("/exclusivity")
async def get_public_exclusivity():
    """Public — returns the currently-enabled exclusivity card config (or empty).

    Used by the landing page to inject the card before the Beach Club pôle.
    """
    doc = await db.config.find_one({"_id": "exclusivity_feature"}) or {}
    if not doc.get("enabled"):
        return {"enabled": False}
    link = await _resolve_exclusivity_link(doc)
    return {
        "enabled": True,
        "title": doc.get("title") or "",
        "subtitle": doc.get("subtitle") or "",
        "description": doc.get("description") or "",
        "image_url": doc.get("image_url") or "",
        "cta_label": doc.get("cta_label") or "Découvrir",
        "link_type": doc.get("link_type") or "custom",
        "link_target_id": doc.get("link_target_id"),
        **link,
    }


@api.get("/staff/exclusivity")
async def staff_get_exclusivity(staff=Depends(get_current_staff)):
    """Staff read of the full exclusivity config (incl. when disabled)."""
    await _require_role(staff, ["admin", "manager"])
    doc = await db.config.find_one({"_id": "exclusivity_feature"}) or {}
    doc.pop("_id", None)
    # Surface the resolved link target name so the admin UI can show a preview
    link = await _resolve_exclusivity_link(doc)
    return {
        "enabled": bool(doc.get("enabled")),
        "title": doc.get("title") or "",
        "subtitle": doc.get("subtitle") or "",
        "description": doc.get("description") or "",
        "image_url": doc.get("image_url") or "",
        "cta_label": doc.get("cta_label") or "Découvrir",
        "link_type": doc.get("link_type") or "special_event",
        "link_target_id": doc.get("link_target_id") or "",
        "link_url": doc.get("link_url") or "",
        **link,
    }


@api.put("/staff/exclusivity")
async def staff_update_exclusivity(body: ExclusivityFeature, staff=Depends(get_current_staff)):
    """Admin upsert of the exclusivity card configuration."""
    await _require_role(staff, ["admin", "manager"])
    # When enabling, ensure the link target resolves (sanity check) so the
    # public card never points to a dangling resource.
    if body.enabled:
        if body.link_type == "special_event" and body.link_target_id:
            ev = await db.special_events.find_one(
                {"id": body.link_target_id, "status": "published"}, {"_id": 0, "id": 1},
            )
            if not ev:
                raise HTTPException(
                    status_code=400,
                    detail="L'événement spécial lié est introuvable ou non publié.",
                )
        elif body.link_type == "offer":
            if not body.link_target_id or body.link_target_id not in OFFERS:
                raise HTTPException(status_code=400, detail="Offre liée invalide.")
        elif body.link_type == "custom":
            if not (body.link_url or "").strip():
                raise HTTPException(status_code=400, detail="URL personnalisée requise.")
    payload = body.model_dump()
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    payload["updated_by"] = (staff.get("email") if isinstance(staff, dict) else None)
    await db.config.update_one(
        {"_id": "exclusivity_feature"},
        {"$set": payload},
        upsert=True,
    )
    return {"ok": True}


# Stub — kept here so the next public endpoint definition keeps its position:
async def _legacy_get_featured_event_passthrough():
    pass



@api.get("/special-events/{event_id}")
async def get_special_event(event_id: str):
    """Public — fetch a single published event by ID (used by the booking tunnel)."""
    ev = await db.special_events.find_one({"id": event_id}, {"_id": 0})
    if not ev:
        raise HTTPException(status_code=404, detail="Événement introuvable")
    today = datetime.now(timezone.utc).date().isoformat()
    if not _event_is_currently_active(ev, today):
        raise HTTPException(status_code=400, detail="Cet événement n'est pas disponible à la réservation")
    out = _public_event(ev)
    # For multi_day events, _public_event() already synthesised event_dates
    # from the programme. Filter to future dates only.
    out["event_dates"] = [d for d in (out.get("event_dates") or []) if d >= today]
    # Per-date remaining seats
    booked_cursor = db.bookings.find(
        {"offer_type": "special_event", "special_event_id": ev["id"], "status": {"$ne": "cancelled"}},
        {"_id": 0, "adults": 1, "children": 1, "date": 1},
    )
    booked_per_date: dict = {}
    async for b in booked_cursor:
        d = b.get("date") or ""
        booked_per_date[d] = booked_per_date.get(d, 0) + int(b.get("adults", 0)) + int(b.get("children", 0))
    out["seats_per_date"] = {d: max(0, int(ev.get("capacity", 0)) - booked_per_date.get(d, 0)) for d in out["event_dates"]}

    # Per-package remaining stock — only meaningful when stock > 0 on the package.
    pkg_sold: dict = {}
    async for b in db.bookings.find(
        {"offer_type": "special_event", "special_event_id": ev["id"], "status": {"$ne": "cancelled"}},
        {"_id": 0, "package_lines": 1},
    ):
        for line in (b.get("package_lines") or []):
            key = (line.get("date") or "", line.get("package_id") or "")
            pkg_sold[key] = pkg_sold.get(key, 0) + 1
    # Annotate each programme item's packages with `sold` / `remaining`.
    for p in (out.get("programme") or []):
        for pkg in (p.get("packages") or []):
            stock = int(pkg.get("stock", 0) or 0)
            sold = pkg_sold.get((p.get("date"), pkg.get("id")), 0)
            pkg["sold"] = sold
            pkg["remaining"] = (stock - sold) if stock > 0 else None
    return {"event": out}


@api.get("/staff/special-events")
async def staff_list_special_events(staff=Depends(get_current_staff)):
    """List all special events, including drafts and archived (for the back-office)."""
    items = await db.special_events.find({}, {"_id": 0}).sort("created_at", -1).to_list(length=500)
    today = datetime.now(timezone.utc).date().isoformat()
    # Hydrate with booked seat counts (sum across all dates)
    by_id: dict = {it["id"]: it for it in items}
    if by_id:
        agg = db.bookings.aggregate([
            {"$match": {
                "offer_type": "special_event",
                "special_event_id": {"$in": list(by_id.keys())},
                "status": {"$ne": "cancelled"},
            }},
            {"$group": {
                "_id": "$special_event_id",
                "guests": {"$sum": {"$add": [{"$ifNull": ["$adults", 0]}, {"$ifNull": ["$children", 0]}]}},
                "bookings": {"$sum": 1},
            }},
        ])
        async for row in agg:
            eid = row["_id"]
            if eid in by_id:
                by_id[eid]["booked_guests"] = int(row.get("guests", 0))
                by_id[eid]["booked_bookings"] = int(row.get("bookings", 0))
    for it in items:
        it.setdefault("booked_guests", 0)
        it.setdefault("booked_bookings", 0)
        it["is_active"] = _event_is_currently_active(it, today)
    return {"items": items}


@api.get("/staff/special-events/{event_id}")
async def staff_get_special_event(event_id: str, staff=Depends(get_current_staff)):
    ev = await db.special_events.find_one({"id": event_id}, {"_id": 0})
    if not ev:
        raise HTTPException(status_code=404, detail="Special event not found")
    return ev


def _validate_event_kind(payload: dict) -> None:
    """Cross-field validation for special events.

    • `single_day` requires at least one `event_dates` entry.
    • `multi_day` requires `start_date` ≤ `end_date` and at least one
      `programme` item with `start_date ≤ date ≤ end_date`.
    """
    kind = payload.get("event_kind") or "single_day"
    if kind == "multi_day":
        sd, ed = payload.get("start_date"), payload.get("end_date")
        if not sd or not ed:
            raise HTTPException(status_code=400, detail="start_date et end_date sont requis pour un événement multi-jours.")
        if sd > ed:
            raise HTTPException(status_code=400, detail="start_date doit être antérieur ou égal à end_date.")
        prog = payload.get("programme") or []
        if not prog:
            raise HTTPException(status_code=400, detail="Au moins une entrée de programme est requise.")
        for p in prog:
            d = p.get("date") if isinstance(p, dict) else p.date
            if not d or d < sd or d > ed:
                raise HTTPException(
                    status_code=400,
                    detail=f"La date '{d}' du programme dépasse l'intervalle de l'événement.",
                )


@api.post("/staff/special-events")
async def staff_create_special_event(body: SpecialEventCreate, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    payload = body.model_dump()
    _validate_event_kind(payload)
    eid = str(uuid.uuid4())
    doc = {
        "id": eid,
        **payload,
        "is_featured": False,
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "created_by_email": staff.get("email"),
    }
    await db.special_events.insert_one(dict(doc))
    doc.pop("_id", None)
    return doc


@api.patch("/staff/special-events/{event_id}")
async def staff_update_special_event(
    event_id: str,
    body: SpecialEventUpdate,
    staff=Depends(get_current_staff),
):
    await _require_role(staff, ["manager", "admin"])
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    # Cross-field validation needs the full picture (merging with stored doc),
    # so we resolve the final shape before persisting.
    if any(k in update for k in ("event_kind", "start_date", "end_date", "programme")):
        existing = await db.special_events.find_one({"id": event_id}, {"_id": 0}) or {}
        merged = {**existing, **update}
        _validate_event_kind(merged)
    update["updated_at"] = now_iso()
    res = await db.special_events.update_one({"id": event_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Special event not found")
    ev = await db.special_events.find_one({"id": event_id}, {"_id": 0})
    return ev


@api.post("/staff/special-events/{event_id}/feature")
async def staff_feature_special_event(event_id: str, staff=Depends(get_current_staff)):
    """Mark the given event as the single featured one (unsets every other event)."""
    await _require_role(staff, ["manager", "admin"])
    ev = await db.special_events.find_one({"id": event_id}, {"_id": 0})
    if not ev:
        raise HTTPException(status_code=404, detail="Special event not found")
    await db.special_events.update_many(
        {"id": {"$ne": event_id}},
        {"$set": {"is_featured": False, "updated_at": now_iso()}},
    )
    await db.special_events.update_one(
        {"id": event_id},
        {"$set": {"is_featured": True, "updated_at": now_iso()}},
    )
    return {"ok": True, "is_featured": True}


@api.post("/staff/special-events/{event_id}/unfeature")
async def staff_unfeature_special_event(event_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    res = await db.special_events.update_one(
        {"id": event_id},
        {"$set": {"is_featured": False, "updated_at": now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Special event not found")
    return {"ok": True, "is_featured": False}


@api.post("/staff/special-events/{event_id}/duplicate")
async def staff_duplicate_special_event(event_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    ev = await db.special_events.find_one({"id": event_id}, {"_id": 0})
    if not ev:
        raise HTTPException(status_code=404, detail="Special event not found")
    clone = dict(ev)
    clone["id"] = str(uuid.uuid4())
    clone["title"] = f"{clone.get('title', '')} (copie)".strip()
    clone["is_featured"] = False
    clone["status"] = "draft"
    clone["created_at"] = now_iso()
    clone["updated_at"] = now_iso()
    clone["created_by_email"] = staff.get("email")
    await db.special_events.insert_one(dict(clone))
    clone.pop("_id", None)
    return clone


@api.delete("/staff/special-events/{event_id}")
async def staff_delete_special_event(event_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    # Guard against deletion when bookings still reference the event
    used = await db.bookings.count_documents({
        "offer_type": "special_event",
        "special_event_id": event_id,
        "status": {"$ne": "cancelled"},
    })
    if used > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cet événement a {used} réservation(s) active(s). Archivez-le plutôt que de le supprimer.",
        )
    res = await db.special_events.delete_one({"id": event_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Special event not found")
    return {"ok": True}


# =================================================================
# ADMIN — Wipe test data (DANGER)
# =================================================================
WIPE_SECTIONS = {
    "bookings": {
        "label": "Réservations & paiements",
        "collections": ["bookings", "fineo_payments", "receipts", "wallets", "deposits", "payments"],
    },
    "notifications": {
        "label": "Notifications (emails + SMS/WhatsApp)",
        "collections": ["email_messages", "twilio_messages"],
    },
    "campaigns": {
        "label": "Campagnes marketing",
        "collections": ["campaigns", "campaign_recipients", "email_campaigns"],
    },
    "feedback": {
        "label": "Retours d'expérience",
        "collections": ["experience_feedback"],
    },
    "loisirs": {
        "label": "Réservations loisirs / événements",
        "collections": ["loisirs_bookings", "event_bookings", "event_requests"],
    },
    "traversees": {
        "label": "Traversées & manifests passagers",
        "collections": ["traversees", "traversee_passengers"],
    },
    "scans": {
        "label": "Scans QR (historique)",
        "collections": ["qr_scans"],
    },
    "clients": {
        "label": "Fiches clients agrégées",
        "collections": ["clients"],
    },
}


class WipeTestDataBody(BaseModel):
    confirmation: str  # must be exactly "VIDER LES DONNEES BBR"
    sections: Optional[List[str]] = None  # None or ["all"] → wipe everything


@api.get("/staff/admin/wipe-sections")
async def admin_get_wipe_sections(staff=Depends(get_current_staff)):
    """Return the list of wipe-able sections with their current document counts.
    Used by the Maintenance tab in StaffConfig."""
    await _require_role(staff, ["admin"])
    existing = await db.list_collection_names()
    out = []
    for key, spec in WIPE_SECTIONS.items():
        total = 0
        for c in spec["collections"]:
            if c in existing:
                total += await db[c].count_documents({})
        out.append({
            "key": key,
            "label": spec["label"],
            "collections": spec["collections"],
            "count": total,
        })
    return {"sections": out}


@api.post("/staff/admin/migrate-data-urls")
async def admin_migrate_data_urls(staff=Depends(get_current_staff)):
    """One-shot migration: any ``data:image/...`` URL stored in special_events
    or offer_overrides is converted to a public ``/api/media/{id}`` URL so
    emails can render the image. Idempotent (re-running is safe)."""
    await _require_role(staff, ["admin"])
    from routers.media import ensure_public_url

    base = FINEO_PUBLIC_BASE_URL.rstrip("/")
    migrated = {"special_events": 0, "offer_overrides": 0}

    async for ev in db.special_events.find({"image_url": {"$regex": "^data:"}}):
        new_url = await ensure_public_url(db, ev["image_url"])
        if new_url and new_url.startswith("/api/media/"):
            new_url = f"{base}{new_url}"
        if new_url:
            await db.special_events.update_one(
                {"id": ev["id"]}, {"$set": {"image_url": new_url}},
            )
            migrated["special_events"] += 1

    async for ov in db.offer_overrides.find({"image_url": {"$regex": "^data:"}}):
        new_url = await ensure_public_url(db, ov["image_url"])
        if new_url and new_url.startswith("/api/media/"):
            new_url = f"{base}{new_url}"
        if new_url:
            await db.offer_overrides.update_one(
                {"offer_id": ov["offer_id"]}, {"$set": {"image_url": new_url}},
            )
            migrated["offer_overrides"] += 1
            # Refresh in-memory OFFERS so the public site reflects immediately
            if ov.get("offer_id") in OFFERS:
                OFFERS[ov["offer_id"]]["image_url"] = new_url

    logging.warning("Admin %s migrated %s data: URLs", staff.get("email"), migrated)
    return {"ok": True, "migrated": migrated}


@api.post("/staff/admin/wipe-test-data")
async def admin_wipe_test_data(body: WipeTestDataBody, staff=Depends(get_current_staff)):
    """Admin-only nuclear button to wipe transactional data while keeping
    catalog + staff + integrations intact. Supports granular wipes per
    ``sections`` (see ``GET /staff/admin/wipe-sections`` for keys).

    Requires the literal confirmation string ``VIDER LES DONNEES BBR`` in the
    request body to avoid accidental clicks.
    """
    await _require_role(staff, ["admin"])
    if body.confirmation != "VIDER LES DONNEES BBR":
        raise HTTPException(
            status_code=400,
            detail="Confirmation invalide. Saisissez exactement : VIDER LES DONNEES BBR",
        )
    # Resolve which collections to actually wipe.
    selected_keys = body.sections or ["all"]
    if "all" in selected_keys:
        selected_keys = list(WIPE_SECTIONS.keys())
    unknown = [k for k in selected_keys if k not in WIPE_SECTIONS]
    if unknown:
        raise HTTPException(
            status_code=400,
            detail=f"Sections inconnues : {', '.join(unknown)}",
        )

    targets = set()
    for k in selected_keys:
        targets.update(WIPE_SECTIONS[k]["collections"])

    existing = await db.list_collection_names()
    summary = {"wiped": {}, "sections": selected_keys}
    total_wiped = 0
    for c in targets:
        if c in existing:
            res = await db[c].delete_many({})
            summary["wiped"][c] = res.deleted_count
            total_wiped += res.deleted_count

    logging.warning(
        "Admin %s wiped %d docs (sections=%s, cols=%s)",
        staff.get("email"), total_wiped, selected_keys, ", ".join(summary["wiped"].keys()),
    )
    return {"ok": True, "total_wiped": total_wiped, "details": summary}





# ----- Event privatization -----
@api.post("/events/privatization")
async def event_privatization(body: EventPrivatization):
    eid = str(uuid.uuid4())
    doc = body.model_dump()
    doc.update({"id": eid, "status": "new", "created_at": now_iso()})
    await db.event_requests.insert_one(doc)
    doc.pop("_id", None)
    return doc


# ----- Startup: seed staff (kept for future back-office) -----
@app.on_event("startup")
async def seed_staff():
    seeds = [
        {"email": "admin@boulay.ci", "name": "Admin Boulay", "role": "admin", "password": "Admin@2026"},
        {"email": "manager@boulay.ci", "name": "Manager Boulay", "role": "manager", "password": "Manager@2026"},
        {"email": "reception@boulay.ci", "name": "Réception Boulay", "role": "receptionist", "password": "Reception@2026"},
    ]
    for s in seeds:
        existing = await db.staff.find_one({"email": s["email"]})
        if not existing:
            await db.staff.insert_one({
                "id": str(uuid.uuid4()),
                "email": s["email"],
                "name": s["name"],
                "role": s["role"],
                "password_hash": hash_password(s["password"]),
                "created_at": now_iso(),
            })
    logging.info("Staff seeding complete")


# Seed/repair the 3 documented test accounts (idempotent — ensures correct roles
# even if the docs got out of sync with the actual DB rows).
@app.on_event("startup")
async def seed_test_accounts():
    test_seeds = [
        {"email": "hotesse.test@boulay.ci", "name": "Hôtesse Test", "role": "hotesse",
         "password": "Hotesse@2026", "pole_id": None},
        {"email": "mgr.pole.test@boulay.ci", "name": "Manager Pôle Test", "role": "manager_pole",
         "password": "MgrPole@2026", "pole_id": "beach_club"},
        {"email": "direction.test@boulay.ci", "name": "Direction Test", "role": "management_general",
         "password": "Direction@2026", "pole_id": None},
    ]
    for s in test_seeds:
        existing = await db.staff.find_one({"email": s["email"]}, {"_id": 0, "id": 1, "role": 1})
        if not existing:
            await db.staff.insert_one({
                "id": str(uuid.uuid4()),
                "email": s["email"],
                "name": s["name"],
                "role": s["role"],
                "pole_id": s["pole_id"],
                "password_hash": hash_password(s["password"]),
                "created_at": now_iso(),
            })
        elif existing.get("role") != s["role"]:
            # Heal drifted role / pole_id without resetting passwords.
            await db.staff.update_one(
                {"email": s["email"]},
                {"$set": {"role": s["role"], "pole_id": s["pole_id"]}},
            )
            logging.info("Repaired role for %s → %s", s["email"], s["role"])


# Seed Beach Club VIP spaces (transats + balinés) if collection is empty.
@app.on_event("startup")
async def seed_vip_spaces_on_startup():
    try:
        from routers.vip_spaces import seed_default_vip_spaces  # noqa: WPS433
        await seed_default_vip_spaces(db)
    except Exception as ex:  # pragma: no cover
        logging.warning("VIP spaces seed failed: %s", ex)


# =================================================================
# BACK-OFFICE (Staff) — Module 1 (Dashboard) & 3 (Embarquement)
# =================================================================

class Bateau(BaseModel):
    name: str
    capacity: int = Field(ge=1, le=300)
    status: Literal["actif", "maintenance"] = "actif"
    charter_price: int = Field(default=0, ge=0)  # FCFA — private charter rate
    # Fuel consumption per single-leg trip (aller OR retour). Used by the
    # advanced stats to compute total litres used across all completed trips.
    fuel_litres_per_trip: int = Field(default=0, ge=0, le=2000)


class BateauUpdate(BaseModel):
    name: Optional[str] = None
    capacity: Optional[int] = None
    status: Optional[Literal["actif", "maintenance"]] = None
    charter_price: Optional[int] = Field(default=None, ge=0)
    fuel_litres_per_trip: Optional[int] = Field(default=None, ge=0, le=2000)


class Traversee(BaseModel):
    bateau_id: str
    date: str  # YYYY-MM-DD
    depart_time: str  # Free-form: "12H", "08H45", "13:30" etc. (staff chooses)
    direction: Literal["aller", "retour"] = "aller"
    skipper_id: Optional[str] = None  # optional skipper assignment at creation
    # Optional: when scheduling an "aller", staff can lock-in a custom return
    # time on the SAME boat. When omitted, no return is auto-created (the staff
    # can program it manually later). Free-form like depart_time.
    return_time: Optional[str] = None
    return_skipper_id: Optional[str] = None  # optional skipper for the return leg


class TraverseeUpdate(BaseModel):
    """Body for PATCH /staff/traversees/{tid} — staff edits depart_time and/or skipper."""
    depart_time: Optional[str] = None
    skipper_id: Optional[str] = None
    skipper_clear: bool = False  # set to true to unassign the skipper


class Skipper(BaseModel):
    """A boat skipper — independent entity that can be assigned to a Traversee."""
    name: str = Field(min_length=2, max_length=80)
    phone: Optional[str] = None
    license_no: Optional[str] = None
    status: Literal["actif", "inactif"] = "actif"


class SkipperUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    license_no: Optional[str] = None
    status: Optional[Literal["actif", "inactif"]] = None


async def _seed_default_bateaux():
    """Seed 3 default boats if none exist."""
    if await db.bateaux.count_documents({}) == 0:
        defaults = [
            {"id": str(uuid.uuid4()), "name": "L'Étoile de Boulay", "capacity": 50, "status": "actif", "charter_price": 350000},
            {"id": str(uuid.uuid4()), "name": "Le Lagon d'Or", "capacity": 40, "status": "actif", "charter_price": 280000},
            {"id": str(uuid.uuid4()), "name": "Le Sunset Express", "capacity": 30, "status": "actif", "charter_price": 220000},
        ]
        await db.bateaux.insert_many(defaults)
        logging.info("Seeded %d default boats", len(defaults))
    # Backfill charter_price field for existing boats (idempotent)
    await db.bateaux.update_many(
        {"charter_price": {"$exists": False}},
        {"$set": {"charter_price": 0}},
    )
    # Backfill fuel_litres_per_trip on existing boats (idempotent)
    await db.bateaux.update_many(
        {"fuel_litres_per_trip": {"$exists": False}},
        {"$set": {"fuel_litres_per_trip": 0}},
    )


@api.get("/bateaux/charter")
async def list_bateaux_charter():
    """Public list of boats available for private charter, sorted by price."""
    await _seed_default_bateaux()
    items = await db.bateaux.find(
        {"status": "actif", "charter_price": {"$gt": 0}},
        {"_id": 0, "id": 1, "name": 1, "capacity": 1, "charter_price": 1},
    ).sort("charter_price", 1).to_list(length=50)
    return {"items": items}


@api.get("/staff/charters")
async def staff_list_charters(
    period: Optional[str] = Query(None, regex="^(day|week|month|all)$"),
    boat_id: Optional[str] = None,
    date_from: Optional[str] = Query(None, regex=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: Optional[str] = Query(None, regex=r"^\d{4}-\d{2}-\d{2}$"),
    staff=Depends(get_current_staff),
):
    """List all bookings that include a private boat charter, with the
    customer, date, boat, amount and payment status. Used by the dashboard
    so management can see who privatised a boat at a glance."""
    await _require_role(staff, ["admin", "manager", "manager_pole", "management_general"])
    filt: dict = {"charter_boat_id": {"$ne": None}}
    if boat_id:
        filt["charter_boat_id"] = boat_id
    # Custom date range takes precedence over the period preset
    if date_from or date_to:
        date_q: dict = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        filt["date"] = date_q
    elif period and period != "all":
        now = datetime.now(timezone.utc)
        if period == "day":
            since = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == "week":
            since = (now - timedelta(days=now.weekday())).replace(
                hour=0, minute=0, second=0, microsecond=0,
            )
        else:
            since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        filt["date"] = {"$gte": since.date().isoformat()}

    items = await db.bookings.find(
        filt,
        {
            "_id": 0, "id": 1, "date": 1, "boat_time": 1, "return_boat_time": 1,
            "charter_boat_id": 1, "charter_boat_name": 1, "charter_amount": 1,
            "total_amount": 1, "paid_amount": 1, "paid_at": 1, "status": 1,
            "name": 1, "email": 1, "phone": 1, "offer_type": 1, "participants": 1,
            "created_at": 1,
        },
    ).sort([("created_at", -1), ("date", -1)]).to_list(length=500)

    # Resolve a clean customer name (some legacy bookings have no top-level "name")
    for b in items:
        if not (b.get("name") or "").strip():
            for p in (b.get("participants") or []):
                if p.get("kind") == "adult":
                    b["name"] = f"{p.get('name','').strip()} {p.get('surname','').strip()}".strip()
                    break
        b.pop("participants", None)  # don't ship raw participant docs to the client

    summary = {
        "count": len(items),
        "total_revenue": sum(int(b.get("charter_amount") or 0) for b in items),
        "paid_count": sum(1 for b in items if b.get("paid_at")),
    }
    return {"items": items, "summary": summary}


async def _require_role(staff: dict, allowed: list):
    """Authorization gate. Supports legacy roles (receptionist, manager, admin)
    AND the new 7-role catalog by expanding each role to its equivalence set."""
    role = staff.get("role", "")
    if role in allowed:
        return
    # Expand via ROLE_INCLUDES so new roles inherit the legacy permissions
    for equiv in ROLE_INCLUDES.get(role, []):
        if equiv in allowed:
            return
    raise HTTPException(status_code=403, detail="Insufficient privileges")


# ============== NEW 7-ROLE CATALOG ==============
# Each new role lists which LEGACY role permissions it inherits, so existing
# endpoints (which check legacy roles via _require_role) keep working without
# touching every callsite.
ROLE_INCLUDES = {
    # Hôtesse — full reservations (read + update status). Includes manager so
    # she can call PATCH /staff/bookings/*/status etc; sidebar restricts UX.
    "hotesse": ["hotesse", "manager", "receptionist"],
    # Serveur & caisse — wallet operations including close. Includes manager
    # to validate payments on the wallet; sidebar shows only /staff/activites.
    "serveur_caisse": ["serveur_caisse", "manager", "receptionist"],
    # Logistique — boats / embarquement / traversees / scanner / boat config
    "logistique": ["logistique", "manager", "receptionist"],
    # Verification — QR scanner only; needs receptionist to perform check-ins
    "verification": ["verification", "receptionist"],
    # Manager pôle — equivalent to legacy manager, but scoped to ONE pole
    # (sidebar + dedicated pole filter enforce the scope)
    "manager_pole": ["manager_pole", "manager", "receptionist"],
    # Management général — read-only consultation. Middleware blocks all writes.
    "management_general": ["management_general", "manager", "receptionist"],
    # Administrator — already covers everything via legacy admin
    "admin": ["admin", "manager", "receptionist"],
    # Legacy roles still self-cover
    "receptionist": ["receptionist"],
    "manager": ["manager", "receptionist"],
}

# Roles that are read-only (any write request → 403 via middleware)
READONLY_ROLES = {"management_general"}

# Roles scoped to a single pole (must have staff.pole_id set)
POLE_SCOPED_ROLES = {"manager_pole"}


def _staff_pole_scope(staff: dict) -> Optional[str]:
    """Return the pole_id the staff is restricted to, or None if no scope."""
    if staff.get("role") in POLE_SCOPED_ROLES:
        return staff.get("pole_id") or None
    return None


# ---------- Dashboard KPIs (Module 1) ----------
@api.get("/staff/dashboard")
async def staff_dashboard(staff=Depends(get_current_staff)):
    """KPIs + planning du jour + alertes pour la page d'accueil staff.

    iter-44: bookings with status='pending' (abandoned carts — no payment, no
    cash-on-arrival commitment) are EXCLUDED from all dashboard counts so the
    chiffre d'affaires and the head-count match what the kitchen / boats will
    really see. They remain visible on the dedicated "Réservations en attente"
    page for follow-up (email relance).
    """
    today = datetime.now(timezone.utc).date().isoformat()
    # iter-44: filter pending out at query level
    cursor = db.bookings.find(
        {"date": today, "status": {"$ne": "pending"}},
        {"_id": 0, "reference_token": 0, "qr_codes": 0},
    )
    bookings_today = await cursor.to_list(length=500)

    revenue_today = sum(b.get("total_amount", 0) for b in bookings_today if b.get("status") in ("confirmed", "arrived", "completed", "pending_cash_payment"))
    guests_today = sum(b.get("adults", 0) + b.get("children", 0) for b in bookings_today)
    crossings = await db.traversees.count_documents({"date": today})

    # Status pipeline counts (iter-44: pending dropped, pending_cash_payment kept)
    pipeline_counts = {
        "confirmed": 0, "arrived": 0, "completed": 0,
        "cancelled": 0, "pending_cash_payment": 0,
    }
    for b in bookings_today:
        s = b.get("status", "confirmed")
        if s == "pending":
            continue
        pipeline_counts[s] = pipeline_counts.get(s, 0) + 1

    # Alerts
    now = datetime.now(timezone.utc)
    imminent = []
    for b in bookings_today:
        bt = b.get("boat_time", "")
        if bt and bt.endswith("H"):
            try:
                hour = int(bt[:-1])
                btime = datetime.combine(now.date(), datetime.min.time()).replace(tzinfo=timezone.utc, hour=hour)
                diff = (btime - now).total_seconds() / 3600
                if 0 <= diff <= 2 and b.get("status") in ("confirmed", "pending_cash_payment"):
                    imminent.append({"booking_id": b["id"], "client": b.get("phone", ""), "offer": b.get("offer_name", ""), "boat_time": bt, "guests": b.get("adults", 0) + b.get("children", 0)})
            except Exception:
                pass

    # iter-44: "unpaid" widget is now the relance funnel — count of pending
    # bookings (any date) still awaiting payment, with the call-to-action.
    unpaid_count = await db.bookings.count_documents({"status": "pending"})
    unpaid = await db.bookings.find(
        {"status": "pending"},
        {"_id": 0, "id": 1, "offer_name": 1, "total_amount": 1, "phone": 1,
         "email": 1, "date": 1, "created_at": 1,
         "payment_link_token": 1},
    ).sort("created_at", -1).limit(20).to_list(length=20)

    # Pôle breakdown — counts + revenue today + last 30 days
    pole_counts_today: dict = {pid: {"count": 0, "guests": 0, "revenue": 0} for pid in POLES}
    for b in bookings_today:
        pole = b.get("pole") or _pole_for_offer(b.get("offer_type", ""))
        if not pole or pole not in pole_counts_today:
            continue
        pole_counts_today[pole]["count"] += 1
        pole_counts_today[pole]["guests"] += int(b.get("adults", 0)) + int(b.get("children", 0))
        if b.get("status") in ("confirmed", "arrived", "completed"):
            pole_counts_today[pole]["revenue"] += int(b.get("total_amount", 0))

    # Last 30 days breakdown
    from datetime import timedelta as _td
    cutoff = (datetime.now(timezone.utc).date() - _td(days=30)).isoformat()
    pole_30d: dict = {pid: {"count": 0, "revenue": 0} for pid in POLES}
    cur30 = db.bookings.find(
        {"date": {"$gte": cutoff}, "status": {"$ne": "cancelled"}},
        {"_id": 0, "pole": 1, "offer_type": 1, "total_amount": 1},
    )
    async for b in cur30:
        pole = b.get("pole") or _pole_for_offer(b.get("offer_type", ""))
        if not pole or pole not in pole_30d:
            continue
        pole_30d[pole]["count"] += 1
        pole_30d[pole]["revenue"] += int(b.get("total_amount", 0) or 0)

    pole_breakdown = []
    for pid, p in sorted(POLES.items(), key=lambda kv: kv[1].get("sort_order", 99)):
        pole_breakdown.append({
            "id": pid,
            "name_fr": p["name_fr"],
            "today": pole_counts_today.get(pid, {"count": 0, "guests": 0, "revenue": 0}),
            "last_30d": pole_30d.get(pid, {"count": 0, "revenue": 0}),
        })

    # Feedback average — last 90 days (rolling). Ratings are stored as flat
    # fields (accueil_arrivee, service_amabilite, restauration_boissons,
    # ambiance_cadre, proprete_confort, experience_globale) each scored 1-5.
    # NPS-style block: promoters = 5★, passives = 4★, detractors ≤ 3★ on
    # experience_globale, score = (promoters - detractors) / total * 100.
    from datetime import timedelta as _td2
    fb_cutoff = (datetime.now(timezone.utc) - _td2(days=90)).isoformat()
    fb_total = 0
    fb_sum = 0.0
    fb_count = 0
    promoters = passives = detractors = 0
    rating_proj = {f: 1 for f in EXP_RATING_FIELDS}
    rating_proj["_id"] = 0
    async for fb in db.experience_feedback.find(
        {"created_at": {"$gte": fb_cutoff}}, rating_proj,
    ):
        vals = [
            fb.get(f) for f in EXP_RATING_FIELDS
            if isinstance(fb.get(f), (int, float)) and fb.get(f) > 0
        ]
        if vals:
            fb_sum += sum(vals) / len(vals)
            fb_count += 1
        eg = fb.get("experience_globale")
        if isinstance(eg, (int, float)):
            if eg >= 5:
                promoters += 1
            elif eg == 4:
                passives += 1
            elif eg >= 1:
                detractors += 1
        fb_total += 1
    fb_average = round(fb_sum / fb_count, 2) if fb_count else None
    total_resp = promoters + passives + detractors
    fb_nps_avg = round(((promoters - detractors) / total_resp) * 100) if total_resp else None

    # Also tag pole_counts_today with pending_cash_payment in revenue
    # (cash-to-collect is real future revenue, not a no-show).
    for b in bookings_today:
        if b.get("status") == "pending_cash_payment":
            pole = b.get("pole") or _pole_for_offer(b.get("offer_type", ""))
            if pole and pole in pole_counts_today:
                pole_counts_today[pole]["revenue"] += int(b.get("total_amount", 0))

    return {
        "kpis": {
            "bookings_today": len(bookings_today),
            "revenue_today": revenue_today,
            "guests_today": guests_today,
            "crossings_today": crossings,
            "feedback_average": fb_average,
            "feedback_count": fb_total,
            "feedback_nps_avg": fb_nps_avg,
            "pending_relance_count": unpaid_count,
        },
        "pipeline": pipeline_counts,
        "bookings_today": bookings_today,
        "alerts": {
            "imminent_arrivals": imminent,
            "unpaid_bookings": unpaid,
        },
        "pole_breakdown": pole_breakdown,
    }


# ---------- Bateaux CRUD (Module 3) ----------
@api.get("/staff/bateaux")
async def list_bateaux(staff=Depends(get_current_staff)):
    await _seed_default_bateaux()
    items = await db.bateaux.find({}, {"_id": 0}).to_list(length=200)
    return items


@api.post("/staff/bateaux")
async def create_bateau(body: Bateau, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    doc = {"id": str(uuid.uuid4()), **body.model_dump()}
    await db.bateaux.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api.patch("/staff/bateaux/{bateau_id}")
async def update_bateau(bateau_id: str, body: BateauUpdate, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    res = await db.bateaux.update_one({"id": bateau_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bateau not found")
    return {"ok": True}


@api.delete("/staff/bateaux/{bateau_id}")
async def delete_bateau(bateau_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    res = await db.bateaux.delete_one({"id": bateau_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bateau not found")
    return {"ok": True}


# ---------- Traversées (Module 3) ----------
@api.get("/staff/traversees")
async def list_traversees(
    date: Optional[str] = None,
    direction: Optional[Literal["aller", "retour"]] = None,
    status: Optional[str] = None,
    staff=Depends(get_current_staff),
):
    """List crossings (default = today) with linked passengers.

    Optional filters used by the scanner workflow:
    - ``direction`` (aller|retour) — narrows by leg
    - ``status`` — single status or comma-separated list (programmé,en_cours,terminée,annulée)
    """
    if date is None:
        date = datetime.now(timezone.utc).date().isoformat()
    query: dict = {"date": date}
    if direction:
        query["direction"] = direction
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        if len(statuses) == 1:
            query["status"] = statuses[0]
        elif statuses:
            query["status"] = {"$in": statuses}
    crossings = await db.traversees.find(query, {"_id": 0}).to_list(length=200)
    # Hydrate with bateau info + passenger count
    bateaux = {b["id"]: b for b in await db.bateaux.find({}, {"_id": 0}).to_list(length=200)}
    for c in crossings:
        c["bateau"] = bateaux.get(c.get("bateau_id"), {})
        passengers = await db.traversee_passengers.find({"traversee_id": c["id"]}, {"_id": 0}).to_list(length=500)
        c["passengers"] = passengers
        c["passenger_count"] = sum(p.get("guests", 1) for p in passengers)
    return sorted(crossings, key=lambda x: (x.get("date", ""), x.get("depart_time", "")))


@api.post("/staff/traversees")
async def create_traversee(body: Traversee, staff=Depends(get_current_staff)):
    # Manual gate — hôtesse inherits 'manager' via ROLE_INCLUDES so we can
    # not rely on _require_role(['manager','admin']) alone here. Traversee
    # scheduling is a logistics responsibility.
    if staff.get("role") not in {"admin", "manager", "logistique"}:
        raise HTTPException(status_code=403, detail="Action réservée aux managers / logistique")
    bateau = await db.bateaux.find_one({"id": body.bateau_id}, {"_id": 0})
    if not bateau:
        raise HTTPException(status_code=404, detail="Bateau not found")
    depart_clean = (body.depart_time or "").strip().upper().replace(":", "H")
    if not depart_clean:
        raise HTTPException(status_code=400, detail="Heure de départ requise")
    # Resolve optional skipper (must exist + be actif)
    skipper_doc = None
    if body.skipper_id:
        skipper_doc = await db.skippers.find_one(
            {"id": body.skipper_id, "status": "actif"}, {"_id": 0},
        )
        if not skipper_doc:
            raise HTTPException(status_code=404, detail="Skipper introuvable ou inactif")
    tid = str(uuid.uuid4())
    doc = {
        "id": tid,
        "bateau_id": body.bateau_id,
        "date": body.date,
        "depart_time": depart_clean,
        "direction": body.direction,
        "status": "programmé",
        "skipper_id": skipper_doc["id"] if skipper_doc else None,
        "skipper_name": skipper_doc["name"] if skipper_doc else None,
        "created_at": now_iso(),
    }
    await db.traversees.insert_one(doc)
    # When an "aller" is created and the staff supplied a return_time, lock-in
    # that exact return on the SAME boat. We NO LONGER auto-generate a return
    # 5h later — staff explicitly schedules each leg now.
    if body.direction == "aller" and (body.return_time or "").strip():
        ret_clean = body.return_time.strip().upper().replace(":", "H")
        ret_skipper = None
        if body.return_skipper_id:
            ret_skipper = await db.skippers.find_one(
                {"id": body.return_skipper_id, "status": "actif"}, {"_id": 0},
            )
        ret_doc = {
            "id": str(uuid.uuid4()),
            "bateau_id": body.bateau_id,
            "date": body.date,
            "depart_time": ret_clean,
            "direction": "retour",
            "status": "programmé",
            "skipper_id": ret_skipper["id"] if ret_skipper else (skipper_doc["id"] if skipper_doc else None),
            "skipper_name": ret_skipper["name"] if ret_skipper else (skipper_doc["name"] if skipper_doc else None),
            "parent_id": tid,
            "created_at": now_iso(),
        }
        await db.traversees.insert_one(ret_doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api.patch("/staff/traversees/{tid}")
async def update_traversee(tid: str, body: TraverseeUpdate, staff=Depends(get_current_staff)):
    """Edit a scheduled traversée: change the departure time and/or the assigned skipper.
    The traversée must still be in 'programmé' status (cannot edit en_cours / terminé)."""
    if staff.get("role") not in {"admin", "manager", "logistique"}:
        raise HTTPException(status_code=403, detail="Action réservée aux managers / logistique")
    existing = await db.traversees.find_one({"id": tid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Traversée introuvable")
    if existing.get("status") not in (None, "programmé"):
        raise HTTPException(status_code=400, detail="Seules les traversées programmées sont modifiables")
    update: dict = {}
    if body.depart_time and body.depart_time.strip():
        update["depart_time"] = body.depart_time.strip().upper().replace(":", "H")
    if body.skipper_clear:
        update["skipper_id"] = None
        update["skipper_name"] = None
    elif body.skipper_id:
        skp = await db.skippers.find_one({"id": body.skipper_id, "status": "actif"}, {"_id": 0})
        if not skp:
            raise HTTPException(status_code=404, detail="Skipper introuvable ou inactif")
        update["skipper_id"] = skp["id"]
        update["skipper_name"] = skp["name"]
    if not update:
        raise HTTPException(status_code=400, detail="Aucun champ à modifier")
    update["updated_at"] = now_iso()
    await db.traversees.update_one({"id": tid}, {"$set": update})
    return {"ok": True, **update}


@api.delete("/staff/traversees/{tid}")
async def delete_traversee(tid: str, staff=Depends(get_current_staff)):
    """Delete a scheduled traversée. Also detaches any boarded bookings. Cannot
    delete a traversée that has already started (status != 'programmé')."""
    if staff.get("role") not in {"admin", "manager", "logistique"}:
        raise HTTPException(status_code=403, detail="Action réservée aux managers / logistique")
    existing = await db.traversees.find_one({"id": tid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Traversée introuvable")
    if existing.get("status") not in (None, "programmé"):
        raise HTTPException(status_code=400, detail="Impossible de supprimer une traversée en cours ou terminée")
    # Unboard passengers + cascade-delete the return leg if this is an 'aller'.
    await db.traversee_passengers.delete_many({"traversee_id": tid})
    children_deleted = 0
    if existing.get("direction") == "aller":
        # Find and remove the matching auto-created return leg, if any.
        child = await db.traversees.find_one({"parent_id": tid}, {"_id": 0})
        if child and child.get("status") in (None, "programmé"):
            await db.traversee_passengers.delete_many({"traversee_id": child["id"]})
            r = await db.traversees.delete_one({"id": child["id"]})
            children_deleted = r.deleted_count
    await db.traversees.delete_one({"id": tid})
    return {"ok": True, "children_deleted": children_deleted}


@api.patch("/staff/traversees/{tid}/status")
async def update_traversee_status(tid: str, status: str = Body(..., embed=True), staff=Depends(get_current_staff)):
    await _require_role(staff, ["receptionist", "manager", "admin"])
    if status not in ("programmé", "en_cours", "terminé"):
        raise HTTPException(status_code=400, detail="Invalid status")
    res = await db.traversees.update_one({"id": tid}, {"$set": {"status": status}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Traversee not found")
    return {"ok": True}


@api.get("/staff/traversees/{tid}/passengers.pdf")
async def export_traversee_passengers_pdf(tid: str, staff=Depends(get_current_staff)):
    """Export the list of passengers boarded on a given traversée as a PDF
    manifest (skipper, boat, time, passengers with name/email/phone).
    Used at quay-side to hand a printable list to the skipper before departure.
    """
    t = await db.traversees.find_one({"id": tid}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Traversée introuvable")
    bateau = await db.bateaux.find_one({"id": t["bateau_id"]}, {"_id": 0}) or {}
    # Fetch boarded passengers
    pax = await db.traversee_passengers.find(
        {"traversee_id": tid}, {"_id": 0},
    ).to_list(length=500)
    booking_ids = [p["booking_id"] for p in pax if p.get("booking_id")]
    bookings = []
    if booking_ids:
        async for b in db.bookings.find(
            {"id": {"$in": booking_ids}},
            {"_id": 0, "id": 1, "offer_name": 1, "adults": 1, "children": 1,
             "participants": 1, "email": 1, "phone": 1, "label": 1},
        ):
            bookings.append(b)
    # Also include visitor_registrations attached to this traversée
    visitors = await db.visitor_registrations.find(
        {"traversee_id": tid}, {"_id": 0},
    ).to_list(length=500)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 textColor=colors.HexColor("#0A0A0A"),
                                 fontSize=18, spaceAfter=4)
    sub = ParagraphStyle("sub", parent=styles["Normal"],
                         textColor=colors.HexColor("#B8922A"),
                         fontSize=8, spaceAfter=4, leading=10)
    label = ParagraphStyle("label", parent=styles["Normal"], fontSize=9,
                           textColor=colors.HexColor("#0A0A0A"), leading=12)
    direction_label = "Aller (vers l'île)" if t.get("direction") == "aller" else "Retour"
    story = [
        Paragraph(f"Manifeste de traversée · {direction_label}", title_style),
        Paragraph(
            f"{t.get('date','')} · {t.get('depart_time','')}",
            sub,
        ),
        Paragraph(
            f"<b>Bateau :</b> {bateau.get('name','—')} ({bateau.get('capacity','?')} places)  "
            f"<b>Skipper :</b> {t.get('skipper_name') or '— Non assigné —'}",
            label,
        ),
        Spacer(1, 4 * mm),
    ]
    rows = [["N°", "Nom complet", "Type", "Email", "Téléphone", "Référence"]]
    n = 0
    for b in bookings:
        parts = b.get("participants") or []
        adults_count = b.get("adults") or 0
        children_count = b.get("children") or 0
        # 1 row per adult, then 1 row "+ N enfants" appended to the booker row
        for i, p in enumerate(parts[:adults_count]):
            n += 1
            extra = ""
            if i == 0 and children_count > 0:
                extra = f" (+{children_count} enfant{'s' if children_count > 1 else ''})"
            rows.append([
                str(n),
                f"{p.get('surname','')} {p.get('name','')}{extra}",
                "Client",
                p.get("email") or b.get("email") or "",
                p.get("phone") or b.get("phone") or "",
                (b.get("id") or "")[:8].upper(),
            ])
    for v in visitors:
        n += 1
        rows.append([
            str(n),
            f"{v.get('surname','')} {v.get('name','')}",
            (v.get("kind") or "").capitalize(),
            v.get("email") or "",
            v.get("phone") or "",
            (v.get("company") or v.get("id") or "")[:18],
        ])
    if len(rows) == 1:
        rows.append(["—", "Aucun passager embarqué", "", "", "", ""])

    table = Table(rows, repeatRows=1, colWidths=[12 * mm, 60 * mm, 22 * mm, 50 * mm, 30 * mm, 22 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5CFC4")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAF7")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(
        f"<i>Édité le {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M UTC')} par {staff.get('email','')}</i>",
        ParagraphStyle("foot", parent=styles["Normal"], fontSize=7,
                       textColor=colors.HexColor("#0A0A0A"), alignment=2),
    ))
    doc.build(story)
    filename = f"traversee-{t.get('date','')}-{t.get('depart_time','')}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@api.get("/staff/traversees/{tid}/passengers.xlsx")
async def export_traversee_passengers_xlsx(tid: str, staff=Depends(get_current_staff)):
    """Same manifest as the PDF endpoint but as an Excel workbook."""
    t = await db.traversees.find_one({"id": tid}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Traversée introuvable")
    bateau = await db.bateaux.find_one({"id": t["bateau_id"]}, {"_id": 0}) or {}
    pax = await db.traversee_passengers.find(
        {"traversee_id": tid}, {"_id": 0},
    ).to_list(length=500)
    booking_ids = [p["booking_id"] for p in pax if p.get("booking_id")]
    bookings = []
    if booking_ids:
        async for b in db.bookings.find(
            {"id": {"$in": booking_ids}},
            {"_id": 0, "id": 1, "offer_name": 1, "adults": 1, "children": 1,
             "participants": 1, "email": 1, "phone": 1, "label": 1},
        ):
            bookings.append(b)
    visitors = await db.visitor_registrations.find(
        {"traversee_id": tid}, {"_id": 0},
    ).to_list(length=500)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Manifeste"

    direction_label = "Aller (vers l'île)" if t.get("direction") == "aller" else "Retour"
    # ----- Header rows -----
    ws.append([f"Manifeste de traversée · {direction_label}"])
    ws["A1"].font = Font(bold=True, size=14, color="0A0A0A")
    ws.append([f"{t.get('date','')} · {t.get('depart_time','')}"])
    ws["A2"].font = Font(italic=True, color="B8922A", size=10)
    ws.append([
        f"Bateau : {bateau.get('name','—')} ({bateau.get('capacity','?')} places) — "
        f"Skipper : {t.get('skipper_name') or '— Non assigné —'}"
    ])
    ws["A3"].font = Font(size=10)
    ws.append([])

    # ----- Table header -----
    headers = ["N°", "Nom complet", "Type", "Email", "Téléphone", "Référence"]
    ws.append(headers)
    header_row = ws.max_row
    for c in ws[header_row]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
        c.alignment = Alignment(horizontal="left", vertical="center")

    # ----- Body rows -----
    n = 0
    for b in bookings:
        parts = b.get("participants") or []
        adults_count = b.get("adults") or 0
        children_count = b.get("children") or 0
        for i, p in enumerate(parts[:adults_count]):
            n += 1
            extra = ""
            if i == 0 and children_count > 0:
                extra = f" (+{children_count} enfant{'s' if children_count > 1 else ''})"
            ws.append([
                n,
                f"{p.get('surname','')} {p.get('name','')}{extra}".strip(),
                "Client",
                p.get("email") or b.get("email") or "",
                p.get("phone") or b.get("phone") or "",
                (b.get("id") or "")[:8].upper(),
            ])
    for v in visitors:
        n += 1
        ws.append([
            n,
            f"{v.get('surname','')} {v.get('name','')}".strip(),
            (v.get("kind") or "").capitalize(),
            v.get("email") or "",
            v.get("phone") or "",
            v.get("company") or "",
        ])
    if n == 0:
        ws.append(["—", "Aucun passager embarqué", "", "", "", ""])

    # ----- Column widths (heuristic) -----
    widths = [6, 32, 14, 32, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Auto-zebra below the header
    thin = Side(style="thin", color="D5CFC4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(header_row + 1, ws.max_row + 1):
        is_alt = (r - header_row) % 2 == 0
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = border
            cell.font = Font(size=9)
            if is_alt:
                cell.fill = PatternFill(start_color="FAFAF7", end_color="FAFAF7", fill_type="solid")

    # Footer
    ws.append([])
    ws.append([f"Édité le {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M UTC')} par {staff.get('email','')}"])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=8, color="6B7280")

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"traversee-{t.get('date','')}-{t.get('depart_time','')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )




@api.post("/staff/traversees/{tid}/board")
async def board_passenger(tid: str, body: dict = Body(...), staff=Depends(get_current_staff)):
    """Mark a booking as boarded on a crossing."""
    await _require_role(staff, ["receptionist", "manager", "admin"])
    booking_id = body.get("booking_id")
    if not booking_id:
        raise HTTPException(status_code=400, detail="booking_id required")
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    crossing = await db.traversees.find_one({"id": tid})
    if not crossing:
        raise HTTPException(status_code=404, detail="Traversee not found")
    # Capacity check
    bateau = await db.bateaux.find_one({"id": crossing["bateau_id"]}, {"_id": 0})
    existing = await db.traversee_passengers.find({"traversee_id": tid}).to_list(length=500)
    booked = sum(p.get("guests", 1) for p in existing)
    guests = booking.get("adults", 0) + booking.get("children", 0)
    if bateau and booked + guests > bateau["capacity"]:
        raise HTTPException(status_code=400, detail=f"Capacity exceeded ({bateau['capacity']})")
    # Upsert
    await db.traversee_passengers.update_one(
        {"traversee_id": tid, "booking_id": booking_id},
        {
            "$set": {
                "traversee_id": tid,
                "booking_id": booking_id,
                "guests": guests,
                "client_name": f"{booking.get('participants', [{}])[0].get('surname', '')} {booking.get('participants', [{}])[0].get('name', '')}".strip(),
                "offer_name": booking.get("offer_name"),
                "boarded_at": now_iso(),
            }
        },
        upsert=True,
    )
    # Mark booking arrived
    await db.bookings.update_one({"id": booking_id}, {"$set": {"status": "arrived"}})
    return {"ok": True, "guests_boarded": guests}


@api.delete("/staff/traversees/{tid}/board/{booking_id}")
async def unboard_passenger(tid: str, booking_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["receptionist", "manager", "admin"])
    await db.traversee_passengers.delete_one({"traversee_id": tid, "booking_id": booking_id})
    return {"ok": True}


# ---------- Traversées — Historique & Rapport PDF ----------
def _resolve_period_range(period: str, ref: str) -> tuple:
    """Convert (period, ref date YYYY-MM-DD) to (start_date_iso, end_date_iso_exclusive, label)."""
    try:
        d = datetime.strptime(ref, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date (YYYY-MM-DD)")
    if period == "day":
        return d.isoformat(), (d + timedelta(days=1)).isoformat(), d.strftime("%A %d %B %Y")
    if period == "week":
        start = d - timedelta(days=d.weekday())  # Monday
        end = start + timedelta(days=7)
        return start.isoformat(), end.isoformat(), f"Semaine du {start.strftime('%d %b')} au {(end - timedelta(days=1)).strftime('%d %b %Y')}"
    if period == "month":
        start = d.replace(day=1)
        next_month = (start + timedelta(days=32)).replace(day=1)
        return start.isoformat(), next_month.isoformat(), start.strftime("%B %Y")
    raise HTTPException(status_code=400, detail="period must be day|week|month")


async def _fetch_history(date_from: str, date_to: str, status: Optional[str] = None, bateau_id: Optional[str] = None, started_only: bool = False) -> dict:
    """Aggregate crossings + passenger counts in a date range. date_to is exclusive.

    Filters:
      • status — explicit single status (`programmé`/`en_cours`/`terminé`)
      • bateau_id — restrict to a specific boat
      • started_only — only include crossings that have actually started
        (status in `en_cours` or `terminé`). Mutually exclusive in spirit with
        `status="programmé"` (which yields an empty set).
    """
    q: dict = {"date": {"$gte": date_from, "$lt": date_to}}
    if status:
        q["status"] = status
    elif started_only:
        q["status"] = {"$in": ["en_cours", "terminé"]}
    if bateau_id:
        q["bateau_id"] = bateau_id
    crossings = await db.traversees.find(q, {"_id": 0}).sort([("date", 1), ("depart_time", 1)]).to_list(length=5000)
    bateaux = await db.bateaux.find({}, {"_id": 0}).to_list(length=200)
    boat_by_id = {b["id"]: b for b in bateaux}

    # Aggregate passengers per crossing
    tids = [c["id"] for c in crossings]
    pax_by_tid: dict = {tid: {"count": 0, "guests": 0} for tid in tids}
    if tids:
        cur = db.traversee_passengers.find({"traversee_id": {"$in": tids}}, {"_id": 0})
        async for p in cur:
            tid = p["traversee_id"]
            pax_by_tid[tid]["count"] += 1
            pax_by_tid[tid]["guests"] += int(p.get("guests", 1))

    by_status = {"programmé": 0, "en_cours": 0, "terminé": 0}
    by_day: dict = {}
    by_boat: dict = {}
    by_direction = {"aller": 0, "retour": 0}
    total_passengers = 0
    total_guests = 0
    items = []

    for c in crossings:
        st = c.get("status") or "programmé"
        by_status[st] = by_status.get(st, 0) + 1
        by_day.setdefault(c["date"], {"date": c["date"], "total": 0, "programmé": 0, "en_cours": 0, "terminé": 0, "guests": 0})
        by_day[c["date"]]["total"] += 1
        by_day[c["date"]][st] = by_day[c["date"]].get(st, 0) + 1
        bid = c["bateau_id"]
        bname = (boat_by_id.get(bid) or {}).get("name", "—")
        by_boat.setdefault(bid, {"bateau_id": bid, "bateau_name": bname, "total": 0, "terminé": 0, "guests": 0})
        by_boat[bid]["total"] += 1
        if st == "terminé":
            by_boat[bid]["terminé"] += 1
        by_direction[c.get("direction", "aller")] = by_direction.get(c.get("direction", "aller"), 0) + 1
        pax = pax_by_tid.get(c["id"], {"count": 0, "guests": 0})
        total_passengers += pax["count"]
        total_guests += pax["guests"]
        by_day[c["date"]]["guests"] += pax["guests"]
        by_boat[bid]["guests"] += pax["guests"]
        items.append({
            **c,
            "bateau_name": bname,
            "passenger_count": pax["count"],
            "guests": pax["guests"],
        })

    return {
        "total": len(crossings),
        "by_status": by_status,
        "by_direction": by_direction,
        "by_day": [by_day[k] for k in sorted(by_day.keys())],
        "by_boat": sorted(by_boat.values(), key=lambda x: x["total"], reverse=True),
        "total_passengers": total_passengers,
        "total_guests": total_guests,
        "items": items,
    }


@api.get("/staff/traversees/history")
async def traversees_history(
    period: str = "day",
    date: Optional[str] = None,
    status: Optional[str] = None,
    bateau_id: Optional[str] = None,
    started_only: bool = False,
    staff=Depends(get_current_staff),
):
    """Crossings history with stats. period=day|week|month, date=YYYY-MM-DD (default today),
    status=programmé|en_cours|terminé (optional filter), bateau_id=optional boat scope,
    started_only=true to exclude programmées (uniquement en_cours+terminé)."""
    await _require_role(staff, ["receptionist", "manager", "admin"])
    ref = date or datetime.now(timezone.utc).date().isoformat()
    date_from, date_to, label = _resolve_period_range(period, ref)
    if status and status not in ("programmé", "en_cours", "terminé"):
        raise HTTPException(status_code=400, detail="Invalid status")
    payload = await _fetch_history(date_from, date_to, status, bateau_id, started_only)
    return {
        "period": period,
        "reference_date": ref,
        "label": label,
        "date_from": date_from,
        "date_to": date_to,
        "status_filter": status,
        "bateau_id": bateau_id,
        "started_only": started_only,
        **payload,
    }


@api.get("/staff/traversees/history/report.pdf")
async def traversees_history_pdf(
    period: str = "day",
    date: Optional[str] = None,
    status: Optional[str] = None,
    bateau_id: Optional[str] = None,
    started_only: bool = True,
    staff=Depends(get_current_staff),
):
    """Generate a luxury-styled PDF report of the crossings for the given period.

    By default (`started_only=true`) the export only contains crossings that
    have actually departed (status en_cours or terminé) — matches the
    operational request "exporter à partir du moment où la traversée est
    considérée comme en cours".
    """
    await _require_role(staff, ["receptionist", "manager", "admin"])
    ref = date or datetime.now(timezone.utc).date().isoformat()
    date_from, date_to, label = _resolve_period_range(period, ref)
    if status and status not in ("programmé", "en_cours", "terminé"):
        raise HTTPException(status_code=400, detail="Invalid status")
    data = await _fetch_history(date_from, date_to, status, bateau_id, started_only)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.pdfgen import canvas as rl_canvas
    from fastapi.responses import StreamingResponse

    GOLD = colors.HexColor("#B8922A")
    DARK = colors.HexColor("#0A0A0A")
    LIGHT = colors.HexColor("#FAFAF7")
    MUTED = colors.HexColor("#888888")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=22, textColor=DARK, alignment=0, spaceAfter=4)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontName="Helvetica", fontSize=10, textColor=GOLD, alignment=0, spaceAfter=16, leading=14)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12, textColor=GOLD, spaceBefore=12, spaceAfter=6)
    body = ParagraphStyle("body", parent=styles["Normal"], fontName="Helvetica", fontSize=9, textColor=DARK)
    small = ParagraphStyle("small", parent=styles["Normal"], fontName="Helvetica", fontSize=8, textColor=MUTED)

    elements = []
    elements.append(Paragraph("Boulay Beach Resort", h1))
    period_label = {"day": "Journalier", "week": "Hebdomadaire", "month": "Mensuel"}.get(period, period)
    filter_label = f" — Statut : {status}" if status else ""
    elements.append(Paragraph(f"Rapport des traversées · {period_label} · {label}{filter_label}", sub))

    # KPI block
    kpi_data = [
        ["Total", "Programmées", "En cours", "Terminées", "Passagers"],
        [
            str(data["total"]),
            str(data["by_status"].get("programmé", 0)),
            str(data["by_status"].get("en_cours", 0)),
            str(data["by_status"].get("terminé", 0)),
            str(data["total_guests"]),
        ],
    ]
    kpi_tbl = Table(kpi_data, colWidths=[3.4 * cm] * 5)
    kpi_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("TEXTCOLOR", (0, 0), (-1, 0), MUTED),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 16),
        ("TEXTCOLOR", (0, 1), (-1, 1), DARK),
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.5, GOLD),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E0D5B5")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_tbl)

    # By boat
    if data["by_boat"]:
        elements.append(Paragraph("Répartition par bateau", h2))
        boat_rows = [["Bateau", "Total", "Terminées", "Passagers"]]
        for b in data["by_boat"]:
            boat_rows.append([b["bateau_name"], str(b["total"]), str(b["terminé"]), str(b["guests"])])
        boat_tbl = Table(boat_rows, colWidths=[7 * cm, 3 * cm, 3 * cm, 3 * cm])
        boat_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(boat_tbl)

    # By day (for week/month)
    if period in ("week", "month") and data["by_day"]:
        elements.append(Paragraph("Détail par jour", h2))
        day_rows = [["Date", "Total", "Programmées", "En cours", "Terminées", "Passagers"]]
        for d in data["by_day"]:
            day_rows.append([d["date"], str(d["total"]), str(d.get("programmé", 0)),
                             str(d.get("en_cours", 0)), str(d.get("terminé", 0)), str(d.get("guests", 0))])
        day_tbl = Table(day_rows, colWidths=[3 * cm, 2 * cm, 2.5 * cm, 2.2 * cm, 2.3 * cm, 2.5 * cm])
        day_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(day_tbl)

    # Detailed crossings
    elements.append(Paragraph("Liste des traversées", h2))
    if not data["items"]:
        elements.append(Paragraph("Aucune traversée sur cette période.", body))
    else:
        rows = [["Date", "Heure", "Direction", "Bateau", "Statut", "Passagers"]]
        status_map = {"programmé": "Programmée", "en_cours": "En cours", "terminé": "Terminée"}
        for it in data["items"]:
            rows.append([
                it.get("date", ""),
                it.get("depart_time", ""),
                (it.get("direction") or "").capitalize(),
                it.get("bateau_name", ""),
                status_map.get(it.get("status"), it.get("status", "")),
                str(it.get("guests", 0)),
            ])
        crossings_tbl = Table(rows, colWidths=[2.5 * cm, 1.6 * cm, 2.2 * cm, 4.5 * cm, 2.8 * cm, 2 * cm], repeatRows=1)
        crossings_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (4, 0), (-1, -1), "CENTER"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.2, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(crossings_tbl)

    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(
        f"Rapport généré le {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M UTC')} · Boulay Beach Resort, Abidjan",
        small,
    ))

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setStrokeColor(GOLD)
        canvas.setLineWidth(0.5)
        canvas.line(2 * cm, 1.5 * cm, A4[0] - 2 * cm, 1.5 * cm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(MUTED)
        canvas.drawString(2 * cm, 1 * cm, "Boulay Beach Resort — Rapport interne — Confidentiel")
        canvas.drawRightString(A4[0] - 2 * cm, 1 * cm, f"Page {doc_.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    filename = f"bbr-traversees-{period}-{ref}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api.get("/staff/traversees/history/report.xlsx")
async def traversees_history_xlsx(
    period: str = "day",
    date: Optional[str] = None,
    status: Optional[str] = None,
    bateau_id: Optional[str] = None,
    started_only: bool = True,
    staff=Depends(get_current_staff),
):
    """Excel export of the crossings history. Same filters as the PDF report.
    Defaults to `started_only=true` so only departed crossings are listed.
    """
    await _require_role(staff, ["receptionist", "manager", "admin"])
    ref = date or datetime.now(timezone.utc).date().isoformat()
    date_from, date_to, label = _resolve_period_range(period, ref)
    if status and status not in ("programmé", "en_cours", "terminé"):
        raise HTTPException(status_code=400, detail="Invalid status")
    data = await _fetch_history(date_from, date_to, status, bateau_id, started_only)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    gold_fill = PatternFill("solid", fgColor="B8922A")
    soft_fill = PatternFill("solid", fgColor="FAFAF7")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(border_style="thin", color="EAE2C9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---------- Sheet 1: Synthèse ----------
    ws = wb.active
    ws.title = "Synthèse"
    ws.append(["Boulay Beach Resort — Historique des traversées"])
    ws["A1"].font = Font(bold=True, size=14, color="0A0A0A")
    period_label = {"day": "Journalier", "week": "Hebdomadaire", "month": "Mensuel"}.get(period, period)
    ws.append([f"{period_label} · {label}" + (f" · Statut: {status}" if status else "") + (f" · Bateau: {bateau_id}" if bateau_id else "")])
    ws.append([])
    ws.append(["Total", "Programmées", "En cours", "Terminées", "Passagers"])
    for c in ws[4]:
        c.font = white_font
        c.fill = gold_fill
        c.alignment = Alignment(horizontal="center")
    ws.append([
        data["total"],
        data["by_status"].get("programmé", 0),
        data["by_status"].get("en_cours", 0),
        data["by_status"].get("terminé", 0),
        data["total_guests"],
    ])
    for c in ws[5]:
        c.alignment = Alignment(horizontal="center")
        c.font = Font(bold=True, size=12)
    for w, col in zip([14, 14, 14, 14, 14], "ABCDE"):
        ws.column_dimensions[col].width = w

    # ---------- Sheet 2: Par bateau ----------
    ws2 = wb.create_sheet("Par bateau")
    headers = ["Bateau", "Total", "Terminées", "Passagers"]
    ws2.append(headers)
    for c in ws2[1]:
        c.font = white_font
        c.fill = gold_fill
        c.alignment = Alignment(horizontal="center")
    for b in data["by_boat"]:
        ws2.append([b["bateau_name"], b["total"], b["terminé"], b["guests"]])
    for w, col in zip([32, 10, 12, 12], "ABCD"):
        ws2.column_dimensions[col].width = w

    # ---------- Sheet 3: Par jour ----------
    ws3 = wb.create_sheet("Par jour")
    headers = ["Jour", "Total", "Programmées", "En cours", "Terminées", "Passagers"]
    ws3.append(headers)
    for c in ws3[1]:
        c.font = white_font
        c.fill = gold_fill
        c.alignment = Alignment(horizontal="center")
    for d in data["by_day"]:
        ws3.append([d["date"], d["total"], d.get("programmé", 0), d.get("en_cours", 0), d.get("terminé", 0), d.get("guests", 0)])
    for w, col in zip([14, 10, 14, 12, 12, 12], "ABCDEF"):
        ws3.column_dimensions[col].width = w

    # ---------- Sheet 4: Détail traversées ----------
    ws4 = wb.create_sheet("Détail")
    headers = ["Date", "Bateau", "Heure", "Direction", "Statut", "Passagers (lignes)", "Convives (total)"]
    ws4.append(headers)
    for c in ws4[1]:
        c.font = white_font
        c.fill = gold_fill
        c.alignment = Alignment(horizontal="center")
    for i, c in enumerate(data["items"], start=2):
        ws4.append([
            c.get("date", ""),
            c.get("bateau_name", "—"),
            c.get("depart_time", ""),
            "Aller" if c.get("direction", "aller") == "aller" else "Retour",
            c.get("status", ""),
            c.get("passenger_count", 0),
            c.get("guests", 0),
        ])
        if i % 2 == 0:
            for cell in ws4[i]:
                cell.fill = soft_fill
    for cell in list(ws4.rows):
        for c in cell:
            c.border = border
    for w, col in zip([12, 28, 10, 12, 14, 18, 16], "ABCDEFG"):
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"bbr-traversees-{period}-{ref}.xlsx"
    from fastapi.responses import Response as _Response
    return _Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api.get("/staff/poles/{pole_id}/report.pdf")
async def staff_pole_report_pdf(pole_id: str, staff=Depends(get_current_staff)):
    """Generate a print-ready monthly PDF report for the given pôle.
    Contains: header + KPIs + sub-offers + analytics breakdowns.
    """
    await _require_role(staff, ["manager", "admin"])
    overview = await staff_pole_overview(pole_id, staff=staff)  # reuse full computation
    pole = overview["pole"]
    kpis = overview["kpis"]
    sub_offers = overview["sub_offers"]
    analytics = overview["analytics"] or {}
    recent = overview["recent_bookings"] or []

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from fastapi.responses import StreamingResponse
    import io

    styles = _pdf_styles()
    GOLD = styles["GOLD"]
    DARK = styles["DARK"]
    LIGHT = styles["LIGHT"]
    MUTED = styles["MUTED"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=2 * cm, bottomMargin=2 * cm, leftMargin=2 * cm, rightMargin=2 * cm,
        title=f"Rapport pôle {pole['name_fr']}",
    )
    el = []
    now_dt = datetime.now(timezone.utc)
    now_human = now_dt.strftime("%d/%m/%Y %H:%M UTC")

    # ===== Header =====
    el.append(Paragraph("BOULAY BEACH RESORT", styles["sub"]))
    el.append(Paragraph(f"Rapport mensuel — Pôle {pole['name_fr']}", styles["h1"]))
    el.append(Paragraph(f"Période glissante : 30 derniers jours · Généré le {now_human}", styles["small"]))
    el.append(Spacer(1, 16))

    # ===== KPIs grid =====
    el.append(Paragraph("Vue d'ensemble", styles["h2"]))
    kpi_rows = [
        ["Réservations 30j", str(kpis["last_30d"].get("count", 0))],
        ["CA 30j", _format_xof(kpis["last_30d"].get("revenue", 0))],
        ["Réservations aujourd'hui", str(kpis["today"].get("count", 0))],
        ["Convives attendus aujourd'hui", str(kpis["today"].get("guests", 0))],
        ["Revenus encaissés aujourd'hui", _format_xof(kpis["today"].get("revenue", 0))],
        ["Panier moyen 30j", _format_xof(analytics.get("avg_basket", 0))],
        ["Délai moyen de réservation", f"{analytics.get('avg_lead_time_days', 0)} jours"],
        ["Taux de paiement", f"{analytics.get('paid_rate', 0)} %"],
        ["Taux d'annulation", f"{analytics.get('cancellation_rate', 0)} %"],
        [
            "Convives 30j",
            f"{(analytics.get('guests_breakdown', {}).get('adults', 0))} adultes · "
            f"{(analytics.get('guests_breakdown', {}).get('children', 0))} enfants",
        ],
    ]
    t = Table(kpi_rows, colWidths=[8 * cm, 8 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("TEXTCOLOR", (0, 0), (0, -1), MUTED),
        ("TEXTCOLOR", (1, 0), (1, -1), DARK),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, -2), 0.25, colors.HexColor("#E5E5E2")),
    ]))
    el.append(t)
    el.append(Spacer(1, 14))

    # ===== Sub-offers =====
    if sub_offers:
        el.append(Paragraph("Sous-offres", styles["h2"]))
        rows = [["Sous-offre", "Réservations", "CA (30j)", "Convives", "Occupation"]]
        for s in sub_offers:
            stats = s.get("stats") or {}
            occ = s.get("occupancy_pct")
            occ_str = "—" if occ is None else f"{occ} %"
            rows.append([
                s.get("name_fr", ""),
                str(stats.get("count", 0)),
                _format_xof(stats.get("revenue", 0)),
                str(stats.get("guests", 0)),
                occ_str,
            ])
        t = Table(rows, colWidths=[6 * cm, 2.5 * cm, 3.8 * cm, 2 * cm, 2.5 * cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GOLD),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5E2")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ]))
        el.append(t)
        el.append(Spacer(1, 14))

    # ===== Status pipeline =====
    by_status = analytics.get("by_status") or []
    if any(s.get("count", 0) for s in by_status):
        el.append(Paragraph("Pipeline des statuts", styles["h2"]))
        STATUS_FR = {
            "pending": "En attente", "confirmed": "Confirmée", "arrived": "Arrivée",
            "completed": "Terminée", "cancelled": "Annulée",
        }
        rows = [["Statut", "Nombre"]]
        for s in by_status:
            rows.append([STATUS_FR.get(s["status"], s["status"]), str(s.get("count", 0))])
        t = Table(rows, colWidths=[10 * cm, 4 * cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GOLD),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5E2")),
        ]))
        el.append(t)
        el.append(Spacer(1, 14))

    # ===== Payment methods =====
    pmethods = analytics.get("by_payment_method") or []
    if pmethods:
        el.append(Paragraph("Répartition par méthode de paiement", styles["h2"]))
        METHOD_FR = {"cash": "Espèces", "card": "Carte", "mobile_money": "Mobile Money", "fineo": "FINEO", "deposit": "Acompte", "unknown": "Non défini"}
        rows = [["Méthode", "Nombre", "Montant total"]]
        for p in sorted(pmethods, key=lambda x: x.get("total", 0), reverse=True):
            rows.append([
                METHOD_FR.get(p.get("method"), p.get("method")),
                str(p.get("count", 0)),
                _format_xof(p.get("total", 0)),
            ])
        t = Table(rows, colWidths=[7 * cm, 4 * cm, 5 * cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GOLD),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ]))
        el.append(t)
        el.append(Spacer(1, 14))

    # ===== Weekday =====
    weekday = analytics.get("by_weekday") or []
    if any(w.get("count", 0) for w in weekday):
        el.append(Paragraph("Répartition par jour de la semaine", styles["h2"]))
        rows = [["Jour", "Réservations", "CA"]]
        for w in weekday:
            rows.append([w["day"], str(w.get("count", 0)), _format_xof(w.get("revenue", 0))])
        t = Table(rows, colWidths=[5 * cm, 4 * cm, 7 * cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GOLD),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ]))
        el.append(t)
        el.append(Spacer(1, 14))

    # ===== Boat times =====
    btimes = analytics.get("by_boat_time") or []
    if btimes:
        el.append(Paragraph("Horaires de traversée les plus demandés", styles["h2"]))
        rows = [["Horaire", "Réservations"]]
        for b in btimes:
            rows.append([b.get("boat_time", "—"), str(b.get("count", 0))])
        t = Table(rows, colWidths=[10 * cm, 4 * cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GOLD),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ]))
        el.append(t)
        el.append(Spacer(1, 14))

    # ===== Top clients =====
    top_clients = analytics.get("top_clients") or []
    if top_clients:
        el.append(Paragraph("Top 5 clients (par CA)", styles["h2"]))
        rows = [["#", "Nom", "Téléphone", "Réservations", "CA"]]
        for i, c in enumerate(top_clients, start=1):
            rows.append([str(i), c.get("name", "—"), c.get("phone", "—"), str(c.get("count", 0)), _format_xof(c.get("total", 0))])
        t = Table(rows, colWidths=[0.8 * cm, 6 * cm, 4 * cm, 2.5 * cm, 3.5 * cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GOLD),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ]))
        el.append(t)
        el.append(Spacer(1, 14))

    # ===== Recent bookings (last 10) =====
    if recent:
        el.append(PageBreak())
        el.append(Paragraph("Réservations récentes", styles["h2"]))
        rows = [["Date", "Heure", "Offre", "Convives", "Montant", "Statut"]]
        STATUS_FR = {
            "pending": "En attente", "confirmed": "Confirmée", "arrived": "Arrivée",
            "completed": "Terminée", "cancelled": "Annulée",
        }
        for b in recent[:15]:
            adults = int(b.get("adults", 0) or 0)
            children = int(b.get("children", 0) or 0)
            convives = f"{adults}A" + (f" +{children}E" if children else "")
            date_iso = b.get("date") or ""
            m = date_iso.split("-") if date_iso else []
            date_fr = f"{m[2]}/{m[1]}/{m[0]}" if len(m) == 3 else "—"
            rows.append([
                date_fr,
                b.get("boat_time") or "—",
                b.get("offer_name", "")[:32],
                convives,
                _format_xof(b.get("total_amount", 0)),
                STATUS_FR.get(b.get("status"), b.get("status", "")),
            ])
        t = Table(rows, colWidths=[2.2 * cm, 1.6 * cm, 5.5 * cm, 2 * cm, 3 * cm, 2.7 * cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GOLD),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("ALIGN", (4, 1), (4, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ]))
        el.append(t)

    footer = _pdf_footer_factory(styles)
    doc.build(el, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    fname = f"bbr-pole-{pole_id}-{now_dt.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------- Shared PDF helpers ----------
def _pdf_styles():
    """Return common reportlab styles used by all BBr PDF reports."""
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    GOLD = colors.HexColor("#B8922A")
    DARK = colors.HexColor("#0A0A0A")
    LIGHT = colors.HexColor("#FAFAF7")
    MUTED = colors.HexColor("#888888")
    base = getSampleStyleSheet()
    return {
        "GOLD": GOLD, "DARK": DARK, "LIGHT": LIGHT, "MUTED": MUTED,
        "h1": ParagraphStyle("h1", parent=base["Title"], fontName="Helvetica-Bold", fontSize=22, textColor=DARK, alignment=0, spaceAfter=4),
        "sub": ParagraphStyle("sub", parent=base["Normal"], fontName="Helvetica", fontSize=10, textColor=GOLD, alignment=0, spaceAfter=16, leading=14),
        "h2": ParagraphStyle("h2", parent=base["Heading2"], fontName="Helvetica-Bold", fontSize=12, textColor=GOLD, spaceBefore=12, spaceAfter=6),
        "body": ParagraphStyle("body", parent=base["Normal"], fontName="Helvetica", fontSize=9, textColor=DARK),
        "small": ParagraphStyle("small", parent=base["Normal"], fontName="Helvetica", fontSize=8, textColor=MUTED),
    }


def _pdf_footer_factory(styles):
    """Return a (canvas, doc) footer drawer using shared style."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    GOLD, MUTED = styles["GOLD"], styles["MUTED"]

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setStrokeColor(GOLD)
        canvas.setLineWidth(0.5)
        canvas.line(2 * cm, 1.5 * cm, A4[0] - 2 * cm, 1.5 * cm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(MUTED)
        canvas.drawString(2 * cm, 1 * cm, "Boulay Beach Resort — Rapport interne — Confidentiel")
        canvas.drawRightString(A4[0] - 2 * cm, 1 * cm, f"Page {doc_.page}")
        canvas.restoreState()
    return _footer


def _format_xof(amount: int) -> str:
    """Format an integer as XOF amount, e.g. 1 500 000 FCFA."""
    try:
        n = int(amount or 0)
    except Exception:
        n = 0
    return f"{n:,}".replace(",", " ") + " FCFA"


async def _resolve_qr_token(raw: str):
    """Resolve a raw user-supplied token into the real qr_token.

    A booking's `qr_codes[].qr_token` is a 32-hex-character lowercase string. However
    the printed PNG ticket only shows a short reference (e.g. "5DF111909C", which is
    `token[:10].upper()`). Receptionists tend to type that visible code into the manual
    scanner. Without normalisation the lookup would 404 because of case + length mismatch.

    Strategy:
      0. If the input is the raw JSON payload that the QR encodes
         (``{"type":"ticket","token":"…","ref":"…"}``), extract ``token`` first.
      1. Try exact match (current behaviour, fastest).
      2. Else lowercase and try again.
      3. Else treat the input as a prefix (>=8 chars) and search via regex.
      4. NEW (iter-41): Else treat the input as a `ref` (booking_id[:8].upper())
         and search by booking id prefix — defends against scanner clients that
         only forward the "ref" portion.
    Returns the booking dict + the matching qr_token, or (None, None).
    """
    if not raw:
        return None, None
    raw = raw.strip()
    # 0. Defensive JSON extraction — supports clients that forward the full QR
    #    payload without parsing it first.
    if raw.startswith("{"):
        try:
            obj = json.loads(raw)
            token_from_json = (obj.get("token") or obj.get("qr_token")
                               or obj.get("guest_token") or "").strip()
            if token_from_json:
                raw = token_from_json
        except Exception:
            pass  # not valid JSON — fall through to the legacy paths
    # 1. Exact
    booking = await db.bookings.find_one({"qr_codes.qr_token": raw}, {"_id": 0})
    if booking:
        return booking, raw
    # 2. Lowercase exact
    low = raw.lower()
    if low != raw:
        booking = await db.bookings.find_one({"qr_codes.qr_token": low}, {"_id": 0})
        if booking:
            return booking, low
    # 3. Prefix match (only if user typed >=8 hex chars, to avoid ambiguity)
    import re as _re
    if _re.fullmatch(r"[0-9a-f]{8,}", low):
        pattern = _re.compile(f"^{_re.escape(low)}")
        booking = await db.bookings.find_one({"qr_codes.qr_token": {"$regex": pattern}}, {"_id": 0})
        if booking:
            real = next(
                (q.get("qr_token") for q in booking.get("qr_codes", []) if q.get("qr_token", "").startswith(low)),
                None,
            )
            if real:
                return booking, real
    # 4. iter-41: fall back to booking-id prefix (the "ref" portion printed on the
    #    ticket: ``booking_id[:8].upper()``). Useful if a 3rd-party scanner app
    #    only forwards the ref code.
    if _re.fullmatch(r"[0-9a-f]{6,}", low):
        booking = await db.bookings.find_one(
            {"id": {"$regex": f"^{_re.escape(low)}", "$options": "i"},
             "qr_codes.0": {"$exists": True}},
            {"_id": 0},
        )
        if booking and booking.get("qr_codes"):
            # Return the booker's qr_token (the first entry is conventionally the booker)
            return booking, booking["qr_codes"][0].get("qr_token")
    return None, None


# ---------- QR Scanner (Module 4) ----------
@api.get("/staff/scan/{qr_token}")
async def scan_qr(qr_token: str, staff=Depends(get_current_staff)):
    """Look up a booking by QR token. Returns the participant + booking summary + scan history.

    Accepts: full 32-hex token (camera scan), lowercase/uppercase variants, OR the 10-char
    reference code (or any >=8-char prefix) printed on the styled PNG ticket.
    """
    booking, real_token = await _resolve_qr_token(qr_token)
    if not booking:
        raise HTTPException(status_code=404, detail="QR code non reconnu")
    guest = next((q for q in booking.get("qr_codes", []) if q.get("qr_token") == real_token), None)
    scans = (guest or {}).get("scans", [])
    next_direction = "aller" if len(scans) == 0 else ("retour" if len(scans) == 1 else None)

    # Fetch wallet to surface participant-level activities + consumptions
    wallet = None
    participant_charges: List[dict] = []
    participant_total = 0
    wallet_total = 0
    wallet_status = None
    if booking.get("wallet_token"):
        wallet = await db.wallets.find_one({"token": booking["wallet_token"]}, {"_id": 0})
        if wallet:
            wallet_status = wallet.get("status")
            for tx in wallet.get("transactions", []) or []:
                if tx.get("status") == "voided":
                    continue
                wallet_total += int(tx.get("amount", 0))
                if (tx.get("participant_token") or "").lower() == real_token.lower():
                    participant_charges.append(tx)
                    participant_total += int(tx.get("amount", 0))
    participant_charges.sort(key=lambda t: t.get("created_at", ""), reverse=True)
    summary = {
        "booking_id": booking["id"],
        "offer_type": booking.get("offer_type"),
        "offer_name": booking["offer_name"],
        "date": booking["date"],
        "checkout_date": booking.get("checkout_date"),
        "nights": booking.get("nights") or 0,
        "boat_time": booking.get("boat_time"),
        "return_boat_time": booking.get("return_boat_time"),
        "rooms": booking.get("rooms") or 1,
        "room_tier": booking.get("room_tier"),
        "room_tier_name": booking.get("room_tier_name"),
        "adults": booking.get("adults"),
        "children": booking.get("children"),
        "status": booking.get("status"),
        "payment_method": booking.get("payment_method"),
        "total_amount": booking.get("total_amount", 0),
        "paid_amount": booking.get("paid_amount", 0),
        "balance_due": booking.get("balance_due", 0),
        "deposit_pct": booking.get("deposit_pct"),
        "phone": booking.get("phone", ""),
        "email": booking.get("email", ""),
        "special_requests": booking.get("special_requests", ""),
        "wallet_token": booking.get("wallet_token"),
        "guest_name": guest.get("guest_name") if guest else "",
        "guest_surname": guest.get("guest_surname") if guest else "",
        "guest_nationality": guest.get("guest_nationality") if guest else "",
        "guest_phone": guest.get("guest_phone") if guest else "",
        "guest_email": guest.get("guest_email") if guest else "",
        "guest_label_fr": guest.get("label_fr") if guest else "",
        # iter-30: full party composition attached to this pass (adults + children breakdown)
        "composition": (guest or {}).get("composition"),
        "booking_code": booking.get("booking_code"),
        "qr_token": real_token,
        "scans": scans,
        "scan_count": len(scans),
        "next_direction": next_direction,
        "fully_used": next_direction is None,
        # Wallet / participant traceability
        "wallet_status": wallet_status,
        "wallet_total_charged": wallet_total,
        "participant_charges": participant_charges,
        "participant_total_charged": participant_total,
    }
    return summary


class WalletCharge(BaseModel):
    activity_id: Optional[str] = None
    label: Optional[str] = None
    amount: int = Field(default=0, ge=0)
    note: Optional[str] = ""
    quantity: int = Field(default=1, ge=1, le=20)
    # Optional participant traceability: when present, the charge is tagged with the
    # ticket QR token of the participant who consumed it (Sky Nautique, Quad,
    # boissons…). Allows per-guest tracing & filtering.
    participant_token: Optional[str] = None


@api.post("/staff/scan/{qr_token}/charge")
async def charge_via_scan(qr_token: str, body: WalletCharge, staff=Depends(get_current_staff)):
    """Charge an activity / consumption directly via the participant's ticket QR.

    The staff scans the participant QR, taps an activity or types a free amount —
    the charge is added to the booking's wallet AND tagged with the participant
    token so we know who consumed what.
    """
    booking, real_token = await _resolve_qr_token(qr_token)
    if not booking:
        raise HTTPException(status_code=404, detail="QR code non reconnu")
    if not booking.get("wallet_token"):
        raise HTTPException(status_code=400, detail="Aucune carte de consommation associée à cette réservation.")
    body.participant_token = real_token
    return await charge_wallet(booking["wallet_token"], body, staff=staff)  # type: ignore


class CheckinOverride(BaseModel):
    """Optional override applied by the staff when the guest didn't take the boat
    originally planned at booking time (missed it / swapped). When omitted, the
    boat is auto-resolved from the booking's planned schedule."""
    boat_time: Optional[str] = None
    boat_id: Optional[str] = None
    boat_name: Optional[str] = None
    skipper_name: Optional[str] = None
    direction: Optional[Literal["aller", "retour"]] = None


@api.post("/staff/scan/{qr_token}/checkin")
async def checkin_qr(
    qr_token: str,
    body: Optional[CheckinOverride] = None,
    staff=Depends(get_current_staff),
):
    """Register an embarkation scan (aller then retour). Max 2 scans per QR.

    Rules:
    - First scan → direction='aller' and booking status becomes 'arrived' if not already
    - Second scan → direction='retour' and booking status becomes 'completed'
    - Third scan → 400 'QR code déjà utilisé entièrement'
    Each scan stores: direction, scanned_at, staff_email, staff_name, boat_*.

    Optional body lets the staff override the boat actually taken (when the guest
    missed the planned one and embarked on a later/earlier crossing).

    Accepts the same flexible token formats as GET /staff/scan/{qr_token}.
    """
    booking, real_token = await _resolve_qr_token(qr_token)
    if not booking:
        raise HTTPException(status_code=404, detail="QR code non reconnu")
    qrs = booking.get("qr_codes", [])
    idx = next((i for i, q in enumerate(qrs) if q.get("qr_token") == real_token), -1)
    if idx == -1:
        raise HTTPException(status_code=404, detail="QR code non reconnu")
    scans = qrs[idx].get("scans") or []
    # Passport tickets: this QR is scannable on each booked date, max 2 scans
    # (aller + retour) per date. Validate that today matches one of the valid
    # dates of the ticket, and count scans for today only.
    valid_dates = qrs[idx].get("valid_dates") or [booking.get("date")]
    today_iso = datetime.now(timezone.utc).date().isoformat()
    is_passport_qr = len(valid_dates) > 1
    if is_passport_qr and today_iso not in valid_dates:
        raise HTTPException(
            status_code=400,
            detail=f"Ce passeport n'est pas valide aujourd'hui. Dates autorisées : {', '.join(valid_dates)}",
        )
    if is_passport_qr:
        today_scans = [s for s in scans if (s.get("scan_date") or s.get("boat_date")) == today_iso]
        if len(today_scans) >= 2:
            raise HTTPException(status_code=400, detail=f"QR code déjà scanné (aller + retour) pour le {today_iso}.")
    else:
        if len(scans) >= 2:
            raise HTTPException(status_code=400, detail="QR code déjà scanné (aller + retour). Plus aucun embarquement possible.")
    # Direction may be forced by the staff (rare: scan a 'retour' before 'aller' has been done).
    forced_dir = (body.direction if body else None) if body else None
    if is_passport_qr:
        prior_today_count = len([s for s in scans if (s.get("scan_date") or s.get("boat_date")) == today_iso])
        direction = forced_dir or ("aller" if prior_today_count == 0 else "retour")
        boat_date = today_iso
    else:
        direction = forced_dir or ("aller" if len(scans) == 0 else "retour")
        boat_date = booking.get("date") if direction == "aller" else (booking.get("checkout_date") or booking.get("date"))
    # Default boat from the booking; can be overridden by the staff (missed boat, etc.)
    planned_boat = booking.get("boat_time") if direction == "aller" else (booking.get("return_boat_time") or booking.get("boat_time"))
    boat_time = (body.boat_time if body else None) or planned_boat
    boat_id = (body.boat_id if body else None)
    boat_name = (body.boat_name if body else None)
    skipper_name = ((body.skipper_name if body else None) or "").strip() or None
    if boat_id and not boat_name:
        bateau = await db.bateaux.find_one({"id": boat_id}, {"_id": 0, "name": 1})
        if bateau:
            boat_name = bateau.get("name")
    boat_label_bits = [b for b in [boat_time, direction] if b]
    boat_label = " ".join(boat_label_bits) or direction
    if boat_name:
        boat_label = f"{boat_label} · {boat_name}"
    overridden = bool(body and (body.boat_time or body.boat_id) and (body.boat_time or "") != (planned_boat or ""))
    entry = {
        "direction": direction,
        "scanned_at": now_iso(),
        "scan_date": today_iso,
        "staff_email": staff.get("email"),
        "staff_name": staff.get("name") or "",
        "boat_time": boat_time,
        "boat_id": boat_id,
        "boat_name": boat_name,
        "boat_date": boat_date,
        "boat_label": boat_label,
        "planned_boat_time": planned_boat,
        "overridden": overridden,
        "skipper_name": skipper_name,
    }
    scans = scans + [entry]
    # Aggregate booking-level status across all QR codes:
    #  - 'arrived' if at least one aller scan and not everyone has done a return
    #  - 'completed' once all participants have done both aller + retour for
    #    every valid date (single QR per adult for passport tickets).
    all_scans_after = [
        (q.get("scans") or []) + ([entry] if i == idx else [])
        for i, q in enumerate(qrs)
    ]

    def _qr_arrived(q_scans: list, q_meta: dict) -> bool:
        """True iff this QR has at least 1 scan (any date)."""
        return len(q_scans) >= 1

    def _qr_completed(q_scans: list, q_meta: dict) -> bool:
        """True iff this QR has aller+retour for each of its valid dates."""
        vd = q_meta.get("valid_dates") or [booking.get("date")]
        if len(vd) <= 1:
            return len(q_scans) >= 2
        from collections import Counter
        per_day = Counter(s.get("scan_date") or s.get("boat_date") for s in q_scans)
        return all(per_day.get(d, 0) >= 2 for d in vd)

    all_arrived = all(_qr_arrived(s, qrs[i]) for i, s in enumerate(all_scans_after))
    all_completed = all(_qr_completed(s, qrs[i]) for i, s in enumerate(all_scans_after))
    new_status = booking.get("status")
    set_ops = {f"qr_codes.{idx}.scans": scans}
    if all_completed:
        new_status = "completed"
        set_ops["status"] = "completed"
        set_ops["completed_at"] = now_iso()
    elif all_arrived and booking.get("status") in (None, "confirmed", "pending"):
        new_status = "arrived"
        set_ops["status"] = "arrived"
        set_ops["arrived_at"] = booking.get("arrived_at") or now_iso()
    await db.bookings.update_one({"id": booking["id"]}, {"$set": set_ops})
    return {
        "ok": True,
        "direction": direction,
        "scanned_at": entry["scanned_at"],
        "staff_email": entry["staff_email"],
        "staff_name": entry["staff_name"],
        "boat_time": entry["boat_time"],
        "boat_id": entry["boat_id"],
        "boat_name": entry["boat_name"],
        "boat_date": entry["boat_date"],
        "boat_label": entry["boat_label"],
        "planned_boat_time": entry["planned_boat_time"],
        "overridden": entry["overridden"],
        "skipper_name": entry["skipper_name"],
        "scan_count": len(scans),
        "next_direction": "retour" if len(scans) == 1 else None,
        "fully_used": len(scans) >= 2,
        "booking_status": new_status,
    }


# /staff/checkins/history MOVED to routers/scanner_history.py in iter-25
# (the remaining /staff/scan/* endpoints stay here for now — to be extracted
#  in a follow-up iteration together with _resolve_qr_token + wallet helpers)


@api.get("/staff/skippers")
async def list_skippers(staff=Depends(get_current_staff)):
    """Return the list of registered skippers (catalog). Each entry has id,
    name, phone, license_no, status. Sorted alphabetically by name. Adds an
    optional `recent_scans_count` field aggregated from past QR scans so the
    UI can hint at activity, but the catalog itself is independent."""
    # Backfill: ensure the `skippers` collection has an _id-style id field
    items = await db.skippers.find({}, {"_id": 0}).sort("name", 1).to_list(length=500)
    return {"items": items}


@api.get("/staff/skippers/recent")
async def list_recent_skipper_names(staff=Depends(get_current_staff)):
    """Return distinct skipper *names* that have been entered on past scans —
    used by the scanner modal autocomplete. Kept separate from the catalog so
    historical free-text entries remain available."""
    pipeline = [
        {"$match": {"qr_codes.scans.skipper_name": {"$nin": [None, ""]}}},
        {"$unwind": "$qr_codes"},
        {"$unwind": "$qr_codes.scans"},
        {"$match": {"qr_codes.scans.skipper_name": {"$nin": [None, ""]}}},
        {"$group": {
            "_id": "$qr_codes.scans.skipper_name",
            "last_used": {"$max": "$qr_codes.scans.scanned_at"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"last_used": -1}},
        {"$limit": 30},
        {"$project": {"_id": 0, "name": "$_id", "last_used": 1, "count": 1}},
    ]
    items = [r async for r in db.bookings.aggregate(pipeline)]
    return {"items": items}


@api.post("/staff/skippers")
async def create_skipper(body: Skipper, staff=Depends(get_current_staff)):
    if staff.get("role") not in {"admin", "manager", "logistique"}:
        raise HTTPException(status_code=403, detail="Action réservée aux managers / logistique")
    name_clean = body.name.strip()
    if await db.skippers.find_one({"name": {"$regex": f"^{name_clean}$", "$options": "i"}}):
        raise HTTPException(status_code=409, detail="Un skipper avec ce nom existe déjà")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name_clean,
        "phone": (body.phone or "").strip() or None,
        "license_no": (body.license_no or "").strip() or None,
        "status": body.status,
        "created_at": now_iso(),
    }
    await db.skippers.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api.patch("/staff/skippers/{skipper_id}")
async def update_skipper(skipper_id: str, body: SkipperUpdate, staff=Depends(get_current_staff)):
    if staff.get("role") not in {"admin", "manager", "logistique"}:
        raise HTTPException(status_code=403, detail="Action réservée aux managers / logistique")
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if "name" in update:
        update["name"] = update["name"].strip()
    if not update:
        raise HTTPException(status_code=400, detail="Aucun champ à modifier")
    update["updated_at"] = now_iso()
    res = await db.skippers.update_one({"id": skipper_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Skipper introuvable")
    # Propagate the name change to scheduled (not yet started) traversées
    if "name" in update:
        await db.traversees.update_many(
            {"skipper_id": skipper_id, "status": "programmé"},
            {"$set": {"skipper_name": update["name"]}},
        )
    return {"ok": True}


@api.delete("/staff/skippers/{skipper_id}")
async def delete_skipper(skipper_id: str, staff=Depends(get_current_staff)):
    if staff.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Action réservée aux administrateurs")
    # Detach from any scheduled traversées first.
    await db.traversees.update_many(
        {"skipper_id": skipper_id, "status": "programmé"},
        {"$set": {"skipper_id": None, "skipper_name": None}},
    )
    res = await db.skippers.delete_one({"id": skipper_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Skipper introuvable")
    return {"ok": True}


@api.post("/staff/bookings/{booking_id}/arrived")
async def mark_arrived(booking_id: str, staff=Depends(get_current_staff)):
    res = await db.bookings.update_one({"id": booking_id}, {"$set": {"status": "arrived", "arrived_at": now_iso()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    return {"ok": True}


# =================================================================
# MODULE 2 — Reservations management (list, filters, detail, actions)
# =================================================================

@api.get("/staff/bookings")
async def list_bookings(
    offer_type: Optional[str] = None,
    pole: Optional[str] = None,
    status: Optional[str] = None,
    payment_status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    include_pending: bool = False,
    limit: int = 200,
    staff=Depends(get_current_staff),
):
    """List bookings with filters. payment_status = paid | unpaid.

    iter-44: by default the result EXCLUDES status='pending' (abandoned carts)
    so the operational lists and the chiffre d'affaires only show real
    bookings (paid online + cash-on-arrival). Pass `include_pending=true`
    or filter explicitly via `status=pending` to surface them — this is what
    the dedicated "Réservations en attente" page does.
    """
    await _require_role(staff, ["manager", "admin"])
    q: dict = {}
    if offer_type:
        q["offer_type"] = offer_type
    if pole:
        # Match bookings tagged with the pole, or fallback to offer_type ∈ pole.offers
        # so legacy bookings (without `pole` field) still surface.
        if pole in POLES:
            offers_in_pole = list(POLES[pole].get("offers", []))
            # special_event lives under activites_events via OFFER_TO_POLE
            if pole == "activites_events":
                offers_in_pole = list(set(offers_in_pole + ["special_event"]))
            q["$or"] = (q.get("$or") or []) + [
                {"pole": pole},
                {"offer_type": {"$in": offers_in_pole}},
            ]
        else:
            q["pole"] = pole
    if status:
        q["status"] = status
    elif not include_pending:
        # iter-44: hide abandoned-cart "pending" by default
        q["status"] = {"$ne": "pending"}
    if date_from or date_to:
        d: dict = {}
        if date_from:
            d["$gte"] = date_from
        if date_to:
            d["$lte"] = date_to
        q["date"] = d
    if payment_status == "paid":
        q["paid_at"] = {"$ne": None}
    elif payment_status == "unpaid":
        q["$or"] = [{"paid_at": None}, {"paid_at": {"$exists": False}}]
    if search:
        s = search.strip()
        q.setdefault("$or", []).extend([
            {"phone": {"$regex": s, "$options": "i"}},
            {"email": {"$regex": s, "$options": "i"}},
            {"participants.name": {"$regex": s, "$options": "i"}},
            {"participants.surname": {"$regex": s, "$options": "i"}},
        ])
    # iter-32: Sort strictly by creation time DESC so the freshest booking is
    # always on top (user request: see latest reservations without scrolling).
    cursor = db.bookings.find(
        q,
        {"_id": 0, "reference_token": 0, "qr_codes.qr_code": 0, "qr_codes.qr_payload": 0, "qr_codes.ticket_image": 0},
    ).sort([("created_at", -1), ("date", -1)]).limit(limit)
    items = await cursor.to_list(length=limit)
    return items


# iter-44: Réservations en attente (abandoned-cart) — dedicated relance funnel
@api.get("/staff/bookings/pending")
async def list_pending_bookings(
    days: int = 90,
    search: Optional[str] = None,
    limit: int = 500,
    staff=Depends(get_current_staff),
):
    """List bookings still in ``pending`` (created via the public tunnel but
    payment never completed). Used by the new staff page
    ``/staff/reservations/en-attente`` for relance / follow-up.

    Filters:
      - ``days``: only show pendings created within the last N days (default 90)
      - ``search``: phone / email / name (case-insensitive substring)
    """
    await _require_role(staff, ["manager", "admin"])
    from datetime import timedelta as _td
    cutoff = (datetime.now(timezone.utc) - _td(days=max(1, days))).isoformat()
    q: dict = {"status": "pending", "created_at": {"$gte": cutoff}}
    if search:
        s = search.strip()
        q["$or"] = [
            {"phone": {"$regex": s, "$options": "i"}},
            {"email": {"$regex": s, "$options": "i"}},
            {"participants.name": {"$regex": s, "$options": "i"}},
            {"participants.surname": {"$regex": s, "$options": "i"}},
            {"name": {"$regex": s, "$options": "i"}},
            {"surname": {"$regex": s, "$options": "i"}},
        ]
    cursor = db.bookings.find(
        q,
        {"_id": 0, "reference_token": 0,
         "qr_codes": 0, "ticket_image": 0},
    ).sort("created_at", -1).limit(limit)
    items = await cursor.to_list(length=limit)
    # Compute a synthetic "relance" status for the UI:
    #   never_relaunched | relaunched (N times) | stale (>14 days)
    now = datetime.now(timezone.utc)
    out = []
    for b in items:
        relance = b.get("relance_log") or []
        last = relance[-1] if relance else None
        age_days = None
        try:
            created = datetime.fromisoformat(b["created_at"].replace("Z", "+00:00"))
            age_days = int((now - created).total_seconds() // 86400)
        except Exception:
            pass
        out.append({
            **b,
            "relance_count": len(relance),
            "last_relance_at": (last or {}).get("at"),
            "age_days": age_days,
            "is_stale": (age_days or 0) > 14,
        })
    # Aggregate total amount stuck in limbo so the UI can show the "CA à
    # récupérer" hint above the table.
    total_pending_amount = sum(int(b.get("total_amount") or 0) for b in out)
    return {"items": out, "total": len(out),
            "total_pending_amount": total_pending_amount}


@api.post("/staff/bookings/{booking_id}/resend-payment-link")
async def resend_payment_link(booking_id: str, staff=Depends(get_current_staff)):
    """Regenerate (if expired) and email a fresh FineoPay payment link to the
    client of a pending booking. Stamps the action in ``relance_log`` so the
    staff can see how many times each client has been contacted.
    """
    await _require_role(staff, ["manager", "admin"])
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Réservation introuvable.")
    if booking.get("status") != "pending":
        raise HTTPException(
            status_code=400,
            detail=f"Cette réservation n'est pas en attente (statut: {booking.get('status')}).",
        )

    # Re-issue / extend the payment-link token (7-day window)
    from secrets import token_urlsafe
    token = booking.get("payment_link_token") or token_urlsafe(24)
    expires_at = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()

    base = os.environ.get("PUBLIC_FRONTEND_URL") or os.environ.get("FRONTEND_URL", "")
    payment_url = f"{base.rstrip('/')}/pay/{token}" if base else f"/pay/{token}"

    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "payment_link_token": token,
            "payment_link_expires_at": expires_at,
        }},
    )

    # Send email
    offer = OFFERS.get(booking.get("offer_type", "")) or {}
    if booking.get("dates") and len(booking["dates"]) > 1:
        date_str = " · ".join(booking["dates"])
    else:
        date_str = booking.get("date", "")
    amount_label = f"{int(booking.get('total_amount', 0)):,}".replace(",", " ") + " FCFA"

    email_sent = False
    email_error = None
    try:
        from services import email_service as _es
        primary = (booking.get("participants") or [{}])[0]
        booker_name = booking.get("name") or primary.get("name") or ""
        booker_surname = booking.get("surname") or primary.get("surname") or ""
        full_name = f"{booker_name} {booker_surname}".strip()
        rendered = _es.render_payment_link(
            name=full_name,
            ref=booking["id"][:8].upper(),
            offer_label=offer.get("name_fr") or booking.get("offer_type", ""),
            date_str=date_str,
            boat_time=booking.get("boat_time"),
            amount_label=amount_label,
            payment_url=payment_url,
            expires_label="dans 7 jours",
        )
        res = await _es.send_email(
            db,
            to_email=booking["email"],
            subject="Relance — " + rendered["subject"],
            html=rendered["html"],
            plain=rendered["plain"],
            purpose="payment_link_relance",
            booking_id=booking["id"],
            to_name=full_name or None,
        )
        email_sent = bool(res.get("ok"))
    except Exception as ex:
        email_error = str(ex)
        logging.warning("Resend payment link email failed for %s: %s", booking_id, ex)

    # Stamp the relance event
    await db.bookings.update_one(
        {"id": booking_id},
        {"$push": {"relance_log": {
            "at": datetime.now(timezone.utc).isoformat(),
            "channel": "email",
            "by_staff_id": staff.get("id"),
            "by_staff_email": staff.get("email"),
            "email_sent": email_sent,
            "error": email_error,
        }}},
    )
    return {
        "ok": True,
        "email_sent": email_sent,
        "email_error": email_error,
        "payment_link": payment_url,
        "expires_at": expires_at,
    }



@api.get("/staff/bookings/calendar")
async def bookings_calendar(month: str, staff=Depends(get_current_staff)):
    """Return all bookings for a month (YYYY-MM) grouped by date for calendar view."""
    await _require_role(staff, ["manager", "admin"])
    if not month or len(month) != 7:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")
    date_from = f"{month}-01"
    next_month = datetime.strptime(date_from, "%Y-%m-%d").date() + timedelta(days=32)
    date_to = next_month.replace(day=1).isoformat()
    cursor = db.bookings.find(
        # iter-44: exclude abandoned-cart pending from the calendar view
        {"date": {"$gte": date_from, "$lt": date_to}, "status": {"$ne": "pending"}},
        {"_id": 0, "id": 1, "date": 1, "offer_type": 1, "offer_name": 1, "status": 1, "adults": 1, "children": 1, "boat_time": 1, "total_amount": 1, "paid_at": 1},
    )
    items = await cursor.to_list(length=2000)
    by_date: dict = {}
    for b in items:
        by_date.setdefault(b["date"], []).append(b)
    return {"month": month, "by_date": by_date, "total": len(items)}


@api.get("/staff/bookings/{booking_id}")
async def booking_detail(booking_id: str, staff=Depends(get_current_staff)):
    """Full booking detail (excludes heavy ticket_image / qr_code base64 payloads)."""
    await _require_role(staff, ["manager", "admin"])
    booking = await db.bookings.find_one(
        {"id": booking_id},
        {"_id": 0, "reference_token": 0, "qr_codes.qr_code": 0, "qr_codes.qr_payload": 0, "qr_codes.ticket_image": 0},
    )
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    return booking


@api.patch("/staff/bookings/{booking_id}/status")
async def update_booking_status(booking_id: str, status: str = Body(..., embed=True), staff=Depends(get_current_staff)):
    """Move booking through the lifecycle: pending → confirmed → arrived → completed → cancelled.

    Cancelled bookings cannot be re-opened directly; the only valid transition out
    of `cancelled` is staying `cancelled` (use a brand-new booking instead).
    """
    await _require_role(staff, ["manager", "admin"])
    if status not in ("pending", "confirmed", "arrived", "completed", "cancelled"):
        raise HTTPException(status_code=400, detail="Invalid status")
    existing = await db.bookings.find_one({"id": booking_id}, {"_id": 0, "status": 1})
    if not existing:
        raise HTTPException(status_code=404, detail="Booking not found")
    if existing.get("status") == "cancelled" and status != "cancelled":
        raise HTTPException(status_code=400, detail="Cancelled bookings cannot be re-opened")
    await db.bookings.update_one({"id": booking_id}, {"$set": {"status": status}})
    return {"ok": True, "status": status}


@api.patch("/staff/bookings/{booking_id}/payment")
async def update_booking_payment(
    booking_id: str,
    payment_method: str = Body(..., embed=True),
    paid: bool = Body(True, embed=True),
    staff=Depends(get_current_staff),
):
    """Mark a booking as paid / unpaid by staff (e.g. cash collected at counter).

    Only auto-confirms when the booking is still `pending` — never regresses an
    already-arrived or already-completed booking.
    """
    await _require_role(staff, ["manager", "admin"])
    existing = await db.bookings.find_one({"id": booking_id}, {"_id": 0, "status": 1})
    if not existing:
        raise HTTPException(status_code=404, detail="Booking not found")
    update: dict = {"payment_method": payment_method}
    if paid:
        update["paid_at"] = now_iso()
        if existing.get("status") == "pending":
            update["status"] = "confirmed"
    else:
        update["paid_at"] = None
    await db.bookings.update_one({"id": booking_id}, {"$set": update})
    return {"ok": True}


@api.post("/staff/bookings/{booking_id}/confirm-cash-payment")
async def confirm_cash_payment(booking_id: str, staff=Depends(get_current_staff)):
    """Validates a cash booking after the staff has physically collected the
    money. Replaces the cream "provisoire" receipt by the styled gold QR
    ticket, books the wallet, emits the fiscal receipt and sends the second
    confirmation email with the final boarding pass attached.
    Idempotent — re-running on an already-confirmed booking returns 400.
    """
    await _require_role(staff, ["hotesse", "receptionist", "manager", "manager_pole", "admin"])
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    if booking.get("payment_method") != "cash":
        raise HTTPException(status_code=400, detail="Cette réservation n'est pas un paiement en espèces")
    if booking.get("status") != "pending_cash_payment":
        raise HTTPException(status_code=400, detail="Encaissement déjà confirmé ou réservation non éligible")

    if booking["offer_type"] == "special_event":
        offer = await _resolve_special_event_offer(booking.get("special_event_id") or "", booking.get("date"))
    else:
        offer = OFFERS[booking["offer_type"]]
    participants = booking.get("participants", [])
    adult_participants = [p for p in participants if p.get("kind", "adult") == "adult"]
    children_count = int(booking.get("children", 0))
    total_amount = int(booking.get("total_amount", 0))
    booking_ref_short = booking_id[:8].upper()

    # Regenerate styled QR tickets — one per adult, booker carries children info.
    qr_codes = []
    for adult_i, p in enumerate(adult_participants, start=1):
        # Preserve the original qr_token issued when the booking was paid so the
        # token printed on the cream receipt still resolves to the same QR.
        existing = next(
            (q for q in (booking.get("qr_codes") or []) if q.get("guest_name") == p["name"] and q.get("guest_surname") == p["surname"]),
            None,
        )
        token = (existing or {}).get("qr_token") or uuid.uuid4().hex
        is_booker = adult_i == 1
        if is_booker and children_count > 0:
            label_fr = f"Réservant · +{children_count} enfant{'s' if children_count > 1 else ''}"
            label_en = f"Booker · +{children_count} child{'ren' if children_count > 1 else ''}"
        elif is_booker:
            label_fr, label_en = "Réservant", "Booker"
        else:
            label_fr, label_en = f"Adulte #{adult_i}", f"Adult #{adult_i}"
        compact_qr = json.dumps(
            {"type": "ticket", "token": token, "ref": booking_ref_short},
            ensure_ascii=False, separators=(",", ":"),
        )
        token_short = token[:10].upper()
        qr_codes.append({
            "label_fr": label_fr,
            "label_en": label_en,
            "kind": "adult",
            "guest_name": p["name"],
            "guest_surname": p["surname"],
            "guest_email": p.get("email", "") or booking.get("email", ""),
            "guest_phone": p.get("phone", "") or booking.get("phone", ""),
            "guest_nationality": p["nationality"],
            "qr_token": token,
            "qr_code": make_qr(compact_qr, styled=True),
            "ticket_image": make_ticket_image(
                offer_id=booking["offer_type"],
                offer_name=offer["name_fr"],
                date_iso=booking["date"],
                boat_time=booking.get("boat_time", ""),
                owner_name=f"{p['name']} {p['surname']}",
                qr_payload=compact_qr,
                ref_code=token_short,
                lang="fr",
                hero_url=offer.get("image_url") or None,
            ),
            "children_attached": children_count if is_booker else 0,
        })

    paid_at = now_iso()
    update = {
        "status": "confirmed",
        "qr_codes": qr_codes,
        "paid_at": paid_at,
        "paid_amount": total_amount,
        "balance_due": 0,
        "cash_confirmed_at": paid_at,
        "cash_confirmed_by": (staff.get("email") if isinstance(staff, dict) else None),
    }
    await db.bookings.update_one({"id": booking_id}, {"$set": update})
    booking.update(update)

    # Fiscal receipt
    try:
        primary = adult_participants[0] if adult_participants else {}
        await _create_receipt(
            source="booking",
            source_id=booking_id,
            customer_name=f"{primary.get('surname','').strip()} {primary.get('name','').strip()}".strip() or "—",
            customer_email=primary.get("email") or booking.get("email", ""),
            customer_phone=primary.get("phone") or booking.get("phone", ""),
            lines=[{
                "description": f"{offer['name_fr']} — {booking['date']}",
                "quantity": 1, "unit_price": total_amount, "total": total_amount,
            }],
            payment_method="cash",
            issued_by=(staff.get("email") if isinstance(staff, dict) else "staff"),
            issued_by_role=(staff.get("role") if isinstance(staff, dict) else "staff"),
            metadata={"offer_type": booking["offer_type"], "cash_confirmed": True},
        )
    except Exception as ex:
        logging.warning("Cash-confirm: receipt creation failed: %s", ex)

    # Twilio + SendGrid notifications (definitive ticket)
    try:
        qr_url = f"{FINEO_PUBLIC_BASE_URL}/api/bookings/{booking_id}/ticket.png?ref={booking['reference_token']}"
        await twilio_service.notify_booking_paid(db, booking, qr_image_url=qr_url)
    except Exception as ex:
        logging.warning("Cash-confirm: Twilio notification failed: %s", ex)
    try:
        # Send the definitive email (purpose=booking_paid). The dedup check
        # only skips if the SAME purpose was already sent — temporary email
        # uses purpose=booking_pending_cash, so this one goes through.
        await _send_booking_confirmation_email(booking, temporary=False)
    except Exception as ex:
        logging.warning("Cash-confirm: confirmation email failed: %s", ex)

    return {"ok": True, "status": "confirmed", "paid_at": paid_at}


@api.get("/staff/payments/summary")
async def payments_summary(
    pole: Optional[str] = None,
    period: Optional[str] = "30d",
    staff=Depends(get_current_staff),
):
    """Payment KPIs + lists for the dedicated /staff/payments page.
    Optional filters:
      - pole : restricts unpaid+paid stats to a given pôle
      - period : 'today' | '7d' | '30d' | 'all' — affects the paid-by-method breakdown
    """
    await _require_role(staff, ["manager", "admin"])
    pole_filter: dict = {}
    if pole and pole in POLES:
        offers_in = list(POLES[pole].get("offers", []))
        if pole == "activites_events":
            offers_in = list(set(offers_in + ["special_event"]))
        pole_filter = {"$or": [{"pole": pole}, {"offer_type": {"$in": offers_in}}]}

    unpaid_match: dict = {
        "$and": [
            {"$or": [{"paid_at": None}, {"paid_at": {"$exists": False}}]},
            {"status": {"$ne": "cancelled"}},
        ],
    }
    if pole_filter:
        unpaid_match["$and"].append(pole_filter)
    unpaid_cursor = db.bookings.find(
        unpaid_match,
        {
            "_id": 0, "id": 1, "offer_type": 1, "offer_name": 1, "date": 1,
            "total_amount": 1, "deposit_amount": 1, "deposit_pct": 1, "paid_amount": 1,
            "phone": 1, "email": 1, "participants": 1, "pole": 1, "status": 1,
            "created_at": 1, "boat_time": 1, "adults": 1, "children": 1,
        },
    ).sort("created_at", -1)
    unpaid = await unpaid_cursor.to_list(length=500)
    unpaid_total = sum(b.get("total_amount", 0) for b in unpaid)

    # Paid breakdown — period-bounded
    paid_match: dict = {"paid_at": {"$ne": None}}
    if period and period != "all":
        days = {"today": 0, "7d": 7, "30d": 30, "90d": 90}.get(period, 30)
        cutoff = (datetime.now(timezone.utc).date() - timedelta(days=days)).isoformat()
        paid_match["date"] = {"$gte": cutoff}
    if pole_filter:
        paid_match = {"$and": [paid_match, pole_filter]}
    paid = await db.bookings.find(
        paid_match,
        {"_id": 0, "id": 1, "offer_name": 1, "date": 1, "payment_method": 1, "total_amount": 1, "paid_amount": 1, "paid_at": 1, "phone": 1, "participants": 1, "pole": 1, "status": 1},
    ).sort("paid_at", -1).to_list(length=2000)
    by_method: dict = {}
    paid_total = 0
    for b in paid:
        m = b.get("payment_method") or "unknown"
        by_method.setdefault(m, {"count": 0, "total": 0})
        by_method[m]["count"] += 1
        by_method[m]["total"] += b.get("total_amount", 0)
        paid_total += int(b.get("total_amount", 0) or 0)

    # Today's paid amount (always, for the "today" KPI tile)
    today_iso = datetime.now(timezone.utc).date().isoformat()
    today_paid = await db.bookings.find(
        {"paid_at": {"$ne": None}, "date": today_iso},
        {"_id": 0, "total_amount": 1},
    ).to_list(length=1000)
    today_revenue = sum(int(b.get("total_amount", 0) or 0) for b in today_paid)

    return {
        "unpaid": unpaid,
        "unpaid_count": len(unpaid),
        "unpaid_total": unpaid_total,
        "paid_count": len(paid),
        "paid_total": paid_total,
        "today_revenue": today_revenue,
        "today_paid_count": len(today_paid),
        "by_method": by_method,
        "recent_paid": paid[:30],
        "period": period or "30d",
        "pole": pole or "",
    }


# =================================================================
# MODULE 5 — CLIENTS (CRM)
# =================================================================

@api.get("/staff/clients")
async def list_clients(search: Optional[str] = None, staff=Depends(get_current_staff)):
    """Aggregate clients from bookings by primary email (contact)."""
    await _require_role(staff, ["manager", "admin"])
    pipeline = [
        {"$match": {"status": {"$ne": "cancelled"}, "email": {"$nin": [None, ""]}}},
        {
            "$group": {
                "_id": {"$toLower": "$email"},
                "email": {"$first": "$email"},
                "phone": {"$first": "$phone"},
                "name": {"$first": {
                    "$let": {
                        "vars": {"adults": {"$filter": {"input": "$participants", "as": "p", "cond": {"$eq": ["$$p.kind", "adult"]}}}},
                        "in": {"$ifNull": [{"$arrayElemAt": ["$$adults.name", 0]}, {"$arrayElemAt": ["$participants.name", 0]}]},
                    }
                }},
                "surname": {"$first": {
                    "$let": {
                        "vars": {"adults": {"$filter": {"input": "$participants", "as": "p", "cond": {"$eq": ["$$p.kind", "adult"]}}}},
                        "in": {"$ifNull": [{"$arrayElemAt": ["$$adults.surname", 0]}, {"$arrayElemAt": ["$participants.surname", 0]}]},
                    }
                }},
                "nationality": {"$first": {
                    "$let": {
                        "vars": {"adults": {"$filter": {"input": "$participants", "as": "p", "cond": {"$eq": ["$$p.kind", "adult"]}}}},
                        "in": {"$ifNull": [{"$arrayElemAt": ["$$adults.nationality", 0]}, {"$arrayElemAt": ["$participants.nationality", 0]}]},
                    }
                }},
                "bookings_count": {"$sum": 1},
                "total_spent": {
                    "$sum": {
                        "$cond": [{"$ne": ["$paid_at", None]}, "$total_amount", 0]
                    }
                },
                "last_visit": {"$max": "$date"},
                "first_visit": {"$min": "$date"},
                "offers": {"$addToSet": "$offer_type"},
            }
        },
        {"$sort": {"last_visit": -1}},
        {"$limit": 1000},
    ]
    items = await db.bookings.aggregate(pipeline).to_list(length=1000)
    for it in items:
        it.pop("_id", None)
    if search:
        s = search.strip().lower()
        items = [
            it for it in items
            if s in (it.get("email") or "").lower()
            or s in (it.get("phone") or "").lower()
            or s in (it.get("name") or "").lower()
            or s in (it.get("surname") or "").lower()
        ]
    return {"items": items, "count": len(items)}


@api.get("/staff/clients/export.csv")
async def export_clients_csv(staff=Depends(get_current_staff)):
    """CSV export of aggregated client list."""
    await _require_role(staff, ["manager", "admin"])
    from fastapi.responses import Response
    import csv
    pipeline = [
        {"$match": {"status": {"$ne": "cancelled"}, "email": {"$nin": [None, ""]}}},
        {
            "$group": {
                "_id": {"$toLower": "$email"},
                "email": {"$first": "$email"},
                "phone": {"$first": "$phone"},
                "name": {"$first": {
                    "$let": {
                        "vars": {"adults": {"$filter": {"input": "$participants", "as": "p", "cond": {"$eq": ["$$p.kind", "adult"]}}}},
                        "in": {"$ifNull": [{"$arrayElemAt": ["$$adults.name", 0]}, {"$arrayElemAt": ["$participants.name", 0]}]},
                    }
                }},
                "surname": {"$first": {
                    "$let": {
                        "vars": {"adults": {"$filter": {"input": "$participants", "as": "p", "cond": {"$eq": ["$$p.kind", "adult"]}}}},
                        "in": {"$ifNull": [{"$arrayElemAt": ["$$adults.surname", 0]}, {"$arrayElemAt": ["$participants.surname", 0]}]},
                    }
                }},
                "nationality": {"$first": {
                    "$let": {
                        "vars": {"adults": {"$filter": {"input": "$participants", "as": "p", "cond": {"$eq": ["$$p.kind", "adult"]}}}},
                        "in": {"$ifNull": [{"$arrayElemAt": ["$$adults.nationality", 0]}, {"$arrayElemAt": ["$participants.nationality", 0]}]},
                    }
                }},
                "bookings_count": {"$sum": 1},
                "total_spent": {"$sum": {"$cond": [{"$ne": ["$paid_at", None]}, "$total_amount", 0]}},
                "last_visit": {"$max": "$date"},
                "first_visit": {"$min": "$date"},
            }
        },
        {"$sort": {"last_visit": -1}},
    ]
    items = await db.bookings.aggregate(pipeline).to_list(length=10000)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Nom", "Prénom", "Email", "Téléphone", "Nationalité", "Réservations", "Total dépensé (FCFA)", "Première visite", "Dernière visite"])
    for it in items:
        writer.writerow([
            it.get("surname") or "",
            it.get("name") or "",
            it.get("email") or "",
            it.get("phone") or "",
            it.get("nationality") or "",
            it.get("bookings_count", 0),
            it.get("total_spent", 0),
            it.get("first_visit") or "",
            it.get("last_visit") or "",
        ])
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="bbr-clients.csv"'},
    )


@api.get("/staff/clients/report.pdf")
async def export_clients_pdf(search: Optional[str] = None, staff=Depends(get_current_staff)):
    """Stylized PDF export of the aggregated clients list."""
    await _require_role(staff, ["manager", "admin"])
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from fastapi.responses import StreamingResponse

    pipeline = [
        {"$match": {"status": {"$ne": "cancelled"}, "email": {"$nin": [None, ""]}}},
        {
            "$group": {
                "_id": {"$toLower": "$email"},
                "email": {"$first": "$email"},
                "phone": {"$first": "$phone"},
                "name": {"$first": {
                    "$let": {
                        "vars": {"adults": {"$filter": {"input": "$participants", "as": "p", "cond": {"$eq": ["$$p.kind", "adult"]}}}},
                        "in": {"$ifNull": [{"$arrayElemAt": ["$$adults.name", 0]}, {"$arrayElemAt": ["$participants.name", 0]}]},
                    }
                }},
                "surname": {"$first": {
                    "$let": {
                        "vars": {"adults": {"$filter": {"input": "$participants", "as": "p", "cond": {"$eq": ["$$p.kind", "adult"]}}}},
                        "in": {"$ifNull": [{"$arrayElemAt": ["$$adults.surname", 0]}, {"$arrayElemAt": ["$participants.surname", 0]}]},
                    }
                }},
                "nationality": {"$first": {
                    "$let": {
                        "vars": {"adults": {"$filter": {"input": "$participants", "as": "p", "cond": {"$eq": ["$$p.kind", "adult"]}}}},
                        "in": {"$ifNull": [{"$arrayElemAt": ["$$adults.nationality", 0]}, {"$arrayElemAt": ["$participants.nationality", 0]}]},
                    }
                }},
                "bookings_count": {"$sum": 1},
                "total_spent": {"$sum": {"$cond": [{"$ne": ["$paid_at", None]}, "$total_amount", 0]}},
                "last_visit": {"$max": "$date"},
                "first_visit": {"$min": "$date"},
            }
        },
        {"$sort": {"total_spent": -1, "last_visit": -1}},
        {"$limit": 2000},
    ]
    items = await db.bookings.aggregate(pipeline).to_list(length=2000)
    if search:
        s = search.strip().lower()
        items = [
            it for it in items
            if s in (it.get("email") or "").lower()
            or s in (it.get("phone") or "").lower()
            or s in (it.get("name") or "").lower()
            or s in (it.get("surname") or "").lower()
        ]

    total_clients = len(items)
    total_revenue = sum(int(it.get("total_spent") or 0) for it in items)
    total_bookings = sum(int(it.get("bookings_count") or 0) for it in items)

    styles = _pdf_styles()
    GOLD, DARK, LIGHT, MUTED = styles["GOLD"], styles["DARK"], styles["LIGHT"], styles["MUTED"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
    elements = []
    elements.append(Paragraph("Boulay Beach Resort", styles["h1"]))
    sub_label = f"Base clients — {total_clients} client(s)"
    if search:
        sub_label += f" — Recherche : {search}"
    elements.append(Paragraph(sub_label, styles["sub"]))

    kpi_rows = [
        ["Clients", "Réservations cumulées", "Revenu cumulé"],
        [str(total_clients), str(total_bookings), _format_xof(total_revenue)],
    ]
    kpi_tbl = Table(kpi_rows, colWidths=[5.6 * cm, 5.6 * cm, 5.6 * cm])
    kpi_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("TEXTCOLOR", (0, 0), (-1, 0), MUTED),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("TEXTCOLOR", (0, 1), (-1, 1), DARK),
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.5, GOLD),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E0D5B5")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_tbl)

    elements.append(Paragraph("Liste des clients", styles["h2"]))
    if not items:
        elements.append(Paragraph("Aucun client.", styles["body"]))
    else:
        rows = [["#", "Nom", "Email", "Téléphone", "Résa", "Total dépensé", "Dernière visite"]]
        for i, it in enumerate(items, start=1):
            rows.append([
                str(i),
                f"{it.get('surname') or ''} {it.get('name') or ''}".strip() or "—",
                it.get("email") or "—",
                it.get("phone") or "—",
                str(it.get("bookings_count") or 0),
                _format_xof(it.get("total_spent") or 0),
                it.get("last_visit") or "—",
            ])
        tbl = Table(rows, colWidths=[0.8 * cm, 3.8 * cm, 4.7 * cm, 2.6 * cm, 1.2 * cm, 2.8 * cm, 2.1 * cm], repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (4, 0), (5, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.2, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(tbl)

    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(
        f"Rapport généré le {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M UTC')} · Boulay Beach Resort, Abidjan",
        styles["small"],
    ))

    doc.build(elements, onFirstPage=_pdf_footer_factory(styles), onLaterPages=_pdf_footer_factory(styles))
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="bbr-clients.pdf"'},
    )


@api.get("/staff/clients/{email}")
async def client_detail(email: str, staff=Depends(get_current_staff)):
    """Full client history for the given email (case-insensitive)."""
    await _require_role(staff, ["manager", "admin"])
    cursor = db.bookings.find(
        {"email": {"$regex": f"^{email}$", "$options": "i"}},
        {"_id": 0, "reference_token": 0, "qr_codes.qr_code": 0, "qr_codes.qr_payload": 0, "qr_codes.ticket_image": 0},
    ).sort([("date", -1)])
    bookings = await cursor.to_list(length=500)
    if not bookings:
        raise HTTPException(status_code=404, detail="Client not found")
    primary = next((p for b in bookings for p in b.get("participants", []) if p.get("kind") == "adult"), None) or {}
    total_spent = sum(b.get("total_amount", 0) for b in bookings if b.get("paid_at"))
    return {
        "email": bookings[0].get("email"),
        "phone": bookings[0].get("phone"),
        "name": primary.get("name", ""),
        "surname": primary.get("surname", ""),
        "nationality": primary.get("nationality", ""),
        "bookings_count": len(bookings),
        "total_spent": total_spent,
        "bookings": bookings,
    }


# =================================================================
# MODULE 7 — REVENUE (CHIFFRE D'AFFAIRES)
# =================================================================

@api.get("/staff/revenue")
async def revenue_overview(
    period: str = "month",  # day | week | month | year | all
    date_from: Optional[str] = Query(None, regex=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: Optional[str] = Query(None, regex=r"^\d{4}-\d{2}-\d{2}$"),
    staff=Depends(get_current_staff),
):
    """Revenue dashboard: KPIs, by offer, by payment method, daily trend, top clients.

    ``date_from``/``date_to`` override the preset ``period`` when provided —
    inclusive custom range used by the "Période personnalisée" picker.
    """
    await _require_role(staff, ["manager", "admin"])
    today = datetime.now(timezone.utc).date()
    if date_from or date_to:
        date_from_iso = date_from
        date_to_iso = date_to
    else:
        if period == "day":
            date_from_iso = today.isoformat()
        elif period == "week":
            date_from_iso = (today - timedelta(days=7)).isoformat()
        elif period == "month":
            date_from_iso = (today - timedelta(days=30)).isoformat()
        elif period == "year":
            date_from_iso = (today - timedelta(days=365)).isoformat()
        else:
            date_from_iso = None
        date_to_iso = None

    q: dict = {"paid_at": {"$ne": None}}
    if date_from_iso or date_to_iso:
        rng: dict = {}
        if date_from_iso:
            rng["$gte"] = date_from_iso
        if date_to_iso:
            rng["$lte"] = date_to_iso
        q["date"] = rng

    paid = await db.bookings.find(
        q,
        {"_id": 0, "offer_type": 1, "offer_name": 1, "date": 1, "total_amount": 1, "payment_method": 1, "email": 1, "phone": 1, "participants": 1, "paid_at": 1},
    ).to_list(length=10000)

    total_revenue = sum(b.get("total_amount", 0) for b in paid)
    total_bookings = len(paid)
    avg_basket = (total_revenue / total_bookings) if total_bookings else 0

    by_offer: dict = {}
    by_pole: dict = {pid: {"id": pid, "name_fr": POLES[pid]["name_fr"], "count": 0, "total": 0} for pid in POLES}
    by_method: dict = {}
    by_day: dict = {}
    by_client: dict = {}

    for b in paid:
        oid = b.get("offer_type", "unknown")
        by_offer.setdefault(oid, {"offer_id": oid, "offer_name": b.get("offer_name", oid), "count": 0, "total": 0})
        by_offer[oid]["count"] += 1
        by_offer[oid]["total"] += b.get("total_amount", 0)

        pole = b.get("pole") or _pole_for_offer(oid)
        if pole in by_pole:
            by_pole[pole]["count"] += 1
            by_pole[pole]["total"] += b.get("total_amount", 0)

        m = b.get("payment_method") or "unknown"
        by_method.setdefault(m, {"method": m, "count": 0, "total": 0})
        by_method[m]["count"] += 1
        by_method[m]["total"] += b.get("total_amount", 0)

        d = b.get("date") or ""
        if d:
            by_day.setdefault(d, 0)
            by_day[d] += b.get("total_amount", 0)

        email = (b.get("email") or "").lower()
        if email:
            participants = b.get("participants", [])
            primary = next((p for p in participants if p.get("kind") == "adult"), participants[0] if participants else {})
            by_client.setdefault(email, {
                "email": email,
                "phone": b.get("phone", ""),
                "name": primary.get("name", "") if primary else "",
                "surname": primary.get("surname", "") if primary else "",
                "count": 0,
                "total": 0,
            })
            by_client[email]["count"] += 1
            by_client[email]["total"] += b.get("total_amount", 0)

    daily_trend = [{"date": d, "amount": amt} for d, amt in sorted(by_day.items())]
    top_clients = sorted(by_client.values(), key=lambda c: c["total"], reverse=True)[:10]

    return {
        "period": period,
        "total_revenue": total_revenue,
        "total_bookings": total_bookings,
        "avg_basket": int(avg_basket),
        "by_offer": list(by_offer.values()),
        "by_pole": list(by_pole.values()),
        "by_method": list(by_method.values()),
        "daily_trend": daily_trend,
        "top_clients": top_clients,
    }


@api.get("/staff/revenue/report.pdf")
async def export_revenue_pdf(
    period: str = "month",  # day | week | month | year | all
    date_from: Optional[str] = Query(None, regex=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: Optional[str] = Query(None, regex=r"^\d{4}-\d{2}-\d{2}$"),
    staff=Depends(get_current_staff),
):
    """Stylized PDF report of the revenue dashboard for the selected period
    (preset) or a custom inclusive date range."""
    await _require_role(staff, ["manager", "admin"])
    # Reuse the revenue aggregator to compute the same payload (no auth re-check)
    data = await revenue_overview(period=period, date_from=date_from, date_to=date_to, staff=staff)  # type: ignore

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from fastapi.responses import StreamingResponse

    period_label = {
        "day": "Aujourd'hui",
        "week": "7 derniers jours",
        "month": "30 derniers jours",
        "year": "12 derniers mois",
        "all": "Depuis le lancement",
    }.get(period, period)
    if date_from or date_to:
        period_label = f"Du {date_from or '—'} au {date_to or '—'}"

    method_label = {
        "fineo": "FINEO",
        "card": "Carte bancaire",
        "mobile_money": "Mobile Money",
        "cash": "Espèces",
        "unknown": "Inconnu",
    }

    styles = _pdf_styles()
    GOLD, DARK, LIGHT, MUTED = styles["GOLD"], styles["DARK"], styles["LIGHT"], styles["MUTED"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
    elements = []
    elements.append(Paragraph("Boulay Beach Resort", styles["h1"]))
    elements.append(Paragraph(f"Rapport de chiffre d'affaires — {period_label}", styles["sub"]))

    # KPIs
    kpi_rows = [
        ["Revenu total", "Réservations payées", "Panier moyen"],
        [_format_xof(data["total_revenue"]), str(data["total_bookings"]), _format_xof(data["avg_basket"])],
    ]
    kpi_tbl = Table(kpi_rows, colWidths=[5.6 * cm, 5.6 * cm, 5.6 * cm])
    kpi_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("TEXTCOLOR", (0, 0), (-1, 0), MUTED),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("TEXTCOLOR", (0, 1), (-1, 1), DARK),
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.5, GOLD),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E0D5B5")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_tbl)

    # By offer
    if data.get("by_offer"):
        elements.append(Paragraph("Répartition par offre", styles["h2"]))
        rows = [["Offre", "Réservations", "Revenu", "Part"]]
        total = data["total_revenue"] or 1
        for o in sorted(data["by_offer"], key=lambda x: x["total"], reverse=True):
            pct = (o["total"] / total * 100) if total else 0
            rows.append([o["offer_name"], str(o["count"]), _format_xof(o["total"]), f"{pct:.1f}%"])
        tbl = Table(rows, colWidths=[7 * cm, 3 * cm, 4 * cm, 2.5 * cm])
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(tbl)

    # By payment method
    if data.get("by_method"):
        elements.append(Paragraph("Répartition par méthode de paiement", styles["h2"]))
        rows = [["Méthode", "Réservations", "Revenu", "Part"]]
        total = data["total_revenue"] or 1
        for m in sorted(data["by_method"], key=lambda x: x["total"], reverse=True):
            pct = (m["total"] / total * 100) if total else 0
            rows.append([method_label.get(m["method"], m["method"]), str(m["count"]), _format_xof(m["total"]), f"{pct:.1f}%"])
        tbl = Table(rows, colWidths=[7 * cm, 3 * cm, 4 * cm, 2.5 * cm])
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(tbl)

    # Top clients
    if data.get("top_clients"):
        elements.append(Paragraph("Top 10 clients", styles["h2"]))
        rows = [["#", "Client", "Email", "Résa", "Total dépensé"]]
        for i, c in enumerate(data["top_clients"], start=1):
            full_name = f"{c.get('surname','')} {c.get('name','')}".strip() or "—"
            rows.append([str(i), full_name, c.get("email") or "—", str(c.get("count") or 0), _format_xof(c.get("total") or 0)])
        tbl = Table(rows, colWidths=[1 * cm, 4.5 * cm, 5.5 * cm, 1.8 * cm, 3.7 * cm], repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(tbl)

    # Daily trend (compact table)
    if data.get("daily_trend"):
        elements.append(Paragraph("Évolution journalière", styles["h2"]))
        rows = [["Date", "Revenu"]]
        for d in data["daily_trend"]:
            rows.append([d["date"], _format_xof(d["amount"])])
        tbl = Table(rows, colWidths=[4 * cm, 4 * cm], repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(tbl)

    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(
        f"Rapport généré le {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M UTC')} · Boulay Beach Resort, Abidjan",
        styles["small"],
    ))

    doc.build(elements, onFirstPage=_pdf_footer_factory(styles), onLaterPages=_pdf_footer_factory(styles))
    buf.seek(0)
    filename = f"bbr-revenue-{period}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =================================================================
# MODULE 6 — LE KAAI (TABLES)
# =================================================================

# ---------- Activities & Wallets (in-resort spend) ----------
class ActivityModel(BaseModel):
    id: str
    name_fr: str
    name_en: Optional[str] = None
    category: Optional[str] = "Activité"
    subcategory: Optional[str] = ""
    price: int = Field(ge=0)
    active: bool = True


class WalletCharge(BaseModel):  # noqa: F811 — declared earlier near scanner endpoints
    activity_id: Optional[str] = None
    label: Optional[str] = None
    amount: int = Field(default=0, ge=0)
    note: Optional[str] = ""
    quantity: int = Field(default=1, ge=1, le=20)
    participant_token: Optional[str] = None


@api.get("/activities")
async def list_activities_public():
    """Public list of activities — used to inform the booking UX of available services."""
    await _seed_default_activities()
    items = await db.activities.find({"active": True}, {"_id": 0}).sort("category", 1).to_list(length=200)
    return {"items": items}


@api.get("/staff/activities")
async def list_activities_staff(staff=Depends(get_current_staff)):
    await _require_role(staff, ["receptionist", "manager", "admin"])
    await _seed_default_activities()
    items = await db.activities.find({}, {"_id": 0}).sort("category", 1).to_list(length=200)
    return {"items": items}


@api.post("/staff/activities")
async def create_activity(body: ActivityModel, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    if await db.activities.find_one({"id": body.id}):
        raise HTTPException(status_code=400, detail="Activity id already exists")
    doc = body.model_dump()
    await db.activities.insert_one(dict(doc))
    return doc


@api.patch("/staff/activities/{activity_id}")
async def update_activity(activity_id: str, body: ActivityModel, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    payload = body.model_dump(exclude={"id"})
    res = await db.activities.update_one({"id": activity_id}, {"$set": payload})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Activity not found")
    return {"ok": True}


@api.delete("/staff/activities/{activity_id}")
async def delete_activity(activity_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    res = await db.activities.delete_one({"id": activity_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Activity not found")
    return {"ok": True}


async def _wallet_summary(wallet: dict) -> dict:
    """Build a serialisable summary of a wallet doc + linked booking info."""
    wallet = dict(wallet)
    wallet.pop("_id", None)
    booking_id = wallet.get("booking_id")
    if booking_id:
        b = await db.bookings.find_one(
            {"id": booking_id},
            {"_id": 0, "id": 1, "offer_name": 1, "date": 1, "checkout_date": 1, "phone": 1,
             "email": 1, "status": 1, "boat_time": 1, "return_boat_time": 1, "balance_due": 1,
             "total_amount": 1, "paid_amount": 1, "deposit_pct": 1, "participants": 1,
             "room_tier_name": 1, "rooms": 1, "adults": 1, "children": 1},
        )
        wallet["booking"] = b
    txs = wallet.get("transactions", [])
    wallet["total_charged"] = sum(t.get("amount", 0) for t in txs if t.get("status") != "voided")
    return wallet


@api.get("/staff/wallets/{token}")
async def get_wallet(token: str, staff=Depends(get_current_staff)):
    """Look up a wallet by:
    - its full wallet UUID token,
    - its short reference (first 10 hex chars of the wallet token, with or without dashes),
    - its short booking_ref (8 chars), or
    - the short/long token of any **ticket QR** attached to the same booking
      (so staff can scan/type either QR — billet or wallet).
    """
    await _require_role(staff, ["receptionist", "manager", "admin"])
    wallet = await db.wallets.find_one({"token": token}, {"_id": 0})
    if not wallet:
        # Normalise: strip whitespace, dashes, dots → compare alnum only, lowercased
        raw = (token or "").strip()
        norm = re.sub(r"[^a-z0-9]", "", raw.lower())
        if not norm:
            raise HTTPException(status_code=404, detail="Wallet not found")
        # 1) Match against wallet_token prefix / booking_ref
        async for w in db.wallets.find({}, {"_id": 0}):
            t_norm = re.sub(r"[^a-z0-9]", "", (w.get("token") or "").lower())
            r_norm = re.sub(r"[^a-z0-9]", "", (w.get("booking_ref") or "").lower())
            if t_norm.startswith(norm) or r_norm == norm:
                wallet = w
                break
        # 2) Fallback: match against any booking.qr_codes[].qr_token (ticket code).
        #    Returns the wallet linked to that booking — staff can use either QR.
        if not wallet:
            async for b in db.bookings.find(
                {"wallet_token": {"$exists": True, "$ne": None}},
                {"_id": 0, "id": 1, "wallet_token": 1, "qr_codes": 1},
            ):
                qrs = b.get("qr_codes") or []
                for q in qrs:
                    q_norm = re.sub(r"[^a-z0-9]", "", (q.get("qr_token") or "").lower())
                    if q_norm and q_norm.startswith(norm):
                        wallet = await db.wallets.find_one({"token": b["wallet_token"]}, {"_id": 0})
                        break
                if wallet:
                    break
    if not wallet:
        raise HTTPException(status_code=404, detail="Wallet not found")
    return await _wallet_summary(wallet)


@api.post("/staff/wallets/{token}/charge")
async def charge_wallet(token: str, body: WalletCharge, staff=Depends(get_current_staff)):
    """Add an activity charge to the wallet. Either ``activity_id`` (catalog
    lookup, amount × quantity) or a custom ``label + amount`` is required."""
    await _require_role(staff, ["receptionist", "manager", "admin"])
    wallet = await db.wallets.find_one({"token": token}, {"_id": 0})
    if not wallet:
        raise HTTPException(status_code=404, detail="Wallet not found")
    if wallet.get("status") == "closed":
        raise HTTPException(status_code=400, detail="Wallet already closed")

    activity_label = body.label or "Prestation"
    unit_price = body.amount
    if body.activity_id:
        activity = await db.activities.find_one({"id": body.activity_id, "active": True}, {"_id": 0})
        if not activity:
            raise HTTPException(status_code=404, detail="Activity not found")
        activity_label = activity["name_fr"]
        unit_price = int(activity["price"])
    if unit_price <= 0:
        raise HTTPException(status_code=400, detail="amount must be > 0 (or provide a valid activity_id)")
    total = unit_price * body.quantity
    # Resolve participant from booking.qr_codes for traceability
    participant_name = None
    if body.participant_token and wallet.get("booking_id"):
        booking_for_p = await db.bookings.find_one(
            {"id": wallet["booking_id"]},
            {"_id": 0, "qr_codes": 1},
        )
        if booking_for_p:
            p = next(
                (q for q in (booking_for_p.get("qr_codes") or [])
                 if (q.get("qr_token") or "").lower() == body.participant_token.lower()),
                None,
            )
            if p:
                participant_name = f"{p.get('guest_surname','').strip()} {p.get('guest_name','').strip()}".strip() or None
    tx = {
        "id": str(uuid.uuid4()),
        "activity_id": body.activity_id,
        "label": activity_label,
        "unit_price": unit_price,
        "quantity": body.quantity,
        "amount": total,
        "note": body.note or "",
        "status": "active",
        "created_at": now_iso(),
        "created_by": staff.get("name"),
        "created_by_role": staff.get("role"),
        "participant_token": body.participant_token,
        "participant_name": participant_name,
    }
    await db.wallets.update_one(
        {"token": token},
        {
            "$push": {"transactions": tx},
            "$inc": {"total_charged": total},
        },
    )
    # Fiscal receipt for this on-site activity charge — tagged with the participant when known
    try:
        receipt_customer = participant_name or wallet.get("customer_name", "")
        await _create_receipt(
            source="activity",
            source_id=token,
            customer_name=receipt_customer,
            customer_email=wallet.get("customer_email", ""),
            customer_phone=wallet.get("customer_phone", ""),
            lines=[{
                "description": activity_label,
                "quantity": int(body.quantity),
                "unit_price": int(unit_price),
                "total": int(total),
            }],
            payment_method="on_site",
            issued_by=staff.get("name") or "",
            issued_by_role=staff.get("role") or "",
            metadata={
                "sub_id": tx["id"],
                "booking_id": wallet.get("booking_id"),
                "participant_token": body.participant_token,
                "participant_name": participant_name,
            },
        )
    except Exception as ex:
        logging.exception("Failed to create activity receipt: %s", ex)
    fresh = await db.wallets.find_one({"token": token}, {"_id": 0})
    return await _wallet_summary(fresh)


@api.delete("/staff/wallets/{token}/charge/{tx_id}")
async def void_wallet_charge(token: str, tx_id: str, staff=Depends(get_current_staff)):
    """Void a charge (kept in history, balance adjusted)."""
    await _require_role(staff, ["manager", "admin"])
    wallet = await db.wallets.find_one({"token": token, "transactions.id": tx_id}, {"_id": 0})
    if not wallet:
        raise HTTPException(status_code=404, detail="Wallet or charge not found")
    tx = next((t for t in wallet.get("transactions", []) if t["id"] == tx_id), None)
    if not tx or tx.get("status") == "voided":
        raise HTTPException(status_code=400, detail="Charge already voided")
    await db.wallets.update_one(
        {"token": token, "transactions.id": tx_id},
        {
            "$set": {
                "transactions.$.status": "voided",
                "transactions.$.voided_at": now_iso(),
                "transactions.$.voided_by": staff.get("name"),
            },
            "$inc": {"total_charged": -tx["amount"]},
        },
    )
    fresh = await db.wallets.find_one({"token": token}, {"_id": 0})
    return await _wallet_summary(fresh)


class WalletCloseBody(BaseModel):
    payment_method: Literal["cash", "card", "mobile_money"]


@api.post("/staff/wallets/{token}/close")
async def close_wallet(token: str, body: WalletCloseBody, staff=Depends(get_current_staff)):
    """Mark the wallet as settled (paid at check-out). The staff must select the
    payment method actually used by the customer (cash / card / mobile money) —
    validating this is the proof the customer has paid on site."""
    await _require_role(staff, ["manager", "admin"])
    wallet = await db.wallets.find_one({"token": token}, {"_id": 0})
    if not wallet:
        raise HTTPException(status_code=404, detail="Wallet not found")
    if wallet.get("status") == "closed":
        raise HTTPException(status_code=400, detail="Wallet already closed")
    if int(wallet.get("total_charged", 0) or 0) <= 0:
        raise HTTPException(status_code=400, detail="Aucune prestation à encaisser")

    paid_amount = int(wallet.get("total_charged", 0) or 0)
    closed_at = now_iso()
    await db.wallets.update_one(
        {"token": token},
        {
            "$set": {
                "status": "closed",
                "closed_at": closed_at,
                "closed_by": staff.get("name"),
                "payment_method": body.payment_method,
                "paid_amount": paid_amount,
                "paid_at": closed_at,
            }
        },
    )
    # Emit a consolidated fiscal receipt for the on-site settlement
    try:
        active_lines = [
            {
                "description": t.get("label") or "Prestation",
                "quantity": int(t.get("quantity", 1) or 1),
                "unit_price": int(t.get("unit_price", t.get("amount", 0)) or 0),
                "total": int(t.get("amount", 0) or 0),
            }
            for t in (wallet.get("transactions") or [])
            if t.get("status") != "voided"
        ]
        if active_lines:
            await _create_receipt(
                source="wallet_settlement",
                source_id=token,
                customer_name=wallet.get("customer_name", "") or wallet.get("owner_name", ""),
                customer_email=wallet.get("customer_email", ""),
                customer_phone=wallet.get("customer_phone", ""),
                lines=active_lines,
                payment_method=body.payment_method,
                issued_by=staff.get("name") or "",
                issued_by_role=staff.get("role") or "",
                metadata={
                    "booking_id": wallet.get("booking_id"),
                    "wallet_token": token,
                },
            )
    except Exception as ex:
        logging.exception("Failed to create settlement receipt: %s", ex)

    fresh = await db.wallets.find_one({"token": token}, {"_id": 0})
    return await _wallet_summary(fresh)


# =================================================================
# MODULE 6 — LE KAAI (TABLES) — historical anchor
# =================================================================

class KaaiTable(BaseModel):
    number: str
    capacity: int = Field(ge=1, le=30)
    zone: Optional[str] = "Salle"
    status: Literal["active", "indisponible"] = "active"


class KaaiTableUpdate(BaseModel):
    number: Optional[str] = None
    capacity: Optional[int] = None
    zone: Optional[str] = None
    status: Optional[Literal["active", "indisponible"]] = None


class KaaiZone(BaseModel):
    """Logical seating zone with a hard capacity cap used as an overbooking guard."""
    name: str = Field(min_length=1, max_length=40)
    capacity: int = Field(ge=0, le=500)
    sort_order: int = 0


class KaaiZoneUpdate(BaseModel):
    capacity: Optional[int] = Field(default=None, ge=0, le=500)
    sort_order: Optional[int] = None


DEFAULT_KAAI_ZONES = [
    {"name": "Terrasse 1", "capacity": 24, "sort_order": 1},
    {"name": "Terrasse 2", "capacity": 24, "sort_order": 2},
    {"name": "Salle", "capacity": 32, "sort_order": 3},
]


async def _seed_default_kaai_zones():
    """Seed default Le Kaai zones if collection is empty AND migrate legacy table zones."""
    if await db.kaai_zones.count_documents({}) == 0:
        seeds = [{**z, "id": str(uuid.uuid4()), "created_at": now_iso()} for z in DEFAULT_KAAI_ZONES]
        await db.kaai_zones.insert_many(seeds)
        logging.info("Seeded %d Le Kaai zones", len(seeds))
        # Migrate legacy zone labels on tables: split 'Terrasse' across Terrasse 1 / Terrasse 2.
        legacy = await db.kaai_tables.find({"zone": "Terrasse"}, {"_id": 0, "id": 1, "number": 1}).sort("number", 1).to_list(length=500)
        if legacy:
            half = max(1, len(legacy) // 2)
            ops = []
            for i, t in enumerate(legacy):
                new_zone = "Terrasse 1" if i < half else "Terrasse 2"
                ops.append((t["id"], new_zone))
            for tid, nz in ops:
                await db.kaai_tables.update_one({"id": tid}, {"$set": {"zone": nz}})
        # Migrate any other unknown zone to 'Salle' to keep capacities consistent.
        known = {z["name"] for z in DEFAULT_KAAI_ZONES}
        await db.kaai_tables.update_many(
            {"zone": {"$nin": list(known)}},
            {"$set": {"zone": "Salle"}},
        )


async def _seed_default_kaai_tables():
    """Seed default Le Kaai tables if none exist."""
    if await db.kaai_tables.count_documents({}) == 0:
        layout = [
            ("Terrasse 1", 6, 2),
            ("Terrasse 2", 6, 4),
            ("Salle", 8, 2),
            ("Salle", 4, 4),
            ("Salle", 4, 6),
        ]
        seeds = []
        i = 1
        for zone, count, cap in layout:
            for _ in range(count):
                seeds.append({"id": str(uuid.uuid4()), "number": f"T{i:02d}", "capacity": cap, "zone": zone, "status": "active"})
                i += 1
        await db.kaai_tables.insert_many(seeds)
        logging.info("Seeded %d Le Kaai tables", len(seeds))


# ----- Zones CRUD -----
@api.get("/staff/kaai/zones")
async def list_kaai_zones(staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    await _seed_default_kaai_zones()
    items = await db.kaai_zones.find({}, {"_id": 0}).sort("sort_order", 1).to_list(length=200)
    return {"items": items}


@api.post("/staff/kaai/zones")
async def create_kaai_zone(body: KaaiZone, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    # Names are unique (case-insensitive).
    import re as _re
    existing = await db.kaai_zones.find_one({"name": {"$regex": f"^{_re.escape(body.name)}$", "$options": "i"}}, {"_id": 0, "id": 1})
    if existing:
        raise HTTPException(status_code=400, detail="Une zone porte déjà ce nom")
    doc = {**body.model_dump(), "id": str(uuid.uuid4()), "created_at": now_iso()}
    await db.kaai_zones.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.patch("/staff/kaai/zones/{zone_id}")
async def update_kaai_zone(zone_id: str, body: KaaiZoneUpdate, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        return {"ok": True}
    res = await db.kaai_zones.update_one({"id": zone_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Zone introuvable")
    return {"ok": True}


@api.delete("/staff/kaai/zones/{zone_id}")
async def delete_kaai_zone(zone_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    res = await db.kaai_zones.delete_one({"id": zone_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Zone introuvable")
    return {"ok": True}


# ----- Helpers: zone occupancy + guard -----
async def _kaai_zone_occupancy(date: str):
    """Return {zone_name: seats_used} on a given day, based on table assignments
    (each assigned booking counts adults + children seats against its table's zone)."""
    tables = await db.kaai_tables.find({}, {"_id": 0, "id": 1, "zone": 1}).to_list(length=1000)
    table_zone = {t["id"]: t.get("zone", "Salle") for t in tables}
    bookings = await db.bookings.find(
        {"offer_type": "le_kaai", "date": date, "status": {"$ne": "cancelled"}, "table_id": {"$exists": True, "$ne": None}},
        {"_id": 0, "table_id": 1, "adults": 1, "children": 1},
    ).to_list(length=1000)
    usage: dict = {}
    for b in bookings:
        z = table_zone.get(b.get("table_id"))
        if not z:
            continue
        usage[z] = usage.get(z, 0) + int(b.get("adults") or 0) + int(b.get("children") or 0)
    return usage


async def _kaai_zone_guard(date: str, table_id: str, adults: int, children: int, exclude_booking_id: Optional[str] = None):
    """Raise HTTP 400 if assigning this booking would push the table's zone above its capacity."""
    table = await db.kaai_tables.find_one({"id": table_id}, {"_id": 0, "zone": 1})
    if not table:
        return  # caller handles 404
    zone_name = table.get("zone") or "Salle"
    zone = await db.kaai_zones.find_one({"name": zone_name}, {"_id": 0, "capacity": 1})
    if not zone:
        return  # no zone configured → no guard
    capacity = int(zone.get("capacity") or 0)
    if capacity <= 0:
        return
    # Current usage minus this booking if it already occupies a table in the same zone
    usage = await _kaai_zone_occupancy(date)
    current = usage.get(zone_name, 0)
    if exclude_booking_id:
        existing = await db.bookings.find_one(
            {"id": exclude_booking_id},
            {"_id": 0, "table_id": 1, "adults": 1, "children": 1},
        )
        if existing and existing.get("table_id"):
            t2 = await db.kaai_tables.find_one({"id": existing["table_id"]}, {"_id": 0, "zone": 1})
            if t2 and (t2.get("zone") == zone_name):
                current -= int(existing.get("adults") or 0) + int(existing.get("children") or 0)
                current = max(0, current)
    new_total = current + int(adults or 0) + int(children or 0)
    if new_total > capacity:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Capacité de la salle « {zone_name} » dépassée "
                f"({new_total}/{capacity} couverts pour le {date})."
            ),
        )


@api.get("/staff/kaai/tables")
async def list_kaai_tables(staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    await _seed_default_kaai_zones()
    await _seed_default_kaai_tables()
    items = await db.kaai_tables.find({}, {"_id": 0}).sort("number", 1).to_list(length=500)
    return {"items": items}


@api.post("/staff/kaai/tables")
async def create_kaai_table(body: KaaiTable, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    doc = body.model_dump()
    doc.update({"id": str(uuid.uuid4()), "created_at": now_iso()})
    await db.kaai_tables.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.patch("/staff/kaai/tables/{table_id}")
async def update_kaai_table(table_id: str, body: KaaiTableUpdate, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        return {"ok": True}
    res = await db.kaai_tables.update_one({"id": table_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Table not found")
    return {"ok": True}


@api.delete("/staff/kaai/tables/{table_id}")
async def delete_kaai_table(table_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    res = await db.kaai_tables.delete_one({"id": table_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Table not found")
    # Also clear assignments referencing this table
    await db.bookings.update_many({"table_id": table_id}, {"$unset": {"table_id": ""}})
    return {"ok": True}


@api.get("/staff/kaai/day")
async def kaai_day(date: str, staff=Depends(get_current_staff)):
    """Le Kaai bookings + table assignments + zones with capacity/occupation for a day."""
    await _require_role(staff, ["manager", "admin"])
    await _seed_default_kaai_zones()
    bookings = await db.bookings.find(
        {"offer_type": "le_kaai", "date": date, "status": {"$ne": "cancelled"}},
        {"_id": 0, "id": 1, "boat_time": 1, "adults": 1, "children": 1, "phone": 1, "email": 1,
         "participants": 1, "status": 1, "special_requests": 1, "table_id": 1, "paid_at": 1, "created_at": 1},
    ).sort("boat_time", 1).to_list(length=500)
    tables = await db.kaai_tables.find({}, {"_id": 0}).sort("number", 1).to_list(length=500)
    zones_raw = await db.kaai_zones.find({}, {"_id": 0}).sort("sort_order", 1).to_list(length=200)
    usage = await _kaai_zone_occupancy(date)
    zones = []
    for z in zones_raw:
        used = int(usage.get(z["name"], 0))
        cap = int(z.get("capacity") or 0)
        zones.append({
            **z,
            "used": used,
            "available": max(0, cap - used),
            "saturation_pct": round((used / cap * 100), 1) if cap else 0,
        })
    return {"date": date, "bookings": bookings, "tables": tables, "zones": zones}


@api.patch("/staff/kaai/bookings/{booking_id}/table")
async def assign_kaai_table(
    booking_id: str,
    table_id: Optional[str] = Body(None, embed=True),
    staff=Depends(get_current_staff),
):
    """Assign or unassign a table to a Le Kaai booking, enforcing zone capacity."""
    await _require_role(staff, ["manager", "admin"])
    booking = await db.bookings.find_one(
        {"id": booking_id, "offer_type": "le_kaai"},
        {"_id": 0, "offer_type": 1, "date": 1, "adults": 1, "children": 1, "table_id": 1},
    )
    if not booking:
        raise HTTPException(status_code=404, detail="Le Kaai booking not found")
    if table_id:
        table = await db.kaai_tables.find_one({"id": table_id}, {"_id": 0, "capacity": 1, "status": 1})
        if not table:
            raise HTTPException(status_code=404, detail="Table not found")
        if table.get("status") == "indisponible":
            raise HTTPException(status_code=400, detail="Cette table est indisponible.")
        guests = int(booking.get("adults") or 0) + int(booking.get("children") or 0)
        if guests > int(table.get("capacity") or 0):
            raise HTTPException(
                status_code=400,
                detail=f"Capacité de la table dépassée ({guests} convives > {table.get('capacity')}).",
            )
        # Zone capacity guard (excludes the current booking if it was already on this zone)
        await _kaai_zone_guard(
            date=booking["date"],
            table_id=table_id,
            adults=int(booking.get("adults") or 0),
            children=int(booking.get("children") or 0),
            exclude_booking_id=booking_id,
        )
        await db.bookings.update_one({"id": booking_id}, {"$set": {"table_id": table_id}})
    else:
        await db.bookings.update_one({"id": booking_id}, {"$unset": {"table_id": ""}})
    return {"ok": True}


# =================================================================
# MODULE HÉBERGEMENT — STAFF CALENDAR & DAILY ARRIVALS/DEPARTURES
# =================================================================

@api.get("/staff/hebergement/today")
async def hebergement_today(date: Optional[str] = None, staff=Depends(get_current_staff)):
    """Arrivals (check-in) and departures (check-out) for a given day. Defaults to today."""
    await _require_role(staff, ["manager", "admin", "receptionist"])
    target = date or datetime.now(timezone.utc).date().isoformat()
    proj = {
        "_id": 0, "id": 1, "boat_time": 1, "return_boat_time": 1,
        "room_tier": 1, "room_tier_name": 1, "room_id": 1, "room_label": 1,
        "rooms": 1, "adults": 1, "children": 1, "phone": 1, "email": 1,
        "name": 1, "first_name": 1, "last_name": 1, "participants": 1,
        "status": 1, "paid_at": 1, "nights": 1, "date": 1, "checkout_date": 1,
        "special_requests": 1, "checked_in_at": 1, "checked_out_at": 1,
    }
    arrivals = await db.bookings.find(
        {"offer_type": "hebergement", "date": target, "status": {"$ne": "cancelled"}},
        proj,
    ).sort("boat_time", 1).to_list(length=500)
    departures = await db.bookings.find(
        {"offer_type": "hebergement", "checkout_date": target, "status": {"$ne": "cancelled"}},
        proj,
    ).sort("return_boat_time", 1).to_list(length=500)
    for lst in (arrivals, departures):
        for b in lst:
            rid = b.get("room_id")
            if rid and not b.get("room_label"):
                meta = HEBERGEMENT_ROOMS_BY_ID.get(rid)
                if meta:
                    b["room_label"] = meta["label"]
    return {
        "date": target,
        "arrivals": arrivals,
        "departures": departures,
        "default_checkin_time": HEBERGEMENT_DEFAULT_CHECKIN,
        "default_checkout_time": HEBERGEMENT_DEFAULT_CHECKOUT,
    }


@api.get("/staff/hebergement/occupancy")
async def hebergement_occupancy(date: Optional[str] = None, staff=Depends(get_current_staff)):
    """Physical-room occupancy snapshot for a given day.

    Returns each room (Supérieures 1001-1020 + 6 Suites nommées) with its
    current status (available / occupied / arriving_today / departing_today)
    and the booking attached if any. Plus per-tier KPIs (total/available/occupied).
    """
    await _require_role(staff, ["manager", "admin", "receptionist"])
    target = date or datetime.now(timezone.utc).date().isoformat()

    # All hebergement bookings that overlap the target day
    cursor = db.bookings.find(
        {
            "offer_type": "hebergement",
            "status": {"$ne": "cancelled"},
            "date": {"$lte": target},
            "checkout_date": {"$gt": target},
        },
        {
            "_id": 0, "id": 1, "date": 1, "checkout_date": 1, "room_id": 1,
            "room_tier": 1, "room_tier_name": 1, "rooms": 1,
            "adults": 1, "children": 1, "first_name": 1, "last_name": 1,
            "name": 1, "email": 1, "phone": 1, "boat_time": 1,
            "return_boat_time": 1, "checked_in_at": 1, "checked_out_at": 1,
        },
    )
    overlapping = await cursor.to_list(length=2000)

    # Bookings that ARRIVE today (date == target) and DEPART today (checkout == target)
    arriving_today_ids = {b["id"] for b in overlapping if b.get("date") == target}
    departing_cursor = db.bookings.find(
        {
            "offer_type": "hebergement",
            "status": {"$ne": "cancelled"},
            "checkout_date": target,
        },
        {"_id": 0, "id": 1, "room_id": 1, "room_tier": 1, "date": 1,
         "adults": 1, "children": 1, "first_name": 1, "last_name": 1, "name": 1,
         "phone": 1, "email": 1, "return_boat_time": 1, "checked_out_at": 1},
    )
    departing_today = await departing_cursor.to_list(length=500)
    departing_today_ids = {b["id"] for b in departing_today}

    # Build room → booking index
    booking_by_room: dict = {}
    for b in overlapping:
        rid = b.get("room_id")
        if rid:
            booking_by_room[rid] = b

    # Compose rooms list with statuses
    rooms_out = []
    for r in HEBERGEMENT_ROOMS:
        bk = booking_by_room.get(r["id"])
        status = "available"
        if bk:
            if bk["id"] in departing_today_ids:
                status = "departing_today"
            elif bk["id"] in arriving_today_ids:
                status = "arriving_today"
            else:
                status = "occupied"
        rooms_out.append({
            **r,
            "status": status,
            "booking": bk,
        })

    # Departures of bookings whose room hasn't yet been assigned still listed
    for b in departing_today:
        if not b.get("room_id"):
            rooms_out.append({
                "id": None, "label": "—", "tier": b.get("room_tier"),
                "status": "departing_today", "booking": b,
            })

    # KPIs per tier (based on the configured inventory)
    tier_inv = {t["id"]: int(t.get("inventory", 0)) for t in OFFERS["hebergement"]["room_tiers"]}
    tier_name = {t["id"]: t.get("name_fr") or t["id"] for t in OFFERS["hebergement"]["room_tiers"]}
    by_tier: dict = {}
    for r in rooms_out:
        if not r.get("id"):
            continue
        t = r["tier"]
        slot = by_tier.setdefault(t, {
            "tier_id": t, "tier_name": tier_name.get(t, t),
            "total": tier_inv.get(t, 0),
            "occupied": 0, "available": 0,
            "arriving_today": 0, "departing_today": 0,
        })
        if r["status"] in ("occupied", "arriving_today", "departing_today"):
            slot["occupied"] += 1
            if r["status"] == "arriving_today":
                slot["arriving_today"] += 1
            elif r["status"] == "departing_today":
                slot["departing_today"] += 1
        else:
            slot["available"] += 1

    # Pending arrivals / departures (no room assigned yet — to be assigned by staff)
    pending_arrivals = [b for b in overlapping if b.get("date") == target and not b.get("room_id")]

    return {
        "date": target,
        "default_checkin_time": HEBERGEMENT_DEFAULT_CHECKIN,
        "default_checkout_time": HEBERGEMENT_DEFAULT_CHECKOUT,
        "rooms": rooms_out,
        "by_tier": sorted(by_tier.values(), key=lambda x: x["tier_id"]),
        "totals": {
            "rooms_total": len(HEBERGEMENT_ROOMS),
            "rooms_occupied": sum(1 for r in rooms_out if r.get("id") and r["status"] != "available"),
            "rooms_available": sum(1 for r in rooms_out if r.get("id") and r["status"] == "available"),
            "arriving_today": len([b for b in overlapping if b.get("date") == target]),
            "departing_today": len(departing_today),
            "pending_arrivals_no_room": len(pending_arrivals),
        },
    }


class AssignRoomBody(BaseModel):
    room_id: Optional[str] = None  # null clears assignment


@api.patch("/staff/bookings/{booking_id}/assign-room")
async def assign_room(booking_id: str, body: AssignRoomBody, staff=Depends(get_current_staff)):
    """Assign (or clear) a physical room for an hebergement booking."""
    await _require_role(staff, ["manager", "admin", "receptionist"])
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    if booking.get("offer_type") != "hebergement":
        raise HTTPException(status_code=400, detail="Only hebergement bookings can be assigned a room")
    if body.room_id is None:
        await db.bookings.update_one(
            {"id": booking_id},
            {"$set": {"room_id": None, "room_label": None}},
        )
        return {"id": booking_id, "room_id": None}
    meta = HEBERGEMENT_ROOMS_BY_ID.get(body.room_id)
    if not meta:
        raise HTTPException(status_code=400, detail="Unknown room_id")
    if meta["tier"] != booking.get("room_tier"):
        raise HTTPException(
            status_code=400,
            detail=f"Cette chambre est de catégorie {meta['tier']} alors que la réservation porte sur {booking.get('room_tier')}.",
        )
    # Check overlap with another booking on the same room
    overlap = await db.bookings.find_one({
        "id": {"$ne": booking_id},
        "offer_type": "hebergement",
        "status": {"$ne": "cancelled"},
        "room_id": body.room_id,
        "date": {"$lt": booking.get("checkout_date")},
        "checkout_date": {"$gt": booking.get("date")},
    })
    if overlap:
        raise HTTPException(
            status_code=409,
            detail=f"La chambre {meta['label']} est déjà occupée sur cette période (réservation {overlap.get('id')}).",
        )
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {"room_id": body.room_id, "room_label": meta["label"]}},
    )
    return {"id": booking_id, "room_id": body.room_id, "room_label": meta["label"]}


@api.get("/staff/hebergement/calendar")
async def hebergement_calendar(month: str, staff=Depends(get_current_staff)):
    """Monthly room occupancy: for each day, how many rooms are occupied per tier."""
    await _require_role(staff, ["manager", "admin"])
    if not month or len(month) != 7:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")
    date_from = datetime.strptime(f"{month}-01", "%Y-%m-%d").date()
    next_month = (date_from + timedelta(days=32)).replace(day=1)
    # Find any booking that overlaps this month
    cursor = db.bookings.find(
        {
            "offer_type": "hebergement",
            "status": {"$ne": "cancelled"},
            "date": {"$lt": next_month.isoformat()},
            "checkout_date": {"$gt": date_from.isoformat()},
        },
        {"_id": 0, "id": 1, "date": 1, "checkout_date": 1, "rooms": 1, "room_tier": 1, "room_tier_name": 1,
         "adults": 1, "children": 1, "email": 1, "phone": 1, "participants": 1},
    )
    items = await cursor.to_list(length=2000)
    # Expand per-night occupancy
    occupancy: dict = {}  # date -> {tier_id: {name, rooms}}
    bookings_by_day: dict = {}  # date -> [booking]
    cur = date_from
    while cur < next_month:
        d = cur.isoformat()
        occupancy[d] = {}
        bookings_by_day[d] = []
        cur += timedelta(days=1)
    for b in items:
        arr = datetime.strptime(b["date"], "%Y-%m-%d").date()
        chk = datetime.strptime(b["checkout_date"], "%Y-%m-%d").date()
        night = arr
        while night < chk:
            key = night.isoformat()
            if key in occupancy:
                tier = b.get("room_tier") or "unknown"
                occupancy[key].setdefault(tier, {"tier_id": tier, "tier_name": b.get("room_tier_name") or tier, "rooms": 0})
                occupancy[key][tier]["rooms"] += int(b.get("rooms", 1))
                bookings_by_day[key].append({"id": b["id"], "rooms": b.get("rooms", 1), "tier": tier, "tier_name": b.get("room_tier_name"), "guests": int(b.get("adults", 0)) + int(b.get("children", 0))})
            night += timedelta(days=1)
    # Format as list per day
    # Build inventory lookup for hebergement tiers (after potential admin overrides)
    heb_offer = OFFERS.get("hebergement", {})
    tier_inventory = {t["id"]: int(t.get("inventory", 0)) for t in heb_offer.get("room_tiers", [])}
    total_inventory = sum(tier_inventory.values())
    days = []
    for d in sorted(occupancy.keys()):
        by_tier_list = []
        any_over = False
        for v in occupancy[d].values():
            cap = tier_inventory.get(v["tier_id"], 0)
            over = cap > 0 and v["rooms"] > cap
            if over:
                any_over = True
            by_tier_list.append({**v, "inventory": cap, "is_overbooked": over})
        total_rooms = sum(v["rooms"] for v in occupancy[d].values())
        days.append({
            "date": d,
            "total_rooms": total_rooms,
            "total_inventory": total_inventory,
            "is_overbooked": any_over,
            "by_tier": by_tier_list,
            "bookings": bookings_by_day[d],
        })
    return {"month": month, "days": days, "tier_inventory": tier_inventory, "total_inventory": total_inventory}


# ---------- Hébergement: history & stats ----------
def _heb_period_window(period: str):
    """Return (date_from_iso, date_to_iso, label) covering a period. Filters on check-in date.
    Periods: day, week, month, year, all. Bounds are inclusive."""
    today = datetime.now(timezone.utc).date()
    if period == "day":
        return today.isoformat(), today.isoformat(), "Aujourd'hui"
    if period == "week":
        return (today - timedelta(days=6)).isoformat(), today.isoformat(), "7 derniers jours"
    if period == "month":
        return (today - timedelta(days=29)).isoformat(), today.isoformat(), "30 derniers jours"
    if period == "year":
        return (today - timedelta(days=364)).isoformat(), today.isoformat(), "12 derniers mois"
    return "1970-01-01", "2999-12-31", "Depuis le lancement"


@api.get("/staff/hebergement/stats")
async def hebergement_stats(period: str = "month", staff=Depends(get_current_staff)):
    """Hébergement statistics over a period: occupancy, revenue, by tier, top guests, history."""
    await _require_role(staff, ["manager", "admin"])
    date_from, date_to, label = _heb_period_window(period)
    cursor = db.bookings.find(
        {
            "offer_type": "hebergement",
            # iter-44: exclude abandoned-cart pending so occupancy & revenue
            # match real (paid + cash-on-arrival) bookings only.
            "status": {"$nin": ["cancelled", "pending"]},
            "date": {"$gte": date_from, "$lte": date_to},
        },
        {"_id": 0, "id": 1, "date": 1, "checkout_date": 1, "nights": 1, "rooms": 1,
         "room_tier": 1, "room_tier_name": 1, "total_amount": 1, "paid_amount": 1,
         "balance_due": 1, "adults": 1, "children": 1, "participants": 1,
         "boat_time": 1, "return_boat_time": 1, "payment_method": 1, "deposit_pct": 1,
         "status": 1, "paid_at": 1, "created_at": 1, "phone": 1, "email": 1},
    )
    bookings = await cursor.to_list(length=2000)

    heb_offer = OFFERS.get("hebergement", {})
    tier_inventory = {t["id"]: int(t.get("inventory", 0)) for t in heb_offer.get("room_tiers", [])}
    tier_name_by_id = {t["id"]: t.get("name_fr", t["id"]) for t in heb_offer.get("room_tiers", [])}
    total_inventory = sum(tier_inventory.values())

    nights_sold = 0
    revenue_total = 0
    revenue_paid = 0
    balance_due_total = 0
    total_stays = len(bookings)
    by_tier_agg: dict = {tid: {"tier_id": tid, "tier_name": tier_name_by_id.get(tid, tid),
                                 "stays": 0, "rooms": 0, "nights": 0, "revenue": 0,
                                 "inventory": tier_inventory.get(tid, 0)} for tid in tier_inventory}
    # Per-day occupancy across the window
    try:
        d_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        d_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except Exception:
        d_from = datetime.now(timezone.utc).date()
        d_to = d_from
    daily_nights: dict = {}
    daily_revenue: dict = {}
    cur = d_from
    while cur <= d_to:
        daily_nights[cur.isoformat()] = 0
        daily_revenue[cur.isoformat()] = 0
        cur += timedelta(days=1)
    days_in_window = max(1, (d_to - d_from).days + 1)
    # Top guests aggregation
    guest_agg: dict = {}
    for b in bookings:
        n = int(b.get("nights") or 0)
        r = int(b.get("rooms") or 1)
        room_nights = n * r
        nights_sold += room_nights
        amount = int(b.get("total_amount") or 0)
        paid = int(b.get("paid_amount") or 0)
        bal = int(b.get("balance_due") or 0)
        revenue_total += amount
        revenue_paid += paid
        balance_due_total += bal
        tid = b.get("room_tier") or "unknown"
        if tid in by_tier_agg:
            by_tier_agg[tid]["stays"] += 1
            by_tier_agg[tid]["rooms"] += r
            by_tier_agg[tid]["nights"] += room_nights
            by_tier_agg[tid]["revenue"] += amount
        # Daily occupancy: each night between arrival and checkout-1
        try:
            arr = datetime.strptime(b["date"], "%Y-%m-%d").date()
            chk = datetime.strptime(b.get("checkout_date") or b["date"], "%Y-%m-%d").date()
        except Exception:
            continue
        per_night_rev = (amount // n) if n > 0 else 0
        nd = arr
        while nd < chk:
            key = nd.isoformat()
            if key in daily_nights:
                daily_nights[key] += r
                daily_revenue[key] += per_night_rev
            nd += timedelta(days=1)
        # Guest
        primary = next((p for p in (b.get("participants") or []) if p.get("kind") == "adult"), None) or {}
        if primary:
            email = (primary.get("email") or b.get("email") or "").lower()
            if email:
                g = guest_agg.setdefault(email, {
                    "email": email,
                    "name": primary.get("name", ""),
                    "surname": primary.get("surname", ""),
                    "nationality": primary.get("nationality", ""),
                    "stays": 0,
                    "nights": 0,
                    "revenue": 0,
                })
                g["stays"] += 1
                g["nights"] += room_nights
                g["revenue"] += amount

    available_room_nights = total_inventory * days_in_window
    occupancy_rate = round((nights_sold / available_room_nights * 100), 1) if available_room_nights else 0
    avg_stay_nights = round((nights_sold / total_stays), 1) if total_stays else 0
    avg_revenue_per_stay = int(revenue_total / total_stays) if total_stays else 0
    avg_revenue_per_night = int(revenue_total / nights_sold) if nights_sold else 0

    by_tier = []
    for tid, agg in by_tier_agg.items():
        share = (agg["revenue"] / revenue_total * 100) if revenue_total else 0
        tier_available = agg["inventory"] * days_in_window
        tier_occ = round((agg["nights"] / tier_available * 100), 1) if tier_available else 0
        by_tier.append({**agg, "revenue_share_pct": round(share, 1), "occupancy_pct": tier_occ})
    by_tier.sort(key=lambda x: x["revenue"], reverse=True)

    daily_trend = [
        {"date": d, "nights": daily_nights[d], "revenue": daily_revenue[d]}
        for d in sorted(daily_nights.keys())
    ]
    top_guests = sorted(guest_agg.values(), key=lambda x: x["nights"], reverse=True)[:10]

    # History (most recent first, limited)
    history = sorted(bookings, key=lambda b: (b.get("date") or "", b.get("created_at") or ""), reverse=True)
    # Strip heavy fields for the history table
    history_lite = [
        {
            "id": b["id"],
            "date": b.get("date"),
            "checkout_date": b.get("checkout_date"),
            "nights": int(b.get("nights") or 0),
            "rooms": int(b.get("rooms") or 1),
            "room_tier": b.get("room_tier"),
            "room_tier_name": b.get("room_tier_name"),
            "adults": int(b.get("adults") or 0),
            "children": int(b.get("children") or 0),
            "total_amount": int(b.get("total_amount") or 0),
            "paid_amount": int(b.get("paid_amount") or 0),
            "balance_due": int(b.get("balance_due") or 0),
            "deposit_pct": b.get("deposit_pct"),
            "payment_method": b.get("payment_method"),
            "status": b.get("status"),
            "boat_time": b.get("boat_time"),
            "return_boat_time": b.get("return_boat_time"),
            "phone": b.get("phone"),
            "email": b.get("email"),
            "primary_name": (
                next((p for p in (b.get("participants") or []) if p.get("kind") == "adult"), None) or {}
            ).get("name", ""),
            "primary_surname": (
                next((p for p in (b.get("participants") or []) if p.get("kind") == "adult"), None) or {}
            ).get("surname", ""),
        }
        for b in history
    ]
    return {
        "period": period,
        "period_label": label,
        "date_from": date_from,
        "date_to": date_to,
        "days_in_window": days_in_window,
        "total_inventory": total_inventory,
        "tier_inventory": tier_inventory,
        "kpis": {
            "total_stays": total_stays,
            "nights_sold": nights_sold,
            "occupancy_rate_pct": occupancy_rate,
            "revenue_total": revenue_total,
            "revenue_paid": revenue_paid,
            "balance_due_total": balance_due_total,
            "avg_stay_nights": avg_stay_nights,
            "avg_revenue_per_stay": avg_revenue_per_stay,
            "avg_revenue_per_night": avg_revenue_per_night,
        },
        "by_tier": by_tier,
        "daily_trend": daily_trend,
        "top_guests": top_guests,
        "history": history_lite,
    }


@api.get("/staff/hebergement/report.pdf")
async def export_hebergement_pdf(period: str = "month", staff=Depends(get_current_staff)):
    """Stylized PDF report of Hébergement statistics for the selected period."""
    await _require_role(staff, ["manager", "admin"])
    data = await hebergement_stats(period=period, staff=staff)  # type: ignore

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from fastapi.responses import StreamingResponse

    styles = _pdf_styles()
    GOLD, DARK, LIGHT, MUTED = styles["GOLD"], styles["DARK"], styles["LIGHT"], styles["MUTED"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
    elements = []
    elements.append(Paragraph("Boulay Beach Resort", styles["h1"]))
    elements.append(Paragraph(f"Rapport Hébergement — {data['period_label']} ({data['date_from']} → {data['date_to']})", styles["sub"]))

    k = data["kpis"]
    kpi_rows = [
        ["Séjours", "Nuitées vendues", "Taux d'occupation", "Revenu total"],
        [str(k["total_stays"]), str(k["nights_sold"]), f"{k['occupancy_rate_pct']}%", _format_xof(k["revenue_total"])],
    ]
    kpi_tbl = Table(kpi_rows, colWidths=[4.2 * cm, 4.2 * cm, 4.2 * cm, 4.2 * cm])
    kpi_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("TEXTCOLOR", (0, 0), (-1, 0), MUTED),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 12),
        ("TEXTCOLOR", (0, 1), (-1, 1), DARK),
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.5, GOLD),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E0D5B5")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_tbl)
    elements.append(Spacer(1, 0.2 * cm))
    sub_rows = [
        ["Séjour moyen", "Revenu / séjour", "Revenu / nuitée", "Encaissé / Solde dû"],
        [
            f"{k['avg_stay_nights']} nuits",
            _format_xof(k["avg_revenue_per_stay"]),
            _format_xof(k["avg_revenue_per_night"]),
            f"{_format_xof(k['revenue_paid'])} / {_format_xof(k['balance_due_total'])}",
        ],
    ]
    sub_tbl = Table(sub_rows, colWidths=[4.2 * cm, 4.2 * cm, 4.2 * cm, 4.2 * cm])
    sub_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("TEXTCOLOR", (0, 0), (-1, 0), MUTED),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 10),
        ("TEXTCOLOR", (0, 1), (-1, 1), DARK),
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.5, GOLD),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E0D5B5")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(sub_tbl)

    # By tier
    if data.get("by_tier"):
        elements.append(Paragraph("Répartition par catégorie", styles["h2"]))
        rows = [["Catégorie", "Séjours", "Nuitées", "Taux occ.", "Revenu", "Part"]]
        for t in data["by_tier"]:
            rows.append([
                t["tier_name"],
                str(t["stays"]),
                str(t["nights"]),
                f"{t['occupancy_pct']}%",
                _format_xof(t["revenue"]),
                f"{t['revenue_share_pct']}%",
            ])
        tbl = Table(rows, colWidths=[5 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm, 3.5 * cm, 1.8 * cm])
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(tbl)

    # Top guests
    if data.get("top_guests"):
        elements.append(Paragraph("Top 10 clients (par nuitées)", styles["h2"]))
        rows = [["#", "Client", "Nationalité", "Séjours", "Nuitées", "Total dépensé"]]
        for i, g in enumerate(data["top_guests"], start=1):
            full = f"{g.get('surname','')} {g.get('name','')}".strip() or "—"
            rows.append([str(i), full, g.get("nationality") or "—", str(g["stays"]), str(g["nights"]), _format_xof(g["revenue"])])
        tbl = Table(rows, colWidths=[1 * cm, 4.5 * cm, 3.5 * cm, 2 * cm, 2 * cm, 3.5 * cm], repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(tbl)

    # History (recent stays)
    if data.get("history"):
        elements.append(Paragraph("Historique des séjours", styles["h2"]))
        rows = [["Arrivée", "Départ", "Client", "Cat.", "Ch.", "Nuits", "Total", "Solde"]]
        for b in data["history"][:60]:
            full = f"{b.get('primary_surname','')} {b.get('primary_name','')}".strip() or "—"
            rows.append([
                b.get("date") or "—",
                b.get("checkout_date") or "—",
                full[:24],
                (b.get("room_tier_name") or "—")[:18],
                str(b.get("rooms") or 1),
                str(b.get("nights") or 0),
                _format_xof(b.get("total_amount") or 0),
                _format_xof(b.get("balance_due") or 0),
            ])
        tbl = Table(rows, colWidths=[2.1 * cm, 2.1 * cm, 4 * cm, 3 * cm, 1 * cm, 1.2 * cm, 2.8 * cm, 2.6 * cm], repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7.5),
            ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
            ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(tbl)

    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(
        f"Rapport généré le {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M UTC')} · Boulay Beach Resort, Abidjan",
        styles["small"],
    ))

    doc.build(elements, onFirstPage=_pdf_footer_factory(styles), onLaterPages=_pdf_footer_factory(styles))
    buf.seek(0)
    filename = f"bbr-hebergement-{period}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------- Staff-created bookings (manager+) ----------
class StaffBookingCreate(BaseModel):
    """Body for POST /staff/bookings — manager creates a booking on behalf of a guest."""
    offer_type: OfferType
    event_id: Optional[str] = None  # required when offer_type='special_event'
    date: str
    checkout_date: Optional[str] = None
    room_tier: Optional[str] = None
    rooms: int = Field(default=1, ge=1, le=20)
    adults: int = Field(ge=0, le=20)
    children: int = Field(ge=0, le=20)
    boat_time: str
    return_boat_time: Optional[str] = None
    participants: List[Participant]
    special_requests: Optional[str] = ""
    # Payment: manager picks the method directly. For 'deposit', supply deposit_pct.
    # 'online' creates a pending booking and emails a payment link to the client.
    payment_method: Literal["card", "mobile_money", "cash", "deposit", "online"] = "cash"
    deposit_pct: Optional[Literal[10, 30, 70]] = None


@api.post("/staff/bookings")
async def staff_create_booking(body: StaffBookingCreate, staff=Depends(get_current_staff)):
    """Manager creates a booking.

    For instant methods (cash / card / mobile_money / deposit) the booking is
    paid immediately (legacy behaviour). For ``online`` the booking is left in
    ``pending_payment`` state with a secure payment-link token, and an email is
    dispatched to the client with a "Payer maintenant" CTA. Once they complete
    the payment, the existing FineoPay settle flow regenerates the QR ticket
    and emails the confirmation.
    """
    await _require_role(staff, ["manager", "admin"])
    # Step 1: create booking (reuses public validator)
    payload = BookingCreate(
        offer_type=body.offer_type,
        special_event_id=body.event_id,
        date=body.date,
        checkout_date=body.checkout_date,
        room_tier=body.room_tier,
        rooms=body.rooms,
        adults=body.adults,
        children=body.children,
        boat_time=body.boat_time,
        return_boat_time=body.return_boat_time,
        participants=body.participants,
        special_requests=body.special_requests or "",
    )
    booking = await create_booking(payload)  # type: ignore
    # Mark as staff-created for audit/reporting
    await db.bookings.update_one(
        {"id": booking["id"]},
        {"$set": {"created_by_staff": True, "created_by_email": staff.get("email")}},
    )

    # ------- Online payment: create a payment link, email it, do NOT pay -------
    if body.payment_method == "online":
        token = uuid.uuid4().hex
        expires_at = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
        await db.bookings.update_one(
            {"id": booking["id"]},
            {"$set": {
                "status": "pending_payment",
                "payment_method": "online",
                "payment_link_token": token,
                "payment_link_expires_at": expires_at,
            }},
        )
        # Build public URL (FINEO_PUBLIC_BASE_URL = production app domain)
        base = FINEO_PUBLIC_BASE_URL or os.environ.get("PUBLIC_BASE_URL", "")
        payment_url = f"{base.rstrip('/')}/pay/{token}" if base else f"/pay/{token}"

        # Resolve offer label for the email
        if booking["offer_type"] == "special_event":
            offer = await _resolve_special_event_offer(booking.get("special_event_id") or "", booking.get("date"))
        else:
            offer = OFFERS.get(booking["offer_type"], {"name_fr": booking["offer_type"]})
        # Format the booking date as DD/MM/YYYY
        try:
            date_str = datetime.fromisoformat(booking["date"]).strftime("%d/%m/%Y")
        except Exception:
            date_str = booking["date"]
        # Amount label
        amount_label = f"{int(booking.get('total_amount', 0)):,}".replace(",", " ") + " FCFA"

        email_sent = False
        try:
            from services import email_service as _es  # type: ignore
            primary = (booking.get("participants") or [{}])[0]
            booker_name = booking.get("name") or primary.get("name") or ""
            booker_surname = booking.get("surname") or primary.get("surname") or ""
            full_name = f"{booker_name} {booker_surname}".strip()
            rendered = _es.render_payment_link(
                name=full_name,
                ref=booking["id"][:8].upper(),
                offer_label=offer.get("name_fr") or booking["offer_type"],
                date_str=date_str,
                boat_time=booking.get("boat_time"),
                amount_label=amount_label,
                payment_url=payment_url,
                expires_label="dans 7 jours",
            )
            res = await _es.send_email(
                db,
                to_email=booking["email"],
                subject=rendered["subject"],
                html=rendered["html"],
                plain=rendered["plain"],
                purpose="payment_link",
                booking_id=booking["id"],
                to_name=full_name or None,
            )
            email_sent = bool(res.get("ok"))
        except Exception as ex:
            logger.warning("Payment-link email failed for booking %s: %s", booking["id"], ex)

        return {
            **booking,
            "status": "pending_payment",
            "payment_method": "online",
            "payment_link": payment_url,
            "payment_link_token": token,
            "email_sent": email_sent,
            "created_by_staff": True,
            "created_by_email": staff.get("email"),
        }

    # ------- Instant methods: legacy immediate-confirmation path -------
    pay = PayBooking(
        reference_token=booking["reference_token"],
        payment_method=body.payment_method,
        deposit_pct=body.deposit_pct,
    )
    paid = await pay_booking(booking["id"], pay)  # type: ignore
    paid["created_by_staff"] = True
    paid["created_by_email"] = staff.get("email")
    return paid


# =================================================================
# MODULE RECEIPTS — Fiscal receipts (activities + bookings/events)
# =================================================================
# A receipt is created automatically whenever real money changes hands:
#   - Wallet activity charge        → source="activity"
#   - Booking paid (public/staff)   → source="booking"
#   - Event privatization paid      → source="event"
# Each receipt carries an HMAC signature derived from the immutable fields so
# the staff app can verify authenticity later (digital seal).

RECEIPT_SECRET = os.environ.get("RECEIPT_SECRET", JWT_SECRET)


async def _next_receipt_number() -> str:
    """Generate sequential daily receipt number BBR-YYYYMMDD-XXXXX (atomic counter)."""
    day = datetime.now(timezone.utc).date().isoformat().replace("-", "")
    counter = await db.counters.find_one_and_update(
        {"id": f"receipt-{day}"},
        {"$inc": {"value": 1}},
        upsert=True,
        return_document=True,
    )
    seq = (counter or {}).get("value") or 1
    return f"BBR-{day}-{seq:05d}"


def _sign_receipt(receipt: dict) -> str:
    """Compact HMAC-SHA256 signature over receipt id + total + source_id + issued_at."""
    import hmac as _hmac
    import hashlib as _hashlib
    msg = f"{receipt['id']}|{receipt['total']}|{receipt.get('source_id') or ''}|{receipt['issued_at']}".encode()
    digest = _hmac.new(RECEIPT_SECRET.encode(), msg, _hashlib.sha256).hexdigest()
    return digest[:16].upper()


async def _create_receipt(
    *,
    source: Literal["activity", "booking", "event"],
    source_id: str,
    customer_name: str,
    customer_email: str = "",
    customer_phone: str = "",
    lines: List[dict],
    payment_method: str,
    currency: str = "XOF",
    issued_by: str = "system",
    issued_by_role: str = "system",
    metadata: Optional[dict] = None,
):
    """Persist a fiscal receipt. Idempotent on (source, source_id, sub_id)."""
    sub_id = (metadata or {}).get("sub_id")
    # Idempotency: don't create twice for the same wallet charge / booking payment
    if sub_id:
        existing = await db.receipts.find_one({"source": source, "source_id": source_id, "metadata.sub_id": sub_id}, {"_id": 0})
        if existing:
            return existing
    elif source == "booking":
        existing = await db.receipts.find_one({"source": source, "source_id": source_id}, {"_id": 0})
        if existing:
            return existing
    subtotal = sum(int(ln.get("total", 0)) for ln in lines)
    rid = str(uuid.uuid4())
    issued_at = now_iso()
    receipt = {
        "id": rid,
        "receipt_number": await _next_receipt_number(),
        "source": source,
        "source_id": source_id,
        "customer_name": customer_name or "—",
        "customer_email": customer_email or "",
        "customer_phone": customer_phone or "",
        "lines": [
            {
                "description": str(ln.get("description", "")),
                "quantity": int(ln.get("quantity", 1)),
                "unit_price": int(ln.get("unit_price", 0)),
                "total": int(ln.get("total", 0)),
            }
            for ln in lines
        ],
        "subtotal": subtotal,
        "total": subtotal,
        "currency": currency,
        "payment_method": payment_method,
        "issued_at": issued_at,
        "issued_by": issued_by,
        "issued_by_role": issued_by_role,
        "metadata": metadata or {},
        "voided": False,
    }
    receipt["signature"] = _sign_receipt(receipt)
    await db.receipts.insert_one(dict(receipt))
    receipt.pop("_id", None)
    return receipt


@api.get("/staff/receipts")
async def list_receipts(
    source: Optional[Literal["activity", "booking", "event"]] = None,
    pole: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    q: Optional[str] = None,
    payment_method: Optional[str] = None,
    page: int = 1,
    page_size: int = 30,
    staff=Depends(get_current_staff),
):
    """List receipts. Manager+admin only. Each receipt is enriched with its
    derived ``pole`` (from the linked booking when applicable, or
    'activites_events' for event receipts). A global ``summary_by_pole`` is
    computed over the current filter (excluding the pole filter itself, so the
    pole tabs always reflect the full distribution)."""
    await _require_role(staff, ["manager", "admin"])
    filter_q: dict = {"voided": {"$ne": True}}
    if source:
        filter_q["source"] = source
    if payment_method:
        filter_q["payment_method"] = payment_method
    if date_from or date_to:
        rng: dict = {}
        if date_from:
            rng["$gte"] = date_from
        if date_to:
            rng["$lte"] = date_to + "T23:59:59Z"
        filter_q["issued_at"] = rng
    if q:
        rgx = re.compile(re.escape(q), re.IGNORECASE)
        filter_q["$or"] = [
            {"receipt_number": rgx},
            {"customer_name": rgx},
            {"customer_email": rgx},
            {"customer_phone": rgx},
        ]
    page = max(1, page)
    page_size = max(1, min(100, page_size))

    # ----- Helper: build a pole index for the full filtered set -----
    # We need to map every receipt → pole regardless of pagination, so we run a
    # light projection first, build a booking_id → pole map (one lookup per
    # distinct booking), then group / paginate in memory after the pole filter.
    proj_cursor = db.receipts.find(filter_q, {
        "_id": 0, "id": 1, "source": 1, "source_id": 1,
        "metadata": 1, "total": 1,
    })
    all_proj = await proj_cursor.to_list(length=5000)
    booking_ids: set = set()
    for r in all_proj:
        if r.get("source") == "booking" and r.get("source_id"):
            booking_ids.add(r["source_id"])
        bid = (r.get("metadata") or {}).get("booking_id")
        if bid:
            booking_ids.add(bid)
    pole_by_booking: dict = {}
    if booking_ids:
        async for b in db.bookings.find(
            {"id": {"$in": list(booking_ids)}},
            {"_id": 0, "id": 1, "pole": 1, "offer_type": 1},
        ):
            p = b.get("pole") or _pole_for_offer(b.get("offer_type", "")) or ""
            if p:
                pole_by_booking[b["id"]] = p

    def _resolve_pole(r: dict) -> str:
        if r.get("source") == "event":
            return "activites_events"
        bid = r.get("source_id") if r.get("source") == "booking" else (r.get("metadata") or {}).get("booking_id")
        return pole_by_booking.get(bid or "", "")

    # Apply pole filter (post-resolution) and compute the global pole summary
    summary_by_pole: dict = {pid: {"count": 0, "total": 0} for pid in POLES}
    filtered_ids: list = []
    for r in all_proj:
        rp = _resolve_pole(r)
        if rp:
            summary_by_pole[rp]["count"] += 1
            summary_by_pole[rp]["total"] += int(r.get("total", 0) or 0)
        if pole and rp != pole:
            continue
        filtered_ids.append(r["id"])

    total = len(filtered_ids)
    # Paginate the filtered list and fetch full docs in order
    start = (page - 1) * page_size
    page_ids = filtered_ids[start:start + page_size] if filtered_ids else []
    items: list = []
    if page_ids:
        docs = await db.receipts.find({"id": {"$in": page_ids}}, {"_id": 0}).to_list(length=page_size)
        docs.sort(key=lambda x: x.get("issued_at", ""), reverse=True)
        for d in docs:
            d["pole"] = _resolve_pole(d)
            items.append(d)

    # by_source aggregation runs on the original filter_q (without pole),
    # mirroring summary_by_pole (so both summaries reflect the same period).
    pipeline = [
        {"$match": filter_q},
        {"$group": {"_id": "$source", "count": {"$sum": 1}, "total": {"$sum": "$total"}}},
    ]
    by_source = {}
    async for row in db.receipts.aggregate(pipeline):
        by_source[row["_id"]] = {"count": row["count"], "total": int(row["total"])}

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size if total else 1,
        "summary_by_source": by_source,
        "summary_by_pole": summary_by_pole,
        "poles": [
            {"id": pid, "name_fr": p["name_fr"], "sort_order": p["sort_order"]}
            for pid, p in sorted(POLES.items(), key=lambda kv: kv[1]["sort_order"])
        ],
    }


@api.get("/staff/receipts/{receipt_id}.pdf")
async def export_receipt_pdf(receipt_id: str, staff=Depends(get_current_staff)):
    """Stylized fiscal receipt PDF (BBR header, line items, subtotal, total, signature)."""
    await _require_role(staff, ["manager", "admin"])
    receipt = await db.receipts.find_one({"id": receipt_id}, {"_id": 0})
    if not receipt:
        raise HTTPException(status_code=404, detail="Reçu introuvable")

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
    from fastapi.responses import StreamingResponse

    styles = _pdf_styles()
    GOLD, DARK, LIGHT, MUTED = styles["GOLD"], styles["DARK"], styles["LIGHT"], styles["MUTED"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
    elements = []
    # BBR logo header — best-effort fetch from CDN with byte cache.
    logo_bytes = _fetch_logo_bytes()
    if logo_bytes:
        try:
            img = RLImage(io.BytesIO(logo_bytes))
            iw, ih = img.imageWidth, img.imageHeight
            target_h = 1.8 * cm
            img.drawHeight = target_h
            img.drawWidth = target_h * iw / max(ih, 1)
            img.hAlign = "CENTER"
            elements.append(img)
            elements.append(Spacer(1, 0.2 * cm))
        except Exception:
            pass
    elements.append(Paragraph("Boulay Beach Resort", styles["h1"]))
    elements.append(Paragraph("Reçu de paiement", styles["sub"]))

    # Receipt meta block
    src_fr = {"activity": "Activité sur place", "booking": "Réservation", "event": "Privatisation / Événement"}.get(receipt["source"], receipt["source"])
    meta_rows = [
        ["N° de reçu", receipt["receipt_number"]],
        ["Date d'émission", receipt["issued_at"].replace("T", " à ").split(".")[0] + " UTC"],
        ["Type", src_fr],
        ["Client", receipt.get("customer_name") or "—"],
    ]
    if receipt.get("customer_email"):
        meta_rows.append(["Email", receipt["customer_email"]])
    if receipt.get("customer_phone"):
        meta_rows.append(["Téléphone", receipt["customer_phone"]])
    meta_rows.append(["Émis par", f"{receipt.get('issued_by','')} ({receipt.get('issued_by_role','')})"])
    meta_tbl = Table(meta_rows, colWidths=[4.5 * cm, 12 * cm])
    meta_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), MUTED),
        ("TEXTCOLOR", (1, 0), (1, -1), DARK),
        ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#EEE")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(meta_tbl)
    elements.append(Spacer(1, 0.4 * cm))

    # Line items
    lines_rows = [["Description", "Qté", "P.U.", "Total"]]
    for ln in receipt["lines"]:
        lines_rows.append([
            ln["description"],
            str(ln.get("quantity", 1)),
            _format_xof(ln.get("unit_price", 0)),
            _format_xof(ln.get("total", 0)),
        ])
    lines_tbl = Table(lines_rows, colWidths=[9 * cm, 1.8 * cm, 3 * cm, 3 * cm], repeatRows=1)
    lines_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
        ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, GOLD),
        ("FONTSIZE", (0, 1), (-1, -1), 9.5),
        ("TEXTCOLOR", (0, 1), (-1, -1), DARK),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("LINEBELOW", (0, 1), (-1, -1), 0.2, colors.HexColor("#EEE")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(lines_tbl)
    elements.append(Spacer(1, 0.3 * cm))

    # Totals
    totals_rows = [
        ["Sous-total", _format_xof(receipt["subtotal"])],
        ["Total payé", _format_xof(receipt["total"])],
        ["Mode de paiement", PAYMENT_METHOD_FR_LABEL.get(receipt["payment_method"], receipt["payment_method"])],
    ]
    totals_tbl = Table(totals_rows, colWidths=[13 * cm, 3.8 * cm])
    totals_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONTSIZE", (0, 1), (-1, 1), 13),
        ("TEXTCOLOR", (0, 0), (0, -1), MUTED),
        ("TEXTCOLOR", (1, 0), (1, -1), DARK),
        ("TEXTCOLOR", (0, 1), (-1, 1), GOLD),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("LINEABOVE", (0, 1), (-1, 1), 0.5, GOLD),
        ("LINEBELOW", (0, 1), (-1, 1), 0.5, GOLD),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(totals_tbl)
    elements.append(Spacer(1, 0.8 * cm))

    # Digital seal
    seal = receipt.get("signature") or ""
    elements.append(Paragraph(
        f"<b>Signature numérique</b>&nbsp;·&nbsp; <font face='Courier'>{seal}</font>",
        styles["small"],
    ))
    elements.append(Paragraph(
        "Ce reçu fait foi du paiement. La signature numérique HMAC-SHA256 garantit son authenticité.",
        styles["small"],
    ))
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(
        "Boulay Beach Resort · Abidjan, Côte d'Ivoire · contact@boulaybeach.ci",
        styles["small"],
    ))

    doc.build(elements, onFirstPage=_pdf_footer_factory(styles), onLaterPages=_pdf_footer_factory(styles))
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{receipt["receipt_number"]}.pdf"'},
    )


PAYMENT_METHOD_FR_LABEL = {
    "card": "Carte bancaire",
    "fineo": "Carte bancaire",
    "mobile_money": "Mobile Money",
    "cash": "Espèces",
    "deposit": "Acompte (carte)",
    "on_site": "Sur place",
    "transfer": "Virement",
}


@api.get("/staff/receipts/{receipt_id}")
async def get_receipt(receipt_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["manager", "admin"])
    receipt = await db.receipts.find_one({"id": receipt_id}, {"_id": 0})
    if not receipt:
        raise HTTPException(status_code=404, detail="Reçu introuvable")
    return receipt


# =================================================================
# MODULE LOISIRS — Event privatization requests
# =================================================================

@api.get("/staff/loisirs/events")
async def list_event_requests(status: Optional[str] = None, staff=Depends(get_current_staff)):
    """List event/privatization requests."""
    await _require_role(staff, ["manager", "admin"])
    q: dict = {}
    if status:
        q["status"] = status
    items = await db.event_requests.find(q, {"_id": 0}).sort("created_at", -1).to_list(length=500)
    return {"items": items, "count": len(items)}


@api.patch("/staff/loisirs/events/{event_id}")
async def update_event_request(
    event_id: str,
    status: Optional[str] = Body(None, embed=True),
    notes: Optional[str] = Body(None, embed=True),
    staff=Depends(get_current_staff),
):
    """Update an event request status / notes."""
    await _require_role(staff, ["manager", "admin"])
    if status and status not in ("new", "contacted", "confirmed", "declined", "completed"):
        raise HTTPException(status_code=400, detail="Invalid status")
    update: dict = {}
    if status:
        update["status"] = status
    if notes is not None:
        update["notes"] = notes
    if not update:
        return {"ok": True}
    res = await db.event_requests.update_one({"id": event_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Event request not found")
    return {"ok": True}


class EventPayment(BaseModel):
    amount: int = Field(ge=1, le=100_000_000)
    payment_method: Literal["card", "mobile_money", "cash", "transfer"] = "cash"
    description: Optional[str] = None


@api.post("/staff/loisirs/events/{event_id}/payment")
async def register_event_payment(event_id: str, body: EventPayment, staff=Depends(get_current_staff)):
    """Register a payment for a privatization request and emit a fiscal receipt."""
    await _require_role(staff, ["manager", "admin"])
    event = await db.event_requests.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Demande introuvable")
    description = body.description or f"Privatisation — {event.get('event_type','Événement')} du {event.get('event_date','')}"
    receipt = await _create_receipt(
        source="event",
        source_id=event_id,
        customer_name=f"{event.get('surname','')} {event.get('name','')}".strip() or "—",
        customer_email=event.get("email", ""),
        customer_phone=event.get("phone", ""),
        lines=[{"description": description, "quantity": 1, "unit_price": int(body.amount), "total": int(body.amount)}],
        payment_method=body.payment_method,
        issued_by=staff.get("name") or "",
        issued_by_role=staff.get("role") or "",
        metadata={
            "event_type": event.get("event_type"),
            "event_date": event.get("event_date"),
            "guest_count": event.get("guest_count"),
            "sub_id": str(uuid.uuid4()),  # allow multiple payments per event (e.g. acompte then solde)
        },
    )
    # Append payment to event for audit/history
    await db.event_requests.update_one(
        {"id": event_id},
        {
            "$push": {"payments": {
                "id": receipt["id"],
                "receipt_number": receipt["receipt_number"],
                "amount": int(body.amount),
                "payment_method": body.payment_method,
                "paid_at": receipt["issued_at"],
                "paid_by": staff.get("email") or "",
            }},
            "$inc": {"total_paid": int(body.amount)},
        },
    )
    return {"ok": True, "receipt": receipt}


# =================================================================
# CONFIG ADMIN — Staff user management & offer price overrides
# =================================================================

class StaffUserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str = Field(min_length=8)
    role: Literal[
        # New 7-role catalog
        "hotesse", "serveur_caisse", "logistique", "verification",
        "manager_pole", "management_general", "admin",
        # iter-42: canteen roles
        "directeur", "rh", "cuisine",
        # iter-46: planning
        "chef_dept",
        # Legacy roles still accepted for backward compatibility
        "receptionist", "manager",
    ]
    pole_id: Optional[Literal["beach_club", "hebergement", "corporate", "activites_events", "le_kaai"]] = None
    # iter-46: dept_id for chef_dept role (planning module)
    dept_id: Optional[str] = None
    # Optional per-user override of the sections visible in the sidebar.
    # When set (non-empty list), this fully replaces the role-defaults. When
    # None or empty, the role's default matrix is used (backward compatible).
    nav_sections: Optional[List[str]] = None


class StaffUserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = Field(default=None, min_length=8)
    role: Optional[Literal[
        "hotesse", "serveur_caisse", "logistique", "verification",
        "manager_pole", "management_general", "admin",
        "directeur", "rh", "cuisine",
        "chef_dept",
        "receptionist", "manager",
    ]] = None
    pole_id: Optional[Literal["beach_club", "hebergement", "corporate", "activites_events", "le_kaai"]] = None
    nav_sections: Optional[List[str]] = None


@api.get("/staff/config/users")
async def list_staff_users(staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    items = await db.staff.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", 1).to_list(length=200)
    return {"items": items}


@api.post("/staff/config/users")
async def create_staff_user(body: StaffUserCreate, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    existing = await db.staff.find_one({"email": body.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    doc = {
        "id": str(uuid.uuid4()),
        "name": body.name.strip(),
        "email": body.email.lower(),
        "role": body.role,
        "pole_id": body.pole_id if body.role == "manager_pole" else None,
        "nav_sections": body.nav_sections or None,
        "password_hash": hash_password(body.password),
        "created_at": now_iso(),
    }
    await db.staff.insert_one(doc)
    doc.pop("_id", None)
    doc.pop("password_hash", None)
    return doc


@api.patch("/staff/config/users/{user_id}")
async def update_staff_user(user_id: str, body: StaffUserUpdate, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    target = await db.staff.find_one({"id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    update: dict = {}
    if body.name is not None:
        update["name"] = body.name.strip()
    if body.email is not None:
        update["email"] = body.email.lower()
    if body.role is not None:
        update["role"] = body.role
        # Reset pole_id when role is no longer manager_pole
        if body.role != "manager_pole":
            update["pole_id"] = None
    if body.pole_id is not None:
        update["pole_id"] = body.pole_id
    if body.password is not None:
        update["password_hash"] = hash_password(body.password)
    if body.nav_sections is not None:
        # Empty list ⇒ clear override (fall back to role defaults).
        update["nav_sections"] = body.nav_sections or None
    if not update:
        return {"ok": True}
    await db.staff.update_one({"id": user_id}, {"$set": update})
    return {"ok": True}


@api.delete("/staff/config/users/{user_id}")
async def delete_staff_user(user_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    if user_id == staff.get("id"):
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    res = await db.staff.delete_one({"id": user_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True}


class OfferPriceOverride(BaseModel):
    # Pricing & capacity
    price_adult: Optional[int] = Field(default=None, ge=0)
    price_child: Optional[int] = Field(default=None, ge=0)
    max_capacity: Optional[int] = Field(default=None, ge=1)
    room_tiers: Optional[List[dict]] = None  # [{id, name_fr, name_en, price}]
    # Editorial content (so admin can rename/retitle offers from the back-office)
    name_fr: Optional[str] = Field(default=None, max_length=120)
    name_en: Optional[str] = Field(default=None, max_length=120)
    schedule_fr: Optional[str] = Field(default=None, max_length=180)
    schedule_en: Optional[str] = Field(default=None, max_length=180)
    tagline_fr: Optional[str] = Field(default=None, max_length=4000)
    tagline_en: Optional[str] = Field(default=None, max_length=4000)
    image_url: Optional[str] = Field(default=None, max_length=600)


# Fields that flow from the override to the in-memory OFFERS dict
_OVERRIDE_SCALAR_FIELDS = (
    "price_adult", "price_child", "max_capacity",
    "name_fr", "name_en", "schedule_fr", "schedule_en",
    "tagline_fr", "tagline_en", "image_url",
)


async def _apply_overrides(offer: dict) -> dict:
    """Merge any stored overrides on top of the static OFFERS dict."""
    override = await db.offer_overrides.find_one({"offer_id": offer["id"]}, {"_id": 0})
    if not override:
        return offer
    merged = dict(offer)
    for k in _OVERRIDE_SCALAR_FIELDS:
        if override.get(k) is not None:
            merged[k] = override[k]
    if override.get("room_tiers"):
        merged["room_tiers"] = override["room_tiers"]
    return merged


@api.get("/staff/config/offers")
async def list_config_offers(staff=Depends(get_current_staff)):
    """All offers with overrides applied — used by admin config screen."""
    await _require_role(staff, ["admin"])
    result = []
    for o in OFFERS.values():
        merged = await _apply_overrides(o)
        result.append(_with_boat_times(merged))
    return {"items": result}


@api.patch("/staff/config/offers/{offer_id}")
async def update_offer_override(offer_id: str, body: OfferPriceOverride, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    if offer_id not in OFFERS:
        raise HTTPException(status_code=404, detail="Offer not found")
    payload = {k: v for k, v in body.model_dump().items() if v is not None}
    if not payload:
        return {"ok": True}
    payload["offer_id"] = offer_id
    payload["updated_at"] = now_iso()
    await db.offer_overrides.update_one(
        {"offer_id": offer_id},
        {"$set": payload},
        upsert=True,
    )
    # Mutate in-memory OFFERS dict so public site reflects immediately
    for k in _OVERRIDE_SCALAR_FIELDS:
        if payload.get(k) is not None:
            OFFERS[offer_id][k] = payload[k]
    if payload.get("room_tiers"):
        OFFERS[offer_id]["room_tiers"] = payload["room_tiers"]
    return {"ok": True}


# =================================================================
# MODULE STATS AVANCÉES — Moved to routers/stats.py during iter-24 refactor.
# The /staff/stats/advanced endpoint is now served by that router.
# =================================================================


@app.on_event("startup")
async def apply_offer_overrides_on_boot():
    try:
        async for ov in db.offer_overrides.find({}, {"_id": 0}):
            oid = ov.get("offer_id")
            if oid in OFFERS:
                for k in _OVERRIDE_SCALAR_FIELDS:
                    if ov.get(k) is not None:
                        OFFERS[oid][k] = ov[k]
                if ov.get("room_tiers"):
                    OFFERS[oid]["room_tiers"] = ov["room_tiers"]
        logging.info("Offer overrides applied on boot")
    except Exception as e:
        logging.warning("Offer override boot failed: %s", e)


@app.on_event("startup")
async def backfill_booking_poles():
    """One-shot retroactive migration: every booking without a `pole` field gets one
    derived from its offer_type. Re-runs are cheap (it filters on missing/empty pole)."""
    try:
        updated = 0
        for offer_id, pole_id in OFFER_TO_POLE.items():
            res = await db.bookings.update_many(
                {"offer_type": offer_id, "$or": [{"pole": {"$exists": False}}, {"pole": ""}, {"pole": None}]},
                {"$set": {"pole": pole_id}},
            )
            updated += res.modified_count
        if updated:
            logging.info("Backfilled `pole` field on %d existing bookings", updated)
    except Exception as e:
        logging.warning("Pole backfill failed: %s", e)




@app.on_event("startup")
async def migrate_hebergement_suite_split():
    """One-shot migration: split the legacy single "suite" tier (445k FCFA) into
    "suite_jardin" (420k FCFA) and "suite_lagune" (470k FCFA).

    Affects the `offer_overrides` document for hebergement if it still has the
    legacy 2-tier shape — replaces room_tiers with the new 3-tier list.
    Past bookings keep their historical `room_tier`/`total_amount` untouched
    (any "suite" value displays its denormalized `room_tier_name`).

    Idempotent: re-runs return immediately when the new shape is detected.
    Self-contained tier definitions so it works even after the boot-time
    `apply_offer_overrides_on_boot()` has mutated the in-memory OFFERS dict.
    """
    canonical_tiers = [
        {"id": "superieure", "name_fr": "Chambre Supérieure", "name_en": "Superior Room",
         "price": 200000, "inventory": 20},
        {"id": "suite_jardin", "name_fr": "Suite côté jardin", "name_en": "Garden-view Suite",
         "price": 420000, "inventory": 3},
        {"id": "suite_lagune", "name_fr": "Suite côté lagune", "name_en": "Lagoon-view Suite",
         "price": 470000, "inventory": 3},
    ]
    try:
        ov = await db.offer_overrides.find_one({"offer_id": "hebergement"}, {"_id": 0})
        if not ov:
            logging.info("hebergement: no override doc, defaults already 3-tier.")
            return
        tiers = ov.get("room_tiers") or []
        tier_ids = {t.get("id") for t in tiers}
        if "suite_jardin" in tier_ids and "suite_lagune" in tier_ids:
            return  # already migrated
        # Preserve any custom price/inventory already set on `superieure`.
        existing_sup = next((t for t in tiers if t.get("id") == "superieure"), None)
        new_tiers = [dict(t) for t in canonical_tiers]
        if existing_sup:
            new_tiers[0].update({
                "price": existing_sup.get("price", new_tiers[0]["price"]),
                "inventory": existing_sup.get("inventory", new_tiers[0]["inventory"]),
                "name_fr": existing_sup.get("name_fr", new_tiers[0]["name_fr"]),
                "name_en": existing_sup.get("name_en", new_tiers[0]["name_en"]),
            })
        await db.offer_overrides.update_one(
            {"offer_id": "hebergement"},
            {"$set": {
                "room_tiers": new_tiers,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }},
        )
        # Also refresh the in-memory OFFERS dict so the running process serves
        # the new tiers immediately (no second restart needed).
        OFFERS["hebergement"]["room_tiers"] = new_tiers
        logging.info("Migrated hebergement override to 3-tier (suite_jardin + suite_lagune).")
    except Exception as e:
        logging.warning("Hebergement suite split migration failed: %s", e)



@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


# ============================================================
# Scheduled notifications (J-1 reminders, J+1 review requests)
# Runs every hour. Idempotent via twilio_messages.purpose+booking_id.
# ============================================================
from apscheduler.schedulers.asyncio import AsyncIOScheduler  # noqa: E402

scheduler = AsyncIOScheduler(timezone="UTC")


async def _sent_already(purpose: str, booking_id: str) -> bool:
    found = await db.twilio_messages.find_one(
        {"purpose": purpose, "booking_id": booking_id, "status": {"$in": ["sent", "queued", "delivered", "read"]}},
    )
    return found is not None


async def _email_sent_already(purpose: str, booking_id: str) -> bool:
    found = await db.email_messages.find_one(
        {"purpose": purpose, "booking_id": booking_id, "status": "accepted"},
    )
    return found is not None


def _fmt_xof(n) -> str:
    try:
        return f"{int(n):,}".replace(",", " ") + " FCFA"
    except Exception:
        return "—"


def _fmt_date_fr(iso_date: str) -> str:
    months_fr = ["", "janvier", "février", "mars", "avril", "mai", "juin",
                 "juillet", "août", "septembre", "octobre", "novembre", "décembre"]
    try:
        y, m, d = iso_date.split("-")
        return f"{int(d)} {months_fr[int(m)]} {y}"
    except Exception:
        return iso_date or "—"


def _offer_label_fr(offer_type: str) -> str:
    return {
        "pass_day": "Day Pass",
        "sunset": "Sunset Experience",
        "brunch": "Brunch Boulay",
        "le_kaai": "Le Kaai",
        "hebergement": "Hébergement",
        "lounge": "Lounge",
        "spa_wellness": "Spa & Wellness",
        "special_event": "Événement spécial",
        "seminaire": "Séminaire résidentiel",
        "team_building": "Team Building",
        "journee_etude": "Journée d'étude",
        "dejeuner_diner_entreprise": "Déjeuner & dîner entreprise",
        "formule_personnalisee": "Formule personnalisée",
        "offres_loisirs": "Activité",
    }.get(offer_type, (offer_type or "Réservation").replace("_", " ").title())


async def _send_individual_ticket_email(booking: dict, qr_entry: dict, booker_name: str) -> None:
    """Send a per-adult ticket email to an adult passenger whose email differs
    from the booker's. The email contains ONLY that passenger's QR PNG plus a
    short personalized note. We don't dedupe via _email_sent_already because
    each recipient is a distinct address (no risk of replays on resend).
    """
    if not email_service.SENDGRID_ENABLED:
        return
    guest_email = (qr_entry.get("guest_email") or "").strip()
    if not guest_email or "@" not in guest_email:
        return
    guest_name = f"{qr_entry.get('guest_name','').strip()} {qr_entry.get('guest_surname','').strip()}".strip() or "Cher invité"
    ref = (booking.get("id", "") or "")[:8].upper()
    offer_label = _offer_label_fr(booking.get("offer_type", ""))
    date_str = _fmt_date_fr(booking.get("date", ""))
    boat = booking.get("boat_time") or ""

    # Decode the ticket PNG
    img_data = qr_entry.get("ticket_image", "")
    attachments = []
    if img_data:
        if img_data.startswith("data:"):
            img_data = img_data.split(",", 1)[1]
        try:
            import base64 as _b64
            png_bytes = _b64.b64decode(img_data)
            first = (qr_entry.get("guest_name") or "").strip().replace(" ", "_") or "passager"
            attachments.append({
                "content": png_bytes,
                "filename": f"BBR-billet-{ref}-{first}.png",
                "mime": "image/png",
                "disposition": "attachment",
            })
        except Exception:
            pass

    subject = f"Votre billet d'embarquement BBr · {ref}"
    html = f"""
    <div style="font-family:Georgia,serif;max-width:560px;margin:0 auto;padding:24px;color:#0A0A0A;">
      <h2 style="font-weight:400;font-size:22px;color:#0A0A0A;margin:0 0 6px;">Votre billet d'embarquement</h2>
      <div style="color:#B8922A;letter-spacing:.18em;text-transform:uppercase;font-size:11px;margin-bottom:18px;">Boulay Beach Resort</div>
      <p style="font-size:14px;line-height:1.6;color:#3a3a3a;">
        Bonjour <strong>{guest_name}</strong>,<br><br>
        Vous figurez sur la réservation effectuée par <strong>{booker_name}</strong> pour
        <strong>{offer_label}</strong> du <strong>{date_str}</strong>{(' — départ ' + boat) if boat else ''}.
      </p>
      <p style="font-size:14px;line-height:1.6;color:#3a3a3a;">
        Vous trouverez en pièce jointe votre <strong>billet personnel</strong> avec QR code.
        Présentez-le à l'embarquement le jour J.
      </p>
      <div style="background:#FAFAF7;border-left:3px solid #B8922A;padding:14px 18px;margin:18px 0;font-size:13px;color:#0A0A0A;">
        <strong>Référence :</strong> {ref}<br>
        <strong>Expérience :</strong> {offer_label}<br>
        <strong>Date :</strong> {date_str}
      </div>
      <p style="font-size:12px;color:#0A0A0A/55;margin-top:20px;">À très bientôt sur la lagune.<br>L'équipe BBr</p>
    </div>
    """
    plain = (
        f"Bonjour {guest_name},\n\n"
        f"Vous figurez sur la réservation BBr de {booker_name} pour {offer_label} "
        f"du {date_str}{(' — départ ' + boat) if boat else ''}.\n"
        f"Référence : {ref}\n\n"
        f"Votre billet personnel avec QR code se trouve en pièce jointe.\n\n"
        f"L'équipe Boulay Beach Resort"
    )
    await email_service.send_email(
        db, to_email=guest_email, to_name=guest_name,
        subject=subject, html=html, plain=plain,
        purpose=f"booking_paid_passenger:{qr_entry.get('qr_token','')[:8]}",
        booking_id=booking.get("id"),
        attachments=attachments,
    )


async def _send_booking_confirmation_email(booking: dict, temporary: bool = False) -> None:
    """Send the post-payment confirmation email with QR PNG attached.

    When `temporary=True` (cash pending validation), the email subject and body
    flag the booking as "EN ATTENTE DE VALIDATION" and the QR attachment is
    replaced by the cream temporary receipt (already in qr_codes[].ticket_image).
    A second confirmation email (with the final styled QR) is sent later by
    `confirm_cash_payment()` once the staff validates the cash collection.
    """
    if not email_service.SENDGRID_ENABLED:
        return
    if not booking or not booking.get("email"):
        return
    purpose = "booking_pending_cash" if temporary else "booking_paid"
    if await _email_sent_already(purpose, booking.get("id", "")):
        return
    name = (booking.get("name") or "").strip()
    if not name:
        # Real customer name lives in participants[0] (booking.name is empty)
        for p in (booking.get("participants") or []):
            if p.get("kind") == "adult":
                name = f"{p.get('name','').strip()} {p.get('surname','').strip()}".strip()
                if name:
                    break
        if not name and (booking.get("participants") or []):
            p = booking["participants"][0]
            name = f"{p.get('name','').strip()} {p.get('surname','').strip()}".strip()
    if not name:
        name = "Cher client"
    ref = (booking.get("id", "") or "")[:8].upper()
    offer_label = _offer_label_fr(booking.get("offer_type", ""))
    date_str = _fmt_date_fr(booking.get("date", ""))
    boat = booking.get("boat_time")
    amount_label = _fmt_xof(booking.get("paid_amount") or booking.get("total_amount", 0))
    ticket_url = f"{FINEO_PUBLIC_BASE_URL}/api/bookings/{booking['id']}/ticket.png?ref={booking.get('reference_token', '')}"

    # ---------- Resolve the actual offer/event image so the email banner and
    # ticket use the right photo (special events were defaulting to "sunset").
    hero_override = ""
    if booking.get("offer_type") == "special_event" and booking.get("special_event_id"):
        ev_doc = await db.special_events.find_one(
            {"id": booking["special_event_id"]}, {"_id": 0, "image_url": 1}
        )
        if ev_doc and (ev_doc.get("image_url") or "").strip():
            hero_override = ev_doc["image_url"]
    if not hero_override:
        ov = await db.offer_overrides.find_one(
            {"_id": booking.get("offer_type")}, {"_id": 0, "image_url": 1}
        )
        if ov and (ov.get("image_url") or "").strip():
            hero_override = ov["image_url"]

    # Pull configured email footer (Dashboard) — falls back to default on first run.
    try:
        from routers.site_config import fetch_email_footer_html
        custom_footer_html = await fetch_email_footer_html(db)
    except Exception:
        custom_footer_html = None

    # iter-33: include the booking_code and the public companion link when
    # there are still companion slots open. Skip for cash 'temporary' receipts
    # since the booking_code is only valid after the payment is confirmed.
    booking_code = (booking.get("booking_code") or "") if not temporary else ""
    slots_remaining = max(
        0,
        int(booking.get("companion_slots_total") or 0)
        - int(booking.get("companion_slots_used") or 0),
    ) if not temporary else 0
    companion_url = (
        f"{(os.environ.get('FINEO_PUBLIC_BASE_URL') or '').rstrip('/')}/companion/{booking_code}"
        if booking_code and slots_remaining > 0 else None
    )

    tpl = email_service.render_booking_confirmation(
        name=name, ref=ref, offer_label=offer_label, date_str=date_str,
        boat_time=boat, amount_label=amount_label, ticket_url=ticket_url,
        offer_type=booking.get("offer_type", ""),
        hero_override=hero_override,
        booking_code=booking_code or None,
        companion_url=companion_url,
        companion_slots_remaining=slots_remaining,
        custom_footer_html=custom_footer_html,
    )
    if temporary:
        # Patch the rendered template with a clearly-marked "PROVISOIRE" wrapper.
        tpl["subject"] = f"[EN ATTENTE] Reçu provisoire — {ref}"
        warning_banner = (
            '<div style="background:#FFF3CD;border-left:4px solid #B8922A;padding:14px 18px;'
            'margin:0 0 18px 0;color:#664D03;">'
            '<strong style="font-size:14px;">Paiement en espèces — validation à l\'arrivée</strong>'
            '<div style="font-size:13px;margin-top:6px;line-height:1.5;">'
            'Vous trouverez ci-joint un <strong>reçu provisoire</strong>. '
            'Votre billet définitif (avec QR code d\'embarquement) vous sera envoyé '
            'par e-mail dès que notre équipe aura encaissé le règlement à votre arrivée.'
            '</div></div>'
        )
        tpl["html"] = warning_banner + tpl["html"]
        tpl["plain"] = (
            "EN ATTENTE — Paiement en espèces\n"
            "Ceci est un reçu provisoire. Le billet définitif avec QR d'embarquement "
            "vous sera envoyé après encaissement à votre arrivée.\n\n" + tpl["plain"]
        )

    # ---- Booker email — full PDF + ALL ticket PNGs (one per adult) ----
    attachments = []
    all_qrs = booking.get("qr_codes") or []
    adult_qrs = [q for q in all_qrs if q.get("kind") == "adult"]
    try:
        # Always attach every adult's ticket so the booker has the full set,
        # not only the first one (legacy behaviour).
        for q in adult_qrs:
            img_data = q.get("ticket_image", "")
            if not img_data:
                continue
            if img_data.startswith("data:"):
                img_data = img_data.split(",", 1)[1]
            import base64 as _b64
            try:
                png_bytes = _b64.b64decode(img_data)
            except Exception:
                continue
            guest_first = (q.get("guest_name") or "").strip().replace(" ", "_") or "passager"
            fname = (f"BBR-recu-provisoire-{ref}-{guest_first}.png" if temporary
                     else f"BBR-billet-{ref}-{guest_first}.png")
            attachments.append({
                "content": png_bytes,
                "filename": fname,
                "mime": "image/png",
                "disposition": "attachment",
            })
        # Fallback: legacy bookings without per-guest images.
        if not attachments:
            png_bytes = await _build_ticket_png(booking)
            if png_bytes:
                fname = (f"BBR-recu-provisoire-{ref}.png" if temporary else f"BBR-billet-{ref}.png")
                attachments.append({
                    "content": png_bytes, "filename": fname,
                    "mime": "image/png", "disposition": "attachment",
                })
    except Exception as ex:
        logging.warning("Could not attach ticket PNGs to email: %s", ex)
    # No styled PDF for cash-pending state (only the provisional receipt).
    if not temporary:
        try:
            pdf_bytes = await _build_booking_confirmation_pdf(booking)
            if pdf_bytes:
                attachments.append({
                    "content": pdf_bytes,
                    "filename": f"BBR-reservation-{ref}.pdf",
                    "mime": "application/pdf",
                    "disposition": "attachment",
                })
        except Exception as ex:
            logging.warning("Could not attach reservation PDF to email: %s", ex)

    await email_service.send_email(
        db, to_email=booking["email"], to_name=name,
        subject=tpl["subject"], html=tpl["html"], plain=tpl["plain"],
        purpose=purpose, booking_id=booking.get("id"),
        attachments=attachments,
        attach_livret=not temporary,  # only attach livret on FINAL confirmation
    )

    # ---- Per-adult tickets — each non-booker adult who provided their own
    # email gets a personalized email with ONLY their own QR ticket. The booker
    # is skipped (they just got the full bundle above).
    if not temporary:  # cash pending → wait for final confirmation
        booker_email = (booking.get("email") or "").strip().lower()
        sent_to = {booker_email}
        for q in adult_qrs:
            guest_email = (q.get("guest_email") or "").strip().lower()
            if not guest_email or guest_email in sent_to or "@" not in guest_email:
                continue
            sent_to.add(guest_email)
            await _send_individual_ticket_email(booking, q, name)


async def _build_ticket_png(booking: dict) -> Optional[bytes]:
    """Return the QR ticket PNG bytes already stored in the booking document."""
    try:
        qrs = booking.get("qr_codes") or []
        img_data = (qrs[0] if qrs else {}).get("ticket_image", "")
        if not img_data:
            return None
        if img_data.startswith("data:"):
            img_data = img_data.split(",", 1)[1]
        from base64 import b64decode
        return b64decode(img_data)
    except Exception:
        return None


async def _build_booking_confirmation_pdf(booking: dict) -> Optional[bytes]:
    """Render a luxury-styled, single-page reservation confirmation PDF.

    The PDF doubles as a printable boarding pass: branded header, booking
    details table, embedded QR code, and a polite info footer. Returns the
    raw PDF bytes, or ``None`` if anything goes wrong.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
        )

        styles = _pdf_styles()
        GOLD, DARK, LIGHT, MUTED = styles["GOLD"], styles["DARK"], styles["LIGHT"], styles["MUTED"]

        # Resolve the customer name: prefer the booking-level "name" if set,
        # otherwise fall back to the first adult participant.
        def _customer_name(b: dict) -> str:
            raw = (b.get("name") or "").strip()
            if raw:
                return raw
            for p in (b.get("participants") or []):
                if p.get("kind") == "adult":
                    full = f"{p.get('name','').strip()} {p.get('surname','').strip()}".strip()
                    if full:
                        return full
            ps = b.get("participants") or []
            if ps:
                full = f"{ps[0].get('name','').strip()} {ps[0].get('surname','').strip()}".strip()
                if full:
                    return full
            return "Cher client"

        name = _customer_name(booking)
        ref = (booking.get("id", "") or "")[:8].upper()
        offer_label = _offer_label_fr(booking.get("offer_type", "") or booking.get("offer_name", ""))
        date_str = _fmt_date_fr(booking.get("date", ""))
        boat_time = booking.get("boat_time")
        return_boat_time = booking.get("return_boat_time")
        amount_label = _fmt_xof(booking.get("paid_amount") or booking.get("total_amount", 0))
        email = booking.get("email") or "—"
        phone = booking.get("phone") or "—"
        # Compute total guests from adults + children (real schema) or fallback
        adults = int(booking.get("adults") or 0)
        children = int(booking.get("children") or 0)
        guests = adults + children or booking.get("guests") or booking.get("nb_guests")
        guests_label = None
        if guests:
            if adults and children:
                guests_label = f"{adults} adulte(s) · {children} enfant(s)"
            else:
                guests_label = f"{int(guests)} personne(s)"
        room_label = booking.get("room_tier_name") or booking.get("room_label") or booking.get("room_id")
        nights = int(booking.get("nights") or 0)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=2 * cm, rightMargin=2 * cm,
            topMargin=2 * cm, bottomMargin=2 * cm,
            title=f"Réservation BBR — {ref}",
            author="Boulay Beach Resort",
        )

        elements = []
        # BBR logo header — same cached fetch as the receipt PDF.
        logo_bytes = _fetch_logo_bytes()
        if logo_bytes:
            try:
                from reportlab.platypus import Image as RLImage
                _logo = RLImage(io.BytesIO(logo_bytes))
                _lh = 1.8 * cm
                _logo.drawHeight = _lh
                _logo.drawWidth = _lh * _logo.imageWidth / max(_logo.imageHeight, 1)
                _logo.hAlign = "CENTER"
                elements.append(_logo)
                elements.append(Spacer(1, 0.2 * cm))
            except Exception:
                pass
        # Brand header
        elements.append(Paragraph("Boulay Beach Resort", styles["h1"]))
        elements.append(Paragraph("Confirmation de réservation", styles["sub"]))

        # Greeting band
        elements.append(Paragraph(
            f"Bonjour <b>{name}</b>,<br/>"
            "Nous avons le plaisir de vous confirmer votre réservation. "
            "Présentez ce document (ou le QR code ci-dessous) à votre arrivée.",
            styles["body"],
        ))
        elements.append(Spacer(1, 0.5 * cm))

        # Details table
        details = [
            ["Référence",   ref],
            ["Prestation",  offer_label],
            ["Date",        date_str],
        ]
        if boat_time:
            details.append(["Traversée aller", str(boat_time)])
        if return_boat_time:
            details.append(["Traversée retour", str(return_boat_time)])
        if nights:
            details.append(["Nuitées", f"{nights} nuit(s)"])
        if guests_label:
            details.append(["Personnes", guests_label])
        if room_label and booking.get("offer_type") in ("hebergement", "seminaire"):
            details.append(["Chambre", str(room_label)])
        details.append(["Montant réglé", amount_label])
        details.append(["Client", name])
        details.append(["Email", email])
        if phone and phone != "—":
            details.append(["Téléphone", phone])

        tbl = Table(details, colWidths=[5 * cm, 11 * cm])
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TEXTCOLOR", (0, 0), (0, -1), MUTED),
            ("TEXTCOLOR", (1, 0), (1, -1), DARK),
            ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(tbl)
        elements.append(Spacer(1, 0.6 * cm))

        # QR code block (embedded — same image as the email PNG attachment)
        qr_bytes = await _build_ticket_png(booking)
        if qr_bytes:
            qr_buf = io.BytesIO(qr_bytes)
            qr_img = Image(qr_buf, width=5.5 * cm, height=5.5 * cm)
            qr_label = Paragraph(
                "<b>Votre QR code d'accès</b><br/>"
                "Scannez ce code à la réception ou à l'embarquement.<br/>"
                f"<font color='#888888'>Code de référence : {ref}</font>",
                styles["body"],
            )
            qr_tbl = Table(
                [[qr_img, qr_label]],
                colWidths=[6 * cm, 10 * cm],
            )
            qr_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
                ("BOX", (0, 0), (-1, -1), 0.5, GOLD),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ]))
            elements.append(qr_tbl)
            elements.append(Spacer(1, 0.5 * cm))

        # Practical info
        elements.append(Paragraph("Informations pratiques", styles["h2"]))
        elements.append(Paragraph(
            "• Accès uniquement par bateau depuis notre quai à Abidjan, Zone 4.<br/>"
            "• Présentez-vous 15 minutes avant l'horaire de traversée indiqué.<br/>"
            "• Un justificatif d'identité peut vous être demandé à l'embarquement.<br/>"
            "• Toute annulation doit nous être communiquée au moins 48h à l'avance.",
            styles["body"],
        ))
        elements.append(Spacer(1, 0.4 * cm))
        elements.append(Paragraph(
            "Pour toute question : contact@boulaybeachresort.com",
            styles["small"],
        ))

        def _footer(canvas, doc_):
            canvas.saveState()
            canvas.setStrokeColor(GOLD)
            canvas.setLineWidth(0.5)
            canvas.line(2 * cm, 1.5 * cm, A4[0] - 2 * cm, 1.5 * cm)
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(MUTED)
            canvas.drawString(2 * cm, 1 * cm, "Boulay Beach Resort — Life Is Here")
            canvas.drawRightString(A4[0] - 2 * cm, 1 * cm, f"Réf. {ref}")
            canvas.restoreState()

        doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
        return buf.getvalue()
    except Exception as ex:
        logging.warning("Failed to build booking confirmation PDF: %s", ex)
        return None


async def _run_j_minus_1():
    if not (twilio_service.TWILIO_ENABLED or email_service.SENDGRID_ENABLED):
        return
    tomorrow = (datetime.now(timezone.utc) + timedelta(days=1)).date().isoformat()
    cursor = db.bookings.find(
        {"date": tomorrow, "paid_at": {"$exists": True}, "status": {"$ne": "cancelled"}},
        {"_id": 0},
    )
    bookings = await cursor.to_list(length=500)
    for b in bookings:
        try:
            if not await _sent_already("j_minus_1", b["id"]):
                await twilio_service.notify_j_minus_1(db, b)
        except Exception as ex:
            logging.warning("J-1 Twilio failed for %s: %s", b.get("id"), ex)
        try:
            if email_service.SENDGRID_ENABLED and b.get("email") \
                    and not await _email_sent_already("j_minus_1", b["id"]):
                tpl = email_service.render_j_minus_1(
                    name=(b.get("name") or "").strip() or "Cher client",
                    ref=(b.get("id", "") or "")[:8].upper(),
                    offer_label=_offer_label_fr(b.get("offer_type", "")),
                    date_str=_fmt_date_fr(b.get("date", "")),
                    boat_time=b.get("boat_time"),
                    offer_type=b.get("offer_type", ""),
                )
                await email_service.send_email(
                    db, to_email=b["email"], to_name=b.get("name"),
                    subject=tpl["subject"], html=tpl["html"], plain=tpl["plain"],
                    purpose="j_minus_1", booking_id=b.get("id"),
                )
        except Exception as ex:
            logging.warning("J-1 email failed for %s: %s", b.get("id"), ex)


async def _run_j_plus_1():
    if not (twilio_service.TWILIO_ENABLED or email_service.SENDGRID_ENABLED):
        return
    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).date().isoformat()
    cursor = db.bookings.find(
        {"date": yesterday, "paid_at": {"$exists": True}, "status": {"$ne": "cancelled"}},
        {"_id": 0},
    )
    bookings = await cursor.to_list(length=500)
    for b in bookings:
        try:
            if not await _sent_already("j_plus_1", b["id"]):
                await twilio_service.notify_j_plus_1(db, b)
        except Exception as ex:
            logging.warning("J+1 Twilio failed for %s: %s", b.get("id"), ex)
        try:
            if email_service.SENDGRID_ENABLED and b.get("email") \
                    and not await _email_sent_already("j_plus_1", b["id"]):
                tpl = email_service.render_j_plus_1(
                    name=(b.get("name") or "").strip() or "Cher client",
                    review_url=None,
                    offer_type=b.get("offer_type", ""),
                    offer_label=_offer_label_fr(b.get("offer_type", "")),
                )
                await email_service.send_email(
                    db, to_email=b["email"], to_name=b.get("name"),
                    subject=tpl["subject"], html=tpl["html"], plain=tpl["plain"],
                    purpose="j_plus_1", booking_id=b.get("id"),
                )
        except Exception as ex:
            logging.warning("J+1 email failed for %s: %s", b.get("id"), ex)


async def _sweep_pending_fineo_payments():
    """Background reconciliation: find pending FineoPay payments whose webhook
    never fired or whose customer never returned to /payment/fineo/result, and
    actively poll Fineo's /transactions to settle them.

    Runs every 30 seconds. Safe to fail silently — exceptions are logged and
    the next tick will retry.
    """
    if not FINEO_ENABLED:
        return
    try:
        # Pick payments that have been pending for at least 30s (let the
        # webhook a fair chance) and at most 24h (anything older is unlikely
        # to settle and would just spam the Fineo API).
        cutoff_floor = datetime.now(timezone.utc) - timedelta(hours=24)
        cutoff_ceil = datetime.now(timezone.utc) - timedelta(seconds=30)
        cursor = db.fineo_payments.find(
            {
                "status": "pending",
                "intent": "booking",  # only customer bookings need this rescue
                "created_at": {"$gte": cutoff_floor.isoformat(), "$lte": cutoff_ceil.isoformat()},
            },
            {"_id": 0, "sync_ref": 1, "booking_id": 1, "intent": 1, "amount": 1, "reference": 1},
        ).limit(50)
        pending = await cursor.to_list(length=50)
        if not pending:
            return

        # Fetch the latest transactions ONCE, then cross-match against all our
        # pending payments — way more efficient than calling per-payment.
        client_ = FineoClient()
        try:
            resp = await client_.list_transactions(page=1, limit=100)
        except Exception as ex:
            logging.warning("Fineo sweeper: list_transactions failed: %s", ex)
            return
        data = (resp or {}).get("data") or {}
        items = data.get("transactions") if isinstance(data, dict) else (data or [])
        items = items or []

        # Build an index by syncRef for O(1) lookup; fall back to amount-based
        # heuristic (last 24h, cashin, success) if Fineo doesn't echo our ref.
        from datetime import datetime as _dt, timezone as _tz
        success_set = {"success", "successful", "completed", "paid"}
        items_by_syncref = {(it.get("syncRef") or it.get("sync_ref") or ""): it for it in items}

        for pay in pending:
            sync_ref = pay["sync_ref"]
            tx = items_by_syncref.get(sync_ref)
            # Heuristic match (amount + recent + success + not already attributed)
            if not tx:
                target_amount = int(pay.get("amount", 0))
                for it in items:
                    if int(it.get("amount", 0)) != target_amount:
                        continue
                    if (it.get("direction") or "") != "cashin":
                        continue
                    if (it.get("status") or "").lower() not in success_set:
                        continue
                    existing = await db.fineo_payments.find_one(
                        {"reference": it.get("reference")}, {"_id": 0, "sync_ref": 1},
                    )
                    if existing and existing.get("sync_ref") != sync_ref:
                        continue
                    tx = it
                    break
            if not tx:
                continue
            raw_status = (tx.get("status") or "").lower()
            new_status = (
                "paid" if raw_status in success_set else
                "failed" if raw_status in ("failed", "declined") else
                "expired" if raw_status in ("expired", "cancelled") else
                "pending"
            )
            if new_status == "pending":
                continue
            fineo_ref = tx.get("reference") or tx.get("transactionReference")
            tx_amount = int(tx.get("amount", 0) or pay.get("amount", 0))
            await db.fineo_payments.update_one(
                {"sync_ref": sync_ref},
                {"$set": {
                    "status": new_status,
                    "reference": fineo_ref,
                    "settled_at": now_iso() if new_status == "paid" else None,
                    "raw_callback": tx,
                    "updated_at": now_iso(),
                    "settled_via": "background_sweep",
                }},
            )
            if new_status == "paid" and fineo_ref:
                try:
                    await _settle_payment(pay["booking_id"], pay["intent"],
                                          sync_ref, fineo_ref, tx_amount)
                    logging.info("Fineo sweeper: settled booking=%s ref=%s amount=%s",
                                 pay.get("booking_id"), fineo_ref, tx_amount)
                except Exception as ex:
                    logging.exception("Fineo sweeper: settle_payment failed for %s: %s",
                                      pay.get("booking_id"), ex)
    except Exception as ex:
        logging.exception("Fineo sweeper iteration failed: %s", ex)


@app.on_event("startup")
async def start_scheduler():
    # J-1 reminders every day at 17:00 UTC (≈18h Abidjan)
    scheduler.add_job(_run_j_minus_1, "cron", hour=17, minute=0, id="j_minus_1", replace_existing=True)
    # J+1 review request every day at 10:00 UTC (≈11h Abidjan)
    scheduler.add_job(_run_j_plus_1, "cron", hour=10, minute=0, id="j_plus_1", replace_existing=True)
    # Campaign dispatcher — every minute
    async def _run_campaigns():
        await campaign_service.run_due_campaigns(db)
    scheduler.add_job(
        _run_campaigns,
        "interval", minutes=1, id="campaigns_runner", replace_existing=True,
    )
    # FineoPay pending sweeper — actively reconciles payments whose webhook
    # never fired or whose customer never returned to our result page.
    scheduler.add_job(
        _sweep_pending_fineo_payments,
        "interval", seconds=30, id="fineo_pending_sweeper", replace_existing=True,
    )
    # iter-42: Cantine — monthly auto-renewal of meal credits on the 1st of
    # every month at 00:05 UTC (≈00:05 Abidjan, after midnight close).
    from routers.cantine import _job_monthly_renew as _cantine_renew  # noqa: WPS433
    from routers.cantine import _job_close_yesterday as _cantine_close  # noqa: WPS433
    async def _cantine_renew_job():
        try:
            await _cantine_renew(db)
        except Exception as ex:  # noqa: BLE001
            logging.exception("Cantine monthly renew failed: %s", ex)
    async def _cantine_close_job():
        try:
            await _cantine_close(db)
        except Exception as ex:  # noqa: BLE001
            logging.exception("Cantine close-yesterday failed: %s", ex)
    scheduler.add_job(_cantine_renew_job, "cron", day=1, hour=0, minute=5,
                      id="cantine_renew", replace_existing=True)
    # Close at 00:01 UTC every day → flag yesterday's still-reserved as absent
    scheduler.add_job(_cantine_close_job, "cron", hour=0, minute=1,
                      id="cantine_close", replace_existing=True)
    scheduler.start()
    logging.info("APScheduler started: J-1 @17:00 UTC, J+1 @10:00 UTC, campaigns_runner @1min, fineo_sweeper @30s")


@app.on_event("shutdown")
async def stop_scheduler():
    try:
        scheduler.shutdown(wait=False)
    except Exception:
        pass


# ============================================================
# FineoPay integration (real payment gateway, replaces mock)
# ============================================================
class FineoCheckoutBody(BaseModel):
    booking_id: str
    intent: Literal["booking", "wallet", "deposit"] = "booking"
    amount: Optional[int] = None  # required for "deposit" intent (custom amount)


class FineoClient:
    """Thin async client for FineoPay's business API.

    Docs: https://devsandbox.fineopay.com/ — 2 headers (businessCode, apiKey),
    JSON bodies. Hosted-checkout flow: POST /checkout-link returns a URL we
    redirect the customer to; the gateway calls our callbackUrl when the
    transaction settles.
    """

    def __init__(self):
        self.base_url = FINEO_BASE_URL
        self.headers = {
            "businessCode": FINEO_BUSINESS_CODE,
            "apiKey": FINEO_API_KEY,
            "Content-Type": "application/json",
        }
        self.timeout = httpx.Timeout(connect=5.0, read=15.0, write=10.0, pool=5.0)

    async def create_checkout_link(self, payload: dict) -> dict:
        async with httpx.AsyncClient(timeout=self.timeout) as cli:
            r = await cli.post(f"{self.base_url}checkout-link", json=payload, headers=self.headers)
            r.raise_for_status()
            return r.json()

    async def get_transaction(self, reference: str) -> dict:
        async with httpx.AsyncClient(timeout=self.timeout) as cli:
            r = await cli.get(f"{self.base_url}transactions/{reference}", headers=self.headers)
            r.raise_for_status()
            return r.json()

    async def list_transactions(self, page: int = 1, limit: int = 50) -> dict:
        """List recent transactions (paginated). Used as a fallback when the
        webhook hasn't fired — we scan the latest page(s) for our syncRef."""
        async with httpx.AsyncClient(timeout=self.timeout) as cli:
            r = await cli.get(
                f"{self.base_url}transactions",
                params={"page": page, "limit": limit},
                headers=self.headers,
            )
            r.raise_for_status()
            return r.json()


def _fineo_callback_url() -> str:
    return f"{FINEO_PUBLIC_BASE_URL}/api/webhooks/fineo?secret={FINEO_CALLBACK_SECRET}"


def _fineo_return_url(booking_id: str, intent: str) -> str:
    return f"{FINEO_PUBLIC_BASE_URL}/payment/fineo/result?booking_id={booking_id}&intent={intent}"


@api.post("/payments/fineo/checkout")
async def fineo_create_checkout(body: FineoCheckoutBody):
    """Create a FineoPay hosted-checkout link for a booking, a wallet
    settlement or a custom deposit. Idempotency: we reuse the same syncRef
    derived from booking_id+intent so retries hit the same Fineo transaction."""
    if not FINEO_ENABLED:
        raise HTTPException(status_code=503, detail="FineoPay non configuré sur cette instance.")

    booking = await db.bookings.find_one({"id": body.booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    # Human-friendly offer label (shown in the FineoPay title).
    _OFFER_LABELS = {
        "pass_day": "Day Pass",
        "sunset": "Sunset Experience",
        "brunch": "Brunch Boulay",
        "kaai": "Le Kaai",
        "le_kaai": "Le Kaai",
        "hebergement": "Hébergement",
        "lounge": "Lounge",
        "corporate": "Séminaire",
        "activites_events": "Activité",
        "special_event": "Événement spécial",
    }
    short_ref = (booking.get('id', '') or '')[:8].upper()
    offer_label = _OFFER_LABELS.get(booking.get("offer_type", ""), (booking.get("offer_type") or "Réservation").replace("_", " ").title())

    def _fmt_xof(n: int) -> str:
        return f"{n:,}".replace(",", " ") + " FCFA"

    if body.intent == "booking":
        total_amount = int(booking.get("total_amount", 0) or 0)
        # Allow a partial amount when paying a deposit through FineoPay (hébergement).
        # The frontend passes amount = round(total * pct/100); _settle_payment
        # auto-detects the deposit ratio and applies the proper booking status.
        if body.amount is not None and 0 < int(body.amount) < total_amount:
            amount = int(body.amount)
            ratio_pct = round(amount * 100 / total_amount)
            title = f"BBR — Acompte {ratio_pct}% · {offer_label} — {_fmt_xof(amount)}"
        else:
            amount = total_amount
            title = f"BBR — {offer_label} — {_fmt_xof(amount)}"
    elif body.intent == "wallet":
        wallet = await db.wallets.find_one({"booking_id": body.booking_id}, {"_id": 0})
        if not wallet:
            raise HTTPException(status_code=404, detail="Wallet not found")
        amount = int(wallet.get("total_charged", 0) or 0)
        title = f"BBR — Consommation sur place — {_fmt_xof(amount)}"
    elif body.intent == "deposit":
        if not body.amount or body.amount <= 0:
            raise HTTPException(status_code=400, detail="Montant d'acompte requis (> 0).")
        amount = int(body.amount)
        title = f"BBR — Acompte hébergement — {_fmt_xof(amount)}"
    else:
        raise HTTPException(status_code=400, detail="Intent invalide")

    if amount <= 0:
        raise HTTPException(status_code=400, detail="Montant à payer invalide ou déjà soldé.")

    sync_ref = f"BBR-{body.intent.upper()}-{body.booking_id}"

    # Idempotency: if we've already created an active checkout for this sync_ref
    # and it's still pending, reuse the URL — avoid double-billing.
    existing = await db.fineo_payments.find_one(
        {"sync_ref": sync_ref, "status": {"$in": ["pending", "processing"]}},
        {"_id": 0},
    )
    if existing and existing.get("checkout_url"):
        return {
            "checkout_url": existing["checkout_url"],
            "sync_ref": sync_ref,
            "amount": existing.get("amount", amount),
            "reused": True,
        }

    payload = {
        "title": title,
        "amount": amount,
        "callbackUrl": _fineo_callback_url(),
        "returnUrl": _fineo_return_url(body.booking_id, body.intent),
        "syncRef": sync_ref,
        "inputs": [
            {"label": "Référence réservation", "value": short_ref},
            {"label": "Offre", "value": offer_label},
            {"label": "Montant à régler", "value": _fmt_xof(amount)},
            {"label": "Client", "value": (booking.get("name") or "").strip() or "Client BBR"},
            {"label": "Téléphone", "value": booking.get("phone") or ""},
            {"label": "Email", "value": booking.get("email") or ""},
        ],
    }

    client_ = FineoClient()
    try:
        resp = await client_.create_checkout_link(payload)
    except httpx.HTTPStatusError as e:
        logging.exception("Fineo checkout failed (%s): %s", e.response.status_code, e.response.text)
        raise HTTPException(status_code=502, detail=f"FineoPay a refusé la demande ({e.response.status_code}).") from e
    except httpx.HTTPError as e:
        logging.exception("Fineo network error: %s", e)
        raise HTTPException(status_code=502, detail="FineoPay injoignable. Réessayez dans un instant.") from e

    if not resp.get("success") or "checkoutLink" not in (resp.get("data") or {}):
        logging.error("Unexpected Fineo response: %s", resp)
        raise HTTPException(status_code=502, detail="Réponse FineoPay inattendue.")

    checkout_url = resp["data"]["checkoutLink"]

    await db.fineo_payments.update_one(
        {"sync_ref": sync_ref},
        {
            "$set": {
                "sync_ref": sync_ref,
                "booking_id": body.booking_id,
                "intent": body.intent,
                "amount": amount,
                "checkout_url": checkout_url,
                "status": "pending",
                "updated_at": now_iso(),
            },
            "$setOnInsert": {"created_at": now_iso()},
        },
        upsert=True,
    )

    return {
        "checkout_url": checkout_url,
        "sync_ref": sync_ref,
        "amount": amount,
        "reused": False,
    }


# ----- Public payment-link endpoints (staff-issued bookings) -----
@api.get("/payment-links/{token}")
async def payment_link_summary(token: str):
    """Public endpoint — returns a booking summary for a payment-link token.
    Used by the /pay/:token public page before triggering checkout.
    """
    booking = await db.bookings.find_one(
        {"payment_link_token": token},
        {"_id": 0, "qr_codes": 0, "wallet_history": 0},
    )
    if not booking:
        raise HTTPException(status_code=404, detail="Lien de paiement introuvable")

    # Expired?
    exp = booking.get("payment_link_expires_at")
    if exp:
        try:
            if datetime.fromisoformat(exp) < datetime.now(timezone.utc):
                raise HTTPException(status_code=410, detail="Lien de paiement expiré")
        except (ValueError, TypeError):
            pass

    if booking.get("status") not in ("pending_payment", "pending"):
        raise HTTPException(status_code=410, detail=f"Réservation déjà {booking.get('status')}")

    # Offer label (catalog or special event)
    if booking["offer_type"] == "special_event":
        offer = await _resolve_special_event_offer(booking.get("special_event_id") or "", booking.get("date"))
    else:
        offer = OFFERS.get(booking["offer_type"], {"name_fr": booking["offer_type"]})

    return {
        "booking_id": booking["id"],
        "ref": booking["id"][:8].upper(),
        "status": booking.get("status"),
        "customer_name": (
            f"{booking.get('name','') or ''} {booking.get('surname','') or ''}".strip()
            or (
                f"{(booking.get('participants') or [{}])[0].get('name','')} "
                f"{(booking.get('participants') or [{}])[0].get('surname','')}".strip()
            )
        ),
        "customer_email": booking.get("email"),
        "offer_type": booking["offer_type"],
        "offer_label": offer.get("name_fr") or booking["offer_type"],
        "date": booking.get("date"),
        "checkout_date": booking.get("checkout_date"),
        "boat_time": booking.get("boat_time"),
        "adults": int(booking.get("adults", 0)),
        "children": int(booking.get("children", 0)),
        "total_amount": int(booking.get("total_amount", 0)),
        "currency": "XOF",
        "expires_at": booking.get("payment_link_expires_at"),
    }


@api.post("/payment-links/{token}/checkout")
async def payment_link_checkout(token: str):
    """Public endpoint — creates a FineoPay checkout for a payment-link booking.
    Returns the same shape as /payments/fineo/checkout so the frontend can
    redirect the customer.
    """
    booking = await db.bookings.find_one({"payment_link_token": token}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Lien de paiement introuvable")
    exp = booking.get("payment_link_expires_at")
    if exp:
        try:
            if datetime.fromisoformat(exp) < datetime.now(timezone.utc):
                raise HTTPException(status_code=410, detail="Lien de paiement expiré")
        except (ValueError, TypeError):
            pass
    if booking.get("status") not in ("pending_payment", "pending"):
        raise HTTPException(status_code=410, detail=f"Réservation déjà {booking.get('status')}")

    # Delegate to the existing FineoPay checkout creator
    return await fineo_create_checkout(FineoCheckoutBody(booking_id=booking["id"], intent="booking"))


# ===== Companion-link endpoints (iter-30) =====
# Allow the other adult passengers of a confirmed booking to register
# themselves with just their name + phone (+ optional email) and the 5-digit
# booking_code shared by the booker. Each registration creates one extra QR
# ticket on the booking and emails it to the new adult.

class CompanionRegisterBody(BaseModel):
    first_name: str = Field(min_length=1, max_length=80)
    last_name: str = Field(min_length=1, max_length=80)
    phone: str = Field(min_length=4, max_length=40)
    email: Optional[str] = Field(default=None, max_length=120)
    nationality: Optional[str] = Field(default=None, max_length=80)


def _companion_summary(booking: dict) -> dict:
    """Public shape returned by the lookup endpoint — no PII besides what the
    user already knows from sharing the booking code."""
    used = int(booking.get("companion_slots_used") or 0)
    total = int(booking.get("companion_slots_total") or 0)
    return {
        "booking_id": booking["id"],
        "ref": booking["id"][:8].upper(),
        "booking_code": booking.get("booking_code"),
        "offer_type": booking.get("offer_type"),
        "offer_label": booking.get("offer_name"),
        "date": booking.get("date"),
        "boat_time": booking.get("boat_time"),
        "booker_name": f"{((booking.get('participants') or [{}])[0].get('name','') or '').strip()} "
                       f"{((booking.get('participants') or [{}])[0].get('surname','') or '').strip()}".strip(),
        "slots_total": total,
        "slots_used": used,
        "slots_remaining": max(0, total - used),
        "status": booking.get("status"),
        "closed": used >= total or booking.get("status") in ("cancelled", "completed"),
    }


@api.get("/companion/{code}")
async def companion_lookup(code: str):
    """Public — resolve a 5-digit code to a booking summary + remaining slots."""
    code = (code or "").strip()
    if not code.isdigit() or len(code) < 4:
        raise HTTPException(status_code=400, detail="Code de réservation invalide")
    booking = await db.bookings.find_one(
        {"booking_code": code, "status": {"$in": ["confirmed", "completed", "arrived"]}},
        {"_id": 0, "qr_codes": 0, "wallet_history": 0},
    )
    if not booking:
        raise HTTPException(status_code=404, detail="Code introuvable ou réservation non confirmée")
    return _companion_summary(booking)


@api.post("/companion/{code}/register")
async def companion_register(code: str, body: CompanionRegisterBody):
    """Public — register an additional adult on a confirmed booking, generating
    a styled QR ticket and emailing it (best-effort).
    """
    code = (code or "").strip()
    booking = await db.bookings.find_one(
        {"booking_code": code, "status": {"$in": ["confirmed", "completed", "arrived"]}},
        {"_id": 0},
    )
    if not booking:
        raise HTTPException(status_code=404, detail="Code introuvable")
    used = int(booking.get("companion_slots_used") or 0)
    total = int(booking.get("companion_slots_total") or 0)
    if used >= total:
        raise HTTPException(status_code=410, detail="Tous les passagers prévus sont déjà enregistrés")

    # Build the new participant doc
    participants = list(booking.get("participants") or [])
    new_index = sum(1 for p in participants if p.get("kind", "adult") == "adult") + 1
    new_part = {
        "name": body.first_name.strip(),
        "surname": body.last_name.strip(),
        "email": (body.email or "").strip(),
        "phone": body.phone.strip(),
        "nationality": (body.nationality or (participants[0] or {}).get("nationality") or "—"),
        "kind": "adult",
    }

    # Resolve offer to build the ticket image
    if booking["offer_type"] == "special_event":
        offer = await _resolve_special_event_offer(booking.get("special_event_id") or "", booking.get("date"))
    else:
        offer = OFFERS.get(booking["offer_type"], {"name_fr": booking["offer_type"]})

    ticket_dates = list(booking.get("multi_day_dates") or []) or [booking.get("date")]
    is_passport = len(ticket_dates) > 1
    primary_date = ticket_dates[0]
    token = uuid.uuid4().hex
    label_fr = f"Adulte #{new_index}"
    compact_qr = json.dumps(
        {"type": "ticket", "token": token, "ref": booking["id"][:8].upper()},
        ensure_ascii=False, separators=(",", ":"),
    )
    token_short = token[:10].upper()
    entry = {
        "label_fr": label_fr, "label_en": f"Adult #{new_index}",
        "kind": "adult", "event_date": primary_date,
        "valid_dates": ticket_dates, "is_passport": is_passport,
        "guest_name": new_part["name"], "guest_surname": new_part["surname"],
        "guest_email": new_part["email"] or booking.get("email", ""),
        "guest_phone": new_part["phone"], "guest_nationality": new_part["nationality"],
        "qr_token": token, "qr_payload": compact_qr,
        "qr_code": make_qr(compact_qr, styled=True),
        "children_attached": 0, "companion_added_at": now_iso(),
    }
    try:
        entry["ticket_image"] = make_ticket_image(
            offer_id=booking["offer_type"], offer_name=offer.get("name_fr", booking["offer_type"]),
            date_iso=primary_date, boat_time=booking.get("boat_time", ""),
            owner_name=f"{new_part['name']} {new_part['surname']}",
            qr_payload=compact_qr, ref_code=token_short, lang="fr",
            hero_url=offer.get("image_url") or None,
            dates_list=ticket_dates if is_passport else None,
        )
    except Exception as _e:
        logger.warning("Companion ticket image build failed for %s: %s", booking["id"], _e)

    # Persist the new participant + QR + bump slot counter atomically
    await db.bookings.update_one(
        {"id": booking["id"], "companion_slots_used": used},
        {
            "$push": {"participants": new_part, "qr_codes": entry},
            "$inc": {"companion_slots_used": 1, "adults_registered": 1},
        },
    )
    # Reload for latest counts
    booking = await db.bookings.find_one({"id": booking["id"]}, {"_id": 0})

    # Best-effort email with the ticket attached
    email_sent = False
    if (body.email or "").strip():
        try:
            from services import email_service as _es  # type: ignore
            html_body = (
                f"<p>Bonjour {new_part['name']},</p>"
                f"<p>Vous êtes maintenant enregistré·e sur la réservation "
                f"<strong>{booking['id'][:8].upper()}</strong> de {booking.get('name','')} "
                f"au Boulay Beach Resort.</p>"
                f"<p>Conservez ce QR : il vous sera demandé au moment de l'embarquement.</p>"
            )
            res = await _es.send_email(
                db,
                to_email=body.email.strip(),
                subject=f"Votre billet BBR · {booking['id'][:8].upper()}",
                html=html_body,
                plain=f"Votre billet BBR · réf {booking['id'][:8].upper()} · {new_part['name']} {new_part['surname']}.",
                purpose="companion_ticket",
                booking_id=booking["id"],
                to_name=f"{new_part['name']} {new_part['surname']}",
                attachments=[{
                    "filename": f"BBR-ticket-{token_short}.png",
                    "content_b64": entry.get("ticket_image", "").split(",", 1)[-1] if entry.get("ticket_image") else "",
                    "mime": "image/png",
                }] if entry.get("ticket_image") else None,
            )
            email_sent = bool(res.get("ok"))
        except Exception as ex:
            logger.warning("Companion ticket email failed: %s", ex)

    return {
        **_companion_summary(booking),
        "registered": {
            "guest_name": new_part["name"],
            "guest_surname": new_part["surname"],
            "qr_token": token,
            "ticket_image": entry.get("ticket_image"),
            "email_sent": email_sent,
        },
    }


# ----- On-site direct payment (no booking) -----
ONSITE_LOCATIONS = {
    "quai_bbr":       "Quai BBr",
    "restaurant":     "Restaurant",
    "bar_beach_club": "Bar Beach Club",
    "reception":      "Réception",
    "lounge":         "Lounge",
    "boutique":       "Boutique",
}


class OnsiteCheckoutBody(BaseModel):
    location: str       # one of ONSITE_LOCATIONS keys
    amount: int         # FCFA, > 0
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None


@api.post("/payments/fineo/onsite-checkout")
async def fineo_onsite_checkout(body: OnsiteCheckoutBody):
    """Generate a FineoPay hosted-checkout link for an on-site (no booking)
    payment. Used by the standalone /accueil hub so guests can settle a
    consumption directly from their phone at one of the 6 BBr points of sale.

    No idempotency: each request creates a fresh syncRef (uuid) — staff
    typically enters a different amount each time.
    """
    if not FINEO_ENABLED:
        raise HTTPException(status_code=503, detail="FineoPay non configuré sur cette instance.")
    if body.location not in ONSITE_LOCATIONS:
        raise HTTPException(status_code=400, detail="Point de paiement inconnu.")
    if body.amount <= 0 or body.amount > 50_000_000:
        raise HTTPException(status_code=400, detail="Montant invalide (1 - 50 000 000 FCFA).")

    label = ONSITE_LOCATIONS[body.location]
    amount = int(body.amount)
    sync_ref = f"BBR-ONSITE-{uuid.uuid4().hex[:12].upper()}"

    def _fmt_xof(n: int) -> str:
        return f"{n:,}".replace(",", " ") + " FCFA"

    title = f"BBR — {label} — {_fmt_xof(amount)}"

    payload = {
        "title": title,
        "amount": amount,
        "callbackUrl": _fineo_callback_url(),
        "returnUrl": f"{FINEO_PUBLIC_BASE_URL}/accueil/paiement/resultat?ref={sync_ref}",
        "syncRef": sync_ref,
        "inputs": [
            {"label": "Point de paiement", "value": label},
            {"label": "Montant", "value": _fmt_xof(amount)},
            {"label": "Client", "value": (body.customer_name or "").strip() or "Client BBR"},
            {"label": "Téléphone", "value": (body.customer_phone or "").strip()},
        ],
    }

    client_ = FineoClient()
    try:
        resp = await client_.create_checkout_link(payload)
    except httpx.HTTPStatusError as e:
        logging.exception("Fineo on-site checkout failed (%s): %s", e.response.status_code, e.response.text)
        raise HTTPException(status_code=502, detail=f"FineoPay a refusé la demande ({e.response.status_code}).") from e
    except httpx.HTTPError as e:
        logging.exception("Fineo network error (on-site): %s", e)
        raise HTTPException(status_code=502, detail="FineoPay injoignable. Réessayez dans un instant.") from e

    if not resp.get("success") or "checkoutLink" not in (resp.get("data") or {}):
        logging.error("Unexpected Fineo on-site response: %s", resp)
        raise HTTPException(status_code=502, detail="Réponse FineoPay inattendue.")

    checkout_url = resp["data"]["checkoutLink"]

    await db.fineo_payments.insert_one({
        "sync_ref": sync_ref,
        "booking_id": None,
        "intent": "onsite",
        "location": body.location,
        "location_label": label,
        "customer_name": body.customer_name,
        "customer_phone": body.customer_phone,
        "amount": amount,
        "checkout_url": checkout_url,
        "status": "pending",
        "created_at": now_iso(),
        "updated_at": now_iso(),
    })

    return {"checkout_url": checkout_url, "sync_ref": sync_ref, "amount": amount}


@api.get("/payments/fineo/status/{booking_id}")
async def fineo_payment_status(booking_id: str, intent: str = "booking"):
    """Polling endpoint used by the frontend result page while awaiting the
    callback. Falls back to a live lookup on FineoPay's /transactions list if
    the callback hasn't been received yet — so we still settle the payment and
    notify the customer even when the webhook fails to fire."""
    sync_ref = f"BBR-{intent.upper()}-{booking_id}"
    pay = await db.fineo_payments.find_one({"sync_ref": sync_ref}, {"_id": 0})
    # Include booking reference_token so the frontend can deep-link the ticket PNG
    booking_lite = await db.bookings.find_one(
        {"id": booking_id}, {"_id": 0, "reference_token": 1, "qr_codes": 1, "status": 1}
    )
    booking_meta = {
        "reference_token": (booking_lite or {}).get("reference_token"),
        "qr_count": len(((booking_lite or {}).get("qr_codes") or [])),
        "booking_status": (booking_lite or {}).get("status"),
    }
    if not pay:
        return {"status": "unknown", "sync_ref": sync_ref, **booking_meta}
    # If callback already settled it: return immediately.
    if pay.get("status") in ("paid", "failed", "expired"):
        return {
            "status": pay["status"],
            "sync_ref": sync_ref,
            "amount": pay.get("amount"),
            "reference": pay.get("reference"),
            "settled_at": pay.get("settled_at"),
            **booking_meta,
        }
    # Live status from Fineo (best-effort) — active fallback when the callback
    # hasn't fired yet. Two paths:
    #   1. We already have the Fineo reference → GET /transactions/{ref}.
    #   2. We don't → scan recent /transactions and match by syncRef or, as a
    #      last resort, by exact amount within the last 6 hours.
    if FINEO_ENABLED:
        try:
            client_ = FineoClient()
            tx = None
            if pay.get("reference"):
                resp = await client_.get_transaction(pay["reference"])
                inner = (resp or {}).get("data") or {}
                # API returns {"data": {"transaction": {...}}} for GET /tx/{ref}
                tx = inner.get("transaction") if isinstance(inner, dict) else None
                tx = tx or inner if isinstance(inner, dict) else None
            else:
                resp = await client_.list_transactions(page=1, limit=50)
                data = (resp or {}).get("data") or {}
                items = data.get("transactions") if isinstance(data, dict) else (data or [])
                items = items or []
                # Strategy A: explicit syncRef (future-proof if Fineo adds it).
                for it in items:
                    if (it.get("syncRef") or it.get("sync_ref") or "") == sync_ref:
                        tx = it
                        break
                # Strategy B: match by exact amount within last 6 hours,
                # cashin direction, success status. Fragile but only used as
                # last resort when Fineo doesn't echo our syncRef.
                if not tx:
                    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
                    cutoff = _dt.now(_tz.utc) - _td(hours=6)
                    target_amount = int(pay.get("amount", 0))
                    for it in items:
                        if int(it.get("amount", 0)) != target_amount:
                            continue
                        if (it.get("direction") or "") != "cashin":
                            continue
                        if (it.get("status") or "").lower() not in ("success", "successful", "completed", "paid"):
                            continue
                        try:
                            tx_date = _dt.fromisoformat((it.get("date") or "").replace("Z", "+00:00"))
                        except Exception:
                            continue
                        if tx_date < cutoff:
                            continue
                        # Make sure THIS transaction hasn't already been bound
                        # to a different sync_ref to avoid double-attribution.
                        existing = await db.fineo_payments.find_one(
                            {"reference": it.get("reference")}, {"_id": 0, "sync_ref": 1},
                        )
                        if existing and existing.get("sync_ref") != sync_ref:
                            continue
                        tx = it
                        break
            if tx:
                raw_status = (tx.get("status") or "").lower()
                fineo_ref = tx.get("reference") or tx.get("transactionReference")
                tx_amount = int(tx.get("amount", 0) or pay.get("amount", 0))
                new_status = (
                    "paid" if raw_status in ("success", "successful", "completed", "paid") else
                    "failed" if raw_status in ("failed", "declined") else
                    "expired" if raw_status in ("expired", "cancelled") else
                    "pending"
                )
                if new_status != "pending":
                    await db.fineo_payments.update_one(
                        {"sync_ref": sync_ref},
                        {"$set": {
                            "status": new_status,
                            "reference": fineo_ref,
                            "settled_at": now_iso() if new_status == "paid" else None,
                            "raw_callback": tx,
                            "updated_at": now_iso(),
                            "settled_via": "active_poll",
                        }},
                    )
                    if new_status == "paid" and fineo_ref:
                        try:
                            await _settle_payment(pay["booking_id"], pay["intent"],
                                                  sync_ref, fineo_ref, tx_amount)
                        except Exception as ex:
                            logging.exception("Active-poll settle_payment failed: %s", ex)
                    # Re-read booking_meta in case the booking now has qr_codes
                    booking_lite = await db.bookings.find_one(
                        {"id": booking_id},
                        {"_id": 0, "reference_token": 1, "qr_codes": 1, "status": 1},
                    )
                    booking_meta = {
                        "reference_token": (booking_lite or {}).get("reference_token"),
                        "qr_count": len(((booking_lite or {}).get("qr_codes") or [])),
                        "booking_status": (booking_lite or {}).get("status"),
                    }
                    return {
                        "status": new_status, "sync_ref": sync_ref,
                        "amount": tx_amount, "reference": fineo_ref,
                        **booking_meta,
                    }
        except Exception as ex:
            logging.warning("Fineo active-poll lookup failed: %s", ex)
    return {"status": pay.get("status", "pending"), "sync_ref": sync_ref, "amount": pay.get("amount"), **booking_meta}


async def _settle_payment(booking_id: str, intent: str, sync_ref: str, fineo_ref: str, amount: int) -> None:
    """Idempotently mark a booking/wallet/deposit as paid via FineoPay and
    emit a fiscal receipt. Safe to call multiple times (callback retries)."""
    now = now_iso()
    if intent == "booking":
        # CRITICAL: re-use the same code path as the legacy /pay endpoint so that
        # QR codes, wallet QR, fiscal receipt, Twilio and email confirmation are
        # all generated. The function is idempotent on its own (status check).
        booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
        if not booking:
            logging.warning("Fineo settle: booking %s not found", booking_id)
            return
        # Pick deposit_pct from the booking itself for overnight stays
        deposit_pct = None
        if booking.get("offer_type") in ("hebergement", "seminaire") and \
                int(booking.get("total_amount", 0)) > 0 and \
                amount > 0 and amount < int(booking["total_amount"]):
            ratio = amount / int(booking["total_amount"])
            for pct in (10, 30, 70):
                if abs(ratio - pct / 100) < 0.01:
                    deposit_pct = pct
                    break
        # Only run the QR-generation path on the first webhook for this booking.
        if booking.get("status") == "pending" and not booking.get("qr_codes"):
            try:
                pay_body = PayBooking(
                    reference_token=booking.get("reference_token", ""),
                    payment_method=("deposit" if deposit_pct else "fineo"),
                    deposit_pct=deposit_pct,
                )
                await pay_booking(booking_id, pay_body)
                logging.info("Fineo settle: pay_booking generated QR codes for %s", booking_id)
            except Exception as ex:
                logging.exception("Fineo settle: pay_booking failed for %s — falling back to manual update: %s",
                                  booking_id, ex)
                # Fallback: mark as paid even without QRs so we don't leave booking stuck.
                await db.bookings.update_one(
                    {"id": booking_id, "paid_at": {"$exists": False}},
                    {"$set": {
                        "paid_at": now,
                        "payment_method": "fineo",
                        "payment_reference": fineo_ref,
                        "paid_amount": amount,
                    }},
                )
        # Always (re)attach the FineoPay reference and ensure paid_at is set.
        await db.bookings.update_one(
            {"id": booking_id},
            {"$set": {
                "payment_method": "fineo",
                "payment_reference": fineo_ref,
                "paid_amount": amount,
                **({"paid_at": now} if not booking.get("paid_at") else {}),
            }},
        )
        # Re-read the booking so downstream notifications see the latest QR codes.
        booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
        if not booking:
            return
        # Fiscal receipt — idempotent on (source, source_id)
        try:
            await _create_receipt(
                source="booking",
                source_id=booking_id,
                customer_name=booking.get("name", ""),
                customer_email=booking.get("email", ""),
                customer_phone=booking.get("phone", ""),
                lines=[{"description": f"Réservation {booking.get('offer_type', '')}",
                        "quantity": 1, "unit_price": amount, "total": amount}],
                payment_method="fineo",
                issued_by="FineoPay",
                issued_by_role="gateway",
                metadata={"booking_id": booking_id, "fineo_ref": fineo_ref, "sync_ref": sync_ref},
            )
        except Exception as ex:
            logging.exception("Booking receipt creation failed: %s", ex)
        # Twilio + email — both idempotent (_email_sent_already / _sent_already)
        try:
            qrs = booking.get("qr_codes") or []
            qr_url = (qrs[0] if qrs else {}).get("ticket_image", "")
            if qr_url and not qr_url.startswith(("http://", "https://")):
                qr_url = f"{FINEO_PUBLIC_BASE_URL}/api/bookings/{booking_id}/ticket.png?ref={booking.get('reference_token','')}"
            await twilio_service.notify_booking_paid(db, booking, qr_image_url=qr_url)
        except Exception as ex:
            logging.warning("Twilio fineo-paid notification failed: %s", ex)
        try:
            await _send_booking_confirmation_email(booking)
        except Exception as ex:
            logging.warning("SendGrid fineo-paid email failed: %s", ex)
    elif intent == "wallet":
        wallet = await db.wallets.find_one({"booking_id": booking_id}, {"_id": 0})
        if wallet and wallet.get("status") != "closed":
            await db.wallets.update_one(
                {"token": wallet["token"]},
                {"$set": {
                    "status": "closed",
                    "closed_at": now,
                    "closed_by": "FineoPay",
                    "payment_method": "fineo",
                    "paid_amount": amount,
                    "paid_at": now,
                    "fineo_reference": fineo_ref,
                }},
            )
    elif intent == "deposit":
        await db.bookings.update_one(
            {"id": booking_id},
            {
                "$inc": {"deposit_paid": amount},
                "$push": {
                    "deposit_payments": {
                        "amount": amount,
                        "method": "fineo",
                        "reference": fineo_ref,
                        "paid_at": now,
                    },
                },
            },
        )


@app.post("/api/webhooks/fineo")
async def fineo_webhook(request: Request):
    """Server-to-server callback from FineoPay. Authenticated via the
    `secret` query string (a 32+ char random token we generated and gave to
    Fineo as part of the callbackUrl). Idempotent."""
    secret = request.query_params.get("secret", "")
    if not secret or not FINEO_CALLBACK_SECRET or secret != FINEO_CALLBACK_SECRET:
        raise HTTPException(status_code=401, detail="Bad webhook secret")
    body = await request.json()
    sync_ref = body.get("syncRef") or body.get("sync_ref")
    fineo_ref = body.get("reference") or body.get("transactionReference")
    status = (body.get("status") or "").lower()
    amount = int(body.get("amount", 0) or 0)
    if not sync_ref or not fineo_ref:
        raise HTTPException(status_code=400, detail="Missing syncRef or reference")

    pay = await db.fineo_payments.find_one({"sync_ref": sync_ref}, {"_id": 0})
    if not pay:
        # Out-of-order callback or unknown payment: log & accept 200 to avoid retries storm
        logging.warning("Fineo callback for unknown sync_ref=%s ref=%s", sync_ref, fineo_ref)
        return {"received": True}

    new_status = "paid" if status in ("success", "successful", "completed", "paid") else (
        "failed" if status in ("failed", "declined") else (
            "expired" if status in ("expired", "cancelled") else "pending"
        )
    )
    await db.fineo_payments.update_one(
        {"sync_ref": sync_ref},
        {"$set": {
            "status": new_status,
            "reference": fineo_ref,
            "settled_at": now_iso() if new_status == "paid" else None,
            "raw_callback": body,
            "updated_at": now_iso(),
        }},
    )

    if new_status == "paid":
        try:
            await _settle_payment(pay["booking_id"], pay["intent"], sync_ref, fineo_ref, amount or pay.get("amount", 0))
        except Exception as ex:
            logging.exception("Fineo settle_payment failed: %s", ex)

    return {"received": True, "status": new_status}


# ============================================================
# Twilio notifications — admin endpoints
# ============================================================
class TwilioTestBody(BaseModel):
    phone: str
    body: str = "Test depuis Boulay Beach Resort \u2728"
    channel: Literal["auto", "sms", "whatsapp"] = "auto"
    trial_safe: Optional[bool] = None  # None = use env default; True = reroute to TWILIO_TEST_RECIPIENT


@api.post("/staff/notifications/test")
async def admin_twilio_test(body: TwilioTestBody, staff=Depends(get_current_staff)):
    """Send a test message to verify Twilio config. Admin-only.

    Performs a *synchronous* status re-fetch ~3s after sending so the admin
    sees the real final status (Twilio WhatsApp Sandbox returns ``queued``
    initially and then ``failed`` if the destination has not opted in)."""
    import asyncio
    await _require_role(staff, ["admin"])
    if not twilio_service.TWILIO_ENABLED:
        raise HTTPException(status_code=503, detail="Twilio non configuré.")
    res = await twilio_service.send_notification(
        db, phone=body.phone, body=body.body, purpose="admin_test",
        trial_safe=body.trial_safe,
    )

    # Re-fetch final status from Twilio so the UI shows the truth.
    cli = twilio_service.get_client()
    if cli:
        await asyncio.sleep(3)
        for channel in ("whatsapp", "sms"):
            sent = res.get(channel)
            if not sent or not sent.get("sid"):
                continue
            try:
                m = cli.messages(sent["sid"]).fetch()
                sent["status"] = m.status
                sent["error_code"] = m.error_code
                sent["error_message"] = m.error_message
                # Persist update in our log
                await db.twilio_messages.update_one(
                    {"sid": sent["sid"]},
                    {"$set": {"status": m.status, "error_code": m.error_code,
                              "error_message": m.error_message}},
                )
                # Map Twilio sandbox opt-in error → human readable explanation
                if m.error_code in (63007, 63015, 63016, 63018):
                    res["errors"].append(
                        f"{channel}:{m.error_code}: Le numéro {sent['to']} n'a pas fait l'opt-in WhatsApp Sandbox. "
                        f"Demandez au destinataire d'envoyer 'join <code>' au +14155238886 depuis WhatsApp."
                    )
                elif m.error_code == 21704:
                    res["errors"].append(
                        f"{channel}:21704: Aucun numéro émetteur configuré dans votre Messaging Service Twilio. "
                        f"Achetez un numéro Twilio SMS ou ajoutez un sender à votre Messaging Service."
                    )
                elif m.status in ("failed", "undelivered"):
                    res["errors"].append(f"{channel}:{m.error_code or '?'}:{m.status}")
            except Exception as e:
                res["errors"].append(f"{channel}:status-check-failed:{e}")
    return res


@api.get("/staff/notifications/outbound")
async def list_outbound_notifications(limit: int = 50, staff=Depends(get_current_staff)):
    """Last N Twilio outbound messages with delivery status."""
    await _require_role(staff, ["manager", "admin"])
    items = await db.twilio_messages.find({}, {"_id": 0}).sort("created_at", -1).limit(min(200, max(1, limit))).to_list(length=200)
    return {"items": items, "count": len(items), "twilio_enabled": twilio_service.TWILIO_ENABLED}


@api.get("/staff/notifications/emails")
async def list_email_notifications(limit: int = 50, purpose: Optional[str] = None,
                                   staff=Depends(get_current_staff)):
    """Last N SendGrid email sends with delivery status."""
    await _require_role(staff, ["manager", "admin"])
    q = {}
    if purpose:
        q["purpose"] = purpose
    items = await db.email_messages.find(q, {"_id": 0}).sort("created_at", -1).limit(min(200, max(1, limit))).to_list(length=200)
    return {"items": items, "count": len(items), "sendgrid_enabled": email_service.SENDGRID_ENABLED}


@api.get("/staff/notifications/new-bookings")
async def list_new_bookings(since: Optional[str] = None,
                            limit: int = 20,
                            staff=Depends(get_current_staff)):
    """In-dashboard notification feed: returns bookings created since `since`
    (ISO timestamp, exclusive). Used by the bell icon to badge unread count
    + dropdown list. No mutation; pure read.

    Returns the latest `limit` bookings sorted by created_at desc, plus
    `latest_created_at` so the client can advance its cursor on next poll.
    """
    q: dict = {"status": {"$ne": "cancelled"}}
    if since:
        q["created_at"] = {"$gt": since}
    cursor = db.bookings.find(
        q,
        {
            "_id": 0,
            "id": 1, "offer_type": 1, "label": 1, "date": 1,
            "boat_time": 1, "adults": 1, "children": 1,
            "total_amount": 1, "status": 1, "payment_method": 1,
            "created_at": 1, "pole": 1,
            "booker_name": 1, "booker_email": 1, "booker_phone": 1,
            "participants": 1,
        },
    ).sort("created_at", -1).limit(min(100, max(1, limit)))
    items: list = []
    async for b in cursor:
        # Derive a friendly customer label
        booker = b.get("booker_name") or ""
        if not booker:
            parts = b.get("participants") or []
            if parts:
                p0 = parts[0]
                booker = f"{(p0.get('name') or '').strip()} {(p0.get('surname') or '').strip()}".strip()
        items.append({
            "id": b.get("id"),
            "offer_type": b.get("offer_type"),
            "label": b.get("label") or b.get("offer_type"),
            "date": b.get("date"),
            "boat_time": b.get("boat_time"),
            "adults": b.get("adults") or 0,
            "children": b.get("children") or 0,
            "guests_total": (b.get("adults") or 0) + (b.get("children") or 0),
            "total_amount": b.get("total_amount") or 0,
            "status": b.get("status"),
            "payment_method": b.get("payment_method"),
            "pole": b.get("pole"),
            "booker": booker or "Client",
            "created_at": b.get("created_at"),
        })
    latest = items[0]["created_at"] if items else since
    return {"items": items, "count": len(items), "latest_created_at": latest}




@api.post("/staff/bookings/{booking_id}/resend-ticket-email")
async def resend_ticket_email(booking_id: str, staff=Depends(get_current_staff)):
    """Resend the booking confirmation email (with QR PNGs attached) — staff action.

    Also **regenerates** all per-adult ticket images using the *current* offer
    / event image, so support can fix tickets where the hero was missing or
    stale (e.g. event image uploaded AFTER booking was paid). Idempotent.
    """
    await _require_role(staff, ["manager", "admin", "hotesse", "receptionist"])
    if not email_service.SENDGRID_ENABLED:
        raise HTTPException(status_code=503, detail="SendGrid non configuré.")
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Réservation introuvable.")
    if not booking.get("email"):
        raise HTTPException(status_code=400, detail="Aucune adresse email sur cette réservation.")

    # ---- Regenerate ticket PNGs against current offer/event data ----
    try:
        if booking["offer_type"] == "special_event":
            offer = await _resolve_special_event_offer(
                booking.get("special_event_id") or "", booking.get("date"),
            )
        else:
            offer = OFFERS[booking["offer_type"]]
        is_paid = bool(booking.get("paid_at"))
        styled_qr = is_paid  # paid → gold QR ticket ; pending cash → cream receipt
        existing_qrs = booking.get("qr_codes") or []
        new_qrs = []
        for q in existing_qrs:
            owner_name = f"{(q.get('guest_name') or '').strip()} {(q.get('guest_surname') or '').strip()}".strip()
            d = q.get("event_date") or booking.get("date") or ""
            ref_code = (q.get("qr_token") or "")[:10].upper()
            qr_payload = (q.get("qr_payload") or "").strip() or json.dumps(
                {"type": "ticket", "token": q.get("qr_token", ""), "ref": booking_id[:8].upper()},
                ensure_ascii=False, separators=(",", ":"),
            )
            if styled_qr:
                q["ticket_image"] = make_ticket_image(
                    offer_id=booking["offer_type"],
                    offer_name=offer["name_fr"],
                    date_iso=d,
                    boat_time=booking.get("boat_time", ""),
                    owner_name=owner_name,
                    qr_payload=qr_payload,
                    ref_code=ref_code,
                    lang="fr",
                    hero_url=offer.get("image_url") or None,
                )
            else:
                q["ticket_image"] = make_cash_receipt_image(
                    offer_id=booking["offer_type"],
                    offer_name=offer["name_fr"],
                    date_iso=d,
                    boat_time=booking.get("boat_time", ""),
                    owner_name=owner_name,
                    ref_code=ref_code,
                    lang="fr",
                    hero_url=offer.get("image_url") or None,
                )
            new_qrs.append(q)
        if new_qrs:
            await db.bookings.update_one({"id": booking_id}, {"$set": {"qr_codes": new_qrs}})
            booking["qr_codes"] = new_qrs
            logging.info("Regenerated %d ticket images for booking %s", len(new_qrs), booking_id)
    except Exception as ex:
        logging.warning("Ticket regeneration failed for %s: %s", booking_id, ex)

    # Build the email payload using shared helpers (do not skip on idempotency
    # check — staff explicitly clicked "resend").
    name = (booking.get("name") or "").strip() or "Cher client"
    ref = (booking.get("id", "") or "")[:8].upper()
    offer_label = _offer_label_fr(booking.get("offer_type", ""))
    date_str = _fmt_date_fr(booking.get("date", ""))
    amount_label = _fmt_xof(booking.get("total_amount", 0))
    ticket_url = f"{FINEO_PUBLIC_BASE_URL}/api/bookings/{booking_id}/ticket.png?ref={booking.get('reference_token', '')}"
    # Resolve actual offer/event image for the email banner
    hero_override = ""
    if booking.get("offer_type") == "special_event" and booking.get("special_event_id"):
        ev_doc = await db.special_events.find_one(
            {"id": booking["special_event_id"]}, {"_id": 0, "image_url": 1}
        )
        if ev_doc and (ev_doc.get("image_url") or "").strip():
            hero_override = ev_doc["image_url"]
    if not hero_override:
        ov = await db.offer_overrides.find_one(
            {"_id": booking.get("offer_type")}, {"_id": 0, "image_url": 1}
        )
        if ov and (ov.get("image_url") or "").strip():
            hero_override = ov["image_url"]
    # iter-33: enrich resend with booking_code + companion link
    bcode = booking.get("booking_code") or ""
    slots_remaining = max(
        0,
        int(booking.get("companion_slots_total") or 0)
        - int(booking.get("companion_slots_used") or 0),
    )
    companion_url = (
        f"{(os.environ.get('FINEO_PUBLIC_BASE_URL') or '').rstrip('/')}/companion/{bcode}"
        if bcode and slots_remaining > 0 else None
    )
    tpl = email_service.render_booking_confirmation(
        name=name, ref=ref, offer_label=offer_label, date_str=date_str,
        boat_time=booking.get("boat_time"), amount_label=amount_label,
        ticket_url=ticket_url,
        offer_type=booking.get("offer_type", ""),
        hero_override=hero_override,
        booking_code=bcode or None,
        companion_url=companion_url,
        companion_slots_remaining=slots_remaining,
    )
    attachments = []
    try:
        png_bytes = await _build_ticket_png(booking)
        if png_bytes:
            attachments.append({
                "content": png_bytes,
                "filename": f"BBR-billet-{ref}.png",
                "mime": "image/png",
                "disposition": "attachment",
            })
    except Exception:
        pass
    return await email_service.send_email(
        db, to_email=booking["email"], to_name=name,
        subject=tpl["subject"], html=tpl["html"], plain=tpl["plain"],
        purpose="booking_resend", booking_id=booking_id,
        attachments=attachments,
    )


class StaffAlertBody(BaseModel):
    phone: str
    title: str
    detail: str


@api.post("/staff/notifications/staff-alert")
async def admin_send_staff_alert(body: StaffAlertBody, staff=Depends(get_current_staff)):
    """Manually broadcast a staff alert (Manager+admin)."""
    await _require_role(staff, ["manager", "admin"])
    return await twilio_service.notify_staff(db, recipient_phone=body.phone, title=body.title, detail=body.detail)


@api.post("/staff/notifications/run-j-minus-1")
async def admin_trigger_j_minus_1(staff=Depends(get_current_staff)):
    """Manually trigger the J-1 reminder job (admin)."""
    await _require_role(staff, ["admin"])
    await _run_j_minus_1()
    return {"triggered": "j_minus_1"}


@api.post("/staff/notifications/run-j-plus-1")
async def admin_trigger_j_plus_1(staff=Depends(get_current_staff)):
    """Manually trigger the J+1 review job (admin)."""
    await _require_role(staff, ["admin"])
    await _run_j_plus_1()
    return {"triggered": "j_plus_1"}


# ============================================================
# Documentation — full app guide PDF (admin/management only)
# ============================================================

@api.get("/staff/docs/guide.pdf")
async def admin_download_guide_pdf(staff=Depends(get_current_staff)):
    """Generate and stream the complete app guide PDF.

    Includes features list, RBAC roles, pôles, data models, security,
    integrations, business flows and improvement roadmap. Accessible to
    admin and management_general (read-only audience).
    """
    await _require_role(staff, ["admin", "management_general"])
    from services.guide_pdf import build_guide_pdf
    from starlette.responses import Response
    pdf_bytes = build_guide_pdf()
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="bbr-guide-{today}.pdf"'},
    )


# ============================================================
# Integrations — connectivity tests (admin only)
# ============================================================

@api.get("/staff/integrations/status")
async def admin_integrations_status(staff=Depends(get_current_staff)):
    """Return high-level config status of every 3rd-party integration."""
    await _require_role(staff, ["admin"])
    return {
        "twilio": {
            "enabled": twilio_service.TWILIO_ENABLED,
            "whatsapp_from": twilio_service.TWILIO_WHATSAPP_FROM or None,
            "sms_from": twilio_service.TWILIO_SMS_FROM or None,
            "messaging_service_sid": (twilio_service.TWILIO_MESSAGING_SERVICE_SID or "")[:8] + "…"
                if twilio_service.TWILIO_MESSAGING_SERVICE_SID else None,
            "trial_safe_default": twilio_service.TWILIO_TRIAL_SAFE_DEFAULT,
        },
        "sendgrid": {
            "enabled": email_service.SENDGRID_ENABLED,
            "from_email": email_service.SENDGRID_FROM_EMAIL or None,
            "from_name": email_service.SENDGRID_FROM_NAME or None,
            "api_key_prefix": (email_service.SENDGRID_API_KEY[:8] + "…") if email_service.SENDGRID_API_KEY else None,
        },
        "fineo": {
            "enabled": FINEO_ENABLED,
            "base_url": FINEO_BASE_URL,
            "business_code": FINEO_BUSINESS_CODE or None,
            "api_key_prefix": (FINEO_API_KEY[:14] + "…") if FINEO_API_KEY else None,
            "public_base_url": FINEO_PUBLIC_BASE_URL or None,
        },
    }


@api.post("/staff/integrations/fineo/test")
async def admin_fineo_test(staff=Depends(get_current_staff)):
    """Probe FineoPay connectivity using a tiny checkout-link request.

    The probe always uses a 100 FCFA dummy amount and a unique syncRef so
    nothing in production is affected. Returns the raw Fineo response so the
    admin can see exactly what the gateway said.
    """
    await _require_role(staff, ["admin"])
    if not FINEO_ENABLED:
        return {
            "ok": False,
            "stage": "config",
            "message": "FineoPay non configuré : vérifiez FINEO_BUSINESS_CODE / FINEO_API_KEY / FINEO_PUBLIC_BASE_URL.",
            "config": {
                "base_url": FINEO_BASE_URL,
                "business_code_set": bool(FINEO_BUSINESS_CODE),
                "api_key_set": bool(FINEO_API_KEY),
                "public_base_url_set": bool(FINEO_PUBLIC_BASE_URL),
            },
        }
    sync_ref = f"BBR-PROBE-{uuid.uuid4().hex[:10]}"
    probe_payload = {
        "title": "Test connectivité BBR ↔ FineoPay",
        "amount": 100,
        "callbackUrl": _fineo_callback_url(),
        "returnUrl": _fineo_return_url("probe", "booking"),
        "syncRef": sync_ref,
        "inputs": [
            {"label": "Type", "value": "Probe connectivité (admin)"},
        ],
    }
    client_ = FineoClient()
    try:
        async with httpx.AsyncClient(timeout=client_.timeout) as cli:
            r = await cli.post(f"{client_.base_url}checkout-link", json=probe_payload, headers=client_.headers)
    except httpx.HTTPError as e:
        return {
            "ok": False,
            "stage": "network",
            "message": f"FineoPay injoignable : {type(e).__name__}",
            "detail": str(e),
            "request": {"url": f"{client_.base_url}checkout-link", "businessCode": FINEO_BUSINESS_CODE,
                        "api_key_prefix": FINEO_API_KEY[:14] + "…"},
        }
    try:
        body_json = r.json()
    except Exception:
        body_json = {"raw": r.text[:500]}

    if r.status_code == 200 and body_json.get("success") and "checkoutLink" in (body_json.get("data") or {}):
        # Success — we got a checkout URL.
        checkout_url = body_json["data"]["checkoutLink"]
        return {
            "ok": True,
            "stage": "checkout_created",
            "message": "Connexion OK — FineoPay a généré un lien de paiement.",
            "checkout_url": checkout_url,
            "sync_ref": sync_ref,
            "amount": 100,
            "raw": body_json,
        }

    # Map known Fineo errors to actionable French messages.
    msg_fineo = (body_json or {}).get("message", "")
    actionable = msg_fineo
    if msg_fineo == "Compte marchand inexistant":
        actionable = (
            "FineoPay ne reconnaît pas votre businessCode dans cet environnement. "
            "Contactez FineoPay pour vérifier que le compte marchand est bien provisionné "
            "sur le sandbox /dev/ et que la clé API est rattachée à ce businessCode."
        )
    elif msg_fineo == "Identifiants requis mais non fourni":
        actionable = "Headers d'authentification absents (businessCode + apiKey)."
    elif "title" in msg_fineo:
        actionable = f"Champ requis manquant côté backend BBR : {msg_fineo}"

    return {
        "ok": False,
        "stage": "fineo_rejected",
        "http_status": r.status_code,
        "message": actionable,
        "fineo_message": msg_fineo,
        "raw": body_json,
        "request": {
            "url": f"{client_.base_url}checkout-link",
            "businessCode": FINEO_BUSINESS_CODE,
            "api_key_prefix": FINEO_API_KEY[:14] + "…",
            "amount": 100,
            "syncRef": sync_ref,
        },
    }


@api.post("/staff/integrations/sendgrid/test")
async def admin_sendgrid_test(body: dict, staff=Depends(get_current_staff)):
    """Send a probe email to verify SendGrid configuration (admin only).

    Renders the full luxury booking-confirmation template so the admin
    previews exactly what a real client receives (image swap, footer,
    clickable CTAs, etc.)."""
    await _require_role(staff, ["admin"])
    if not email_service.SENDGRID_ENABLED:
        return {"ok": False, "message": "SendGrid non configuré (SENDGRID_API_KEY / SENDGRID_FROM_EMAIL absents)."}
    to_email = (body.get("to_email") or "").strip()
    if not to_email:
        raise HTTPException(status_code=400, detail="to_email requis")
    offer_type = (body.get("offer_type") or "pass_day").strip()
    tpl = email_service.render_booking_confirmation(
        name=(body.get("name") or "Cher client"),
        ref="TEST" + datetime.now(timezone.utc).strftime("%H%M%S"),
        offer_label=_offer_label_fr(offer_type),
        date_str=_fmt_date_fr(datetime.now(timezone.utc).date().isoformat()),
        boat_time="10H",
        amount_label=_fmt_xof(50000),
        ticket_url=f"{FINEO_PUBLIC_BASE_URL}/",
        offer_type=offer_type,
        booking_code="12345",
        companion_url=f"{FINEO_PUBLIC_BASE_URL}/companion/12345",
        companion_slots_remaining=2,
    )
    res = await email_service.send_email(
        db, to_email=to_email, subject="[TEST] " + tpl["subject"],
        html=tpl["html"], plain=tpl["plain"],
        purpose="admin_test",
    )
    return res


# =================================================================
# EMAIL CAMPAIGNS — Import CSV/XLSX + schedule bulk send
# =================================================================
from fastapi import UploadFile, File  # noqa: E402


@api.post("/staff/campaigns/parse-list")
async def staff_campaigns_parse_list(file: UploadFile = File(...),
                                     staff=Depends(get_current_staff)):
    """Parse an uploaded CSV/XLSX file and return the list of detected
    recipients. Does not store anything."""
    await _require_role(staff, ["admin", "manager"])
    content = await file.read()
    if len(content) > 4 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Fichier trop volumineux (max 4 Mo).")
    name = (file.filename or "").lower()
    try:
        if name.endswith(".xlsx"):
            recipients = campaign_service.parse_recipients_xlsx(content)
        else:
            recipients = campaign_service.parse_recipients_csv(content)
    except Exception as ex:  # noqa: BLE001
        logging.exception("Failed to parse campaign list")
        raise HTTPException(status_code=400, detail=f"Format de fichier non reconnu : {ex}")
    return {
        "filename": file.filename,
        "total": len(recipients),
        "recipients": recipients[:500],  # cap UI preview
        "truncated": len(recipients) > 500,
    }


class CampaignCreateBody(BaseModel):
    name: Optional[str] = None
    subject: str
    title: Optional[str] = None
    body: str
    offer_type: Optional[str] = "pass_day"
    special_event_id: Optional[str] = None  # If set, used to resolve hero image + CTA defaults
    hero_image_url: Optional[str] = None     # explicit override, takes precedence
    cta_label: Optional[str] = None
    cta_url: Optional[str] = None
    recipients: List[dict]  # [{email, name?}]
    scheduled_at: Optional[str] = None  # ISO UTC; if None → draft (manual send)


async def _resolve_campaign_hero(body: "CampaignCreateBody") -> Optional[str]:
    """Resolve the hero image URL for a campaign. Order of precedence:
    1. Explicit ``hero_image_url`` provided by the admin
    2. The image of the bound special event (when ``offer_type='special_event'``)
    3. ``None`` → falls back to OFFER_HERO_IMAGES[offer_type] in the renderer

    The returned URL is guaranteed to be a publicly fetchable URL (any
    ``data:`` URL is migrated to ``/api/media/{id}`` on the fly because Gmail
    / Outlook block inline ``data:`` images in email bodies).
    """
    from routers.media import ensure_public_url
    candidate: Optional[str] = None
    if (body.hero_image_url or "").strip():
        candidate = body.hero_image_url.strip()
    elif body.offer_type == "special_event" and body.special_event_id:
        ev = await db.special_events.find_one(
            {"id": body.special_event_id}, {"_id": 0, "image_url": 1},
        )
        if ev and (ev.get("image_url") or "").strip():
            candidate = ev["image_url"].strip()
    if not candidate:
        return None
    public_url = await ensure_public_url(db, candidate)
    # Make it absolute (email clients need fully-qualified URLs)
    if public_url and public_url.startswith("/api/media/"):
        return f"{FINEO_PUBLIC_BASE_URL.rstrip('/')}{public_url}"
    return public_url


@api.post("/staff/campaigns")
async def staff_campaigns_create(body: CampaignCreateBody,
                                 staff=Depends(get_current_staff)):
    """Create a campaign. ``scheduled_at`` ISO UTC string. Without it, the
    campaign stays as ``draft`` until the admin clicks ``send-now``."""
    await _require_role(staff, ["admin", "manager"])
    # Sanitize recipients
    cleaned: list[dict] = []
    seen: set = set()
    for r in body.recipients:
        e = (r.get("email") or "").lower().strip()
        if "@" not in e or e in seen:
            continue
        seen.add(e)
        cleaned.append({"email": e, "name": (r.get("name") or "").strip() or None})
    if not cleaned:
        raise HTTPException(status_code=400, detail="Aucune adresse e-mail valide.")
    payload = body.model_dump()
    payload["hero_image_url"] = await _resolve_campaign_hero(body)
    doc = campaign_service.new_campaign_doc(
        payload=payload,
        recipients=cleaned,
        created_by=staff.get("email") or staff.get("id") or "admin",
    )
    await db.email_campaigns.insert_one(doc)
    return {"ok": True, "id": doc["id"], "total": len(cleaned), "status": doc["status"]}


@api.get("/staff/campaigns")
async def staff_campaigns_list(staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin", "manager"])
    cursor = db.email_campaigns.find({}, {"_id": 0}).sort("created_at", -1).limit(100)
    items = await cursor.to_list(length=100)
    # Strip the heavy recipients list from the index response
    for it in items:
        it.pop("recipients", None)
        it.pop("sent_emails", None)
    return {"items": items}


@api.get("/staff/campaigns/{campaign_id}")
async def staff_campaigns_get(campaign_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin", "manager"])
    c = await db.email_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Campagne introuvable.")
    return c


@api.post("/staff/campaigns/{campaign_id}/send-now")
async def staff_campaigns_send_now(campaign_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    if not email_service.SENDGRID_ENABLED:
        raise HTTPException(status_code=503, detail="SendGrid non configuré.")
    # Trigger in the background so the request returns immediately.
    import asyncio
    asyncio.create_task(campaign_service.send_campaign_now(db, campaign_id))
    return {"ok": True, "status": "dispatching"}


@api.post("/staff/campaigns/{campaign_id}/cancel")
async def staff_campaigns_cancel(campaign_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    res = await db.email_campaigns.update_one(
        {"id": campaign_id, "status": {"$in": ["scheduled", "draft"]}},
        {"$set": {"status": "cancelled"}},
    )
    if res.modified_count == 0:
        raise HTTPException(status_code=409, detail="Impossible d'annuler cette campagne.")
    return {"ok": True}


@api.delete("/staff/campaigns/{campaign_id}")
async def staff_campaigns_delete(campaign_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    res = await db.email_campaigns.delete_one({"id": campaign_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Campagne introuvable.")
    return {"ok": True}


@api.post("/staff/campaigns/preview")
async def staff_campaigns_preview(body: CampaignCreateBody,
                                  staff=Depends(get_current_staff)):
    """Return the rendered HTML so the admin can preview before scheduling."""
    await _require_role(staff, ["admin", "manager"])
    sample_name = (body.recipients[0].get("name") if body.recipients else None) or "Prénom"
    hero_override = await _resolve_campaign_hero(body)
    html, _ = campaign_service._render_campaign_html(
        title=body.title or body.subject,
        body=body.body or "",
        recipient_name=sample_name,
        cta_label=body.cta_label, cta_url=body.cta_url,
        offer_type=body.offer_type or "pass_day",
        hero_image_override=hero_override,
    )
    return {"html": html, "sample_recipient": sample_name}


# =================================================================
# RETOUR EXPÉRIENCE — Public feedback form + Staff analytics
# =================================================================

EXP_CATEGORIES = [
    "pass_day", "sunset", "brunch", "lounge", "restaurant", "hebergement", "evenement_prive", "autre",
]

EXP_RATING_FIELDS = [
    "accueil_arrivee",
    "service_amabilite",
    "restauration_boissons",
    "ambiance_cadre",
    "proprete_confort",
    "experience_globale",
]


class FeedbackBody(BaseModel):
    experience_type: str
    other_label: Optional[str] = None
    visit_date: Optional[str] = None  # ISO date string
    full_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    accueil_arrivee: int
    service_amabilite: int
    restauration_boissons: int
    ambiance_cadre: int
    proprete_confort: int
    experience_globale: int
    most_appreciated: Optional[str] = None
    improvement_suggestion: Optional[str] = None
    staff_member_mention: Optional[str] = None


@api.post("/feedback")
async def public_submit_feedback(body: FeedbackBody, request: Request):
    """Public endpoint — no auth required. Stores a customer feedback entry."""
    if body.experience_type not in EXP_CATEGORIES:
        raise HTTPException(status_code=400, detail="Type d'expérience invalide.")
    for f in EXP_RATING_FIELDS:
        v = getattr(body, f, None)
        if not isinstance(v, int) or v < 1 or v > 5:
            raise HTTPException(status_code=400, detail=f"Note {f} invalide (1-5).")
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    # Light spam guard: source IP/UA for moderation later
    client_ip = (request.headers.get("x-forwarded-for") or request.client.host or "")
    doc["source_ip"] = client_ip.split(",")[0].strip()[:64]
    doc["user_agent"] = (request.headers.get("user-agent") or "")[:300]
    await db.experience_feedback.insert_one(doc)
    return {"ok": True, "id": doc["id"]}


@api.get("/staff/feedback")
async def staff_feedback_list(limit: int = 200, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin", "manager", "management_general"])
    cursor = db.experience_feedback.find({}, {"_id": 0}).sort("created_at", -1).limit(limit)
    items = await cursor.to_list(length=limit)
    return {"items": items}


@api.get("/staff/feedback/analytics")
async def staff_feedback_analytics(staff=Depends(get_current_staff)):
    """Aggregated metrics: counts per experience type, average rating per
    criterion, NPS-style breakdown of overall rating, weekly trend & most-
    appreciated keywords.

    The "global score" (`avg_globale`) is computed **per-client then averaged**:
    for each feedback document, we first compute the mean of the 6 criteria
    (accueil, service, restauration, ambiance, propreté, expérience_globale),
    then we average those per-client means. Pooling all rows for a single
    criterion would skew results when clients fill different criteria
    differently — the current method correctly weights each client equally.
    """
    await _require_role(staff, ["admin", "manager", "management_general"])
    # Common $addFields stage: build the per-row "client_score" as the mean
    # of the 6 criteria (ignoring 0 / null with $filter), then reuse it
    # in every downstream aggregation.
    criteria_fields = [
        "$accueil_arrivee", "$service_amabilite", "$restauration_boissons",
        "$ambiance_cadre", "$proprete_confort", "$experience_globale",
    ]
    add_client_score = {
        "$addFields": {
            "_valid_scores": {
                "$filter": {
                    "input": criteria_fields,
                    "as": "v",
                    "cond": {"$and": [
                        {"$ne": ["$$v", None]},
                        {"$gt": ["$$v", 0]},
                    ]},
                },
            },
        },
    }
    add_mean = {
        "$addFields": {
            "client_score": {
                "$cond": [
                    {"$gt": [{"$size": "$_valid_scores"}, 0]},
                    {"$avg": "$_valid_scores"},
                    None,
                ],
            },
        },
    }

    pipeline = [
        add_client_score, add_mean,
        {"$group": {
            "_id": None,
            "total": {"$sum": 1},
            "avg_accueil": {"$avg": "$accueil_arrivee"},
            "avg_service": {"$avg": "$service_amabilite"},
            "avg_restau":  {"$avg": "$restauration_boissons"},
            "avg_ambiance": {"$avg": "$ambiance_cadre"},
            "avg_proprete": {"$avg": "$proprete_confort"},
            # avg_globale = per-client average of 6 criteria, then averaged
            # (was previously the simple avg of `experience_globale` only)
            "avg_globale": {"$avg": "$client_score"},
        }},
    ]
    agg = await db.experience_feedback.aggregate(pipeline).to_list(length=1)
    overall = agg[0] if agg else {}
    overall.pop("_id", None)

    by_type_cursor = db.experience_feedback.aggregate([
        add_client_score, add_mean,
        {"$group": {
            "_id": "$experience_type",
            "count": {"$sum": 1},
            # Per-client mean averaged within each experience_type
            "avg_globale": {"$avg": "$client_score"},
            "avg_accueil": {"$avg": "$accueil_arrivee"},
            "avg_service": {"$avg": "$service_amabilite"},
            "avg_restau":  {"$avg": "$restauration_boissons"},
            "avg_ambiance": {"$avg": "$ambiance_cadre"},
            "avg_proprete": {"$avg": "$proprete_confort"},
        }},
        {"$sort": {"count": -1}},
    ])
    by_type = await by_type_cursor.to_list(length=20)
    for r in by_type:
        r["type"] = r.pop("_id")

    distrib_cursor = db.experience_feedback.aggregate([
        {"$group": {"_id": "$experience_globale", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
    ])
    distrib = await distrib_cursor.to_list(length=10)
    for r in distrib:
        r["rating"] = r.pop("_id")

    # ---- Trend: feedback count + avg rating per day for last 30 days ----
    from datetime import timedelta
    horizon = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    trend_cursor = db.experience_feedback.aggregate([
        {"$match": {"created_at": {"$gte": horizon}}},
        add_client_score, add_mean,
        {"$addFields": {"day": {"$substr": ["$created_at", 0, 10]}}},
        {"$group": {
            "_id": "$day",
            "count": {"$sum": 1},
            "avg_globale": {"$avg": "$client_score"},
        }},
        {"$sort": {"_id": 1}},
    ])
    trend = await trend_cursor.to_list(length=60)
    for r in trend:
        r["day"] = r.pop("_id")

    # ---- NPS-like (promoters 5★, passives 4★, detractors ≤3★) ----
    promoters = sum(1 for d in distrib if d["rating"] == 5) and \
                sum(d["count"] for d in distrib if d["rating"] == 5)
    passives = sum(d["count"] for d in distrib if d["rating"] == 4)
    detractors = sum(d["count"] for d in distrib if d["rating"] <= 3)
    total_resp = (promoters or 0) + passives + detractors
    nps = round(((promoters - detractors) / total_resp) * 100) if total_resp else 0
    nps_block = {
        "promoters": promoters or 0,
        "passives": passives,
        "detractors": detractors,
        "score": nps,
    }

    return {
        "overall": overall,
        "by_type": by_type,
        "distribution": distrib,
        "trend": trend,
        "nps": nps_block,
    }


@api.delete("/staff/feedback/{fb_id}")
async def staff_feedback_delete(fb_id: str, staff=Depends(get_current_staff)):
    await _require_role(staff, ["admin"])
    res = await db.experience_feedback.delete_one({"id": fb_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Retour introuvable.")
    return {"ok": True}


app.include_router(api)


# ===== Self-service guest registration module =====
from routers import registrations as _registrations_mod  # noqa: E402
from routers import media as _media_mod  # noqa: E402
from routers import corporate as _corporate_mod  # noqa: E402

app.include_router(
    _media_mod.build_router(
        db=db,
        require_role=_require_role,
        get_current_staff=get_current_staff,
        public_base_url=FINEO_PUBLIC_BASE_URL,
    ),
    prefix="/api",
)

app.include_router(
    _corporate_mod.build_router(
        db=db,
        offers_catalog=OFFERS,
        require_role=_require_role,
        get_current_staff=get_current_staff,
        email_service=email_service,
        public_base_url=FINEO_PUBLIC_BASE_URL,
    ),
    prefix="/api",
)

app.include_router(
    _registrations_mod.build_router(
        db=db,
        offers_catalog=OFFERS,
        require_role=_require_role,
        get_current_staff=get_current_staff,
        email_service=email_service,
        public_base_url=FINEO_PUBLIC_BASE_URL,
    ),
    prefix="/api",
)

# Public photo gallery — albums auto-derived from OFFERS + published events.
from routers import gallery as _gallery_mod  # noqa: E402
app.include_router(
    _gallery_mod.build_gallery_router(
        db=db,
        OFFERS=OFFERS,
        get_current_staff=get_current_staff,
        require_role=_require_role,
    ),
    prefix="/api",
)

# Beach Club VIP spaces — numbered transats & balinés (unique per date).
from routers import vip_spaces as _vip_spaces_mod  # noqa: E402
app.include_router(
    _vip_spaces_mod.build_vip_spaces_router(
        db=db,
        require_role=_require_role,
        get_current_staff=get_current_staff,
    ),
    prefix="/api",
)

# Site-wide configuration (email footer + livret BBR PDF).
from routers import site_config as _site_config_mod  # noqa: E402
app.include_router(
    _site_config_mod.build_site_config_router(
        db=db,
        require_role=_require_role,
        get_current_staff=get_current_staff,
    ),
    prefix="/api",
)


# Iteration 21 — Loisirs activities CRUD, Corporate request links, Visitor enregistrement
from routers import loisirs as _loisirs_mod  # noqa: E402
from routers import corporate_requests as _corp_req_mod  # noqa: E402
from routers import visitor_registrations as _visitor_mod  # noqa: E402

app.include_router(
    _loisirs_mod.build_router(db=db, get_current_staff=get_current_staff, require_role=_require_role),
    prefix="/api",
)
app.include_router(
    _corp_req_mod.build_router(
        db=db,
        get_current_staff=get_current_staff,
        require_role=_require_role,
        # Inject the helpers required by the corporate ticket generator
        make_qr=make_qr,
        make_ticket_image=make_ticket_image,
    ),
    prefix="/api",
)
app.include_router(
    _visitor_mod.build_router(db=db, get_current_staff=get_current_staff, require_role=_require_role),
    prefix="/api",
)


# Iteration 24 — Stats router extracted from server.py monolith
from routers import stats as _stats_mod  # noqa: E402

app.include_router(
    _stats_mod.build_router(
        db=db,
        get_current_staff=get_current_staff,
        require_role=_require_role,
        OFFERS=OFFERS,
    ),
    prefix="/api",
)


# Iteration 25 — Scanner history extracted (first slice of the scanner module)
from routers import scanner_history as _scanner_hist_mod  # noqa: E402

app.include_router(
    _scanner_hist_mod.build_router(
        db=db,
        get_current_staff=get_current_staff,
        require_role=_require_role,
    ),
    prefix="/api",
)


# Iteration 28 — Configurable Reports (5 types, selectable columns, PDF/XLSX)
from routers import custom_reports as _custom_reports_mod  # noqa: E402

app.include_router(
    _custom_reports_mod.build_router(
        db=db,
        get_current_staff=get_current_staff,
        require_role=_require_role,
    ),
    prefix="/api",
)


# Iteration 34 — Central passenger registry (cross-source aggregation)
from routers import passengers as _passengers_mod  # noqa: E402

app.include_router(
    _passengers_mod.build_router(
        db=db,
        get_current_staff=get_current_staff,
        require_role=_require_role,
        poles_catalog=POLES,
    ),
    prefix="/api",
)


# Iteration 42 — Cantine du personnel (Phase A)
from routers import cantine as _cantine_mod  # noqa: E402
from routers import planning as _planning_mod  # noqa: E402

app.include_router(
    _cantine_mod.build_router(
        db=db,
        get_current_staff=get_current_staff,
        require_role=_require_role,
    ),
    prefix="/api",
)
app.include_router(
    _planning_mod.build_router(
        db=db,
        get_current_staff=get_current_staff,
        require_role=_require_role,
    ),
    prefix="/api",
)






# ----- Read-only role enforcement -----
@app.middleware("http")
async def readonly_role_middleware(request, call_next):
    """Block write operations (POST/PATCH/PUT/DELETE) for management_general
    role. Login is exempted so they can still authenticate."""
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return await call_next(request)
    # Always allow auth endpoints (login, refresh)
    path = request.url.path or ""
    if path.startswith("/api/auth/"):
        return await call_next(request)
    auth_header = request.headers.get("authorization", "")
    if auth_header.lower().startswith("bearer "):
        try:
            payload = decode_token(auth_header[7:])
            if payload.get("type") == "staff" and payload.get("role") in READONLY_ROLES:
                from starlette.responses import JSONResponse
                return JSONResponse(
                    {"detail": "Compte consultation : lecture seule, modification interdite."},
                    status_code=403,
                )
        except Exception:
            pass  # let downstream auth handlers respond with 401
    return await call_next(request)


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
