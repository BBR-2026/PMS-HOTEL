"""Self-service guest registration at the resort entrance.

A guest fills in a short form (name, email, nationality, chosen offer or
"other") and immediately gets a boarding pass (PDF + QR) emailed to them
and downloadable on the confirmation screen. Staff sees the full list on a
dedicated dashboard with CSV/XLSX/PDF export.

This is intentionally decoupled from the main /api/bookings flow because
it does NOT trigger a payment — it is a check-in / self-registration
endpoint used at the on-site /accueil hub.
"""
from __future__ import annotations

import base64
import csv
import io
import logging
import uuid
from datetime import datetime, timezone, timedelta
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel, EmailStr, Field

logger = logging.getLogger(__name__)

# --- Public-facing labels ---
COLOR_GOLD = "#B8922A"
COLOR_DARK = "#0A0A0A"


# ============== Pydantic ==============
class RegistrationCreate(BaseModel):
    first_name: str = Field(min_length=1, max_length=80)
    last_name: str = Field(min_length=1, max_length=80)
    email: EmailStr
    phone: str = Field(min_length=4, max_length=40)
    nationality: str = Field(min_length=1, max_length=80)
    offer_id: str = Field(min_length=1, max_length=60)
    offer_other: Optional[str] = Field(default=None, max_length=120)


class Registration(BaseModel):
    id: str
    first_name: str
    last_name: str
    email: str
    phone: str
    nationality: str
    offer_id: str
    offer_label: str
    offer_other: Optional[str] = None
    pass_token: str
    created_at: str


def _resolve_offer_label(offer_id: str, offer_other: Optional[str], offers_catalog: dict) -> str:
    if offer_id == "autre":
        return f"Autre — {offer_other}" if offer_other else "Autre"
    o = offers_catalog.get(offer_id)
    if o:
        return o.get("name_fr") or o.get("name") or offer_id
    return offer_id.replace("_", " ").title()


# ============== PDF generator ==============
def _make_qr_png(text: str) -> bytes:
    """Generate a clean square QR PNG."""
    import qrcode
    img = qrcode.make(text)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def build_boarding_pass_pdf(reg: dict, public_base_url: str = "") -> bytes:
    """A4 boarding pass with brand header, QR code and registration details.

    Re-uses the same visual language as the booking confirmation PDF so
    customers get a coherent experience between the two flows.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
    )
    from reportlab.lib.styles import ParagraphStyle

    gold = colors.HexColor(COLOR_GOLD)
    dark = colors.HexColor(COLOR_DARK)
    light = colors.HexColor("#FAF7F2")
    muted = colors.HexColor("#666666")

    h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=20,
                        textColor=dark, spaceAfter=4, leading=24)
    sub = ParagraphStyle("sub", fontName="Helvetica", fontSize=10,
                         textColor=gold, spaceAfter=14, leading=12,
                         fontStyle=None)
    h2 = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=12,
                        textColor=dark, spaceBefore=8, spaceAfter=6)
    body = ParagraphStyle("body", fontName="Helvetica", fontSize=10,
                          textColor=dark, leading=14)

    ref = reg["id"][:8].upper()
    full_name = f"{reg['first_name']} {reg['last_name']}".strip()
    created_dt = datetime.fromisoformat(reg["created_at"].replace("Z", "+00:00"))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=f"Pass d'embarquement BBR — {ref}",
        author="Boulay Beach Resort",
    )

    elements = []
    elements.append(Paragraph("Boulay Beach Resort", h1))
    elements.append(Paragraph("PASS D'EMBARQUEMENT", sub))
    elements.append(Paragraph(
        f"Bonjour <b>{reg['first_name']}</b>, voici votre pass d'embarquement personnel. "
        "Présentez ce document (ou le QR ci-dessous) à notre équipe d'accueil au quai. "
        "Bienvenue à bord !",
        body,
    ))
    elements.append(Spacer(1, 0.5 * cm))

    # Details table
    details = [
        ["Référence",   ref],
        ["Nom complet", full_name],
        ["Nationalité", reg["nationality"]],
        ["Email",       reg["email"]],
        ["Téléphone",   reg["phone"]],
        ["Expérience",  reg["offer_label"]],
        ["Enregistré le", created_dt.strftime("%d/%m/%Y à %Hh%M")],
    ]
    tbl = Table(details, colWidths=[5 * cm, 11 * cm])
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (0, -1), muted),
        ("TEXTCOLOR", (1, 0), (1, -1), dark),
        ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#EEE")),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 0.6 * cm))

    # QR block
    try:
        qr_payload = f"BBR-REG:{reg['id']}:{reg['pass_token']}"
        qr_bytes = _make_qr_png(qr_payload)
        qr_buf = io.BytesIO(qr_bytes)
        qr_img = Image(qr_buf, width=5.5 * cm, height=5.5 * cm)
        qr_label = Paragraph(
            "<b>Votre QR d'embarquement</b><br/>"
            "Présentez ce code à l'accueil au quai pour valider votre arrivée.<br/>"
            f"<font color='#888'>Référence : {ref}</font>",
            body,
        )
        qr_tbl = Table([[qr_img, qr_label]], colWidths=[6 * cm, 10 * cm])
        qr_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), light),
            ("BOX", (0, 0), (-1, -1), 0.5, gold),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ]))
        elements.append(qr_tbl)
    except Exception as ex:
        logger.warning("QR generation failed for registration %s: %s", reg["id"], ex)

    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph("Informations pratiques", h2))
    elements.append(Paragraph(
        "• Présentez-vous au quai 15 minutes avant l'horaire prévu.<br/>"
        "• Un justificatif d'identité peut vous être demandé à l'embarquement.<br/>"
        "• En cas de question : contact@boulaybeachresort.com",
        body,
    ))

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setStrokeColor(gold)
        canvas.setLineWidth(0.5)
        canvas.line(2 * cm, 1.5 * cm, A4[0] - 2 * cm, 1.5 * cm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(muted)
        canvas.drawString(2 * cm, 1 * cm, "Boulay Beach Resort — Life Is Here")
        canvas.drawRightString(A4[0] - 2 * cm, 1 * cm, f"Pass {ref}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


# ============== Email ==============
async def send_registration_email(email_service, to: str, reg: dict, pdf_bytes: bytes) -> None:
    """Send the boarding pass PDF as an attachment. Best-effort."""
    try:
        from services.email_service import _formal_greeting
        greet = _formal_greeting(f"{reg['first_name']} {reg['last_name']}")
    except Exception:
        greet = reg.get("first_name") or "cher client"

    ref = reg["id"][:8].upper()
    subject = f"Votre pass d'embarquement BBR — {ref}"
    html = f"""
    <div style="font-family:Helvetica,Arial,sans-serif;max-width:560px;margin:auto;color:#0A0A0A;">
      <h2 style="color:#0A0A0A;font-weight:600;margin:0 0 4px;">Boulay Beach Resort</h2>
      <div style="color:#B8922A;text-transform:uppercase;letter-spacing:3px;font-size:10px;margin-bottom:18px;">
        Pass d'embarquement
      </div>
      <p>Bonjour {greet},</p>
      <p>Merci pour votre enregistrement. Votre pass d'embarquement est joint à ce
      message au format PDF (avec QR code). Présentez-le à notre équipe d'accueil au quai.</p>
      <p><strong>Référence :</strong> {ref}<br/>
         <strong>Expérience :</strong> {reg['offer_label']}</p>
      <p style="color:#666;font-size:13px;margin-top:24px;">À très vite,<br/>L'équipe BBr</p>
    </div>
    """
    plain = (
        f"Bonjour {greet},\n\n"
        f"Merci pour votre enregistrement. Votre pass d'embarquement (référence {ref}) "
        f"est joint en pièce jointe au format PDF.\n\nÀ très vite,\nL'équipe BBr"
    )
    try:
        await email_service.send_email(
            to=to,
            subject=subject,
            html=html,
            plain=plain,
            attachments=[{
                "content": pdf_bytes,
                "filename": f"BBR-pass-{ref}.pdf",
                "mime": "application/pdf",
                "disposition": "attachment",
            }],
            tag="registration_pass",
        )
    except Exception as ex:
        logger.warning("Failed to email registration pass to %s: %s", to, ex)


# ============== Router factory ==============
def build_router(*, db, offers_catalog: dict, require_role, get_current_staff,
                 email_service, public_base_url: str = "") -> APIRouter:
    """Return an APIRouter bound to the host application's deps.

    Splitting it into a factory keeps server.py decoupled from this module
    and removes the need for circular imports.
    """
    router = APIRouter()

    @router.get("/registration-offers")
    async def list_public_offers():
        """Lightweight offer list for the registration dropdown. Adds the
        synthetic "autre" entry handled by the frontend."""
        out = []
        for oid, o in offers_catalog.items():
            out.append({"id": oid, "label": o.get("name_fr") or o.get("name") or oid})
        out.append({"id": "autre", "label": "Autre (préciser)"})
        return {"offers": out}

    @router.post("/registrations")
    async def create_registration(body: RegistrationCreate):
        """Public endpoint — anyone with the form can register themselves."""
        if body.offer_id != "autre" and body.offer_id not in offers_catalog:
            raise HTTPException(status_code=400, detail="Offre inconnue")
        if body.offer_id == "autre" and not (body.offer_other or "").strip():
            raise HTTPException(status_code=400, detail="Précisez l'offre dans le champ 'Autre'.")

        offer_label = _resolve_offer_label(body.offer_id, body.offer_other, offers_catalog)
        now = datetime.now(timezone.utc).isoformat()
        reg = {
            "id": str(uuid.uuid4()),
            "first_name": body.first_name.strip(),
            "last_name": body.last_name.strip(),
            "email": body.email,
            "phone": body.phone.strip(),
            "nationality": body.nationality.strip(),
            "offer_id": body.offer_id,
            "offer_label": offer_label,
            "offer_other": (body.offer_other or "").strip() or None,
            "pass_token": uuid.uuid4().hex,
            "created_at": now,
        }
        await db.registrations.insert_one({**reg})

        # Build the PDF & email it (best-effort, non-blocking failures)
        try:
            pdf_bytes = build_boarding_pass_pdf(reg, public_base_url=public_base_url)
            await send_registration_email(email_service, reg["email"], reg, pdf_bytes)
        except Exception as ex:
            logger.warning("Boarding pass build/email failed for %s: %s", reg["id"], ex)

        return {
            "id": reg["id"],
            "pass_token": reg["pass_token"],
            "first_name": reg["first_name"],
            "offer_label": reg["offer_label"],
            "ref": reg["id"][:8].upper(),
        }

    @router.get("/registrations/{reg_id}/pass.pdf")
    async def download_pass(reg_id: str, token: str = Query(...)):
        """Public download — token-protected (one-time-known by the customer)."""
        reg = await db.registrations.find_one({"id": reg_id}, {"_id": 0})
        if not reg or reg.get("pass_token") != token:
            raise HTTPException(status_code=404, detail="Pass introuvable")
        pdf_bytes = build_boarding_pass_pdf(reg, public_base_url=public_base_url)
        ref = reg["id"][:8].upper()
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="BBR-pass-{ref}.pdf"'},
        )

    # ============== Staff endpoints ==============
    @router.get("/staff/registrations")
    async def staff_list_registrations(
        page: int = Query(1, ge=1),
        limit: int = Query(50, ge=1, le=500),
        q: Optional[str] = None,
        period: Optional[str] = Query(None, regex="^(day|week|month|all)$"),
        offer_id: Optional[str] = None,
        staff=Depends(get_current_staff),
    ):
        await require_role(staff, ["admin", "manager", "manager_pole", "management_general", "hotesse"])
        filt = {}
        if q:
            rx = {"$regex": q, "$options": "i"}
            filt["$or"] = [
                {"first_name": rx}, {"last_name": rx}, {"email": rx},
                {"phone": rx}, {"nationality": rx}, {"offer_label": rx},
            ]
        if offer_id:
            filt["offer_id"] = offer_id
        if period and period != "all":
            now = datetime.now(timezone.utc)
            if period == "day":
                since = now.replace(hour=0, minute=0, second=0, microsecond=0)
            elif period == "week":
                # Start of current week (Monday 00:00 UTC)
                since = (now - timedelta(days=now.weekday())).replace(
                    hour=0, minute=0, second=0, microsecond=0,
                )
            else:  # month
                since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            filt["created_at"] = {"$gte": since.isoformat()}
        total = await db.registrations.count_documents(filt)
        cursor = (
            db.registrations.find(filt, {"_id": 0, "pass_token": 0})
            .sort("created_at", -1)
            .skip((page - 1) * limit)
            .limit(limit)
        )
        items = await cursor.to_list(length=limit)
        return {"total": total, "page": page, "limit": limit, "items": items}

    @router.delete("/staff/registrations/{reg_id}")
    async def staff_delete_registration(reg_id: str, staff=Depends(get_current_staff)):
        await require_role(staff, ["admin"])
        res = await db.registrations.delete_one({"id": reg_id})
        if res.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Enregistrement introuvable")
        return {"ok": True}

    def _build_filter(q: Optional[str], period: Optional[str], offer_id: Optional[str]) -> dict:
        filt: dict = {}
        if q:
            rx = {"$regex": q, "$options": "i"}
            filt["$or"] = [
                {"first_name": rx}, {"last_name": rx}, {"email": rx},
                {"phone": rx}, {"nationality": rx}, {"offer_label": rx},
            ]
        if offer_id:
            filt["offer_id"] = offer_id
        if period and period != "all":
            now = datetime.now(timezone.utc)
            if period == "day":
                since = now.replace(hour=0, minute=0, second=0, microsecond=0)
            elif period == "week":
                since = (now - timedelta(days=now.weekday())).replace(
                    hour=0, minute=0, second=0, microsecond=0,
                )
            else:
                since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            filt["created_at"] = {"$gte": since.isoformat()}
        return filt

    @router.get("/staff/registrations/export.csv")
    async def staff_export_csv(
        q: Optional[str] = None,
        period: Optional[str] = Query(None, regex="^(day|week|month|all)$"),
        offer_id: Optional[str] = None,
        staff=Depends(get_current_staff),
    ):
        await require_role(staff, ["admin", "manager", "manager_pole", "management_general"])
        filt = _build_filter(q, period, offer_id)
        items = await db.registrations.find(filt, {"_id": 0, "pass_token": 0}).sort("created_at", -1).to_list(length=10000)
        buf = io.StringIO()
        w = csv.writer(buf, delimiter=";")
        w.writerow(["Date", "Référence", "Nom", "Prénom", "Email", "Téléphone", "Nationalité", "Offre"])
        for r in items:
            ref = (r.get("id", "") or "")[:8].upper()
            dt = (r.get("created_at", "") or "")[:19].replace("T", " ")
            w.writerow([dt, ref, r.get("last_name", ""), r.get("first_name", ""),
                        r.get("email", ""), r.get("phone", ""), r.get("nationality", ""),
                        r.get("offer_label", "")])
        # Excel-compatible UTF-8 with BOM
        content = ("\ufeff" + buf.getvalue()).encode("utf-8")
        return Response(
            content=content,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="bbr-enregistrements.csv"'},
        )

    @router.get("/staff/registrations/export.xlsx")
    async def staff_export_xlsx(
        q: Optional[str] = None,
        period: Optional[str] = Query(None, regex="^(day|week|month|all)$"),
        offer_id: Optional[str] = None,
        staff=Depends(get_current_staff),
    ):
        await require_role(staff, ["admin", "manager", "manager_pole", "management_general"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        filt = _build_filter(q, period, offer_id)
        items = await db.registrations.find(filt, {"_id": 0, "pass_token": 0}).sort("created_at", -1).to_list(length=10000)
        wb = Workbook()
        ws = wb.active
        ws.title = "Enregistrements"
        headers = ["Date", "Référence", "Nom", "Prénom", "Email", "Téléphone", "Nationalité", "Offre"]
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor=COLOR_GOLD.lstrip("#"))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
        for r in items:
            ref = (r.get("id", "") or "")[:8].upper()
            dt = (r.get("created_at", "") or "")[:19].replace("T", " ")
            ws.append([dt, ref, r.get("last_name", ""), r.get("first_name", ""),
                       r.get("email", ""), r.get("phone", ""), r.get("nationality", ""),
                       r.get("offer_label", "")])
        # Column widths
        widths = [18, 12, 18, 18, 28, 16, 18, 28]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="bbr-enregistrements.xlsx"'},
        )

    @router.get("/staff/registrations/export.pdf")
    async def staff_export_pdf(
        q: Optional[str] = None,
        period: Optional[str] = Query(None, regex="^(day|week|month|all)$"),
        offer_id: Optional[str] = None,
        staff=Depends(get_current_staff),
    ):
        await require_role(staff, ["admin", "manager", "manager_pole", "management_general"])
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        )
        from reportlab.lib.styles import ParagraphStyle

        filt = _build_filter(q, period, offer_id)
        items = await db.registrations.find(filt, {"_id": 0, "pass_token": 0}).sort("created_at", -1).to_list(length=10000)

        gold = colors.HexColor(COLOR_GOLD)
        dark = colors.HexColor(COLOR_DARK)
        muted = colors.HexColor("#666666")

        h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=18,
                            textColor=dark, spaceAfter=4, leading=22)
        sub = ParagraphStyle("sub", fontName="Helvetica", fontSize=10,
                             textColor=gold, spaceAfter=14)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
            topMargin=1.5 * cm, bottomMargin=1.5 * cm,
            title="BBR — Enregistrements",
        )

        els = [
            Paragraph("Boulay Beach Resort", h1),
            Paragraph(f"Liste des enregistrements ({len(items)})", sub),
        ]

        rows = [["Date", "Réf.", "Nom", "Prénom", "Email", "Téléphone", "Nationalité", "Offre"]]
        for r in items:
            ref = (r.get("id", "") or "")[:8].upper()
            dt = (r.get("created_at", "") or "")[:16].replace("T", " ")
            rows.append([
                dt, ref, r.get("last_name", "")[:18], r.get("first_name", "")[:18],
                r.get("email", "")[:28], r.get("phone", "")[:18],
                r.get("nationality", "")[:18], r.get("offer_label", "")[:28],
            ])
        tbl = Table(rows, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), gold),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAF7F2")]),
            ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#EEE")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        els.append(tbl)
        els.append(Spacer(1, 0.5 * cm))
        els.append(Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %Hh%M')}",
            ParagraphStyle("foot", fontName="Helvetica", fontSize=8, textColor=muted),
        ))

        doc.build(els)
        return Response(
            content=buf.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="bbr-enregistrements.pdf"'},
        )

    return router
