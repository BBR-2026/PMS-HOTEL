"""Planning des Équipes — iter-46.

Phase 1 :
    - 14 départements officiels BBr seedés automatiquement
    - Gestion de l'équipe par département (CRUD employés)
    - Planning hebdomadaire (grille L→D × employés × T/R)
    - Click cellule = bascule Travail ↔ Repos
    - Validation hebdomadaire avec horodatage + auteur
    - Vue RH consolidée (compteurs + liste départements)
    - Exports Excel + PDF par département/semaine
    - Recherche/filtrage RH par département & employé

Phase 2 (backlog) : horaires (08-17h, 14-22h), rotations, congés/absences,
remplacements, validation RH bloquante, stats présence, vue mensuelle.
"""
from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone, date, timedelta
from typing import Optional, Literal, List

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from fastapi.responses import StreamingResponse


PLANNING_DEPARTMENTS = [
    "Ressources Humaines", "Logistique et Moyens Généraux", "Achats",
    "Technique", "Hébergement", "Food & Beverage", "Beach Club", "Cuisine",
    "Finance", "Informatique", "Guest Relationship", "Commercial",
    "Marketing et com", "Sécurité",
]

# Roles
HR_ROLES = ["admin", "management_general", "directeur", "rh"]
MANAGER_ROLES = HR_ROLES + ["chef_dept"]   # chef_dept can only edit own dept
DAYS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
DAY_LABELS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]


# ── Models ──────────────────────────────────────────────────────────────────
class DeptUpsert(BaseModel):
    name: str = Field(min_length=1, max_length=80)
    manager_staff_id: Optional[str] = None


class EmployeeUpsert(BaseModel):
    last_name: str = Field(min_length=1, max_length=60)
    first_name: str = Field(min_length=1, max_length=60)
    position: str = Field(min_length=1, max_length=120)


class CellToggle(BaseModel):
    dept_id: str
    week_iso: str = Field(pattern=r"^\d{4}-W\d{2}$",
                          description="ISO week, e.g. 2026-W25")
    employee_id: str
    day: Literal["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
    status: Literal["T", "R"]


# ── Helpers ─────────────────────────────────────────────────────────────────
def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _week_iso(d: date) -> str:
    y, w, _ = d.isocalendar()
    return f"{y}-W{w:02d}"


def _monday_of_week(week_iso: str) -> date:
    y, w = week_iso.split("-W")
    return date.fromisocalendar(int(y), int(w), 1)


def _week_dates(week_iso: str) -> List[date]:
    monday = _monday_of_week(week_iso)
    return [monday + timedelta(days=i) for i in range(7)]


async def _seed_departments(db) -> None:
    n = await db.planning_departments.count_documents({})
    if n == 0:
        docs = [
            {"id": str(uuid.uuid4()), "name": name, "sort_order": i,
             "manager_staff_id": None, "created_at": _now_iso()}
            for i, name in enumerate(PLANNING_DEPARTMENTS)
        ]
        await db.planning_departments.insert_many(docs)


def _user_can_manage(staff: dict, dept: dict) -> bool:
    role = staff.get("role")
    if role in HR_ROLES:
        return True
    if role == "chef_dept" and staff.get("dept_id") == dept["id"]:
        return True
    if dept.get("manager_staff_id") == staff.get("id"):
        return True
    return False


# ── Router ──────────────────────────────────────────────────────────────────
def build_router(db, get_current_staff, require_role, hash_password=None) -> APIRouter:
    r = APIRouter()

    @r.get("/staff/planning/departments")
    async def list_departments(staff=Depends(get_current_staff)):
        await require_role(staff, MANAGER_ROLES)
        await _seed_departments(db)
        items = await db.planning_departments.find({}, {"_id": 0}).sort("sort_order", 1).to_list(length=200)
        # If the user is a chef_dept, surface only their dept (RH/admin see all)
        if staff.get("role") == "chef_dept" and staff.get("dept_id"):
            items = [d for d in items if d["id"] == staff["dept_id"]]
        # Decorate with employee count + planning status of current week
        cur_week = _week_iso(datetime.now(timezone.utc).date())
        for d in items:
            d["employee_count"] = await db.planning_employees.count_documents({"dept_id": d["id"]})
            wk = await db.planning_weeks.find_one(
                {"dept_id": d["id"], "week_iso": cur_week},
                {"_id": 0, "validated_at": 1},
            )
            d["current_week_validated"] = bool(wk and wk.get("validated_at"))
        return {"items": items, "current_week": cur_week}

    @r.put("/staff/planning/departments/{dept_id}")
    async def update_department(dept_id: str, payload: DeptUpsert,
                                staff=Depends(get_current_staff)):
        await require_role(staff, HR_ROLES)
        upd = {k: v for k, v in payload.dict().items() if v is not None}
        upd["updated_at"] = _now_iso()
        await db.planning_departments.update_one({"id": dept_id}, {"$set": upd})
        return {"ok": True}

    # ── Employees CRUD ──
    @r.get("/staff/planning/departments/{dept_id}/employees")
    async def list_employees(dept_id: str, staff=Depends(get_current_staff)):
        await require_role(staff, MANAGER_ROLES)
        dept = await db.planning_departments.find_one({"id": dept_id}, {"_id": 0})
        if not dept:
            raise HTTPException(404, "Département introuvable.")
        if not _user_can_manage(staff, dept):
            raise HTTPException(403, "Accès refusé à ce département.")
        items = await db.planning_employees.find(
            {"dept_id": dept_id}, {"_id": 0},
        ).sort([("last_name", 1), ("first_name", 1)]).to_list(length=500)
        return {"items": items, "department": dept}

    @r.post("/staff/planning/departments/{dept_id}/employees")
    async def create_employee(dept_id: str, payload: EmployeeUpsert,
                              staff=Depends(get_current_staff)):
        await require_role(staff, MANAGER_ROLES)
        dept = await db.planning_departments.find_one({"id": dept_id}, {"_id": 0})
        if not dept:
            raise HTTPException(404, "Département introuvable.")
        if not _user_can_manage(staff, dept):
            raise HTTPException(403, "Accès refusé.")
        doc = {
            "id": str(uuid.uuid4()),
            "dept_id": dept_id,
            "last_name": payload.last_name.strip(),
            "first_name": payload.first_name.strip(),
            "position": payload.position.strip(),
            "created_at": _now_iso(),
        }
        await db.planning_employees.insert_one(doc)
        # Pop the _id added by insert_one (not JSON-serializable)
        doc.pop("_id", None)
        return doc

    @r.patch("/staff/planning/employees/{emp_id}")
    async def update_employee(emp_id: str, payload: EmployeeUpsert,
                              staff=Depends(get_current_staff)):
        await require_role(staff, MANAGER_ROLES)
        emp = await db.planning_employees.find_one({"id": emp_id}, {"_id": 0})
        if not emp:
            raise HTTPException(404, "Employé introuvable.")
        dept = await db.planning_departments.find_one({"id": emp["dept_id"]}, {"_id": 0})
        if not _user_can_manage(staff, dept):
            raise HTTPException(403, "Accès refusé.")
        upd = payload.dict()
        upd["updated_at"] = _now_iso()
        await db.planning_employees.update_one({"id": emp_id}, {"$set": upd})
        return {"ok": True}

    @r.delete("/staff/planning/employees/{emp_id}")
    async def delete_employee(emp_id: str, staff=Depends(get_current_staff)):
        await require_role(staff, MANAGER_ROLES)
        emp = await db.planning_employees.find_one({"id": emp_id}, {"_id": 0})
        if not emp:
            raise HTTPException(404, "Employé introuvable.")
        dept = await db.planning_departments.find_one({"id": emp["dept_id"]}, {"_id": 0})
        if not _user_can_manage(staff, dept):
            raise HTTPException(403, "Accès refusé.")
        await db.planning_employees.delete_one({"id": emp_id})
        # Cleanup cells referencing this employee in any week
        await db.planning_weeks.update_many(
            {"dept_id": emp["dept_id"]},
            {"$unset": {f"cells.{emp_id}": ""}},
        )
        return {"ok": True}

    # ── Weekly planning ──
    @r.get("/staff/planning/week")
    async def get_week(dept_id: str, week_iso: str = Query(..., pattern=r"^\d{4}-W\d{2}$"),
                       staff=Depends(get_current_staff)):
        await require_role(staff, MANAGER_ROLES)
        dept = await db.planning_departments.find_one({"id": dept_id}, {"_id": 0})
        if not dept:
            raise HTTPException(404, "Département introuvable.")
        if not _user_can_manage(staff, dept):
            raise HTTPException(403, "Accès refusé.")
        wk = await db.planning_weeks.find_one(
            {"dept_id": dept_id, "week_iso": week_iso}, {"_id": 0},
        ) or {"dept_id": dept_id, "week_iso": week_iso, "cells": {},
              "validated_at": None, "validated_by": None}
        employees = await db.planning_employees.find(
            {"dept_id": dept_id}, {"_id": 0},
        ).sort([("last_name", 1), ("first_name", 1)]).to_list(length=500)
        return {
            "department": dept,
            "week_iso": week_iso,
            "week_dates": [d.isoformat() for d in _week_dates(week_iso)],
            "employees": employees,
            "cells": wk.get("cells", {}),
            "validated_at": wk.get("validated_at"),
            "validated_by": wk.get("validated_by"),
            "can_edit": _user_can_manage(staff, dept) and staff.get("role") != "rh",
        }

    @r.post("/staff/planning/week/cell")
    async def toggle_cell(payload: CellToggle, staff=Depends(get_current_staff)):
        await require_role(staff, MANAGER_ROLES)
        dept = await db.planning_departments.find_one({"id": payload.dept_id}, {"_id": 0})
        if not dept:
            raise HTTPException(404, "Département introuvable.")
        if not _user_can_manage(staff, dept) or staff.get("role") == "rh":
            raise HTTPException(403, "Vous ne pouvez pas éditer ce planning.")
        # Upsert cell value
        key = f"cells.{payload.employee_id}.{payload.day}"
        await db.planning_weeks.update_one(
            {"dept_id": payload.dept_id, "week_iso": payload.week_iso},
            {"$set": {
                key: payload.status,
                "dept_id": payload.dept_id,
                "week_iso": payload.week_iso,
                "updated_at": _now_iso(),
                "updated_by": staff.get("email"),
            }},
            upsert=True,
        )
        return {"ok": True}

    @r.post("/staff/planning/week/validate")
    async def validate_week(payload: dict, staff=Depends(get_current_staff)):
        await require_role(staff, MANAGER_ROLES)
        dept_id = payload.get("dept_id")
        week_iso = payload.get("week_iso")
        if not dept_id or not week_iso:
            raise HTTPException(400, "dept_id et week_iso requis.")
        dept = await db.planning_departments.find_one({"id": dept_id}, {"_id": 0})
        if not dept or not _user_can_manage(staff, dept):
            raise HTTPException(403, "Accès refusé.")
        await db.planning_weeks.update_one(
            {"dept_id": dept_id, "week_iso": week_iso},
            {"$set": {
                "validated_at": _now_iso(),
                "validated_by": staff.get("email"),
                "dept_id": dept_id, "week_iso": week_iso,
            }},
            upsert=True,
        )
        return {"ok": True}

    # ── HR overview ──
    @r.get("/staff/planning/hr/summary")
    async def hr_summary(week_iso: Optional[str] = None,
                        staff=Depends(get_current_staff)):
        await require_role(staff, HR_ROLES)
        await _seed_departments(db)
        cur_week = week_iso or _week_iso(datetime.now(timezone.utc).date())
        total_depts = await db.planning_departments.count_documents({})
        total_emps = await db.planning_employees.count_documents({})
        validated = await db.planning_weeks.count_documents(
            {"week_iso": cur_week, "validated_at": {"$ne": None}},
        )
        return {
            "current_week": cur_week,
            "total_departments": total_depts,
            "total_employees": total_emps,
            "validated_count": validated,
            "pending_count": max(0, total_depts - validated),
        }

    # ── HR: chef_dept account management (iter-47) ─────────────────────────
    def _slugify(name: str) -> str:
        import re
        import unicodedata
        n = unicodedata.normalize("NFKD", name).encode("ASCII", "ignore").decode()
        n = re.sub(r"[^a-zA-Z0-9]+", ".", n.lower()).strip(".")
        return n

    def _gen_password() -> str:
        import secrets
        import string
        alphabet = string.ascii_letters + string.digits
        return "".join(secrets.choice(alphabet) for _ in range(10))

    @r.get("/staff/planning/hr/chefs")
    async def list_chefs(staff=Depends(get_current_staff)):
        await require_role(staff, HR_ROLES)
        await _seed_departments(db)
        depts = await db.planning_departments.find({}, {"_id": 0}).sort("sort_order", 1).to_list(length=200)
        chefs = await db.staff.find(
            {"role": "chef_dept"},
            {"_id": 0, "id": 1, "name": 1, "email": 1, "dept_id": 1,
             "created_at": 1, "active": 1, "password_rotated_at": 1},
        ).to_list(length=200)
        by_dept = {c.get("dept_id"): c for c in chefs}
        out = []
        for d in depts:
            c = by_dept.get(d["id"])
            out.append({**d, "chef": c, "has_chef": bool(c)})
        return {"items": out, "orphan_chefs":
                [c for c in chefs if c.get("dept_id") not in {d["id"] for d in depts}]}

    @r.post("/staff/planning/hr/chefs/generate")
    async def generate_chef(dept_id: str, staff=Depends(get_current_staff)):
        await require_role(staff, HR_ROLES)
        if hash_password is None:
            raise HTTPException(500, "Password hasher not wired.")
        dept = await db.planning_departments.find_one({"id": dept_id}, {"_id": 0})
        if not dept:
            raise HTTPException(404, "Département introuvable.")
        existing = await db.staff.find_one({"role": "chef_dept", "dept_id": dept_id})
        if existing:
            raise HTTPException(409, "Ce département a déjà un chef.")
        base_slug = _slugify(dept["name"])
        email = f"chef.{base_slug}@boulay.ci"
        suffix = 0
        while await db.staff.find_one({"email": email}, {"_id": 0, "email": 1}):
            suffix += 1
            email = f"chef.{base_slug}{suffix}@boulay.ci"
        password = _gen_password()
        new_user = {
            "id": str(uuid.uuid4()),
            "name": f"Chef {dept['name']}",
            "email": email,
            "password_hash": hash_password(password),
            "role": "chef_dept",
            "dept_id": dept_id,
            "pole_id": None,
            "active": True,
            "created_at": _now_iso(),
            "created_by": staff.get("email"),
        }
        await db.staff.insert_one(new_user)
        await db.planning_departments.update_one(
            {"id": dept_id}, {"$set": {"manager_staff_id": new_user["id"]}},
        )
        return {
            "ok": True,
            "department": dept["name"],
            "email": email,
            "password": password,
            "user_id": new_user["id"],
            "warning": "Conservez ce mot de passe. Il ne sera plus jamais affiché en clair.",
        }

    @r.post("/staff/planning/hr/chefs/{user_id}/regenerate-password")
    async def regen_chef_password(user_id: str, staff=Depends(get_current_staff)):
        await require_role(staff, HR_ROLES)
        if hash_password is None:
            raise HTTPException(500, "Password hasher not wired.")
        u = await db.staff.find_one(
            {"id": user_id, "role": "chef_dept"}, {"_id": 0, "email": 1, "dept_id": 1},
        )
        if not u:
            raise HTTPException(404, "Chef introuvable.")
        password = _gen_password()
        await db.staff.update_one(
            {"id": user_id},
            {"$set": {"password_hash": hash_password(password),
                      "password_rotated_at": _now_iso(),
                      "password_rotated_by": staff.get("email")}},
        )
        return {"ok": True, "email": u["email"], "password": password,
                "warning": "Conservez ce mot de passe. Il ne sera plus jamais affiché en clair."}

    @r.delete("/staff/planning/hr/chefs/{user_id}")
    async def delete_chef(user_id: str, staff=Depends(get_current_staff)):
        await require_role(staff, HR_ROLES)
        u = await db.staff.find_one(
            {"id": user_id, "role": "chef_dept"}, {"_id": 0, "dept_id": 1, "email": 1},
        )
        if not u:
            raise HTTPException(404, "Chef introuvable.")
        await db.staff.delete_one({"id": user_id})
        if u.get("dept_id"):
            await db.planning_departments.update_one(
                {"id": u["dept_id"]}, {"$set": {"manager_staff_id": None}},
            )
        return {"ok": True, "deleted_email": u["email"]}


    # ── Exports ──
    @r.get("/staff/planning/exports/xlsx")
    async def export_xlsx(dept_id: str,
                          week_iso: str = Query(..., pattern=r"^\d{4}-W\d{2}$"),
                          staff=Depends(get_current_staff)):
        await require_role(staff, MANAGER_ROLES)
        dept = await db.planning_departments.find_one({"id": dept_id}, {"_id": 0})
        if not dept or not _user_can_manage(staff, dept):
            raise HTTPException(403, "Accès refusé.")
        employees = await db.planning_employees.find(
            {"dept_id": dept_id}, {"_id": 0},
        ).sort([("last_name", 1), ("first_name", 1)]).to_list(length=500)
        wk = await db.planning_weeks.find_one(
            {"dept_id": dept_id, "week_iso": week_iso}, {"_id": 0},
        ) or {"cells": {}}
        cells = wk.get("cells", {})

        wb = Workbook()
        ws = wb.active
        ws.title = "Planning"
        gold = PatternFill("solid", fgColor="0C0E12")
        green = PatternFill("solid", fgColor="86EFAC")
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="DCDEE2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.cell(row=1, column=1, value=f"Planning — {dept['name']}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Semaine {week_iso}")
        dates = _week_dates(week_iso)
        ws.cell(row=3, column=1, value=f"Du {dates[0].isoformat()} au {dates[6].isoformat()}")

        headers = ["Employé", "Poste"] + DAY_LABELS_FR
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=5, column=col, value=h)
            c.fill = gold
            c.font = Font(bold=True, color="D4B256")
            c.alignment = center
            c.border = border

        for row_idx, emp in enumerate(employees, start=6):
            ws.cell(row=row_idx, column=1, value=f"{emp['last_name']} {emp['first_name']}").border = border
            ws.cell(row=row_idx, column=2, value=emp.get("position", "")).border = border
            for col_idx, day in enumerate(DAYS, start=3):
                val = cells.get(emp["id"], {}).get(day, "T")
                cell = ws.cell(row=row_idx, column=col_idx,
                               value="Repos" if val == "R" else "Travail")
                cell.alignment = center
                cell.border = border
                if val == "R":
                    cell.fill = green

        for col_letter in "ABCDEFGHI":
            ws.column_dimensions[col_letter].width = 14
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"planning_{dept['name']}_{week_iso}.xlsx".replace(" ", "_")
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @r.get("/staff/planning/exports/pdf")
    async def export_pdf(dept_id: str,
                         week_iso: str = Query(..., pattern=r"^\d{4}-W\d{2}$"),
                         staff=Depends(get_current_staff)):
        await require_role(staff, MANAGER_ROLES)
        dept = await db.planning_departments.find_one({"id": dept_id}, {"_id": 0})
        if not dept or not _user_can_manage(staff, dept):
            raise HTTPException(403, "Accès refusé.")
        employees = await db.planning_employees.find(
            {"dept_id": dept_id}, {"_id": 0},
        ).sort([("last_name", 1), ("first_name", 1)]).to_list(length=500)
        wk = await db.planning_weeks.find_one(
            {"dept_id": dept_id, "week_iso": week_iso}, {"_id": 0},
        ) or {"cells": {}}
        cells = wk.get("cells", {})
        dates = _week_dates(week_iso)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Planning BBR",
                                leftMargin=20, rightMargin=20,
                                topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        elems = [
            Paragraph(f"<b>Planning des équipes — {dept['name']}</b>", styles["Title"]),
            Spacer(1, 6),
            Paragraph(f"Semaine du {dates[0].isoformat()} au {dates[6].isoformat()}", styles["Normal"]),
            Paragraph(f"Édité le {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC", styles["Normal"]),
        ]
        if wk.get("validated_at"):
            elems.append(Paragraph(f"Planning validé le {wk['validated_at'][:16]} par {wk.get('validated_by','—')}", styles["Italic"]))
        elems.append(Spacer(1, 12))

        header = ["Employé", "Poste"] + [f"{lbl}\n{d.strftime('%d/%m')}" for lbl, d in zip(DAY_LABELS_FR, dates)]
        data = [header]
        for emp in employees:
            row = [f"{emp['last_name']} {emp['first_name']}", emp.get("position", "")]
            for day in DAYS:
                v = cells.get(emp["id"], {}).get(day, "T")
                row.append("Repos" if v == "R" else "Travail")
            data.append(row)

        t = Table(data, repeatRows=1, colWidths=[110, 110, 60, 60, 60, 60, 60, 60, 60])
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0c0e12")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#d4b256")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dcdee2")),
        ])
        # Color "Repos" cells in light green
        for r_idx, emp in enumerate(employees, start=1):
            for c_idx, day in enumerate(DAYS, start=2):
                if cells.get(emp["id"], {}).get(day) == "R":
                    style.add("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), colors.HexColor("#86efac"))
        t.setStyle(style)
        elems.append(t)
        doc.build(elems)
        buf.seek(0)
        fname = f"planning_{dept['name']}_{week_iso}.pdf".replace(" ", "_")
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    return r
