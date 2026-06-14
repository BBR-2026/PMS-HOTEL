"""Custom Reports — staff-side configurable reports module.

Five report types:
- reservation         (db.bookings, filtered by created_at within period)
- embarquement        (flattened qr_codes.scans across bookings, filtered by scanned_at)
- traversee           (db.traversees, filtered by `date`)
- enregistrement      (db.registrations, filtered by created_at)
- chiffre_affaires    (db.bookings with paid_at, filtered by paid_at)

Each report exposes a fixed catalog of columns; the staff member picks a
subset and a date window, then chooses PDF or XLSX. PDF includes a small
synthetic bar-chart at the top (totals per day, top 5 categories, etc.).
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Literal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

# ----------------------------- Catalog -----------------------------
# Each entry: column_key -> (Human label FR, accessor lambda)
# The accessor receives a dict (booking / traversee / registration / scan) and
# must return a string-renderable value (or None).

COL_DATE = "date"                       # service date (date of activity)
COL_CLIENT_NAME = "client_name"
COL_CLIENT_PHONE = "client_phone"       # "Numéro client"
COL_ADULTS = "adults"
COL_KIDS = "kids"
COL_DATE_RESERVATION = "date_reservation"
COL_DATE_PAIEMENT = "date_paiement"
COL_PAYMENT_METHOD = "payment_method"
COL_DATE_HEURE_EMBARQUEMENT = "embarquement_at"
# Extra columns that make each report self-sufficient (not in the user's
# explicit ask, but the report would be unreadable without them):
COL_REF = "ref"
COL_OFFER = "offer"
COL_STATUS = "status"
COL_TOTAL = "total"
COL_BOAT = "boat"
COL_SKIPPER = "skipper"
COL_DEPART_TIME = "depart_time"
COL_PASSENGERS_COUNT = "passengers_count"
COL_DIRECTION = "direction"
COL_KIND = "kind"                       # registration kind
COL_NATIONALITY = "nationality"
COL_EMAIL = "email"

COL_LABELS_FR = {
    COL_DATE: "Date",
    COL_CLIENT_NAME: "Nom du client",
    COL_CLIENT_PHONE: "Numéro client",
    COL_ADULTS: "Nb adultes",
    COL_KIDS: "Nb enfants",
    COL_DATE_RESERVATION: "Date de réservation",
    COL_DATE_PAIEMENT: "Date de paiement",
    COL_PAYMENT_METHOD: "Moyen de paiement",
    COL_DATE_HEURE_EMBARQUEMENT: "Embarquement (date & heure)",
    COL_REF: "Référence",
    COL_OFFER: "Offre",
    COL_STATUS: "Statut",
    COL_TOTAL: "Montant",
    COL_BOAT: "Bateau",
    COL_SKIPPER: "Skipper",
    COL_DEPART_TIME: "Heure départ",
    COL_PASSENGERS_COUNT: "Passagers",
    COL_DIRECTION: "Sens",
    COL_KIND: "Statut visiteur",
    COL_NATIONALITY: "Nationalité",
    COL_EMAIL: "Email",
}

# Per-report-type whitelist of columns (kept in display order)
REPORT_COLUMNS = {
    "reservation": [
        COL_REF, COL_DATE_RESERVATION, COL_DATE, COL_CLIENT_NAME,
        COL_CLIENT_PHONE, COL_ADULTS, COL_KIDS, COL_OFFER, COL_STATUS,
        COL_TOTAL, COL_PAYMENT_METHOD, COL_DATE_PAIEMENT,
    ],
    "embarquement": [
        COL_DATE, COL_DATE_HEURE_EMBARQUEMENT, COL_CLIENT_NAME, COL_CLIENT_PHONE,
        COL_DIRECTION, COL_BOAT, COL_SKIPPER, COL_OFFER, COL_REF,
    ],
    "traversee": [
        COL_DATE, COL_DEPART_TIME, COL_BOAT, COL_SKIPPER, COL_PASSENGERS_COUNT, COL_STATUS,
    ],
    "enregistrement": [
        COL_DATE_RESERVATION, COL_CLIENT_NAME, COL_CLIENT_PHONE, COL_EMAIL,
        COL_KIND, COL_OFFER, COL_NATIONALITY,
    ],
    "chiffre_affaires": [
        COL_DATE_PAIEMENT, COL_REF, COL_CLIENT_NAME, COL_CLIENT_PHONE,
        COL_OFFER, COL_PAYMENT_METHOD, COL_TOTAL, COL_DATE_RESERVATION,
    ],
}

REPORT_TITLES_FR = {
    "reservation": "Rapport des réservations",
    "embarquement": "Rapport des embarquements",
    "traversee": "Rapport des traversées",
    "enregistrement": "Rapport des enregistrements",
    "chiffre_affaires": "Rapport de chiffre d'affaires",
}


# ----------------------------- Request body -----------------------------
class CustomReportRequest(BaseModel):
    report_type: Literal["reservation", "embarquement", "traversee", "enregistrement", "chiffre_affaires"]
    period: Literal["day", "week", "month", "custom"] = "month"
    date_from: Optional[str] = Field(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$")
    date_to: Optional[str] = Field(default=None, pattern=r"^\d{4}-\d{2}-\d{2}$")
    columns: List[str] = Field(min_length=1, max_length=15)
    format: Literal["pdf", "xlsx"] = "pdf"


# ----------------------------- Helpers -----------------------------
def _resolve_window(req: CustomReportRequest) -> tuple[datetime, datetime]:
    """Return [start, end) datetimes in UTC for the chosen period."""
    now = datetime.now(timezone.utc)
    if req.period == "day":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)
    elif req.period == "week":
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=7)
    elif req.period == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        # naive month-end: jump 32 days then back to first
        next_month = (start + timedelta(days=32)).replace(day=1)
        end = next_month
    else:  # custom
        if not req.date_from or not req.date_to:
            raise HTTPException(status_code=400, detail="Pour une période personnalisée, date_from et date_to sont requis.")
        start = datetime.fromisoformat(req.date_from).replace(tzinfo=timezone.utc)
        end = datetime.fromisoformat(req.date_to).replace(tzinfo=timezone.utc) + timedelta(days=1)
    return start, end


def _fmt_dt(iso: Optional[str], include_time: bool = True) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %Hh%M") if include_time else dt.strftime("%d/%m/%Y")
    except Exception:
        return iso[:16]


def _booker(b: dict) -> dict:
    """Extract the booker dict from a booking, supporting legacy shapes."""
    parts = b.get("participants") or []
    if parts:
        return parts[0]
    bk = b.get("booker") or {}
    return bk


def _client_name(b: dict) -> str:
    bk = _booker(b)
    return f"{bk.get('name','').strip()} {bk.get('surname','').strip()}".strip() or "—"


def _client_phone(b: dict) -> str:
    bk = _booker(b)
    return (bk.get("phone") or bk.get("whatsapp") or "").strip()


def _fmt_money(v) -> str:
    try:
        n = int(v or 0)
        return f"{n:,}".replace(",", " ") + " FCFA"
    except Exception:
        return str(v or "")


def _adults(b: dict) -> int:
    return int(b.get("adults") or b.get("pax_adults") or 0)


def _kids(b: dict) -> int:
    return int(b.get("children") or b.get("pax_kids") or 0)


# Accessor map: per-column extraction by record TYPE.
# We use different extractors per source so we can reuse columns across reports.

def _val_booking(col: str, b: dict) -> str:
    if col == COL_REF: return b.get("ref") or (b.get("id") or "")[:8].upper()
    if col == COL_DATE: return b.get("date") or ""
    if col == COL_CLIENT_NAME: return _client_name(b)
    if col == COL_CLIENT_PHONE: return _client_phone(b)
    if col == COL_ADULTS: return str(_adults(b))
    if col == COL_KIDS: return str(_kids(b))
    if col == COL_OFFER: return b.get("offer_label") or b.get("offer_type") or ""
    if col == COL_STATUS: return b.get("status") or ""
    if col == COL_TOTAL: return _fmt_money(b.get("total_amount") or b.get("total") or 0)
    if col == COL_DATE_RESERVATION: return _fmt_dt(b.get("created_at"))
    if col == COL_DATE_PAIEMENT: return _fmt_dt(b.get("paid_at"))
    if col == COL_PAYMENT_METHOD: return b.get("payment_method") or "—"
    return ""


def _val_scan(col: str, ctx: dict) -> str:
    # ctx = {booking, qr, scan}
    b = ctx["booking"]; sc = ctx["scan"]
    if col == COL_DATE: return sc.get("boat_date") or b.get("date") or ""
    if col == COL_DATE_HEURE_EMBARQUEMENT: return _fmt_dt(sc.get("scanned_at"))
    if col == COL_CLIENT_NAME: return _client_name(b)
    if col == COL_CLIENT_PHONE: return _client_phone(b)
    if col == COL_DIRECTION: return sc.get("direction") or ""
    if col == COL_BOAT: return sc.get("boat_name") or sc.get("boat_label") or sc.get("boat_time") or ""
    if col == COL_SKIPPER: return sc.get("skipper_name") or ""
    if col == COL_OFFER: return b.get("offer_label") or b.get("offer_type") or ""
    if col == COL_REF: return b.get("ref") or (b.get("id") or "")[:8].upper()
    return ""


def _val_traversee(col: str, t: dict) -> str:
    if col == COL_DATE: return t.get("date") or ""
    if col == COL_DEPART_TIME: return t.get("depart_time") or ""
    if col == COL_BOAT: return t.get("bateau_name") or t.get("boat_name") or ""
    if col == COL_SKIPPER: return t.get("skipper_name") or ""
    if col == COL_PASSENGERS_COUNT: return str(t.get("passengers_count") or t.get("pax_count") or 0)
    if col == COL_STATUS: return t.get("status") or ""
    return ""


def _val_registration(col: str, r: dict) -> str:
    if col == COL_DATE_RESERVATION: return _fmt_dt(r.get("created_at"))
    if col == COL_CLIENT_NAME: return f"{r.get('first_name','')} {r.get('last_name','')}".strip()
    if col == COL_CLIENT_PHONE: return r.get("phone") or ""
    if col == COL_EMAIL: return r.get("email") or ""
    if col == COL_KIND:
        labels = {"client": "Client", "personnel": "Personnel", "prestataire": "Prestataire", "invite": "Invité"}
        return labels.get(r.get("kind") or "client", "Client")
    if col == COL_OFFER: return r.get("offer_label") or ""
    if col == COL_NATIONALITY: return r.get("nationality") or ""
    return ""


# ----------------------------- Data fetchers -----------------------------
async def _fetch_reservation(db, start, end) -> list[dict]:
    cursor = db.bookings.find(
        {"created_at": {"$gte": start.isoformat(), "$lt": end.isoformat()}},
        {"_id": 0, "qr_codes": 0, "wallet_history": 0},
    ).sort("created_at", -1)
    return [b async for b in cursor]


async def _fetch_embarquement(db, start, end) -> list[dict]:
    """Flatten qr_codes.scans whose scanned_at falls in [start, end)."""
    rows = []
    cursor = db.bookings.find(
        {"qr_codes.scans.0": {"$exists": True}},
        {"_id": 0, "id": 1, "ref": 1, "date": 1, "offer_label": 1, "offer_type": 1,
         "participants": 1, "booker": 1, "qr_codes": 1},
    )
    start_iso = start.isoformat()
    end_iso = end.isoformat()
    async for b in cursor:
        for qr in b.get("qr_codes") or []:
            for sc in qr.get("scans") or []:
                ts = sc.get("scanned_at") or ""
                if start_iso <= ts < end_iso:
                    rows.append({"booking": b, "qr": qr, "scan": sc, "_sort": ts})
    rows.sort(key=lambda r: r["_sort"], reverse=True)
    return rows


async def _fetch_traversee(db, start, end) -> list[dict]:
    start_d = start.date().isoformat()
    end_d = end.date().isoformat()  # exclusive
    cursor = db.traversees.find(
        {"date": {"$gte": start_d, "$lt": end_d}},
        {"_id": 0},
    ).sort([("date", -1), ("depart_time", 1)])
    return [t async for t in cursor]


async def _fetch_enregistrement(db, start, end) -> list[dict]:
    cursor = db.registrations.find(
        {"created_at": {"$gte": start.isoformat(), "$lt": end.isoformat()}},
        {"_id": 0, "pass_token": 0},
    ).sort("created_at", -1)
    return [r async for r in cursor]


async def _fetch_chiffre_affaires(db, start, end) -> list[dict]:
    cursor = db.bookings.find(
        {
            "paid_at": {"$gte": start.isoformat(), "$lt": end.isoformat()},
            "status": {"$in": ["confirmed", "completed", "arrived"]},
        },
        {"_id": 0, "qr_codes": 0, "wallet_history": 0},
    ).sort("paid_at", -1)
    return [b async for b in cursor]


# ----------------------------- Synthetic stats -----------------------------
def _synthesise(report_type: str, rows: list[dict]) -> dict:
    """Build a small chart-friendly summary (label -> int)."""
    counter: dict[str, int] = {}
    if report_type == "reservation":
        for b in rows:
            label = (b.get("offer_label") or b.get("offer_type") or "—")[:18]
            counter[label] = counter.get(label, 0) + 1
        title = "Réservations par offre"
    elif report_type == "embarquement":
        for ctx in rows:
            d = (ctx["scan"].get("scan_date") or ctx["scan"].get("boat_date") or "")[:10]
            counter[d] = counter.get(d, 0) + 1
        title = "Embarquements par jour"
    elif report_type == "traversee":
        for t in rows:
            counter[t.get("status") or "—"] = counter.get(t.get("status") or "—", 0) + 1
        title = "Traversées par statut"
    elif report_type == "enregistrement":
        labels = {"client": "Client", "personnel": "Personnel", "prestataire": "Prestataire", "invite": "Invité"}
        for r in rows:
            k = labels.get(r.get("kind") or "client", "Client")
            counter[k] = counter.get(k, 0) + 1
        title = "Enregistrements par statut"
    else:  # chiffre_affaires
        for b in rows:
            d = (b.get("paid_at") or "")[:10]
            if d:
                counter[d] = counter.get(d, 0) + int(b.get("total_amount") or b.get("total") or 0)
        title = "Chiffre d'affaires par jour (FCFA)"
    # Keep top 8 entries for readability
    items = sorted(counter.items(), key=lambda kv: kv[0])[:14] if report_type in ("embarquement", "chiffre_affaires") \
            else sorted(counter.items(), key=lambda kv: -kv[1])[:8]
    return {"title": title, "items": items, "total": sum(counter.values())}


# ----------------------------- PDF builder -----------------------------
def _build_pdf(report_type: str, rows: list[dict], columns: list[str], window: tuple[datetime, datetime]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.shapes import Drawing, String

    gold = colors.HexColor("#B8922A")
    dark = colors.HexColor("#0A0A0A")
    muted = colors.HexColor("#6B7280")
    light = colors.HexColor("#FAF7F2")

    buf = io.BytesIO()
    is_wide = len(columns) > 6
    pagesize = landscape(A4) if is_wide else A4
    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title=REPORT_TITLES_FR[report_type], author="Boulay Beach Resort",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=dark, fontSize=18, leading=22, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], textColor=muted, fontSize=10, leading=13, spaceAfter=8)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=gold, fontSize=11, leading=14, spaceBefore=6, spaceAfter=4)

    elements: list = []
    elements.append(Paragraph(REPORT_TITLES_FR[report_type], h1))
    period_label = f"Du {window[0].strftime('%d/%m/%Y')} au {(window[1] - timedelta(seconds=1)).strftime('%d/%m/%Y')}"
    elements.append(Paragraph(f"{period_label} — {len(rows)} ligne(s)", sub))

    # ---------- Synthetic bar chart ----------
    syn = _synthesise(report_type, rows)
    if syn["items"]:
        elements.append(Paragraph(syn["title"], h2))
        chart_w = (26 if is_wide else 17) * cm
        chart_h = 5.5 * cm
        drawing = Drawing(chart_w, chart_h + 0.8 * cm)
        chart = VerticalBarChart()
        chart.x = 30
        chart.y = 18
        chart.width = chart_w - 50
        chart.height = chart_h - 10
        chart.data = [[v for _, v in syn["items"]]]
        chart.categoryAxis.categoryNames = [k[:14] for k, _ in syn["items"]]
        chart.categoryAxis.labels.fontSize = 7
        chart.categoryAxis.labels.angle = 25
        chart.categoryAxis.labels.boxAnchor = "ne"
        chart.categoryAxis.labels.dx = -3
        chart.categoryAxis.labels.dy = -2
        chart.valueAxis.valueMin = 0
        chart.valueAxis.labels.fontSize = 7
        chart.bars[0].fillColor = gold
        chart.bars[0].strokeColor = gold
        chart.barWidth = 8
        drawing.add(chart)
        elements.append(drawing)
        elements.append(Spacer(1, 0.3 * cm))

    # ---------- Data table ----------
    elements.append(Paragraph("Détails", h2))
    header = [COL_LABELS_FR.get(c, c) for c in columns]
    body_rows = []
    for rec in rows:
        if report_type == "reservation" or report_type == "chiffre_affaires":
            body_rows.append([_val_booking(c, rec) for c in columns])
        elif report_type == "embarquement":
            body_rows.append([_val_scan(c, rec) for c in columns])
        elif report_type == "traversee":
            body_rows.append([_val_traversee(c, rec) for c in columns])
        elif report_type == "enregistrement":
            body_rows.append([_val_registration(c, rec) for c in columns])

    if not body_rows:
        elements.append(Paragraph("<i>Aucune donnée sur la période sélectionnée.</i>",
                                  ParagraphStyle("empty", parent=styles["Normal"], textColor=muted)))
    else:
        tbl = Table([header, *body_rows], repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), dark),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light]),
            ("BOX", (0, 0), (-1, -1), 0.4, gold),
            ("LINEBELOW", (0, 0), (-1, 0), 0.6, gold),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(tbl)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(muted)
        canvas.drawString(1.5 * cm, 0.8 * cm, "Boulay Beach Resort — Rapport généré le " + datetime.now(timezone.utc).strftime("%d/%m/%Y à %Hh%M UTC"))
        canvas.drawRightString(pagesize[0] - 1.5 * cm, 0.8 * cm, f"Page {doc_.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


# ----------------------------- XLSX builder -----------------------------
def _build_xlsx(report_type: str, rows: list[dict], columns: list[str], window: tuple[datetime, datetime]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = report_type[:30]

    title = REPORT_TITLES_FR[report_type]
    period_label = f"Du {window[0].strftime('%d/%m/%Y')} au {(window[1] - timedelta(seconds=1)).strftime('%d/%m/%Y')}"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14, color="0A0A0A")
    ws.append([period_label])
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.append([])

    header = [COL_LABELS_FR.get(c, c) for c in columns]
    ws.append(header)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for rec in rows:
        if report_type in ("reservation", "chiffre_affaires"):
            ws.append([_val_booking(c, rec) for c in columns])
        elif report_type == "embarquement":
            ws.append([_val_scan(c, rec) for c in columns])
        elif report_type == "traversee":
            ws.append([_val_traversee(c, rec) for c in columns])
        elif report_type == "enregistrement":
            ws.append([_val_registration(c, rec) for c in columns])

    # Auto-size columns (heuristic: 1.2 × max length, capped at 40)
    for idx, _ in enumerate(columns, start=1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        max_len = max(
            (len(str(ws.cell(row=r, column=idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max(12, int(max_len * 1.2)), 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------- Router -----------------------------
def build_router(*, db, get_current_staff, require_role):
    router = APIRouter()

    @router.get("/staff/reports/custom/schema")
    async def custom_report_schema(staff=Depends(get_current_staff)):
        """Return the catalog of report types and columns, for the UI to render."""
        await require_role(staff, [
            "admin", "manager", "manager_pole", "management_general",
            "hotesse", "receptionist", "logistique", "verification", "serveur_caisse",
        ])
        return {
            "reports": [
                {
                    "key": k,
                    "label": REPORT_TITLES_FR[k],
                    "columns": [
                        {"key": c, "label": COL_LABELS_FR[c]} for c in cols
                    ],
                }
                for k, cols in REPORT_COLUMNS.items()
            ],
        }

    @router.post("/staff/reports/custom")
    async def custom_report(body: CustomReportRequest, staff=Depends(get_current_staff)):
        await require_role(staff, [
            "admin", "manager", "manager_pole", "management_general",
            "hotesse", "receptionist", "logistique", "verification", "serveur_caisse",
        ])
        # Validate columns against catalog
        allowed = set(REPORT_COLUMNS[body.report_type])
        bad = [c for c in body.columns if c not in allowed]
        if bad:
            raise HTTPException(
                status_code=400,
                detail=f"Colonnes invalides pour ce type de rapport : {', '.join(bad)}",
            )
        # Preserve user order, drop dupes
        seen = set()
        ordered_cols = [c for c in body.columns if not (c in seen or seen.add(c))]

        window = _resolve_window(body)

        fetcher = {
            "reservation": _fetch_reservation,
            "embarquement": _fetch_embarquement,
            "traversee": _fetch_traversee,
            "enregistrement": _fetch_enregistrement,
            "chiffre_affaires": _fetch_chiffre_affaires,
        }[body.report_type]
        rows = await fetcher(db, *window)

        if body.format == "pdf":
            data = _build_pdf(body.report_type, rows, ordered_cols, window)
            mime = "application/pdf"
            ext = "pdf"
        else:
            data = _build_xlsx(body.report_type, rows, ordered_cols, window)
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ext = "xlsx"

        ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
        filename = f"BBR-{body.report_type}-{ts}.{ext}"
        return Response(
            content=data,
            media_type=mime,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    return router
