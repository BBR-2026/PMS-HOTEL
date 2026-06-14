"""Passenger registry (iter-34) — central passenger view aggregating every
person who appears in a confirmed booking, regardless of how they were
registered (online payment, cash settlement, companion link, manual accueil).

Each row corresponds to ONE physical passenger:
- The booker (always counted as "enregistré" once the booking is paid).
- Each companion that registered themselves via /companion/{code}.
- Each visitor that walked-in at the accueil (db.registrations).

Aggregation status pipeline:
- ``en_attente``  : the booking is confirmed but this companion slot is still
                    waiting for the passenger to register themselves.
- ``enregistré`` : the passenger has a QR ticket attached to the booking.
- ``embarqué``   : the QR ticket has at least 1 scan (aller).
- ``finalisé``   : the QR ticket has scans for every required leg.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timedelta, timezone
from typing import Optional, Literal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response

logger = logging.getLogger(__name__)


STATUS_LABELS = {
    "en_attente":  "En attente",
    "enregistre":  "Enregistré",
    "embarque":    "Embarqué",
    "finalise":    "Finalisé",
}


def _pole_label(p: Optional[str], poles_catalog: dict) -> str:
    if not p:
        return "—"
    meta = (poles_catalog or {}).get(p)
    return (meta.get("name_fr") if isinstance(meta, dict) else None) or p.title()


def _passenger_status(qr: Optional[dict], booking_status: str) -> str:
    """Resolve the unified status for one passenger row."""
    if not qr:
        # No QR yet → this is a still-open companion slot.
        return "en_attente"
    scans = qr.get("scans") or []
    if not scans:
        return "enregistre"
    # Final iff every required leg has been scanned. For passport (multi-day)
    # tickets `_qr_completed` requires 2 scans per valid date.
    vd = qr.get("valid_dates") or []
    if len(vd) >= 2:
        from collections import Counter
        per_day = Counter(s.get("scan_date") or s.get("boat_date") for s in scans)
        if all(per_day.get(d, 0) >= 2 for d in vd):
            return "finalise"
    elif len(scans) >= 2:
        return "finalise"
    return "embarque"


async def _enumerate_passengers(db, filt: dict, poles_catalog: dict) -> list[dict]:
    """Walk through bookings + registrations and build flat passenger rows.

    `filt` already contains the date / status / search filters that apply to
    the SQL-like layer. We post-filter on status (which is derived).
    """
    rows: list[dict] = []
    booking_filter = filt.get("booking", {})
    cursor = db.bookings.find(
        booking_filter,
        {
            "_id": 0, "id": 1, "status": 1, "offer_type": 1, "offer_name": 1,
            "pole": 1, "date": 1, "boat_time": 1, "created_at": 1, "paid_at": 1,
            "total_amount": 1, "booking_code": 1, "adults": 1, "children": 1,
            "children_paid": 1, "children_free": 1, "participants": 1,
            "qr_codes": 1, "companion_slots_total": 1, "companion_slots_used": 1,
            "name": 1, "surname": 1, "email": 1, "phone": 1,
        },
    )
    async for b in cursor:
        if b.get("status") not in ("confirmed", "arrived", "completed", "pending_cash_payment"):
            continue
        # Match bookings registered QR codes to participants by their index.
        # `participants` is ordered: [booker, adult_2, …]. `qr_codes` is
        # generated only for adults (iter-30). So qr[i] corresponds to
        # participants[i] for i < len(qr_codes).
        qrs = b.get("qr_codes") or []
        participants = b.get("participants") or []
        adult_parts = [p for p in participants if (p.get("kind") or "adult") == "adult"]
        n_adults_total = int(b.get("adults") or len(adult_parts) or 1)
        pole_lbl = _pole_label(b.get("pole") or "", poles_catalog)
        # Build one row per adult — registered ones come first (in QR order),
        # then placeholder rows for still-open companion slots.
        for idx in range(n_adults_total):
            qr = qrs[idx] if idx < len(qrs) else None
            participant = adult_parts[idx] if idx < len(adult_parts) else None
            scans = (qr or {}).get("scans") or []
            first_scan = scans[0] if scans else None
            row = {
                "source": "booking",
                "booking_id": b["id"],
                "booking_ref": b["id"][:8].upper(),
                "booking_code": b.get("booking_code"),
                "category": pole_lbl,
                "offer_type": b.get("offer_type"),
                "offer_label": b.get("offer_name") or b.get("offer_type") or "",
                "first_name": (participant or {}).get("name") or b.get("name") or "",
                "last_name":  (participant or {}).get("surname") or b.get("surname") or "",
                "email":      (participant or {}).get("email") or b.get("email") or "",
                "phone":      (participant or {}).get("phone") or b.get("phone") or "",
                "booking_date": b.get("date"),
                "boat_time":   b.get("boat_time"),
                "created_at":  b.get("created_at"),
                "registered_at": (qr or {}).get("companion_added_at") or b.get("paid_at") if qr else None,
                "registration_status": _passenger_status(qr, b.get("status", "")),
                "boarding_scanned_at": (first_scan or {}).get("scanned_at") if first_scan else None,
                "boarding_boat_name": (first_scan or {}).get("boat_name") if first_scan else None,
                "boarding_boat_time": (first_scan or {}).get("boat_time") if first_scan else None,
                "is_booker": idx == 0,
                "adults": int(b.get("adults") or 0),
                "children": int(b.get("children") or 0),
                "children_paid": int(b.get("children_paid") or 0),
                "children_free": int(b.get("children_free") or 0),
                "total_amount": int(b.get("total_amount") or 0),
            }
            # Children attached only to the booker's row
            if idx != 0:
                row["children_paid"] = 0
                row["children_free"] = 0
                row["children"] = 0
            rows.append(row)

    # Walk-in registrations from the public /accueil/enregistrement form.
    reg_cursor = db.registrations.find(filt.get("registration", {}), {"_id": 0})
    async for r in reg_cursor:
        rows.append({
            "source": "walk_in",
            "booking_id": r.get("id"),
            "booking_ref": (r.get("id") or "")[:8].upper(),
            "booking_code": None,
            "category": (r.get("kind") or "client").capitalize(),
            "offer_type": "registration",
            "offer_label": r.get("offer_label") or "—",
            "first_name": r.get("first_name") or "",
            "last_name":  r.get("last_name") or "",
            "email":      r.get("email") or "",
            "phone":      r.get("phone") or "",
            "booking_date": (r.get("created_at") or "")[:10],
            "boat_time": None,
            "created_at": r.get("created_at"),
            "registered_at": r.get("created_at"),
            "registration_status": "enregistre",
            "boarding_scanned_at": None,
            "boarding_boat_name": None,
            "boarding_boat_time": None,
            "is_booker": True,
            "adults": 1, "children": 0, "children_paid": 0, "children_free": 0,
            "total_amount": 0,
        })

    rows.sort(key=lambda r: r.get("created_at") or "", reverse=True)
    return rows


def _resolve_period(period: Optional[str], date_from: Optional[str],
                    date_to: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    """Return [from_iso, to_iso) bounds (UTC) for the chosen period."""
    if date_from and date_to:
        return f"{date_from}T00:00:00", f"{date_to}T23:59:59.999999"
    now = datetime.now(timezone.utc)
    if period == "today":
        s = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return s.isoformat(), (s + timedelta(days=1)).isoformat()
    if period == "yesterday":
        s = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        return s.isoformat(), (s + timedelta(days=1)).isoformat()
    if period == "week":
        s = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        return s.isoformat(), (s + timedelta(days=7)).isoformat()
    return None, None


def build_router(*, db, get_current_staff, require_role, poles_catalog: dict):
    router = APIRouter()

    @router.get("/staff/passengers")
    async def list_passengers(
        q: Optional[str] = None,
        status: Optional[Literal["en_attente", "enregistre", "embarque", "finalise", "all"]] = "all",
        period: Optional[Literal["today", "yesterday", "week", "all"]] = "all",
        date_from: Optional[str] = Query(None, regex=r"^\d{4}-\d{2}-\d{2}$"),
        date_to: Optional[str] = Query(None, regex=r"^\d{4}-\d{2}-\d{2}$"),
        pole: Optional[str] = None,
        staff=Depends(get_current_staff),
    ):
        await require_role(staff, [
            "admin", "manager", "manager_pole", "management_general",
            "hotesse", "receptionist", "logistique", "verification",
        ])
        bf: dict = {}
        rf: dict = {}
        f_, t_ = _resolve_period(period, date_from, date_to)
        if f_ and t_:
            bf["created_at"] = {"$gte": f_, "$lt": t_}
            rf["created_at"] = {"$gte": f_, "$lt": t_}
        if pole:
            bf["pole"] = pole
        rows = await _enumerate_passengers(db, {"booking": bf, "registration": rf}, poles_catalog)
        if status and status != "all":
            rows = [r for r in rows if r["registration_status"] == status]
        if q:
            ql = q.strip().lower()
            rows = [
                r for r in rows
                if ql in (r["first_name"] or "").lower()
                or ql in (r["last_name"] or "").lower()
                or ql in (r["email"] or "").lower()
                or ql in (r["phone"] or "").lower()
                or ql in (r["booking_ref"] or "").lower()
                or ql in (r.get("booking_code") or "").lower()
            ]
        return {
            "total": len(rows),
            "items": rows,
            "summary": {
                "en_attente": sum(1 for r in rows if r["registration_status"] == "en_attente"),
                "enregistre": sum(1 for r in rows if r["registration_status"] == "enregistre"),
                "embarque":   sum(1 for r in rows if r["registration_status"] == "embarque"),
                "finalise":   sum(1 for r in rows if r["registration_status"] == "finalise"),
            },
        }

    @router.get("/staff/passengers/export.xlsx")
    async def export_passengers_xlsx(
        status: Optional[str] = "finalise",
        period: Optional[str] = "today",
        date_from: Optional[str] = Query(None, regex=r"^\d{4}-\d{2}-\d{2}$"),
        date_to: Optional[str] = Query(None, regex=r"^\d{4}-\d{2}-\d{2}$"),
        pole: Optional[str] = None,
        staff=Depends(get_current_staff),
    ):
        await require_role(staff, [
            "admin", "manager", "manager_pole", "management_general",
            "hotesse", "receptionist", "logistique", "verification",
        ])
        bf: dict = {}
        rf: dict = {}
        f_, t_ = _resolve_period(period, date_from, date_to)
        if f_ and t_:
            bf["created_at"] = {"$gte": f_, "$lt": t_}
            rf["created_at"] = {"$gte": f_, "$lt": t_}
        if pole:
            bf["pole"] = pole
        rows = await _enumerate_passengers(db, {"booking": bf, "registration": rf}, poles_catalog)
        if status and status != "all":
            rows = [r for r in rows if r["registration_status"] == status]

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Passagers"
        ws.append([f"Liste des passagers · statut={STATUS_LABELS.get(status or 'all', status)}"])
        ws["A1"].font = Font(bold=True, size=14, color="0A0A0A")
        ws.append([f"Période : {period or '—'} · Total : {len(rows)} passagers"])
        ws["A2"].font = Font(italic=True, color="6B7280")
        ws.append([])
        headers = [
            "Nom", "Prénom", "Référence", "Catégorie", "Offre",
            "Date réservation", "Date embarquement", "Heure scan",
            "Adultes", "Enfants 6-12", "Enfants <6", "Montant (FCFA)",
        ]
        ws.append(headers)
        hr = ws.max_row
        for c in ws[hr]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
            c.alignment = Alignment(horizontal="left", vertical="center")
        for r in rows:
            scan_at = r.get("boarding_scanned_at") or ""
            ws.append([
                r["last_name"], r["first_name"], r["booking_ref"], r["category"],
                r["offer_label"],
                (r.get("booking_date") or "")[:10],
                scan_at[:10] if scan_at else "",
                scan_at[11:16] if scan_at else "",
                r.get("adults", 0), r.get("children_paid", 0), r.get("children_free", 0),
                r.get("total_amount", 0),
            ])
        for i, w in enumerate([18, 16, 12, 16, 22, 14, 14, 10, 9, 12, 12, 14], start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        buf = io.BytesIO()
        wb.save(buf)
        ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="BBR-passagers-{status}-{ts}.xlsx"'},
        )

    return router
